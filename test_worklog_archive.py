"""월별 일지(…/일지/2026/9월.xlsx) 시트 저장·삭제·연월 자동생성."""
from __future__ import annotations

import os
import tempfile
import unittest
from datetime import date
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


class _FakeSS(dict):
    def pop(self, key, default=None):
        return dict.pop(self, key, default)


def _write_day_xlsx(path: str, label: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws["C5"] = label
    wb.save(path)
    wb.close()


class WorklogMonthArchiveTest(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.root = os.path.join(self._tmp.name, "Desktop", "업무", "일지")
        os.makedirs(self.root, exist_ok=True)
        self.ss = _FakeSS()
        import worklog_tab as wt

        self.wt = wt
        self._p_root = patch.object(wt, "resolve_worklog_archive_root", return_value=self.root)
        self._p_ss = patch.object(wt.st, "session_state", self.ss)
        self._p_root.start()
        self._p_ss.start()

    def tearDown(self):
        self._p_ss.stop()
        self._p_root.stop()
        self._tmp.cleanup()

    def test_save_creates_year_and_month_file_with_day_sheet(self):
        d = date(2026, 9, 3)
        day_xlsx = os.path.join(self._tmp.name, "2026-09-03.xlsx")
        _write_day_xlsx(day_xlsx, "day-3")

        year_dir = os.path.join(self.root, "2026")
        month_path = os.path.join(year_dir, "9월.xlsx")
        self.assertFalse(os.path.isdir(year_dir))
        self.assertFalse(os.path.exists(month_path))

        got = self.wt.upsert_worklog_archive_sheet(d, day_xlsx)
        self.assertEqual(got, month_path)
        self.assertTrue(os.path.isdir(year_dir))
        self.assertTrue(os.path.isfile(month_path))

        wb = load_workbook(month_path, read_only=True)
        try:
            self.assertIn("3", wb.sheetnames)
            self.assertEqual(wb["3"]["C5"].value, "day-3")
        finally:
            wb.close()

    def test_second_day_adds_another_sheet_same_month_file(self):
        d3 = date(2026, 9, 3)
        d5 = date(2026, 9, 5)
        p3 = os.path.join(self._tmp.name, "2026-09-03.xlsx")
        p5 = os.path.join(self._tmp.name, "2026-09-05.xlsx")
        _write_day_xlsx(p3, "day-3")
        _write_day_xlsx(p5, "day-5")
        self.wt.upsert_worklog_archive_sheet(d3, p3)
        self.wt.upsert_worklog_archive_sheet(d5, p5)

        month_path = os.path.join(self.root, "2026", "9월.xlsx")
        wb = load_workbook(month_path, read_only=True)
        try:
            self.assertEqual(sorted(wb.sheetnames, key=lambda n: int(n) if n.isdigit() else n), ["3", "5"])
        finally:
            wb.close()

    def test_delete_removes_sheet_keeps_other_days(self):
        d3 = date(2026, 9, 3)
        d5 = date(2026, 9, 5)
        p3 = os.path.join(self._tmp.name, "2026-09-03.xlsx")
        p5 = os.path.join(self._tmp.name, "2026-09-05.xlsx")
        _write_day_xlsx(p3, "day-3")
        _write_day_xlsx(p5, "day-5")
        self.wt.upsert_worklog_archive_sheet(d3, p3)
        self.wt.upsert_worklog_archive_sheet(d5, p5)

        removed = self.wt.delete_worklog_archive_sheet(d3)
        month_path = os.path.join(self.root, "2026", "9월.xlsx")
        self.assertEqual(removed, month_path)
        self.assertTrue(os.path.isfile(month_path))
        wb = load_workbook(month_path, read_only=True)
        try:
            self.assertNotIn("3", wb.sheetnames)
            self.assertIn("5", wb.sheetnames)
        finally:
            wb.close()

    def test_delete_last_sheet_removes_month_file(self):
        d = date(2026, 9, 3)
        p = os.path.join(self._tmp.name, "2026-09-03.xlsx")
        _write_day_xlsx(p, "day-3")
        self.wt.upsert_worklog_archive_sheet(d, p)
        month_path = os.path.join(self.root, "2026", "9월.xlsx")
        self.assertTrue(os.path.isfile(month_path))
        self.wt.delete_worklog_archive_sheet(d)
        self.assertFalse(os.path.isfile(month_path))

    def test_new_month_creates_separate_file(self):
        d9 = date(2026, 9, 30)
        d10 = date(2026, 10, 1)
        p9 = os.path.join(self._tmp.name, "2026-09-30.xlsx")
        p10 = os.path.join(self._tmp.name, "2026-10-01.xlsx")
        _write_day_xlsx(p9, "sep")
        _write_day_xlsx(p10, "oct")
        self.wt.upsert_worklog_archive_sheet(d9, p9)
        self.wt.upsert_worklog_archive_sheet(d10, p10)
        self.assertTrue(os.path.isfile(os.path.join(self.root, "2026", "9월.xlsx")))
        self.assertTrue(os.path.isfile(os.path.join(self.root, "2026", "10월.xlsx")))

    def test_exists_lookup_does_not_create_year_folder(self):
        d = date(2027, 1, 1)
        self.assertFalse(self.wt.worklog_date_exists_in_archive(d))
        self.assertFalse(os.path.isdir(os.path.join(self.root, "2027")))

    def test_create_worklog_day_local_makes_sheet_without_overwrite(self):
        d = date(2026, 9, 4)
        tpl = os.path.join(self._tmp.name, "template.xlsx")
        cache = os.path.join(self._tmp.name, "cache")
        os.makedirs(cache, exist_ok=True)
        _write_day_xlsx(tpl, "tpl")
        with patch.object(self.wt, "WORKLOG_TEMPLATE", tpl), patch.object(self.wt, "WORKLOG_DIR", cache):
            info = self.wt.create_worklog_day_local(d)
            self.assertTrue(info["created"])
            month_path = os.path.join(self.root, "2026", "9월.xlsx")
            self.assertTrue(os.path.isfile(month_path))
            self.assertTrue(os.path.isfile(os.path.join(cache, "2026-09-04.xlsx")))
            again = self.wt.create_worklog_day_local(d)
            self.assertFalse(again["created"])
        wb = load_workbook(os.path.join(self.root, "2026", "9월.xlsx"), read_only=True)
        try:
            self.assertIn("4", wb.sheetnames)
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
