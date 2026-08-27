"""일일업무일지 탭 — 엑셀 양식 그대로 표시/편집/날짜별 저장/출력."""
from __future__ import annotations

import calendar
import html
import io
import json
import os
import platform
import re
import shutil
import subprocess
import time
from datetime import date
from functools import lru_cache
from typing import Any

import streamlit as st
import streamlit.components.v1 as components
from streamlit.errors import StreamlitAPIException

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    load_workbook = None
    get_column_letter = None


def _wl_rerun(*, full: bool = False) -> None:
    """업무일지 fragment 안이면 fragment만 다시 실행 (전체 앱 로딩 방지).

    full=True 는 저장·날짜변경처럼 왼쪽 요약도 같이 갱신해야 할 때 사용.
    """
    if full:
        st.rerun()
        return
    try:
        st.rerun(scope="fragment")
    except (StreamlitAPIException, RuntimeError):
        st.rerun()


def _wl_quiet_ui() -> bool:
    try:
        if st.session_state.get("force_touch_ui") is True: return True
        v = st.query_params.get("touch_ui", "")
        if isinstance(v, (list, tuple)): v = v[0] if v else ""
        if str(v).strip() in ("1", "true", "True"): return True
    except Exception: pass
    try: return platform.system() != "Darwin"
    except Exception: return True


def _wl_is_streamlit_cloud() -> bool:
    try:
        env = (os.environ.get("STREAMLIT_RUNTIME_ENVIRONMENT") or "").strip().lower()
        if env == "cloud":
            return True
    except Exception:
        pass
    try:
        if os.path.abspath(os.getcwd()).startswith("/mount/src"):
            return True
    except Exception:
        pass
    return False


def _wl_is_ipad_ui() -> bool:
    """iPad/touch UI — app.is_touch_ui()와 동일 기준. 맥·데스크톱 브라우저는 False."""
    try:
        if st.session_state.get("force_touch_ui") is True:
            return True
        v = st.query_params.get("touch_ui", "")
        if isinstance(v, (list, tuple)):
            v = v[0] if v else ""
        if str(v).strip() in ("1", "true", "True"):
            st.session_state["force_touch_ui"] = True
            return True
    except Exception:
        pass
    try:
        cookies = getattr(st.context, "cookies", None)
        if cookies is not None and str(cookies.get("dashboard_touch", "")) == "1":
            st.session_state["force_touch_ui"] = True
            return True
    except Exception:
        pass
    try:
        headers = getattr(st.context, "headers", None)
        ua = ""
        if headers is not None:
            ua = str(headers.get("User-Agent") or headers.get("user-agent") or "")
        if ua and (
            re.search(r"iPad|iPhone|iPod", ua)
            or ("Macintosh" in ua and "Mobile" in ua)
        ):
            st.session_state["force_touch_ui"] = True
            return True
    except Exception:
        pass
    return bool(st.session_state.get("force_touch_ui"))


def _invalidate_saved_dates_cache() -> None:
    st.session_state.pop("wl_saved_dates_cache", None)

WORKLOG_DIR = os.path.join("uploaded_cache", "worklog")
WORKLOG_TEMPLATE = os.path.join(WORKLOG_DIR, "template.xlsx")
WORKLOG_TEMPLATE_SRC = os.path.expanduser("~/Desktop/업무일지.xlsx")
WORKLOG_ARCHIVE_REL = os.path.join("Desktop", "업무", "일지")

# =====================================================================
# 💡 원본.xlsx 양식 기준 행/열 매핑 (인쇄·미리보기 일치)
# 제목 E2 / 날짜 C5 / 내용헤더 G7 / 본문 8~39 / 익일 40~43 / 특이 44~47
# =====================================================================
WL_MIN_ROW, WL_MAX_ROW = 1, 47  # 본문·익일·특이까지(원본 로고는 47행 아래 여백)
WL_MIN_COL, WL_MAX_COL = 3, 28  # C ~ AB

WL_DATE_CELL = "C5"
WL_CLIENT_ROWS = list(range(8, 40))   # 8행 ~ 39행 (총 32줄)
WL_CONTENT_ROWS = list(range(8, 40))  # 8행 ~ 39행 (총 32줄)
WL_NEXT_ROWS = list(range(40, 44))    # 40행 ~ 43행 (총 4줄)
WL_NOTE_ROWS = list(range(44, 48))    # 44행 ~ 47행 (총 4줄)

WL_CONTENT_COL_START = 7  # G
WL_CONTENT_COL_END = 24   # X
WL_CLIENT_COL_START = 3   # C
WL_CLIENT_COL_END = 6     # F

# 화면 엑셀 미리보기 배율 (인쇄 print_mode=True 와 무관 — 인쇄는 Excel pageSetup 그대로)
_WL_PREVIEW_SCALE = 0.65
# 웹/Mac: Batang(윈도우) 미설치 → 한글 글리프가 고딕으로 떨어짐.
# Nanum Myeongjo(CDN)를 최우선으로 두어 바탕체와 같은 명조 계열을 강제.
_WL_FONT_STACK = "'Nanum Myeongjo','Apple Myungjo','Batang','BatangChe','바탕체','바탕','바탕글',serif"
_WL_FONT_FACE_CSS = "@import url('https://fonts.googleapis.com/css2?family=Nanum+Myeongjo:wght@400;700&display=swap');"
# 로컬 반영 확인용 (탭 상단에 표시)
_WL_UI_BUILD = "2026-08-26y · 워크시트일+인쇄레이아웃"


class WorklogSaveBlockedError(Exception):
    """이미 저장된 날짜에 후입력 저장 시도."""


# =====================================================================
# 💡 [컴포넌트 복구] 기존 환경에 등록된 안전한 v2 컴포넌트 이름 유지
# =====================================================================

_WL_LINES_HTML = """
<div class="wl-lines"></div>
"""

# 거래처·내용 입력칸만 글씨를 줄여 한 줄이 잘리지 않게 (인쇄 글씨와 무관)
_WL_LINES_CSS = """
@import url('https://fonts.googleapis.com/css2?family=Nanum+Myeongjo:wght@400;700&display=swap');
.wl-lines { display: flex; flex-direction: column; width: 100%; border: 1px solid #94A3B8; border-radius: 4px; overflow: hidden; background: #fff; box-sizing: border-box; }
.wl-row { display: flex; align-items: center; width: 100%; border-bottom: 1px solid #E2E8F0; box-sizing: border-box; }
.wl-row:last-child { border-bottom: none; }
.wl-row input {
  flex: 1 1 auto;
  min-width: 0;
  width: 100%;
  height: 28px;
  padding: 0 6px;
  border: none;
  background: transparent;
  color: #0F172A;
  font-family: 'Nanum Myeongjo','Apple Myungjo','Batang','BatangChe','바탕체','바탕',serif !important;
  font-size: 11pt;  /* 입력칸 전용: 글씨가 칸 안에 다 보이도록 */
  line-height: 28px;
  outline: none;
  box-sizing: border-box;
}
.wl-lines.client .wl-row input { background: #F8FAFC; text-align: center; }
.wl-row input:focus { background: #E0F2FE; }
.wl-row button {
  flex: 0 0 2rem;
  height: 28px;
  border: none; border-left: 1px solid #E2E8F0;
  background: #F1F5F9; color: #64748B; font-size: 0.65rem;
  padding: 0;
  cursor: pointer; display: inline-flex; align-items: center; justify-content: center;
}
.wl-row button:hover { background: #E2E8F0; color: #DC2626; }
.wl-row button.hidden { display: none; }
"""

_WL_LINES_JS = r"""
const __wlLinesInst = new WeakMap();

export default function (component) {
  const { data, parentElement, setStateValue } = component;
  const root = parentElement.querySelector(".wl-lines");
  if (!root) return;

  const maxU = Number((data && data.max_u) || 64);
  const cellW = Number((data && data.cell_w) || 666);
  const variant = String((data && data.variant) || "content");
  const rev = Number((data && data.rev) || 0);
  const focusReq = Number((data && data.focus));
  const incoming = Array.isArray(data && data.lines)
    ? data.lines.map((x) => String(x ?? ""))
    : [""];

  try {
    root.className = "wl-lines" + (variant === "client" ? " client" : "");
    root.style.width = "100%";
  } catch (e0) {}

  let inst = __wlLinesInst.get(root);
  if (!inst) {
    inst = { lines: null, rev: null, rebuilding: false };
    __wlLinesInst.set(root, inst);
  }

  function charUnits(ch) {
    const o = ch.charCodeAt(0);
    if (
      (o >= 0xac00 && o <= 0xd7a3) ||
      (o >= 0x1100 && o <= 0x11ff) ||
      (o >= 0x3130 && o <= 0x318f) ||
      (o >= 0x2e80 && o <= 0x9fff) ||
      (o >= 0xff00 && o <= 0xffef)
    )
      return 2;
    return 1;
  }
  function displayUnits(s) {
    let w = 0;
    s = s || "";
    for (let i = 0; i < s.length; i++) w += charUnits(s.charAt(i));
    return w;
  }
  function fitByUnits(s, max) {
    if (!s) return { head: "", tail: "" };
    if (displayUnits(s) <= max) return { head: s, tail: "" };
    let acc = 0;
    for (let i = 0; i < s.length; i++) {
      const cu = charUnits(s.charAt(i));
      if (acc + cu > max) {
        if (i === 0) return { head: s.slice(0, 1), tail: s.slice(1) };
        return { head: s.slice(0, i), tail: s.slice(i) };
      }
      acc += cu;
    }
    return { head: s, tail: "" };
  }
  function normalize(arr) {
    const out = (arr || []).map((x) => String(x ?? ""));
    if (!out.length) out.push("");
    if (out[out.length - 1] !== "") out.push("");
    return out;
  }
  function readDomLines() {
    const inputs = root.querySelectorAll("input[data-idx]");
    if (!inputs.length) return normalize(inst.lines || incoming);
    const out = [];
    inputs.forEach((inp) => out.push(inp.value || ""));
    return normalize(out);
  }
  function emit(next, focusIdx) {
    const out = normalize(next);
    inst.lines = out;
    setStateValue("lines", out);
    if (typeof focusIdx === "number") setStateValue("focus", focusIdx);
    return out;
  }
  function localOnly(next) {
    inst.lines = normalize(next);
    return inst.lines;
  }
  function focusAt(idx) {
    requestAnimationFrame(() => {
      const el = root.querySelector('input[data-idx="' + idx + '"]');
      if (!el) return;
      try {
        el.focus({ preventScroll: false });
        const n = (el.value || "").length;
        el.setSelectionRange(n, n);
      } catch (e) {
        try {
          el.focus();
        } catch (e2) {}
      }
    });
  }

  function insertLineAfter(j, inputEl) {
    const raw = [];
    root.querySelectorAll("input[data-idx]").forEach((inp) => {
      raw.push(inp.value || "");
    });
    if (!raw.length) raw.push("");
    while (raw.length <= j) raw.push("");
    raw[j] = (inputEl && inputEl.value) || raw[j] || "";
    raw.splice(j + 1, 0, "");
    if (raw[raw.length - 1] !== "") raw.push("");
    emit(raw, j + 1);
    rebuild(j + 1);
  }

  function rebuild(focusIdx) {
    inst.rebuilding = true;
    root.innerHTML = "";
    const lines = normalize(inst.lines || incoming);
    inst.lines = lines;
    lines.forEach((line, j) => {
      const row = document.createElement("div");
      row.className = "wl-row";
      const input = document.createElement("input");
      input.type = "text";
      input.value = line;
      input.placeholder = "";
      input.dataset.idx = String(j);
      const commitValue = (mode) => {
        if (inst.rebuilding) return;
        const cur = readDomLines();
        let j0 = j;
        let v = cur[j0] || "";
        if (displayUnits(v) <= maxU) {
          // type/composition도 Python에 동기화 — 저장 클릭 시 직전 입력이 누락되지 않게
          if (mode === "blur" || mode === "force" || mode === "type") emit(cur, null);
          else localOnly(cur);
          return;
        }
        while (j0 < cur.length && displayUnits(cur[j0] || "") > maxU) {
          const ft = fitByUnits(cur[j0] || "", maxU);
          cur[j0] = ft.head;
          if (j0 + 1 < cur.length) cur[j0 + 1] = ft.tail + (cur[j0 + 1] || "");
          else cur.splice(j0 + 1, 0, ft.tail);
          j0 += 1;
        }
        const focusIdx = Math.min(j0, Math.max(cur.length - 1, 0));
        emit(cur, focusIdx);
        rebuild(focusIdx);
      };
      input.addEventListener("input", (e) => {
        if (e.isComposing) return;
        commitValue("type");
      });
      input.addEventListener("compositionend", () => {
        commitValue("type");
      });
      input.addEventListener("focus", () => {
        if (inst.rebuilding) return;
        let s = 0;
        let e = 0;
        try {
          s = typeof input.selectionStart === "number" ? input.selectionStart : 0;
          e = typeof input.selectionEnd === "number" ? input.selectionEnd : s;
        } catch (err) {}
        setStateValue("focus", j);
        setStateValue("caret", { s: s, e: e, j: j });
      });
      input.addEventListener("keyup", () => {
        if (inst.rebuilding) return;
        let s = 0;
        let e = 0;
        try {
          s = typeof input.selectionStart === "number" ? input.selectionStart : 0;
          e = typeof input.selectionEnd === "number" ? input.selectionEnd : s;
        } catch (err) {}
        setStateValue("caret", { s: s, e: e, j: j });
      });
      input.addEventListener("click", () => {
        if (inst.rebuilding) return;
        let s = 0;
        let e = 0;
        try {
          s = typeof input.selectionStart === "number" ? input.selectionStart : 0;
          e = typeof input.selectionEnd === "number" ? input.selectionEnd : s;
        } catch (err) {}
        setStateValue("focus", j);
        setStateValue("caret", { s: s, e: e, j: j });
      });
      input.addEventListener("blur", () => {
        commitValue("blur");
      });
      input.addEventListener("keydown", (e) => {
        if (e.key !== "Enter") return;
        const typing = (input.value || "") !== "";
        if (e.isComposing && typing) return;
        e.preventDefault();
        e.stopPropagation();
        insertLineAfter(j, input);
      });
      const del = document.createElement("button");
      del.type = "button";
      del.textContent = "×";
      const isLastEmpty = j === lines.length - 1 && (line || "") === "";
      if (isLastEmpty) del.classList.add("hidden");
      del.addEventListener("click", () => {
        const cur = readDomLines();
        cur.splice(j, 1);
        if (!cur.length) cur.push("");
        const fj = Math.min(j, Math.max(cur.length - 1, 0));
        emit(cur, fj);
        rebuild(fj);
      });
      row.appendChild(input);
      row.appendChild(del);
      root.appendChild(row);
    });
    inst.rebuilding = false;
    if (typeof focusIdx === "number" && focusIdx >= 0) focusAt(focusIdx);
  }

  if (inst.rev !== rev) {
    inst.rev = rev;
    inst.lines = normalize(incoming);
    rebuild(Number.isFinite(focusReq) ? focusReq : -1);
  } else if (!root.childElementCount) {
    if (!inst.lines) inst.lines = normalize(incoming);
    rebuild(Number.isFinite(focusReq) ? focusReq : -1);
  } else {
    const dom = readDomLines();
    inst.lines = normalize(dom);
  }
}
"""

_WL_LINES_EDITOR = st.components.v2.component(
    "worklog_entry_lines_v15",
    html=_WL_LINES_HTML,
    css=_WL_LINES_CSS,
    js=_WL_LINES_JS,
)

_WL_ENTER_HOOK_JS = r"""
export default function (component) {
  const { data, setTriggerValue } = component;
  const iso = (data && data.iso) || "";
  const focusKey = (data && data.focus_key) || "";
  const focusCaret = data && data.focus_caret != null && data.focus_caret !== "" ? Number(data.focus_caret) : null;
  let lastSent = ""; let lastSig = ""; let lastAt = 0;

  function resolveKey(t) {
    if (!t) return null;
    const tag = String(t.tagName || "").toUpperCase();
    if (tag === "TEXTAREA") {
      const wrap = t.closest ? t.closest('[class*="st-key-wl_next_area_"],[class*="st-key-wl_notes_area_"]') : null;
      if (!wrap) return null;
      const cls = Array.prototype.find.call(wrap.classList || [], (c) => {
        const s = String(c);
        return s.indexOf("st-key-wl_next_area_") !== -1 || s.indexOf("st-key-wl_notes_area_") !== -1;
      });
      if (!cls) return null;
      const key = String(cls).replace(/^st-key-/, "");
      const m = /^(wl_(?:next|notes)_area)_(\d{4}-\d{2}-\d{2})$/.exec(key);
      if (!m || m[2] !== iso) return null;
      return { key: key, kind: m[1], ei: -1, lj: -1 };
    }
    if (tag !== "INPUT") return null;
    let wrap = t.closest ? t.closest('[class*="st-key-wl_ent_ln_"],[class*="st-key-wl_ent_cl_"]') : null;
    if (wrap) {
      const cls = Array.prototype.find.call(wrap.classList || [], (c) => {
        const s = String(c);
        return s.indexOf("st-key-wl_ent_ln_") !== -1 || s.indexOf("st-key-wl_ent_cl_") !== -1;
      });
      if (cls) {
        const key = String(cls).replace(/^st-key-/, "");
        const m = /^(wl_ent_ln|wl_ent_cl)_(\d{4}-\d{2}-\d{2})_(\d+)_(\d+)(?:_g\d+)?$/.exec(key);
        if (m && m[2] === iso) return { key: key, kind: m[1], ei: Number(m[3]), lj: Number(m[4]) };
      }
    }
    wrap = t.closest ? t.closest('[class*="st-key-wl_lines_comp_"],[class*="st-key-wl_clients_comp_"]') : null;
    if (!wrap) return null;
    const cls2 = Array.prototype.find.call(wrap.classList || [], (c) => {
      const s = String(c);
      return s.indexOf("st-key-wl_lines_comp_") !== -1 || s.indexOf("st-key-wl_clients_comp_") !== -1;
    });
    if (!cls2) return null;
    const compKey = String(cls2).replace(/^st-key-/, "");
    const m2 = /^(wl_(?:lines|clients)_comp)_(\d{4}-\d{2}-\d{2})_(\d+)$/.exec(compKey);
    if (!m2 || m2[2] !== iso) return null;
    const kind = m2[1] === "wl_clients_comp" ? "wl_ent_cl" : "wl_ent_ln";
    const ei = Number(m2[3]);
    const lj = Number(t.dataset && t.dataset.idx != null ? t.dataset.idx : 0);
    const logicalKey = kind + "_" + iso + "_" + ei + "_" + lj;
    return { key: logicalKey, kind: kind, ei: ei, lj: lj };
  }

  function listKindInputs(kind) {
    const needle = kind === "wl_ent_cl" ? "st-key-wl_ent_cl_" : "st-key-wl_ent_ln_";
    const compNeedle = kind === "wl_ent_cl" ? "st-key-wl_clients_comp_" : "st-key-wl_lines_comp_";
    const nodes = document.querySelectorAll(
      'div[class*="' + needle + '"] input, div[class*="' + compNeedle + '"] .wl-lines input[data-idx]'
    );
    const out = [];
    for (let i = 0; i < nodes.length; i++) {
      const info = resolveKey(nodes[i]);
      if (info) out.push(nodes[i]);
    }
    return out;
  }

  function findInputForFocusKey(focusKey) {
    if (!focusKey) return null;
    let m = /^(wl_ent_ln|wl_ent_cl)_(\d{4}-\d{2}-\d{2})_(\d+)_(\d+)(?:_g\d+)?$/.exec(focusKey);
    if (m) {
      const kind = m[1];
      const isoVal = m[2];
      const ei = m[3];
      const lj = m[4];
      const compNeedle =
        (kind === "wl_ent_cl" ? "st-key-wl_clients_comp_" : "st-key-wl_lines_comp_") + isoVal + "_" + ei;
      const wraps = document.querySelectorAll('div[class*="' + compNeedle + '"]');
      for (let i = 0; i < wraps.length; i++) {
        const el = wraps[i].querySelector('.wl-lines input[data-idx="' + lj + '"]');
        if (el) return el;
      }
      const legacy = document.querySelector('div[class*="st-key-' + focusKey + '"] input');
      if (legacy) return legacy;
    }
    return document.querySelector(
      'div[class*="st-key-' + focusKey + '"] input, div[class*="st-key-' + focusKey + '"] textarea'
    );
  }

  function focusInput(el, caret) {
    if (!el) return false;
    try {
      el.focus({ preventScroll: false });
      const n = (el.value || "").length;
      let pos = n;
      if (caret === "start") pos = 0;
      else if (caret === "end") pos = n;
      else if (typeof caret === "number" && Number.isFinite(caret)) {
        pos = Math.max(0, Math.min(n, Math.floor(caret)));
      }
      el.setSelectionRange(pos, pos);
    } catch (e1) {
      try { el.focus(); } catch (e2) {}
    }
    return true;
  }

  function findPeer(info, targetKind) {
    const list = listKindInputs(targetKind);
    let best = null; let bestScore = 1e9;
    for (let i = 0; i < list.length; i++) {
      const p = resolveKey(list[i]);
      if (!p) continue;
      const score = Math.abs(p.ei - info.ei) * 1000 + Math.abs(p.lj - info.lj);
      if (score < bestScore) {
        bestScore = score;
        best = list[i];
      }
    }
    return best;
  }

  function emit(key, value) {
    const now = Date.now();
    const v = String(value || "");
    const sig = key + "\\0" + v;
    if (sig === lastSig && now - lastAt < 320) return;
    lastSig = sig; lastAt = now;
    const payload = JSON.stringify({ key: key, t: now, v: v });
    if (payload === lastSent) return;
    lastSent = payload;
    setTriggerValue("enter", payload);
  }

  const onKey = (e) => {
    if (e.isComposing || e.keyCode === 229) return;
    const info = resolveKey(e.target);
    if (!info) return;
    const t = e.target;
    const start = typeof t.selectionStart === "number" ? t.selectionStart : 0;
    const end = typeof t.selectionEnd === "number" ? t.selectionEnd : start;
    const len = String(t.value || "").length;
    const caretAll = start === end;

    if (e.key === "Enter") {
      const tag = String(t.tagName || "").toUpperCase();
      // 익일업무·특이사항 textarea: Enter = 다음 줄, ⌘/Ctrl+Enter = 반영
      if (tag === "TEXTAREA") {
        if (e.metaKey || e.ctrlKey) {
          e.preventDefault();
          e.stopPropagation();
          emit(info.key, t.value || "");
        }
        return;
      }
      e.preventDefault();
      e.stopPropagation();
      emit(info.key, t.value || "");
      return;
    }

    if (e.key === "ArrowUp") {
      const list = listKindInputs(info.kind);
      const idx = list.indexOf(t);
      if (idx > 0) {
        e.preventDefault();
        e.stopPropagation();
        focusInput(list[idx - 1], "end");
      } else {
        const peerKind = info.kind === "wl_ent_cl" ? "wl_ent_ln" : "wl_ent_cl";
        const peer = findPeer(info, peerKind);
        if (peer && peer !== t) {
          e.preventDefault();
          e.stopPropagation();
          focusInput(peer, "end");
        }
      }
      return;
    }
    if (e.key === "ArrowDown") {
      const list = listKindInputs(info.kind);
      const idx = list.indexOf(t);
      if (idx >= 0 && idx < list.length - 1) {
        e.preventDefault();
        e.stopPropagation();
        focusInput(list[idx + 1], "end");
      } else {
        const peerKind = info.kind === "wl_ent_cl" ? "wl_ent_ln" : "wl_ent_cl";
        const peer = findPeer({ ...info, lj: info.lj + 1 }, peerKind);
        if (peer && peer !== t) {
          e.preventDefault();
          e.stopPropagation();
          focusInput(peer, "start");
        }
      }
      return;
    }
    if (e.key === "ArrowLeft" && caretAll && start === 0) {
      const peer = info.kind === "wl_ent_ln" ? findPeer(info, "wl_ent_cl") : (() => {
        const list = listKindInputs("wl_ent_cl");
        const idx = list.indexOf(t);
        return idx > 0 ? list[idx - 1] : null;
      })();
      if (peer) {
        e.preventDefault();
        e.stopPropagation();
        focusInput(peer, "end");
      }
      return;
    }
    if (e.key === "ArrowRight" && caretAll && start === len) {
      const peer = info.kind === "wl_ent_cl" ? findPeer(info, "wl_ent_ln") : (() => {
        const list = listKindInputs("wl_ent_ln");
        const idx = list.indexOf(t);
        return idx >= 0 && idx < list.length - 1 ? list[idx + 1] : null;
      })();
      if (peer) {
        e.preventDefault();
        e.stopPropagation();
        focusInput(peer, "start");
      }
      return;
    }
  };

  document.addEventListener("keydown", onKey, true);

  function emitFocus(key) {
    try { setTriggerValue("focus", String(key || "")); } catch (e) {}
  }
  function emitCaret(key, s, e) {
    try {
      setTriggerValue("caret", JSON.stringify({ key: key, s: s, e: e, t: Date.now() }));
    } catch (e) {}
  }
  const onFocusIn = (e) => {
    const info = resolveKey(e.target);
    if (!info) return;
    let s = 0;
    let en = 0;
    try {
      s = typeof e.target.selectionStart === "number" ? e.target.selectionStart : 0;
      en = typeof e.target.selectionEnd === "number" ? e.target.selectionEnd : s;
    } catch (err) {}
    emitFocus(info.key);
    emitCaret(info.key, s, en);
  };
  // 익일업무·특이사항 textarea: keyup마다 caret emit → Streamlit rerun → 커서 점프
  // click/focus만 추적 (특수기호 삽입용), 타이핑 중에는 emit 안 함
  const onSel = (e) => {
    const info = resolveKey(e.target);
    if (!info) return;
    const tag = String(e.target.tagName || "").toUpperCase();
    if (tag === "TEXTAREA") return;
    let s = 0;
    let en = 0;
    try {
      s = typeof e.target.selectionStart === "number" ? e.target.selectionStart : 0;
      en = typeof e.target.selectionEnd === "number" ? e.target.selectionEnd : s;
    } catch (err) {}
    emitCaret(info.key, s, en);
  };
  document.addEventListener("focusin", onFocusIn, true);
  document.addEventListener("keyup", onSel, true);
  document.addEventListener("click", onSel, true);

  if (focusKey) {
    const go = () => {
      const el = findInputForFocusKey(focusKey);
      if (!el) return false;
      const caret = focusCaret != null && Number.isFinite(focusCaret) ? focusCaret : "end";
      focusInput(el, caret);
      return true;
    };
    go(); setTimeout(go, 50); setTimeout(go, 150); setTimeout(go, 350); setTimeout(go, 600);
  }

  return () => {
    document.removeEventListener("keydown", onKey, true);
    document.removeEventListener("focusin", onFocusIn, true);
    document.removeEventListener("keyup", onSel, true);
    document.removeEventListener("click", onSel, true);
  };
}
"""

_WL_ENTER_HOOK = st.components.v2.component(
    "worklog_cell_nav_hook_v22",
    js=_WL_ENTER_HOOK_JS,
)

# 자주 쓰는 순 (앞쪽 = 우선 표시)
_WL_SPECIAL_CHARS = [
    "※", "★", "☆", "○", "●", "◎", "→", "·", "△", "▲", "■", "□",
    "「", "」", "【", "】", "～", "←", "✓", "①", "②", "③", "◇", "◆",
]

_WL_SPECIAL_BAR_HTML = """
<div class="wl-sp"></div>
"""

_WL_SPECIAL_BAR_CSS = """
.wl-sp {
  display: flex;
  flex-wrap: nowrap;
  align-items: center;
  gap: 5px;
  width: 100%;
  max-width: 100%;
  overflow-x: auto;
  overflow-y: hidden;
  -webkit-overflow-scrolling: touch;
  scroll-behavior: smooth;
  padding: 2px 2px 6px;
  box-sizing: border-box;
}
.wl-sp button {
  flex: 0 0 auto;
  min-width: 2.1rem;
  width: 2.1rem;
  height: 2.1rem;
  margin: 0;
  padding: 0;
  border: 1px solid #cbd5e1;
  border-radius: 0.4rem;
  background: #ffffff;
  color: #0f172a;
  font-size: 1.05rem;
  line-height: 2.1rem;
  text-align: center;
  cursor: pointer;
  touch-action: manipulation;
  -webkit-user-select: none;
  user-select: none;
}
.wl-sp button:hover { background: #e2e8f0; }
.wl-sp button:active { background: #dbeafe; }
"""

_WL_SPECIAL_BAR_JS = r"""
export default function (component) {
  const { data, parentElement, setTriggerValue } = component;
  const chars = (data && data.chars) || [];
  const token = String((data && data.token) || "");

  let root = parentElement.querySelector(".wl-sp");
  if (!root) {
    root = document.createElement("div");
    root.className = "wl-sp";
    parentElement.appendChild(root);
  }
  if (root.dataset.token !== token || root.childElementCount !== chars.length) {
    root.dataset.token = token;
    root.innerHTML = "";
    for (let i = 0; i < chars.length; i++) {
      const ch = String(chars[i] || "");
      const btn = document.createElement("button");
      btn.type = "button";
      btn.textContent = ch;
      btn.title = ch + " 삽입";
      const fire = (function (symbol) {
        let lastAt = 0;
        return function (ev) {
          const now = Date.now();
          if (now - lastAt < 280) return;
          lastAt = now;
          if (ev) {
            try { ev.preventDefault(); } catch (e1) {}
            try { ev.stopPropagation(); } catch (e2) {}
          }
          try {
            setTriggerValue("pick", JSON.stringify({ ch: symbol, t: Date.now() }));
          } catch (err) {}
        };
      })(ch);
      btn.addEventListener("pointerup", fire);
      btn.addEventListener("click", fire);
      root.appendChild(btn);
    }
  }

  return () => {};
}
"""

_WL_SPECIAL_BAR = st.components.v2.component(
    "worklog_special_bar_v9",
    html=_WL_SPECIAL_BAR_HTML,
    css=_WL_SPECIAL_BAR_CSS,
    js=_WL_SPECIAL_BAR_JS,
)

def _entry_lines_inst_key(iso: str, entry_i: int) -> str: return f"wl_lines_inst_{iso}_{entry_i}"
def _entry_clients_inst_key(iso: str, entry_i: int) -> str: return f"wl_clients_inst_{iso}_{entry_i}"

def _entry_lines_comp_key(iso: str, entry_i: int) -> str:
    # 인스턴스 번호로 CCv2 위젯을 재마운트 — 동일 키에 남은 구 result.lines 부활 차단
    g = int(st.session_state.get(_entry_lines_inst_key(iso, entry_i), 0) or 0)
    return f"wl_lines_comp_{iso}_{entry_i}_i{g}"

def _entry_lines_rev_key(iso: str, entry_i: int) -> str: return f"wl_ent_rev_{iso}_{entry_i}"
def _entry_lines_live_key(iso: str, entry_i: int) -> str: return f"wl_lines_live_{iso}_{entry_i}"

def _entry_clients_comp_key(iso: str, entry_i: int) -> str:
    g = int(st.session_state.get(_entry_clients_inst_key(iso, entry_i), 0) or 0)
    return f"wl_clients_comp_{iso}_{entry_i}_i{g}"

def _entry_clients_rev_key(iso: str, entry_i: int) -> str: return f"wl_clients_rev_{iso}_{entry_i}"
def _entry_clients_live_key(iso: str, entry_i: int) -> str: return f"wl_clients_live_{iso}_{entry_i}"

def _bump_entry_lines_comp_inst(iso: str, entry_i: int) -> None:
    st.session_state.pop(_entry_lines_comp_key(iso, entry_i), None)
    st.session_state.pop(f"wl_lines_comp_{iso}_{entry_i}", None)  # legacy key
    st.session_state.pop(f"wl_lines_user_edit_{iso}_{entry_i}", None)
    k = _entry_lines_inst_key(iso, entry_i)
    st.session_state[k] = int(st.session_state.get(k, 0) or 0) + 1

def _bump_entry_clients_comp_inst(iso: str, entry_i: int) -> None:
    st.session_state.pop(_entry_clients_comp_key(iso, entry_i), None)
    st.session_state.pop(f"wl_clients_comp_{iso}_{entry_i}", None)  # legacy key
    st.session_state.pop(f"wl_clients_user_edit_{iso}_{entry_i}", None)
    k = _entry_clients_inst_key(iso, entry_i)
    st.session_state[k] = int(st.session_state.get(k, 0) or 0) + 1

def _scrub_dummy_label(val: str) -> str:
    s = (val or "").strip()
    if re.fullmatch(r"거래처\d+", s) or re.fullmatch(r"내용\d+", s): return ""
    return val or ""

def _char_units(ch: str) -> int:
    if not ch: return 0
    o = ord(ch)
    if (0xAC00 <= o <= 0xD7A3 or 0x1100 <= o <= 0x11FF or 0x3130 <= o <= 0x318F or 0x2E80 <= o <= 0x9FFF or 0xF900 <= o <= 0xFAFF or 0xFF00 <= o <= 0xFFEF): return 2
    return 1

def _display_units(s: str) -> int: return sum(_char_units(ch) for ch in (s or ""))

_WL_BODY_FONT_NAME = "바탕체"
_WL_BODY_FONT_PT = 14.0  # 원본.xlsx 본문 글자 크기와 동일

def _set_body_font(cell) -> None:
    try: cell.font = cell.font.copy(name=_WL_BODY_FONT_NAME, size=float(_WL_BODY_FONT_PT))
    except Exception: pass

# =====================================================================
# 💡 [핵심] 글자 넘침 현상 원천 차단 (엄격한 max_units 설정)
# =====================================================================
# 14pt 바탕체 기준, 엑셀 너비를 넘지 않도록 한계치를 하향 조정
@lru_cache(maxsize=1)
def _content_line_units() -> int: return 72  # 한글 34자 제한

@lru_cache(maxsize=1)
def _client_line_units() -> int: return 16  # 👈 한글 8자(16 단위)로 증가

def _fit_by_units(s: str, max_units: int | None = None) -> tuple[str, str]:
    if max_units is None: max_units = _content_line_units()
    if not s: return "", ""
    if _display_units(s) <= max_units: return s, ""
    acc = 0
    for i, ch in enumerate(s):
        cu = _char_units(ch)
        if acc + cu > max_units: return s[:i] if i else s[:1], s[i:] if i else s[1:]
        acc += cu
    return s, ""

def _chunk_text(text: str, max_units: int | None = None) -> list[str]:
    if max_units is None: max_units = _content_line_units()
    s = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    if not s: return []
    out: list[str] = []
    for para in s.split("\n"):
        rest = para
        if rest == "" and not out: continue
        if rest == "": continue
        while rest:
            head, rest = _fit_by_units(rest, max_units)
            if not head and rest: head, rest = rest[:1], rest[1:]
            out.append(head)
            if not rest: break
    return out

def _spill_column(cells: dict, rows: list[int], col: str) -> dict:
    max_u = _content_line_units()
    vals = [str(cells.get(f"{col}{r}", "") or "") for r in rows]
    for i in range(len(vals)):
        while _display_units(vals[i]) > max_u and i + 1 < len(vals):
            head, tail = _fit_by_units(vals[i], max_u)
            vals[i] = head
            vals[i + 1] = tail + vals[i + 1]
        if _display_units(vals[i]) > max_u:
            head, _tail = _fit_by_units(vals[i], max_u)
            vals[i] = head
    out = dict(cells)
    for r, v in zip(rows, vals): out[f"{col}{r}"] = v
    return out

def _spill_all_content(cells: dict) -> dict:
    cells = _spill_column(cells, WL_CONTENT_ROWS, "G")
    cells = _spill_column(cells, WL_NEXT_ROWS, "D")
    cells = _spill_column(cells, WL_NOTE_ROWS, "D")
    return cells

# 💡 템플릿 준비: git의 uploaded_cache/worklog/template.xlsx 를 우선 사용
# (예전엔 ~/Desktop/업무일지.xlsx mtime이 더 新し면 덮어써서 로고·양식 반영이 깨짐)
def _ensure_dirs() -> None:
    os.makedirs(WORKLOG_DIR, exist_ok=True)
    # Desktop 파일은 "캐시 템플릿이 없을 때만" 보충. 있으면 절대 덮어쓰지 않음.
    if (not os.path.exists(WORKLOG_TEMPLATE)) and os.path.exists(WORKLOG_TEMPLATE_SRC):
        try:
            shutil.copy2(WORKLOG_TEMPLATE_SRC, WORKLOG_TEMPLATE)
        except Exception:
            pass

def _iter_google_drive_roots() -> list[str]:
    cloud = os.path.join(os.path.expanduser("~"), "Library", "CloudStorage")
    if not os.path.isdir(cloud): return []
    roots: list[str] = []
    try:
        for name in sorted(os.listdir(cloud)):
            if name.startswith("GoogleDrive"): roots.append(os.path.join(cloud, name))
    except OSError: return []
    return roots

def resolve_worklog_archive_root() -> str | None:
    candidates: list[str] = []
    home = os.path.expanduser("~")
    candidates.append(os.path.join(home, "Desktop", "업무", "일지"))
    for groot in _iter_google_drive_roots():
        for other_name in ("다른 컴퓨터", "Computers"):
            other = os.path.join(groot, other_name)
            if not os.path.isdir(other): continue
            try: pcs = sorted(os.listdir(other))
            except OSError: continue
            pcs.sort(key=lambda n: (0 if "(1)" in n else 1, n))
            for pc in pcs: candidates.append(os.path.join(other, pc, WORKLOG_ARCHIVE_REL))
    existing = [p for p in candidates if os.path.isdir(p)]
    if existing: return existing[0]
    for p in candidates:
        parent = os.path.dirname(p)
        grand = os.path.dirname(parent)
        if os.path.isdir(grand):
            try: os.makedirs(p, exist_ok=True); return p
            except OSError: continue
    return None

def worklog_archive_path(d: date) -> str | None:
    """호환용: 월별 통합 파일 경로 (…/일지/2026/8월.xlsx)."""
    return worklog_archive_month_path(d)


def worklog_archive_year_dir(d: date) -> str | None:
    """연도 폴더: …/일지/2026"""
    root = resolve_worklog_archive_root()
    if not root:
        return None
    year_dir = os.path.join(root, str(d.year))
    try:
        os.makedirs(year_dir, exist_ok=True)
    except OSError:
        return None
    return year_dir


def worklog_archive_month_dir(d: date) -> str | None:
    """구버전 월 하위폴더(…/2026/8월). 레거시 정리용."""
    year_dir = worklog_archive_year_dir(d)
    if not year_dir:
        return None
    return os.path.join(year_dir, f"{d.month}월")


def worklog_archive_month_path(d: date) -> str | None:
    """달이 바뀌면 9월.xlsx 신규. 경로: …/일지/2026/8월.xlsx (연도 폴더 직하)."""
    year_dir = worklog_archive_year_dir(d)
    if not year_dir:
        return None
    return os.path.join(year_dir, f"{d.month}월.xlsx")


def _cleanup_legacy_day_archive_files(d: date, month_path: str | None) -> None:
    """예전 일자별 xlsx·구 월폴더 안 일자파일을 정리."""
    year_dir = worklog_archive_year_dir(d)
    candidates: list[str] = []
    if year_dir:
        candidates.append(os.path.join(year_dir, f"{d.isoformat()}.xlsx"))
        # 구경로: …/2026/8월/YYYY-MM-DD.xlsx , …/2026/8월/8월.xlsx
        old_month_dir = os.path.join(year_dir, f"{d.month}월")
        candidates.append(os.path.join(old_month_dir, f"{d.isoformat()}.xlsx"))
        candidates.append(os.path.join(old_month_dir, f"{d.month}월.xlsx"))
    month_abs = os.path.abspath(month_path) if month_path else ""
    for path in candidates:
        try:
            if month_abs and os.path.abspath(path) == month_abs:
                continue
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass
    # 구 월 폴더가 비면 제거
    if year_dir:
        old_month_dir = os.path.join(year_dir, f"{d.month}월")
        try:
            if os.path.isdir(old_month_dir) and not os.listdir(old_month_dir):
                os.rmdir(old_month_dir)
        except OSError:
            pass


def worklog_archive_sheet_title(d: date) -> str:
    """월별 파일 안 워크시트명 = 일 (예: 27)."""
    return str(d.day)


def _worklog_archive_legacy_sheet_title(d: date) -> str:
    """구버전 워크시트명 (YYYY-MM-DD). 존재·삭제 조회용."""
    return d.isoformat()


def _worklog_archive_sheet_titles_for_lookup(d: date) -> tuple[str, ...]:
    """현재(일) + 레거시(YYYY-MM-DD) 워크시트명."""
    cur = worklog_archive_sheet_title(d)
    leg = _worklog_archive_legacy_sheet_title(d)
    return (cur,) if cur == leg else (cur, leg)


def _resolve_archive_sheet_name(sheetnames: list[str] | tuple[str, ...], d: date) -> str | None:
    for title in _worklog_archive_sheet_titles_for_lookup(d):
        if title in sheetnames:
            return title
    return None


def _archive_sheet_sort_key(name: str) -> tuple:
    if name.isdigit():
        return (0, int(name))
    try:
        return (1, date.fromisoformat(name))
    except ValueError:
        return (2, name)


def _invalidate_worklog_presence_cache(d: date | None = None) -> None:
    """날짜 존재 캐시 무효화 (저장·삭제 후)."""
    if d is not None:
        iso = d.isoformat()
        st.session_state.pop(f"wl_arch_exists_{iso}", None)
        st.session_state.pop(f"wl_presence_{iso}", None)
        return
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and (k.startswith("wl_arch_exists_") or k.startswith("wl_presence_")):
            st.session_state.pop(k, None)


def worklog_date_exists_in_archive(d: date) -> bool:
    """월별 xlsx에 해당 날짜 시트가 있는지 (구 일자파일 포함)."""
    iso = d.isoformat()
    ck = f"wl_arch_exists_{iso}"
    cached = st.session_state.get(ck)
    if isinstance(cached, bool):
        return cached
    titles = _worklog_archive_sheet_titles_for_lookup(d)
    found = False
    month_path = worklog_archive_month_path(d)
    if month_path and os.path.exists(month_path) and load_workbook is not None:
        try:
            wb = load_workbook(month_path, read_only=True)
            try:
                if _resolve_archive_sheet_name(wb.sheetnames, d):
                    found = True
            finally:
                wb.close()
        except Exception:
            pass
    if not found:
        year_dir = worklog_archive_year_dir(d)
        if year_dir:
            for p in (
                os.path.join(year_dir, f"{d.isoformat()}.xlsx"),
                os.path.join(year_dir, f"{d.month}월", f"{d.isoformat()}.xlsx"),
                os.path.join(year_dir, f"{d.month}월", f"{d.month}월.xlsx"),
            ):
                if not os.path.exists(p):
                    continue
                if p.endswith(f"{d.month}월.xlsx") and load_workbook is not None:
                    try:
                        wb = load_workbook(p, read_only=True)
                        try:
                            if _resolve_archive_sheet_name(wb.sheetnames, d):
                                found = True
                                break
                        finally:
                            wb.close()
                    except Exception:
                        pass
                elif os.path.basename(p) == f"{d.isoformat()}.xlsx":
                    found = True
                    break
    st.session_state[ck] = found
    return found


def worklog_date_exists_on_drive(d: date) -> bool:
    try:
        from drive_autoload import resolve_drive_worklog_archive_dir, resolve_drive_worklog_dir

        drv = resolve_drive_worklog_dir()
        if drv and os.path.isfile(os.path.join(drv, f"{d.isoformat()}.xlsx")):
            return True
        # Drive에 올린 월별 통합 파일 시트도 존재로 간주
        arch_dir = resolve_drive_worklog_archive_dir(d.year)
        if not arch_dir or load_workbook is None:
            return False
        month_p = os.path.join(arch_dir, f"{d.month}월.xlsx")
        if not os.path.isfile(month_p):
            return False
        wb = load_workbook(month_p, read_only=True)
        try:
            return _resolve_archive_sheet_name(wb.sheetnames, d) is not None
        finally:
            wb.close()
    except Exception:
        return False


def worklog_date_exists_on_cloud(d: date) -> bool:
    try:
        from worklog_remote_sync import worklog_date_exists_on_cloud as _cloud_exists

        return bool(_cloud_exists(d, WORKLOG_DIR))
    except Exception:
        return False


def detect_worklog_date_presence(d: date, *, include_remote: bool = True) -> dict:
    """로컬 캐시·월별보관·(선택) Drive·Cloud Gist 에 해당 날짜 일지 존재 여부."""
    iso = d.isoformat()
    cache_k = f"wl_presence_{iso}_{'all' if include_remote else 'fast'}"
    cached = st.session_state.get(cache_k)
    if isinstance(cached, dict):
        return cached
    local = os.path.isfile(worklog_path(d))
    archive = False
    drive = False
    cloud = False
    if not local:
        archive = worklog_date_exists_in_archive(d)
        if include_remote:
            if not archive:
                drive = worklog_date_exists_on_drive(d)
            if not archive and not drive:
                cloud = worklog_date_exists_on_cloud(d)
    locs: list[str] = []
    if local:
        locs.append("로컬 캐시")
    if archive:
        root = resolve_worklog_archive_root()
        if root:
            locs.append(f"일지/{d.year}/{d.month}월.xlsx")
        else:
            locs.append(f"월별파일({d.month}월.xlsx)")
    if drive:
        locs.append("Drive worklog")
    if cloud:
        locs.append("Cloud Gist")
    out = {
        "local": local,
        "archive": archive,
        "drive": drive,
        "cloud": cloud,
        "any": bool(local or archive or drive or cloud),
        "locations": locs,
    }
    st.session_state[cache_k] = out
    return out


def check_worklog_save_allowed(d: date, *, had_local_at_open: bool) -> tuple[bool, str]:
    """날짜 중복 시 후입력 저장 차단.

    로컬 캐시·월별 xlsx(Desktop/업무/일지)만 차단.
    Cloud Gist·Drive에만 있으면 맥 로컬 저장은 허용 (저장 시 로컬+아카이브 반영).
    """
    if had_local_at_open and os.path.isfile(worklog_path(d)):
        return True, ""
    local = os.path.isfile(worklog_path(d))
    archive = worklog_date_exists_in_archive(d) if not local else False
    if not local and not archive:
        return True, ""
    locs: list[str] = []
    if local:
        locs.append("로컬 캐시")
    if archive:
        root = resolve_worklog_archive_root()
        if root:
            locs.append(f"일지/{d.year}/{d.month}월.xlsx")
        else:
            locs.append(f"월별파일({d.month}월.xlsx)")
    detail = ", ".join(locs) if locs else "저장소"
    return (
        False,
        f"{d.isoformat()} 일지가 이미 있습니다 ({detail}). "
        "수정하려면 달력에서 해당 날짜(•)를 선택하거나 삭제 후 다시 저장하세요.",
    )


def describe_worklog_archive_target(d: date) -> str:
    """UI용 저장 경로 설명 (Desktop/업무/일지/YYYY/N월.xlsx#일)."""
    root = resolve_worklog_archive_root()
    month_path = worklog_archive_month_path(d)
    if root and month_path:
        return f"{root}/{d.year}/{d.month}월.xlsx#{worklog_archive_sheet_title(d)}"
    if month_path:
        return f"{month_path}#{worklog_archive_sheet_title(d)}"
    home = os.path.expanduser("~")
    return f"{home}/Desktop/업무/일지/{d.year}/{d.month}월.xlsx#{worklog_archive_sheet_title(d)}"


def _copy_worksheet_cross_workbook(src_ws, dst_ws) -> None:
    """다른 통합문서 간 시트 복사 — openpyxl WorksheetCopy + 인쇄·화면 설정."""
    from copy import copy as _cpy

    for (row, col), source_cell in src_ws._cells.items():
        target_cell = dst_ws.cell(column=col, row=row)
        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type
        if source_cell.has_style:
            target_cell._style = _cpy(source_cell._style)
        if source_cell.hyperlink:
            target_cell._hyperlink = _cpy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = _cpy(source_cell.comment)

    for attr in ("row_dimensions", "column_dimensions"):
        src_dims = getattr(src_ws, attr)
        dst_dims = getattr(dst_ws, attr)
        for key, dim in src_dims.items():
            dst_dims[key] = _cpy(dim)
            dst_dims[key].worksheet = dst_ws

    dst_ws.sheet_format = _cpy(src_ws.sheet_format)
    dst_ws.sheet_properties = _cpy(src_ws.sheet_properties)
    dst_ws.merged_cells = _cpy(src_ws.merged_cells)
    dst_ws.page_margins = _cpy(src_ws.page_margins)
    dst_ws.page_setup = _cpy(src_ws.page_setup)
    dst_ws.print_options = _cpy(src_ws.print_options)

    try:
        dst_ws.print_area = src_ws.print_area
    except Exception:
        pass
    try:
        dst_ws.print_title_rows = src_ws.print_title_rows
        dst_ws.print_title_cols = src_ws.print_title_cols
    except Exception:
        pass
    try:
        dst_ws.sheet_view = _cpy(src_ws.sheet_view)
    except Exception:
        pass
    try:
        if getattr(src_ws, "views", None):
            dst_ws.views = _cpy(src_ws.views)
    except Exception:
        pass

    try:
        from openpyxl.drawing.image import Image as XLImage

        for img in list(getattr(src_ws, "_images", []) or []):
            try:
                data = img._data()
                bio = io.BytesIO(data)
                new_img = XLImage(bio)
                if getattr(img, "width", None):
                    new_img.width = img.width
                if getattr(img, "height", None):
                    new_img.height = img.height
                if getattr(img, "anchor", None) is not None:
                    new_img.anchor = img.anchor
                dst_ws.add_image(new_img)
            except Exception:
                continue
    except Exception:
        pass


def _clone_worksheet_to_workbook(src_ws, dst_wb, title: str):
    """날짜일지 시트를 월별 통합 파일로 복사 (원본 인쇄·열 너비·화면 배율 유지)."""
    if title in dst_wb.sheetnames:
        del dst_wb[title]
    dst = dst_wb.create_sheet(title)
    _copy_worksheet_cross_workbook(src_ws, dst)
    return dst


def _migrate_legacy_month_workbook(d: date, month_path: str) -> None:
    """구경로 …/2026/8월/8월.xlsx → …/2026/8월.xlsx 로 1회 이동."""
    if not month_path:
        return
    if os.path.exists(month_path):
        return
    year_dir = worklog_archive_year_dir(d)
    if not year_dir:
        return
    legacy_month = os.path.join(year_dir, f"{d.month}월", f"{d.month}월.xlsx")
    if not os.path.exists(legacy_month):
        return
    try:
        os.makedirs(os.path.dirname(month_path), exist_ok=True)
        shutil.move(legacy_month, month_path)
    except OSError:
        try:
            shutil.copy2(legacy_month, month_path)
            os.remove(legacy_month)
        except OSError:
            pass


def upsert_worklog_archive_sheet(d: date, day_xlsx_path: str, *, allow_overwrite: bool = True) -> str | None:
    """일자 파일을 월별 xlsx의 날짜 시트로 반영. 달이 바뀌면 N월.xlsx 신규 생성."""
    if load_workbook is None:
        return None
    month_path = worklog_archive_month_path(d)
    if not month_path or not day_xlsx_path or not os.path.exists(day_xlsx_path):
        return None
    sheet_title = worklog_archive_sheet_title(d)
    legacy_title = _worklog_archive_legacy_sheet_title(d)
    if not allow_overwrite and worklog_date_exists_in_archive(d):
        return month_path if os.path.exists(month_path) else None
    _migrate_legacy_month_workbook(d, month_path)

    # 첫 월 파일: 일지 xlsx를 그대로 복사 → 인쇄 미리보기·열 너비 100% 유지
    if not os.path.exists(month_path):
        shutil.copy2(day_xlsx_path, month_path)
        month_wb = load_workbook(month_path)
        try:
            ws = month_wb.active
            if ws.title != sheet_title:
                ws.title = sheet_title
            month_wb.save(month_path)
        finally:
            month_wb.close()
        _cleanup_legacy_day_archive_files(d, month_path)
        return month_path

    day_wb = load_workbook(day_xlsx_path)
    try:
        src_ws = day_wb.active
        month_wb = load_workbook(month_path)
        try:
            # 레거시 YYYY-MM-DD 시트 → 일(27) 시트로 통일
            if legacy_title != sheet_title and legacy_title in month_wb.sheetnames:
                del month_wb[legacy_title]
            _clone_worksheet_to_workbook(src_ws, month_wb, sheet_title)
            # 시트 이름 일자순 정렬 (일-only·레거시 ISO 모두)
            names = [
                n
                for n in month_wb.sheetnames
                if n != sheet_title and not n.startswith("_")
            ]
            names.append(sheet_title)
            names.sort(key=_archive_sheet_sort_key)
            for i, name in enumerate(names):
                month_wb.move_sheet(name, offset=i - month_wb.sheetnames.index(name))
            month_wb.save(month_path)
        finally:
            month_wb.close()
    finally:
        day_wb.close()

    # 예전 일자별 파일·구 월폴더 경로 정리
    _cleanup_legacy_day_archive_files(d, month_path)
    return month_path


def delete_worklog_archive_sheet(d: date) -> str | None:
    """월별 파일에서 해당 날짜 시트 삭제. 시트가 없으면 파일 삭제."""
    if load_workbook is None:
        return None
    month_path = worklog_archive_month_path(d)
    if not month_path:
        return None
    removed = None
    # 레거시 일자 파일도 삭제
    _cleanup_legacy_day_archive_files(d, month_path)
    if not os.path.exists(month_path):
        return removed
    try:
        wb = load_workbook(month_path)
        try:
            for sheet_title in _worklog_archive_sheet_titles_for_lookup(d):
                if sheet_title in wb.sheetnames:
                    del wb[sheet_title]
                    removed = month_path
            remaining = [n for n in wb.sheetnames if n and not n.startswith("_")]
            if not remaining:
                wb.close()
                try:
                    os.remove(month_path)
                except OSError:
                    pass
                return month_path
            wb.save(month_path)
        finally:
            try:
                wb.close()
            except Exception:
                pass
    except Exception:
        pass
    return removed


def worklog_path(d: date) -> str: return os.path.join(WORKLOG_DIR, f"{d.isoformat()}.xlsx")

def list_saved_worklog_dates() -> set[str]:
    cached = st.session_state.get("wl_saved_dates_cache")
    if isinstance(cached, set): return cached
    _ensure_dirs()
    out: set[str] = set()
    for name in os.listdir(WORKLOG_DIR):
        if name.endswith(".xlsx") and len(name) >= 15 and name[0:4].isdigit() and name not in {"template.xlsx"} and not name.startswith("_preview_") and "_인쇄" not in name:
            out.add(name.replace(".xlsx", ""))
    st.session_state["wl_saved_dates_cache"] = out
    return out

def format_worklog_date(d: date) -> str:
    weeks = "월화수목금토일"
    return f"{d.strftime('%Y-%m-%d')} ({weeks[d.weekday()]})"

def _clear_content_cells(ws) -> None:
    for r in WL_CLIENT_ROWS:
        try: ws.cell(r, 3).value = None
        except AttributeError: pass
    for r in WL_CONTENT_ROWS:
        try: ws.cell(r, 7).value = None
        except AttributeError: pass
    for r in WL_NEXT_ROWS + WL_NOTE_ROWS:
        try: ws.cell(r, 4).value = None
        except AttributeError: pass

def _empty_cells(d: date) -> dict:
    cells = {"date": format_worklog_date(d)}
    for r in WL_CLIENT_ROWS: cells[f"C{r}"] = ""
    for r in WL_CONTENT_ROWS: cells[f"G{r}"] = ""
    for r in WL_NEXT_ROWS + WL_NOTE_ROWS: cells[f"D{r}"] = ""
    return cells

def read_worklog_cells(d: date) -> dict:
    path = worklog_path(d)
    if not os.path.exists(path) or load_workbook is None: return _empty_cells(d)
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    cells = {"date": format_worklog_date(d)}
    for r in WL_CLIENT_ROWS:
        v = ws.cell(r, 3).value
        cells[f"C{r}"] = "" if v is None else str(v)
    for r in WL_CONTENT_ROWS:
        v = ws.cell(r, 7).value
        cells[f"G{r}"] = "" if v is None else str(v)
    for r in WL_NEXT_ROWS + WL_NOTE_ROWS:
        v = ws.cell(r, 4).value
        cells[f"D{r}"] = "" if v is None else str(v)
    c_date = ws[WL_DATE_CELL].value
    if c_date is not None and not str(c_date).startswith("="): cells["date"] = str(c_date)
    wb.close()
    return cells

# 💡 강제 템플릿 덮어쓰기 로직 적용
def write_cells_to_path(path: str, d: date, cells: dict, *, force_template: bool = False) -> None:
    if load_workbook is None: raise RuntimeError("openpyxl 이 필요합니다.")
    _ensure_dirs()
    if force_template or not os.path.exists(path):
        if not os.path.exists(WORKLOG_TEMPLATE): raise FileNotFoundError("업무일지 템플릿이 없습니다.")
        shutil.copy2(WORKLOG_TEMPLATE, path)
    wb = load_workbook(path)
    ws = wb.active
    if force_template: _clear_content_cells(ws)
    try:
        date_cell = ws[WL_DATE_CELL]
        date_cell.value = cells.get("date") or format_worklog_date(d)
        _set_body_font(date_cell)
    except AttributeError:
        pass

    for r in WL_CLIENT_ROWS:
        try:
            cell = ws.cell(r, 3)
            cell.value = (cells.get(f"C{r}", "") or None)
            _set_body_font(cell)
            try: cell.alignment = cell.alignment.copy(horizontal="center", vertical="center", wrapText=False)
            except Exception: pass
        except AttributeError: pass
    for r in WL_CONTENT_ROWS:
        try:
            cell = ws.cell(r, 7)
            cell.value = (cells.get(f"G{r}", "") or None)
            _set_body_font(cell)
            try: cell.alignment = cell.alignment.copy(wrapText=False, shrinkToFit=False)
            except Exception: pass
        except AttributeError: pass
    for r in WL_NEXT_ROWS + WL_NOTE_ROWS:
        try:
            cell = ws.cell(r, 4)
            cell.value = (cells.get(f"D{r}", "") or None)
            _set_body_font(cell)
            try: cell.alignment = cell.alignment.copy(wrapText=False, shrinkToFit=False)
            except Exception: pass
        except AttributeError: pass
    wb.save(path)
    wb.close()

def save_worklog_cells(d: date, cells: dict, *, force: bool = False, allow_overwrite: bool = False) -> str:
    """저장: 로컬 캐시 + 월별 일지 + Drive + (가능하면) Cloud Gist.

    allow_overwrite=False(기본) — 해당 날짜가 이미 있으면 후입력 저장 차단.
    force=True — Drive/Cloud push 시 원격 덮어쓰기(기존 일지 수정 시).
    """
    ok, block_msg = check_worklog_save_allowed(d, had_local_at_open=allow_overwrite)
    if not ok:
        raise WorklogSaveBlockedError(block_msg)
    path = worklog_path(d)
    cells = _spill_all_content(cells)
    write_cells_to_path(path, d, cells, force_template=True)
    _invalidate_saved_dates_cache()
    _invalidate_worklog_presence_cache(d)
    try:
        from worklog_remote_sync import clear_worklog_day_deleted, invalidate_gist_days_cache
        clear_worklog_day_deleted(d.isoformat(), WORKLOG_DIR)
        invalidate_gist_days_cache()
    except Exception:
        pass
    st.session_state.pop("wl_last_archive_path", None)
    st.session_state.pop("wl_last_archive_err", None)
    st.session_state.pop("wl_last_drive_path", None)
    st.session_state.pop("wl_last_drive_month_path", None)
    st.session_state.pop("wl_last_drive_conflict", None)
    st.session_state.pop("wl_last_cloud_gist", None)
    st.session_state.pop("wl_last_cloud_err", None)
    st.session_state.pop("wl_last_archive_sheet", None)
    st.session_state.pop("wl_last_archive_target", None)
    try:
        archive = upsert_worklog_archive_sheet(d, path, allow_overwrite=allow_overwrite or force)
        if archive:
            st.session_state["wl_last_archive_path"] = archive
            st.session_state["wl_last_archive_sheet"] = worklog_archive_sheet_title(d)
            st.session_state["wl_last_archive_target"] = describe_worklog_archive_target(d)
    except Exception as e:
        st.session_state["wl_last_archive_err"] = str(e)
    try:
        from drive_autoload import push_worklog_day_to_drive, push_worklog_month_archive_to_drive

        drv = push_worklog_day_to_drive(path, WORKLOG_DIR, force=force)
        if drv:
            st.session_state["wl_last_drive_path"] = drv
        elif not force:
            pres = detect_worklog_date_presence(d)
            if pres.get("drive"):
                st.session_state["wl_last_drive_conflict"] = d.isoformat()
        arch_path = st.session_state.get("wl_last_archive_path")
        if arch_path and os.path.isfile(arch_path):
            mdrv = push_worklog_month_archive_to_drive(arch_path, year=d.year, force=force)
            if mdrv:
                st.session_state["wl_last_drive_month_path"] = mdrv
    except Exception:
        pass
    try:
        from worklog_remote_sync import push_worklog_day_remote, resolve_github_token

        if resolve_github_token():
            gid, cerr = push_worklog_day_remote(path, WORKLOG_DIR, force=force)
            if gid:
                st.session_state["wl_last_cloud_gist"] = gid
            elif cerr:
                st.session_state["wl_last_cloud_err"] = cerr
        else:
            st.session_state["wl_last_cloud_err"] = "github_token 없음 (secrets.toml 확인)"
    except Exception as e:
        st.session_state["wl_last_cloud_err"] = str(e)
    return path

def delete_worklog_day(d: date) -> list[str]:
    _ensure_dirs()
    iso = d.isoformat()
    removed: list[str] = []
    targets = [worklog_path(d), os.path.join(WORKLOG_DIR, f"_preview_{iso}.xlsx"), os.path.join(WORKLOG_DIR, f"일일업무일지_{iso}_인쇄.xlsx")]
    try:
        from drive_autoload import resolve_drive_worklog_dir
        _ddr = resolve_drive_worklog_dir()
        if _ddr: targets.append(os.path.join(_ddr, f"{iso}.xlsx"))
    except Exception: pass
    for name in os.listdir(WORKLOG_DIR):
        if iso in name and name.endswith(".xlsx") and name != "template.xlsx": targets.append(os.path.join(WORKLOG_DIR, name))
    seen: set[str] = set()
    for path in targets:
        if path in seen: continue
        seen.add(path)
        if os.path.exists(path):
            try: os.remove(path); removed.append(os.path.basename(path))
            except OSError: pass
    # 월별 통합 파일에서 해당 날짜 시트 제거
    try:
        arch = delete_worklog_archive_sheet(d)
        if arch:
            removed.append(f"{os.path.basename(arch)}#{iso}")
    except Exception:
        pass
    # Cloud/Drive에서도 삭제 — 동기화가 구 파일을 되살리는 원인
    cloud_note = ""
    try:
        from worklog_remote_sync import delete_worklog_day_remote, mark_worklog_day_deleted

        mark_worklog_day_deleted(iso, WORKLOG_DIR)
        ok, cerr = delete_worklog_day_remote(d, WORKLOG_DIR)
        if ok:
            removed.append(f"Cloud:{iso}.xlsx")
        elif cerr and cerr not in ("github_token 없음", "gist 없음"):
            cloud_note = f" Cloud:{cerr}"
    except Exception as e:
        try:
            from worklog_remote_sync import mark_worklog_day_deleted
            mark_worklog_day_deleted(iso, WORKLOG_DIR)
        except Exception:
            pass
        cloud_note = f" Cloud:{e}"
    try:
        from drive_autoload import delete_worklog_day_from_drive

        for dn in delete_worklog_day_from_drive(d, WORKLOG_DIR):
            removed.append(f"Drive:{dn}")
    except Exception:
        pass
    _invalidate_saved_dates_cache()
    _invalidate_worklog_presence_cache(d)
    try:
        from worklog_remote_sync import invalidate_gist_days_cache
        invalidate_gist_days_cache()
    except Exception:
        pass
    _clear_date_widget_state(d)
    st.session_state.pop(f"wl_open_ctx_{iso}", None)
    st.session_state.pop(f"wl_saved_ok_{iso}", None)
    empty = [{"client": "", "content": "", "lines": [], "blank_after": 1}]
    st.session_state[_boot_key(d)] = True
    st.session_state[_entries_key(d)] = empty
    st.session_state[_next_key(d)] = ""
    st.session_state[_notes_key(d)] = ""
    st.session_state[f"wl_entry_count_{iso}"] = 1
    msg = f"삭제 완료" + (f": {', '.join(removed)}" if removed else " (저장본 없음, 입력만 초기화)") + cloud_note
    st.session_state[f"wl_pending_sync_{iso}"] = {"entries": empty, "next": "", "notes": "", "msg": msg}
    try: _publish_view_cells(d, _empty_cells(d))
    except Exception: st.session_state.pop(_view_cells_key(d), None)
    st.session_state["_wl_drive_sync_ts"] = time.time()
    return removed

def reassign_worklog_date(old: date, new: date) -> str:
    if old == new: return "same"
    ok, block_msg = check_worklog_save_allowed(new, had_local_at_open=False)
    if not ok:
        raise FileExistsError(block_msg)
    try: cells = _cells_from_widgets(old)
    except Exception: cells = read_worklog_cells(old)
    cells["date"] = format_worklog_date(new)
    old_saved = os.path.exists(worklog_path(old))
    if old_saved or any(str(cells.get(f"G{r}", "") or "").strip() or str(cells.get(f"C{r}", "") or "").strip() for r in WL_CONTENT_ROWS) or any(str(cells.get(f"D{r}", "") or "").strip() for r in WL_NEXT_ROWS + WL_NOTE_ROWS):
        save_worklog_cells(new, cells, force=False, allow_overwrite=False)
    if old_saved:
        for path in (worklog_path(old), _preview_path(old), _print_xlsx_path(old)):
            if os.path.exists(path):
                try: os.remove(path)
                except OSError: pass
    entries = _grouped_entries_from_cells(cells) or [{"client": "", "content": "", "lines": [], "blank_after": 1}]
    _, nd, nt = _entries_from_cells(cells)
    _clear_date_widget_state(old)
    st.session_state["worklog_selected"] = new
    st.session_state["worklog_month"] = date(new.year, new.month, 1)
    st.session_state[_boot_key(new)] = True
    st.session_state[_entries_key(new)] = entries
    st.session_state[_next_key(new)] = "\n".join(nd)
    st.session_state[_notes_key(new)] = "\n".join(nt)
    st.session_state[f"wl_entry_count_{new.isoformat()}"] = len(entries)
    st.session_state[f"wl_pending_sync_{new.isoformat()}"] = {"entries": entries, "next": "\n".join(nd), "notes": "\n".join(nt), "msg": ""}
    _invalidate_saved_dates_cache()
    return "moved" if old_saved else "retargeted"

def _cell_fill_color(cell) -> str | None:
    try:
        fill = cell.fill
        if not fill or fill.fill_type is None: return None
        fg = fill.fgColor
        if fg is None: return None
        if getattr(fg, "type", None) == "rgb" and fg.rgb and fg.rgb != "00000000":
            rgb = str(fg.rgb)
            if len(rgb) == 8: rgb = rgb[2:]
            return f"#{rgb}"
    except Exception: return None
    return None

def _border_css(cell) -> str:
    parts = []
    try:
        b = cell.border
        for side_name, css_side in (("left", "border-left"), ("right", "border-right"), ("top", "border-top"), ("bottom", "border-bottom")):
            side = getattr(b, side_name, None)
            if side and side.style:
                color = "#111"
                try:
                    c = side.color
                    if c is not None and getattr(c, "type", None) == "rgb" and c.rgb:
                        rgb = str(c.rgb)
                        if len(rgb) == 8: rgb = rgb[2:]
                        if rgb and rgb != "00000000": color = f"#{rgb}"
                except Exception: pass
                parts.append(f"{css_side}:1px solid {color};")
    except Exception: pass
    return "".join(parts)

def _excel_col_width(ws, col_idx: int) -> float:
    best = None
    for dim in ws.column_dimensions.values():
        if dim.min is None or dim.max is None or dim.width is None: continue
        if dim.min <= col_idx <= dim.max:
            span = dim.max - dim.min
            if best is None or span < best[0]: best = (span, float(dim.width))
    if best: return best[1]
    return float(ws.sheet_format.defaultColWidth or 8.43)

def _excel_row_height_px(ws, row: int) -> int:
    """엑셀 행 높이 → px (인위적 늘리기 없음)."""
    h = ws.row_dimensions[row].height
    if h:
        return max(1, int(round(float(h) * 96 / 72)))
    return 20  # Excel 기본 ~15pt

def _wl_cell_font(cell, *, is_content: bool, is_client: bool, is_body_d: bool, is_date: bool) -> tuple[str, float]:
    """셀 font → (font-stack, pt). 본문 계열은 바탕체·엑셀 pt 그대로."""
    font = cell.font
    force_batang = is_content or is_client or is_body_d or is_date
    fname = (font.name or "").strip()
    if force_batang or fname in ("바탕", "바탕체", "바탕글", "Batang", "BatangChe"):
        stack = _WL_FONT_STACK
    elif fname in ("맑은 고딕", "Malgun Gothic"):
        stack = "'Malgun Gothic','Apple SD Gothic Neo',sans-serif"
    elif fname:
        stack = f"'{html.escape(fname)}',{_WL_FONT_STACK}"
    else:
        stack = _WL_FONT_STACK
    if font.size:
        fsize_pt = float(font.size)
    elif force_batang:
        fsize_pt = float(_WL_BODY_FONT_PT)
    else:
        fsize_pt = 11.0
    return stack, fsize_pt

def _excel_width_to_px(width: float) -> int:
    """엑셀 열 너비 → px (화면 100% 기준, 원본과 동일 체감)."""
    try: w = float(width)
    except (TypeError, ValueError): w = 8.43
    return max(10, int(w * 7 + 5))

def _worklog_sheet_pixel_size(path: str) -> tuple[int, int]:
    if load_workbook is None or not path or not os.path.exists(path): return 900, 1312
    try:
        wb = load_workbook(path, data_only=False)
        ws = wb.active
        total_w = sum(_excel_width_to_px(_excel_col_width(ws, c)) for c in range(WL_MIN_COL, WL_MAX_COL + 1))
        total_h = 0
        for r in range(WL_MIN_ROW, WL_MAX_ROW + 1):
            total_h += _excel_row_height_px(ws, r) + 1
        # 원본 로고(특이사항 하단, Excel 표시 높이 61 + 간격)
        total_h += 61 + 10
        wb.close()
        return max(1, total_w), max(1, total_h)
    except Exception: return 900, 1312


def _excel_page_scale(path: str) -> float | None:
    """원본.xlsx pageSetup.scale(예: 69) → 0.69. 없으면 None."""
    if load_workbook is None or not path or not os.path.exists(path):
        return None
    try:
        wb = load_workbook(path, data_only=False)
        ws = wb.active
        sc = ws.page_setup.scale
        wb.close()
        if sc is None:
            return None
        v = float(sc)
        if v > 1.5:  # Excel stores percent (69)
            v = v / 100.0
        if 0.3 <= v <= 1.5:
            return v
    except Exception:
        return None
    return None


def _excel_page_margins_css(path: str) -> str:
    """원본 pageMargins → @page margin CSS (인치). 기본 0.55in 0.51in."""
    top = right = bottom = left = None
    if load_workbook is not None and path and os.path.exists(path):
        try:
            wb = load_workbook(path, data_only=False)
            ws = wb.active
            pm = ws.page_margins
            top, right, bottom, left = pm.top, pm.right, pm.bottom, pm.left
            wb.close()
        except Exception:
            pass
    # 원본.xlsx 기본값(인치)
    t = float(top if top is not None else 0.55)
    r = float(right if right is not None else 0.51)
    b = float(bottom if bottom is not None else 0.55)
    l = float(left if left is not None else 0.51)
    return f"{t:.2f}in {r:.2f}in {b:.2f}in {l:.2f}in"


def _excel_print_scale(path: str | None = None) -> float:
    """인쇄 배율 = 원본 Excel pageSetup.scale 그대로(추가 축소 없음)."""
    s = _excel_page_scale(path) if path else None
    if s is not None:
        return float(s)
    return 0.69  # 원본.xlsx 기본


def _scaled_view_frame_size(path: str, scale: float) -> tuple[int, int]:
    w, h = _worklog_sheet_pixel_size(path)
    s = float(scale) if scale and scale > 0 else 1.0
    return int(w * s) + 28, int(h * s) + 64

def _worklog_logo_bytes(xlsx_path: str | None = None) -> tuple[bytes, str, int, int] | None:
    """원본 템플릿 내장 로고(320×61) 우선. (bytes, mime, w, h)"""
    _here = os.path.dirname(os.path.abspath(__file__))
    if xlsx_path and os.path.exists(xlsx_path):
        try:
            import zipfile
            with zipfile.ZipFile(xlsx_path) as zf:
                for name in ("xl/media/image1.jpeg", "xl/media/image1.jpg", "xl/media/image1.png"):
                    if name in zf.namelist():
                        data = zf.read(name)
                        mime = "image/jpeg" if name.endswith((".jpeg", ".jpg")) else "image/png"
                        return data, mime, 320, 61  # 원본 Excel 표시 크기
        except Exception:
            pass
    for p, mime in (
        (os.path.join(WORKLOG_DIR, "shinil_logo.jpeg"), "image/jpeg"),
        (os.path.join(_here, "logo.png"), "image/png"),
        ("logo.png", "image/png"),
    ):
        if os.path.exists(p):
            try:
                with open(p, "rb") as f:
                    return f.read(), mime, 320, 61
            except Exception:
                continue
    return None


def workbook_to_html(path: str, *, include_logo: bool = True, layout_scale: float = 1.0) -> str:
    if load_workbook is None: return "<p>openpyxl 필요</p>"
    wb = load_workbook(path, data_only=False)
    ws = wb.active

    merge_map: dict[tuple[int, int], tuple[int, int]] = {}
    skip: set[tuple[int, int]] = set()
    for mr in ws.merged_cells.ranges:
        if mr.max_row < WL_MIN_ROW or mr.min_row > WL_MAX_ROW: continue
        if mr.max_col < WL_MIN_COL or mr.min_col > WL_MAX_COL: continue
        max_r, max_c = min(mr.max_row, WL_MAX_ROW), min(mr.max_col, WL_MAX_COL)
        min_r, min_c = max(mr.min_row, WL_MIN_ROW), max(mr.min_col, WL_MIN_COL)
        if min_r > max_r or min_c > max_c: continue
        rs, cs = max_r - min_r + 1, max_c - min_c + 1
        tl_r, tl_c = mr.min_row, mr.min_col
        if tl_r < WL_MIN_ROW or tl_c < WL_MIN_COL: tl_r, tl_c = min_r, min_c
        merge_map[(tl_r, tl_c)] = (rs, cs)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (tl_r, tl_c): skip.add((r, c))

    col_widths = []
    total_w = 0.0
    ls = float(layout_scale) if layout_scale and layout_scale > 0 else 1.0
    for c in range(WL_MIN_COL, WL_MAX_COL + 1):
        w = _excel_col_width(ws, c)
        px = max(1, int(round(_excel_width_to_px(w) * ls)))
        col_widths.append(px)
        total_w += px

    rows_html = []
    for r in range(WL_MIN_ROW, WL_MAX_ROW + 1):
        height_px = max(1, int(round(_excel_row_height_px(ws, r) * ls)))
        tds = []
        for c in range(WL_MIN_COL, WL_MAX_COL + 1):
            if (r, c) in skip: continue
            cell = ws.cell(r, c)
            rs, cs = merge_map.get((r, c), (1, 1))
            val = cell.value
            text = "" if isinstance(val, str) and val.startswith("=") else ("" if val is None else str(val))

            if c == 3 and r == WL_NEXT_ROWS[0] and not text.strip(): text = "익일업무"
            if c == 3 and r == WL_NOTE_ROWS[0] and not text.strip(): text = "특이사항"

            is_content = c == WL_CONTENT_COL_START and WL_CONTENT_ROWS[0] <= r <= WL_CONTENT_ROWS[-1]
            is_client = c == WL_CLIENT_COL_START and WL_CLIENT_ROWS[0] <= r <= WL_CLIENT_ROWS[-1]
            is_body_d = c == 4 and (r in WL_NEXT_ROWS or r in WL_NOTE_ROWS)
            is_date = (c == 3 and r == 5) or (str(cell.coordinate) == WL_DATE_CELL)
            font_stack, fsize_pt = _wl_cell_font(cell, is_content=is_content, is_client=is_client, is_body_d=is_body_d, is_date=is_date)
            # 인쇄 시 Excel pageSetup.scale을 글자 pt·셀 크기에 미리 반영(브라우저 추가 축소 방지)
            if ls != 1.0:
                fsize_pt = round(float(fsize_pt) * ls, 2)
            font = cell.font
            bold = "bold" if font.bold else "normal"
            align = cell.alignment
            ha = align.horizontal or "left"
            va = align.vertical or "middle"
            if ha == "general": ha = "left"
            if va == "center": va = "middle"

            _text_clean = str(text).replace(" ", "").replace("\u3000", "").replace("\n", "") if text else ""
            is_side_label = c == 3 and (_text_clean in ("익일업무", "특이사항") or (r == WL_NEXT_ROWS[0] and rs >= 3) or (r == WL_NOTE_ROWS[0] and rs >= 3))
            is_vertical = is_side_label or getattr(align, "textRotation", 0) in (255, 90) or (_text_clean == "결재" and rs >= 2 and cs == 1)

            if is_vertical: ha, va = "center", "middle"
            elif is_client: ha, va = "center", "middle"  # 거래처 칸 항상 가운데
            elif is_content: ha = "left"
                
            fill = _cell_fill_color(cell) or "#FFFFFF"
            border = _border_css(cell)
            span = ""
            if rs > 1: span += f' rowspan="{rs}"'
            if cs > 1: span += f' colspan="{cs}"'
                
            if text in (_WL_SOFT_BLANK, "\u00a0"): text = ""
            elif is_content and text.strip() == "" and text != "": text = ""
                
            if is_vertical and text.strip():
                chars = [ch for ch in _text_clean if ch]
                esc = "<br>".join(html.escape(ch) for ch in chars)
            else:
                esc = html.escape(text).replace(" ", "&nbsp;").replace("\n", "<br>")
                
            if is_content or is_client: white, overflow, text_overflow = "nowrap", "visible", "clip"
            elif is_vertical: white, overflow, text_overflow = "normal", "hidden", "clip"
            else: white, overflow, text_overflow = "pre-wrap", "visible", "clip"
                
            c0 = c - WL_MIN_COL
            span_w = sum(col_widths[c0 : c0 + max(cs, 1)]) if c0 >= 0 else 0
            width_css = f"width:{span_w}px;min-width:{span_w}px;max-width:{span_w}px;" if span_w else ""

            if is_vertical:
                pad_css = "padding:4px 1px;"
            else:
                pad_css = "padding:0;"
            
            if r < WL_CLIENT_ROWS[0] and not is_vertical:
                style = (
                    f"box-sizing:border-box;{width_css}"
                    f"font-family:{font_stack};font-size:{fsize_pt}pt;font-weight:{bold};"
                    f"text-align:center; vertical-align:middle;"
                    f"background:{fill}; {border} {pad_css}"
                )
            else:
                line_css = "line-height:1.35;" if is_vertical else "line-height:1.25;"
                
                style = (
                    f"box-sizing:border-box;{width_css}"
                    f"font-family:{font_stack};font-size:{fsize_pt}pt;font-weight:{bold};"
                    f"text-align:{ha};vertical-align:{va};"
                    f"background:{fill};{border}"
                    f"{pad_css}white-space:{white};overflow:{overflow};"
                    f"text-overflow:{text_overflow};word-break:keep-all;"
                    f"{line_css}"
                )
            tds.append(f'<td{span} style="{style}">{esc}</td>')
        rows_html.append(f'<tr style="height:{height_px}px;box-sizing:border-box;">{"".join(tds)}</tr>')

    colgroup = "".join(f'<col style="width:{w}px">' for w in col_widths)
    wb.close()
    
    import base64
    logo_tr = ""
    # 원본: 특이사항(47행) 바로 아래 중앙, Excel 표시 크기 320×61
    if include_logo:
        logo = _worklog_logo_bytes(path)
        if logo:
            try:
                data, mime, nat_w, nat_h = logo
                b64_img = base64.b64encode(data).decode("utf-8")
                ls = float(layout_scale) if layout_scale and layout_scale > 0 else 1.0
                logo_w = max(1, int(round(nat_w * ls)))
                logo_h = max(1, int(round(nat_h * ls)))
                pad_top = max(4, int(round(8 * ls)))  # 원본과 비슷한 위쪽 간격
                logo_tr = (
                    f'<tr style="border:none;background:#fff;height:{logo_h + pad_top + 2}px;">'
                    f'<td colspan="26" style="text-align:center;padding:{pad_top}px 0 0 0;border:none;vertical-align:top;">'
                    f'<img src="data:{mime};base64,{b64_img}" alt="신일가스 로고" '
                    f'width="{logo_w}" height="{logo_h}" '
                    f'style="width:{logo_w}px;height:{logo_h}px;max-width:none;display:block;margin:0 auto;">'
                    "</td></tr>"
                )
            except Exception:
                pass

    return f"""
    <table class="wl-sheet" style="border-collapse:collapse;table-layout:fixed;width:{int(total_w)}px;background:#fff;box-sizing:border-box;">
      <colgroup>{colgroup}</colgroup>
      <tbody>
        {"".join(rows_html)}
        {logo_tr}
      </tbody>
    </table>
    """

def _a4_print_fit(raw_w: int, raw_h: int, *, path: str | None = None) -> float:
    """인쇄 배율 = 원본 Excel pageSetup.scale 그대로(추가 맞춤 축소 없음)."""
    return _excel_print_scale(path)

def render_worklog_view_html(path: str, *, print_mode: bool = False, scale: float | None = None, auto_print: bool = False, wrap_height: str | None = None) -> str:
    # 인쇄: Excel 배율(69%)을 HTML 글씨·셀에 직접 반영 → CSS transform으로 또 줄이지 않음
    excel_print_s = _excel_print_scale(path)
    layout_s = float(excel_print_s) if print_mode else 1.0
    sheet = workbook_to_html(path, include_logo=True, layout_scale=layout_s)
    raw_w, raw_h = _worklog_sheet_pixel_size(path)
    print_fit = excel_print_s
    # print_mode에서는 이미 layout_scale 적용된 크기
    if print_mode:
        scaled_w = max(1, int(round(raw_w * layout_s)))
        scaled_h = max(1, int(round(raw_h * layout_s)))
        page_margins = _excel_page_margins_css(path)
        pct = int(round(excel_print_s * 100))
    else:
        scaled_w, scaled_h = max(1, int(round(raw_w * print_fit))), max(1, int(round(raw_h * print_fit)))
        page_margins = "8mm"
        pct = int(round(excel_print_s * 100))

    toolbar = ""
    if print_mode:
        view_scale = 1.0
        toolbar = f"""<div class="toolbar no-print"><button type="button" id="wl-print-btn">인쇄하기</button><span class="hint">원본 Excel 인쇄 설정 그대로입니다 (배율 {pct}%, 여백 동일). 브라우저 인쇄창에서 「용지에 맞춤」을 끄세요.</span></div>"""
    else:
        view_scale = float(scale if scale is not None else _WL_PREVIEW_SCALE)
    frame_w, frame_h = _scaled_view_frame_size(path, view_scale)
    if print_mode:
        # 이미 69%로 렌더됨 → 추가 zoom/transform 없음
        scale_css = f"zoom:1;width:{scaled_w}px;max-width:none;"
        scale_css_fallback = f"transform:none;width:{scaled_w}px;"
        wrap_h, wrap_w, wrap_overflow, body_overflow, body_h = f"{scaled_h}px", f"{scaled_w}px", "hidden", "auto", "auto"
    elif view_scale >= 1:
        scale_css = "zoom:1;width:fit-content;"
        scale_css_fallback = "transform:scale(1);transform-origin:top left;width:fit-content;"
        wrap_h, wrap_w, wrap_overflow, body_overflow, body_h = "auto", "100%", "visible", "visible", "auto"
    else:
        s = float(view_scale)
        scale_css = f"zoom:{s};width:fit-content;"
        scale_css_fallback = f"transform:scale({s});transform-origin:top left;margin-bottom:{(s - 1) * raw_h:.1f}px;width:fit-content;"
        wrap_h, wrap_w, wrap_overflow, body_overflow, body_h = "auto", f"{frame_w}px", "visible", "hidden", f"{frame_h}px"
    if wrap_height is not None: wrap_h = wrap_height

    fit_print_js = f"""
        var wlZoom = 1;
        var wlOrigFit = 1;
        function wlApplyZoom(s) {{
          var sheet = document.querySelector('.sheet-scale');
          var wrap = document.querySelector('.wrap');
          var table = document.querySelector('.wl-sheet');
          if (!sheet || !wrap || !table) return;
          wlZoom = 1;
          sheet.style.transform = 'none';
          sheet.style.zoom = '1';
          wrap.style.maxWidth = 'none';
          wrap.style.width = '{scaled_w}px';
          wrap.style.height = '{scaled_h}px';
          wrap.style.overflow = 'hidden';
          wrap.style.margin = '0 auto';
        }}
        function wlFitToA4() {{ wlApplyZoom(1); }}
        try {{ window.wlFitToA4 = wlFitToA4; }} catch (e0) {{}}
    """
    go_print_js = """
        function goPrint() {
          try { wlFitToA4(); } catch (e0) {}
          var fire = function() {
            setTimeout(function() {
              try { window.focus(); window.print(); } catch (e) {}
            }, 150);
          };
          if (document.fonts && document.fonts.ready) {
            document.fonts.ready.then(fire).catch(fire);
          } else {
            setTimeout(fire, 400);
          }
        }
    """
    auto_script = f"""<script>(function() {{ {fit_print_js} {go_print_js} var btn = document.getElementById('wl-print-btn'); if (btn) btn.addEventListener('click', function(ev) {{ ev.preventDefault(); goPrint(); }}); window.addEventListener('beforeprint', function() {{ try {{ wlFitToA4(); }} catch (e2) {{}} }}); function boot() {{ try {{ wlFitToA4(); }} catch (e3) {{}} {"setTimeout(goPrint, 500);" if auto_print else ""} }} if (document.readyState === 'complete') setTimeout(boot, 250); else window.addEventListener('load', function() {{ setTimeout(boot, 250); }}); }})();</script>""" if print_mode else ""
    fallback_block = f"@supports not (zoom: 1) {{ .sheet-scale {{ {scale_css_fallback} }} }}" if scale_css_fallback else ""
    if print_mode:
        print_media = f"""@media print {{ html, body {{ overflow:visible !important; height:auto !important; width:auto !important; margin:0 !important; padding:0 !important; -webkit-print-color-adjust:exact; print-color-adjust:exact; }} .no-print, .toolbar {{ display:none !important; }} .wrap {{ overflow:visible !important; max-width:none !important; width:{scaled_w}px !important; height:auto !important; border:none !important; margin:0 auto !important; padding:0 !important; }} .sheet-scale {{ zoom:1 !important; transform:none !important; width:{scaled_w}px !important; margin:0 !important; }} .wl-sheet, .wl-sheet td, .wl-sheet tr {{ font-family:{_WL_FONT_STACK} !important; -webkit-print-color-adjust:exact; print-color-adjust:exact; }} }}"""
    else:
        print_media = "@media print { html, body { overflow:visible !important; } }"
    
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>일일업무일지</title><meta name="viewport" content="width=device-width, initial-scale=1"><link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Nanum+Myeongjo:wght@400;700&display=swap" rel="stylesheet"><style>{_WL_FONT_FACE_CSS} @page {{ size: A4 portrait; margin: {page_margins}; }} html, body {{ margin:0; padding:0; background:#fff; overflow:{body_overflow} !important; height:{body_h}; }} body {{ padding:{"6px" if not print_mode else "0"}; box-sizing:border-box; font-family:{_WL_FONT_STACK} !important; }} .toolbar {{ margin-bottom:10px; display:flex; gap:10px; align-items:center; flex-wrap:wrap; }} .toolbar button {{ padding:8px 14px; font-size:14px; border:1px solid #334155; border-radius:6px; background:#1E293B; color:#fff; cursor:pointer; }} .toolbar button.secondary {{ background:#F8FAFC; color:#334155; border-color:#CBD5E1; cursor:default; }} .toolbar .hint {{ font:12px/1.45 sans-serif; color:#64748B; max-width:42rem; }} .wrap {{ overflow:{wrap_overflow} !important; height:{wrap_h}; width:{wrap_w}; max-width:{"none" if print_mode else "100%"}; border:{"none" if print_mode else "1px solid #94A3B8"}; background:#fff; box-sizing:border-box; padding:0; }} .sheet-scale {{ {scale_css} }} .wl-sheet {{ border-collapse:collapse; table-layout:fixed; font-family:{_WL_FONT_STACK} !important; }} .wl-sheet, .wl-sheet td, .wl-sheet tr {{ box-sizing:border-box; font-family:{_WL_FONT_STACK} !important; }} {fallback_block} {print_media}</style></head><body>{toolbar}<div class="wrap"><div class="sheet-scale">{sheet}</div></div>{auto_script}</body></html>"""

def _entry_blank_after(ent: dict | None, default: int = 1) -> int:
    try: n = int((ent or {}).get("blank_after", default))
    except (TypeError, ValueError): n = default
    return max(0, min(10, n))

_WL_SOFT_BLANK = " "

def _grouped_entries_from_cells(cells: dict) -> list[dict]:
    entries: list[dict] = []
    blank_run = 0
    for r in WL_CLIENT_ROWS:
        raw_c, raw_g = str(cells.get(f"C{r}", "") or ""), str(cells.get(f"G{r}", "") or "")
        client_raw = _scrub_dummy_label(raw_c)
        client_stripped = client_raw.strip()
        soft_blank = raw_g == _WL_SOFT_BLANK or raw_g == "\u00a0"
        g_whitespace_only = (not client_stripped and not soft_blank and raw_g != "" and raw_g.strip() == "")
        fully_empty = (not client_stripped and not soft_blank and not g_whitespace_only and raw_c.strip() == "" and raw_g.strip() == "")
        content = "" if (soft_blank or g_whitespace_only) else _scrub_dummy_label(raw_g)
        
        if fully_empty: 
            blank_run += 1
            continue
            
        if g_whitespace_only or (soft_blank and not client_stripped):
            if not entries or blank_run > 0:
                if entries: entries[-1]["blank_after"] = max(0, min(10, blank_run))
                blank_run = 0
                entries.append({
                    "client": client_raw, 
                    "client_lines": [client_raw], 
                    "content": "", 
                    "lines": [""], 
                    "blank_after": 1
                })
            else:
                blank_run = 0
                entries[-1].setdefault("lines", []).append("")
                entries[-1].setdefault("client_lines", []).append(client_raw)
                entries[-1]["content"] = "\n".join(entries[-1].get("lines") or [])
                entries[-1]["client"] = "\n".join(entries[-1].get("client_lines") or [])
            continue
            
        if not entries or blank_run > 0:
            if entries: entries[-1]["blank_after"] = max(0, min(10, blank_run))
            blank_run = 0
            client_lines, lines = [client_raw], [content or ""]
            entries.append({
                "client": "\n".join(client_lines), 
                "client_lines": client_lines, 
                "content": "\n".join(lines), 
                "lines": lines, 
                "blank_after": 1
            })
        else:
            ent = entries[-1]
            cl = ent.setdefault("client_lines", [])
            cl.append(client_raw)
            ent["client"] = "\n".join(cl)
            ent.setdefault("lines", []).append(content or "")
            ent["content"] = "\n".join(ent.get("lines") or [])
    return entries

def _entry_client_lines(ent: dict | None) -> list[str]:
    if not ent: return []
    raw = ent.get("client_lines")
    if isinstance(raw, list) and raw: src = [str(x or "") for x in raw]
    else: 
        raw_c = str(ent.get("client") or "")
        src = raw_c.split('\n') if raw_c else []
    max_u, out = _client_line_units(), []
    for line in src:
        if not str(line or "").strip(): 
            out.append(str(line or ""))
            continue
        out.extend(_chunk_text(str(line or ""), max_u) or [str(line or "")])
    return out

def _entry_pack_lines(ent: dict) -> list[str]:
    max_u = _content_line_units()
    raw = ent.get("lines")
    src = [str(x or "") for x in raw] if isinstance(raw, list) else _chunk_text(str(ent.get("content") or ""), max_u) or []
    out: list[str] = []
    for line in src:
        if not str(line or "").strip(): 
            out.append(str(line or ""))
            continue
        out.extend(_chunk_text(line, max_u) or [line])
    if not out and _entry_client_lines(ent): out = [""]
    return out

def _content_row_usage(entries: list[dict] | None) -> dict:
    total, used, wrote_any, prev_gap, per_entry = len(WL_CONTENT_ROWS), 0, False, 0, []
    for ent in entries or []:
        clients, pack_lines = _entry_client_lines(ent), _entry_pack_lines(ent)
        if not any(str(x).strip() for x in clients) and not any((x or "").strip() for x in pack_lines): per_entry.append(0); continue
        gap = prev_gap if wrote_any else 0
        lines = max(len(clients), len(pack_lines), 1)
        need = gap + lines
        per_entry.append(need)
        used += need; wrote_any = True; prev_gap = _entry_blank_after(ent, 1)
    return {"total": total, "used": used, "remaining": max(0, total - used), "per_entry": per_entry, "last_row": WL_CONTENT_ROWS[-1] if WL_CONTENT_ROWS else 41, "next_row": WL_CONTENT_ROWS[used] if used < total else None, "overflow": used > total}

def _render_row_remain_gauge(usage: dict, *, height_px: int = 980) -> None:
    total, used, rem, overflow = max(1, int(usage.get("total") or 1)), max(0, int(usage.get("used") or 0)), max(0, int(usage.get("remaining") or 0)), bool(usage.get("overflow"))
    if overflow: accent, used_color, rem_color, label, big, rem_show, used_show = "#DC2626", "#FECACA", "#FEE2E2", "초과", f"+{used - total}", 0, total
    elif rem <= 3: accent, used_color, rem_color, label, big, rem_show, used_show = "#D97706", "#E2E8F0", "#FBBF24", "남음", str(rem), rem, min(used, total)
    else: accent, used_color, rem_color, label, big, rem_show, used_show = "#0F766E", "#E2E8F0", "#14B8A6", "남음", str(rem), rem, min(used, total)
    segs = [f'<div style="flex:1 1 0;min-height:3px;border-radius:3px;background:{used_color if i < used_show else rem_color};margin:0 0 {max(1, int(2 if total <= 20 else 1))}px 0;opacity:{0.55 if i < used_show else 1};"></div>' for i in range(total)]
    if segs: segs[-1] = segs[-1].replace(f"margin:0 0 {max(1, int(2 if total <= 20 else 1))}px 0;", "margin:0;")
    next_txt = f"다음 G{usage.get('next_row')}" if usage.get("next_row") else "칸 끝"
    st.markdown(f"""<div style="display:flex;flex-direction:column;align-items:center;justify-content:flex-start;gap:6px;padding:2px 2px 0;min-height:{height_px}px;height:{height_px}px;font-family:'Pretendard','Apple SD Gothic Neo',sans-serif;"><div style="text-align:center;line-height:1.1;flex:0 0 auto;"><div style="font-size:10px;font-weight:700;letter-spacing:.06em;color:#64748B;">{label}</div><div style="font-size:22px;font-weight:800;color:{accent};margin-top:1px;">{big}</div><div style="font-size:10px;color:#94A3B8;">칸</div></div><div style="display:flex;flex-direction:column;justify-content:flex-start;flex:1 1 auto;width:26px;min-height:{max(520, height_px - 120)}px;height:{max(520, height_px - 120)}px;padding:4px 3px;border-radius:10px;background:#F8FAFC;box-shadow:inset 0 0 0 1px #E2E8F0;" title="전체 {total}칸 · 사용 {used}칸 · 남음 {rem}칸">{"".join(segs)}</div><div style="text-align:center;line-height:1.25;flex:0 0 auto;"><div style="font-size:11px;font-weight:800;color:#0F172A;"><span style="color:{accent};">{rem_show}</span><span style="color:#94A3B8;font-weight:600;"> / {total}</span></div><div style="font-size:10px;color:#64748B;">남은칸 / 전체칸</div><div style="font-size:10px;color:#94A3B8;margin-top:2px;">사용 {used_show}칸 · {next_txt}</div></div></div>""", unsafe_allow_html=True)

def _pack_entries_to_cells(d: date, entries: list[dict], next_day: list[str] | None = None, notes: list[str] | None = None) -> dict:
    cells, max_u, row_i, rows, wrote_any, prev_gap = _empty_cells(d), _content_line_units(), 0, WL_CONTENT_ROWS, False, 0
    for ent in entries or []:
        clients, chunks = _entry_client_lines(ent), _entry_pack_lines(ent)
        if not any(str(x).strip() for x in clients) and not any((x or "").strip() for x in chunks): continue
        if wrote_any: row_i += max(0, int(prev_gap))
        for j in range(max(len(clients), len(chunks), 1)):
            if row_i >= len(rows): break
            r = rows[row_i]
            c_val = clients[j] if j < len(clients) else ""
            g_val = chunks[j] if j < len(chunks) else ""
            
            cells[f"C{r}"] = c_val
            if (j < len(chunks) and not str(chunks[j]).strip()) or (j < len(clients) and not str(clients[j]).strip() and j >= len(chunks)):
                cells[f"G{r}"] = _WL_SOFT_BLANK
            else:
                cells[f"G{r}"] = g_val
            row_i += 1
        wrote_any, prev_gap = True, _entry_blank_after(ent, 1)
    for t_list, t_rows in ((next_day, WL_NEXT_ROWS), (notes, WL_NOTE_ROWS)):
        chunks = []
        for t in t_list or []: chunks.extend(_chunk_text(t, max_u))
        for i, r in enumerate(t_rows): cells[f"D{r}"] = chunks[i] if i < len(chunks) else ""
    return cells

def _entries_from_cells(cells: dict) -> tuple[list[tuple[str, str]], list[str], list[str]]:
    rows = [(e.get("client") or "", e.get("content") or "") for e in _grouped_entries_from_cells(cells)]
    next_day = [x for x in [(cells.get(f"D{r}") or "").strip() for r in WL_NEXT_ROWS] if x]
    notes = [x for x in [(cells.get(f"D{r}") or "").strip() for r in WL_NOTE_ROWS] if x]
    return rows, next_day, notes

_WL_SUMMARY_PREVIEW_CSS = (
    "@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.min.css');"
    " .wl-sum-preview { font-family:'Pretendard','Apple SD Gothic Neo',sans-serif; color:#0F172A; }"
    " .wl-sum-preview .card { background:linear-gradient(180deg,#F8FAFC 0%,#FFFFFF 48px); border:1px solid #E2E8F0;"
    " border-radius:14px; box-shadow:0 1px 2px rgba(15,23,42,.04); overflow:hidden; }"
    " .wl-sum-preview .head { padding:16px 18px 14px; border-bottom:1px solid #E2E8F0;"
    " background:linear-gradient(135deg,#0F766E 0%,#0E7490 55%,#0369A1 100%); color:#fff; }"
    " .wl-sum-preview .head .title { font-size:18px; font-weight:750; letter-spacing:-.02em; }"
    " .wl-sum-preview .head .sub { margin-top:4px; font-size:13px; opacity:.92; }"
    " .wl-sum-preview .sec { padding:14px 16px 8px; }"
    " .wl-sum-preview .sec h3 { margin:0 0 10px; font-size:12px; font-weight:700; letter-spacing:.06em;"
    " color:#64748B; text-transform:uppercase; }"
    " .wl-sum-preview .item { display:flex; gap:12px; align-items:flex-start; padding:12px; margin-bottom:8px;"
    " background:#fff; border:1px solid #E2E8F0; border-radius:10px; }"
    " .wl-sum-preview .idx { flex:0 0 28px; height:28px; border-radius:8px; background:#CCFBF1; color:#0F766E;"
    " font-weight:700; font-size:13px; display:flex; align-items:center; justify-content:center; }"
    " .wl-sum-preview .client { font-size:15px; font-weight:700; color:#134E4A; margin-bottom:4px; white-space:pre-wrap; }"
    " .wl-sum-preview .content { font-size:14px; line-height:1.55; color:#334155; white-space:pre-wrap; word-break:break-word; }"
    " .wl-sum-preview .muted { color:#94A3B8; font-weight:500; }"
    " .wl-sum-preview .empty { padding:18px; text-align:center; color:#94A3B8; font-size:13px;"
    " border:1px dashed #CBD5E1; border-radius:10px; background:#F8FAFC; }"
    " .wl-sum-preview .panel { margin:0 16px 14px; padding:12px 14px; border-radius:12px; border:1px solid #E2E8F0; background:#fff; }"
    " .wl-sum-preview .panel.next { border-left:4px solid #2563EB; }"
    " .wl-sum-preview .panel.note { border-left:4px solid #D97706; }"
    " .wl-sum-preview .panel h3 { margin:0 0 8px; font-size:13px; font-weight:700; color:#1E293B; }"
    " .wl-sum-preview .line { display:flex; gap:8px; align-items:flex-start; padding:6px 0; font-size:14px;"
    " line-height:1.5; color:#334155; border-bottom:1px solid #F1F5F9; }"
    " .wl-sum-preview .line:last-child { border-bottom:none; }"
    " .wl-sum-preview .dot { width:7px; height:7px; margin-top:7px; border-radius:50%; background:#94A3B8; flex:0 0 auto; }"
    " .wl-sum-preview .panel.next .dot { background:#2563EB; }"
    " .wl-sum-preview .panel.note .dot { background:#D97706; }"
    " .wl-sum-preview .foot { padding:10px 16px 14px; font-size:11px; color:#94A3B8; }"
)

def render_readable_preview_html(d: date, cells: dict) -> str:
    rows, next_day, notes = _entries_from_cells(cells)
    date_label = html.escape(cells.get("date") or format_worklog_date(d))
    if rows:
        work_items = []
        for i, (client, content) in enumerate(rows, 1):
            c = html.escape(client).replace("\n", "<br>") if client else "<span class='muted'>(거래처 없음)</span>"
            t = html.escape(content) if content else "<span class='muted'>—</span>"
            work_items.append(f"""<div class="item"><div class="idx">{i}</div><div class="body"><div class="client">{c}</div><div class="content">{t}</div></div></div>""")
        work_html = "".join(work_items)
    else: work_html = "<div class='empty'>등록된 업무 내용이 없습니다.</div>"
    def _lines(items: list[str], empty_msg: str) -> str: return "".join(f"<div class='line'><span class='dot'></span><span>{html.escape(x)}</span></div>" for x in items) if items else f"<div class='empty'>{empty_msg}</div>"
    _iframe_css = _WL_SUMMARY_PREVIEW_CSS.replace(".wl-sum-preview ", "") + " html, body { margin:0; padding:0; background:transparent; } body { padding:4px; }"
    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>{_iframe_css}</style></head><body>'
        f'<div class="card"><div class="head"><div class="title">일일업무일지</div><div class="sub">{date_label}</div></div>'
        f'<div class="sec"><h3>거래처 · 내용</h3>{work_html}</div>'
        f'<div class="panel next"><h3>익일업무</h3>{_lines(next_day, "익일업무 없음")}</div>'
        f'<div class="panel note"><h3>특 이 사 항</h3>{_lines(notes, "특이사항 없음")}</div>'
        f'<div class="foot">인쇄는 상단 「인쇄창열기」를 사용하세요.</div></div></body></html>'
    )

def _entries_key(d: date) -> str: return f"wl_entries_{d.isoformat()}"
def _next_key(d: date) -> str: return f"wl_next_{d.isoformat()}"
def _notes_key(d: date) -> str: return f"wl_notes_{d.isoformat()}"
def _boot_key(d: date) -> str: return f"worklog_booted_{d.isoformat()}"

def _init_widget_state(d: date) -> dict:
    bk, ek = _boot_key(d), _entries_key(d)
    if st.session_state.get(bk) and ek in st.session_state: return {}
    cells = read_worklog_cells(d)
    entries = _grouped_entries_from_cells(cells)
    if not entries: entries = [{"client": "", "content": ""}]
    st.session_state[ek] = entries
    _, next_day, notes = _entries_from_cells(cells)
    st.session_state[_next_key(d)] = "\n".join(next_day)
    st.session_state[_notes_key(d)] = "\n".join(notes)
    st.session_state[bk] = True
    return cells

def _entry_line_count_key(iso: str, entry_i: int) -> str: return f"wl_ent_lc_{iso}_{entry_i}"
def _entry_line_gen_key(iso: str, entry_i: int) -> str: return f"wl_ent_gen_{iso}_{entry_i}"
def _entry_line_key(iso: str, entry_i: int, line_j: int) -> str: return f"wl_ent_ln_{iso}_{entry_i}_{line_j}_g{int(st.session_state.get(_entry_line_gen_key(iso, entry_i), 0) or 0)}"

def _bump_entry_line_gen(iso: str, entry_i: int) -> None:
    k = _entry_line_gen_key(iso, entry_i)
    old_g, old_n = int(st.session_state.get(k, 0) or 0), int(st.session_state.get(_entry_line_count_key(iso, entry_i), 0) or 0)
    for j in range(max(old_n, 0) + 8):
        st.session_state.pop(f"wl_ent_ln_{iso}_{entry_i}_{j}_g{old_g}", None); st.session_state.pop(f"wl_ent_ln_{iso}_{entry_i}_{j}", None)
    st.session_state[k] = old_g + 1

def _lines_from_entry_widgets(iso: str, entry_i: int, *, keep_trailing_empty: bool = True) -> list[str]:
    live, cs = st.session_state.get(_entry_lines_live_key(iso, entry_i)), st.session_state.get(_entry_lines_comp_key(iso, entry_i))
    if isinstance(live, list): parts = [str(x or "") for x in live]
    elif isinstance(cs, dict) and isinstance(cs.get("lines"), list): parts = [str(x or "") for x in cs.get("lines") or []]
    else:
        lc = int(st.session_state.get(_entry_line_count_key(iso, entry_i), 0) or 0)
        if lc <= 0: parts = [str(st.session_state.get(f"wl_ent_t_{iso}_{entry_i}", "") or "")] if str(st.session_state.get(f"wl_ent_t_{iso}_{entry_i}", "") or "") else ([""] if keep_trailing_empty else [])
        else: parts = [str(st.session_state.get(_entry_line_key(iso, entry_i, j), "") or "") for j in range(lc)]
    if not keep_trailing_empty:
        while parts and parts[-1] == "": parts.pop()
    elif parts and parts[-1] != "": parts = list(parts) + [""]
    elif not parts and keep_trailing_empty: parts = [""]
    return parts

def _content_from_entry_lines(iso: str, entry_i: int) -> str: return "\n".join(_lines_from_entry_widgets(iso, entry_i, keep_trailing_empty=False))

def _set_comp_lines_state(iso: str, entry_i: int, chunks: list[str], *, focus_j: int | None = None) -> None:
    ck = _entry_lines_comp_key(iso, entry_i)
    prev = st.session_state.get(ck) if isinstance(st.session_state.get(ck), dict) else {}
    fj = prev.get("focus", -1) if focus_j is None else max(0, min(int(focus_j), max(len(chunks) - 1, 0)))
    new_state = {"lines": list(chunks), "focus": fj}
    if isinstance(prev.get("caret"), dict):
        new_state["caret"] = dict(prev.get("caret") or {})
    # 위젯 키는 중첩 mutate 금지 — 통째로 교체. 이전 값 잔존 시 pop 후 재설정.
    st.session_state.pop(ck, None)
    st.session_state[ck] = new_state
    if focus_j is not None:
        st.session_state[f"wl_focus_ln_{iso}"] = _entry_line_key(iso, entry_i, int(fj))
    st.session_state[_entry_lines_live_key(iso, entry_i)] = list(chunks)


def _norm_editor_lines(lines: list | None) -> list[str]:
    out = [str(x or "") for x in (lines or [])]
    if not out or out[-1] != "":
        out = list(out) + [""]
    return out


def _result_lines_list(result: Any) -> list[str] | None:
    if result is None:
        return None
    if hasattr(result, "lines") and isinstance(result.lines, list):
        return _norm_editor_lines(result.lines)
    if isinstance(result, dict) and isinstance(result.get("lines"), list):
        return _norm_editor_lines(result.get("lines"))
    return None


def _pick_editor_out_lines(
    *,
    session_lines: list[str],
    result: Any,
    forced: Any,
    changed_key: str,
) -> list[str]:
    """CCv2 result.lines가 저장/시드 이후 구값으로 session을 덮지 않게 선택.

    - forced(프로그램 시드) 우선
    - 사용자 편집 콜백이 난 런만 result 채택
    - 그 외에는 session(live) 유지 ← 저장 후 구값 부활 차단
    """
    session_n = _norm_editor_lines(session_lines)
    if isinstance(forced, list):
        st.session_state.pop(changed_key, None)
        return _norm_editor_lines(forced)
    user_changed = bool(st.session_state.pop(changed_key, None))
    result_n = _result_lines_list(result)
    if user_changed and result_n is not None:
        return result_n
    return session_n


def _apply_entry_lines(iso: str, entry_i: int, lines: list[str], *, focus_j: int | None = None, bump_gen: bool = False, remount_comp: bool = False) -> None:
    if bump_gen: _bump_entry_line_gen(iso, entry_i)
    if remount_comp: _bump_entry_lines_comp_inst(iso, entry_i)
    chunks = [str(x or "") for x in (lines or [])]
    if not chunks or chunks[-1] != "": chunks.append("")
    old = int(st.session_state.get(_entry_line_count_key(iso, entry_i), 0) or 0)
    for j in range(max(old, len(chunks)) + 3): st.session_state.pop(_entry_line_key(iso, entry_i, j), None)
    st.session_state[_entry_line_count_key(iso, entry_i)] = len(chunks)
    for j, line in enumerate(chunks): st.session_state[_entry_line_key(iso, entry_i, j)] = line
    st.session_state[f"wl_ent_t_{iso}_{entry_i}"] = "\n".join(chunks)
    _set_comp_lines_state(iso, entry_i, chunks, focus_j=focus_j)
    st.session_state[_entry_lines_rev_key(iso, entry_i)] = int(st.session_state.get(_entry_lines_rev_key(iso, entry_i), 0) or 0) + 1
    # 컴포넌트가 이전 result.lines로 덮어쓰지 않도록 가드
    st.session_state[f"wl_force_comp_lines_{iso}_{entry_i}"] = list(chunks)
    if focus_j is not None:
        fj = max(0, min(int(focus_j), len(chunks) - 1))
        st.session_state[f"wl_focus_ln_{iso}"] = _entry_line_key(iso, entry_i, fj)
        st.session_state[f"wl_focus_caret_{iso}"] = len(str(chunks[fj] or ""))

def _insert_line_after(iso: str, entry_i: int, line_j: int) -> None:
    cur = _lines_from_entry_widgets(iso, entry_i, keep_trailing_empty=True)
    if not cur: cur = [""]
    while len(cur) <= line_j: cur.append("")
    key = _entry_line_key(iso, entry_i, line_j)
    if key in st.session_state: cur[line_j] = str(st.session_state.get(key) or "")
    cur.insert(line_j + 1, "")
    _apply_entry_lines(iso, entry_i, cur, focus_j=line_j + 1)

def _entry_client_count_key(iso: str, entry_i: int) -> str: return f"wl_ent_clc_{iso}_{entry_i}"
def _entry_client_key(iso: str, entry_i: int, line_j: int) -> str: return f"wl_ent_cl_{iso}_{entry_i}_{line_j}"

def _clients_from_widgets(iso: str, entry_i: int, *, keep_trailing_empty: bool = False) -> list[str]:
    live, cs = st.session_state.get(_entry_clients_live_key(iso, entry_i)), st.session_state.get(_entry_clients_comp_key(iso, entry_i))
    if isinstance(live, list): parts = [str(x or "") for x in live]
    elif isinstance(cs, dict) and isinstance(cs.get("lines"), list): parts = [str(x or "") for x in cs.get("lines") or []]
    else:
        lc = int(st.session_state.get(_entry_client_count_key(iso, entry_i), 0) or 0)
        if lc > 0: parts = [str(st.session_state.get(_entry_client_key(iso, entry_i, j), "") or "") for j in range(lc)]
        else:
            raw = str(st.session_state.get(f"wl_ent_c_{iso}_{entry_i}", "") or "")
            parts = raw.splitlines() if raw else ([""] if keep_trailing_empty else [])
    if not keep_trailing_empty:
        while parts and not str(parts[-1]).strip(): parts.pop()
    elif parts and str(parts[-1]).strip() != "": parts = list(parts) + [""]
    elif not parts and keep_trailing_empty: parts = [""]
    return parts

def _set_comp_clients_state(iso: str, entry_i: int, chunks: list[str], *, focus_j: int | None = None) -> None:
    ck = _entry_clients_comp_key(iso, entry_i)
    prev = st.session_state.get(ck) if isinstance(st.session_state.get(ck), dict) else {}
    fj = prev.get("focus", -1) if focus_j is None else max(0, min(int(focus_j), max(len(chunks) - 1, 0)))
    new_state = {"lines": list(chunks), "focus": fj}
    if isinstance(prev.get("caret"), dict):
        new_state["caret"] = dict(prev.get("caret") or {})
    st.session_state.pop(ck, None)
    st.session_state[ck] = new_state
    if focus_j is not None:
        st.session_state[f"wl_focus_ln_{iso}"] = _entry_client_key(iso, entry_i, int(fj))
    st.session_state[_entry_clients_live_key(iso, entry_i)] = list(chunks)

def _apply_entry_clients(iso: str, entry_i: int, lines: list[str], *, focus_j: int | None = None, remount_comp: bool = False) -> None:
    if remount_comp: _bump_entry_clients_comp_inst(iso, entry_i)
    chunks = [str(x or "") for x in (lines or [])]
    if not chunks or chunks[-1] != "": chunks.append("")
    old = int(st.session_state.get(_entry_client_count_key(iso, entry_i), 0) or 0)
    for j in range(max(old, len(chunks)) + 3): st.session_state.pop(_entry_client_key(iso, entry_i, j), None)
    st.session_state[_entry_client_count_key(iso, entry_i)] = len(chunks)
    for j, line in enumerate(chunks): st.session_state[_entry_client_key(iso, entry_i, j)] = line
    filled = list(chunks)
    while filled and filled[-1] == "": filled.pop()
    st.session_state[f"wl_ent_c_{iso}_{entry_i}"] = "\n".join(filled)
    _set_comp_clients_state(iso, entry_i, chunks, focus_j=focus_j)
    st.session_state[_entry_clients_rev_key(iso, entry_i)] = int(st.session_state.get(_entry_clients_rev_key(iso, entry_i), 0) or 0) + 1
    st.session_state[f"wl_force_comp_clients_{iso}_{entry_i}"] = list(chunks)
    if focus_j is not None:
        fj = max(0, min(int(focus_j), len(chunks) - 1))
        st.session_state[f"wl_focus_ln_{iso}"] = _entry_client_key(iso, entry_i, fj)
        st.session_state[f"wl_focus_caret_{iso}"] = len(str(chunks[fj] or ""))

def _seed_entry_clients(iso: str, entry_i: int, client: str | list[str]) -> None:
    max_u = _client_line_units()
    src = [str(x or "") for x in client] if isinstance(client, list) else (str(client or "").split('\n') if str(client or "") else [""])
    chunks: list[str] = []
    for line in src:
        s = str(line or "")
        if not s.strip():
            chunks.append(s)
            continue
        chunks.extend(_chunk_text(s, max_u) or [s])
    if not chunks: chunks = [""]
    fj = max(0, len(chunks) - 1)
    for j, line in enumerate(chunks):
        if _display_units(line) >= max_u: fj = min(j + 1, len(chunks))
    _apply_entry_clients(iso, entry_i, chunks, focus_j=fj, remount_comp=True)

def _insert_client_after(iso: str, entry_i: int, line_j: int) -> None:
    cur = _clients_from_widgets(iso, entry_i, keep_trailing_empty=True)
    if not cur: cur = [""]
    while len(cur) <= line_j: cur.append("")
    key = _entry_client_key(iso, entry_i, line_j)
    if key in st.session_state: cur[line_j] = str(st.session_state.get(key) or "")
    cur.insert(line_j + 1, "")
    _apply_entry_clients(iso, entry_i, cur, focus_j=line_j + 1)

def _split_overflow_parts(parts: list[str], max_u: int) -> list[str]:
    out, i, n = len(parts), 0, len(parts)
    out = []
    while i < n:
        s = str(parts[i] or "")
        if _display_units(s) > max_u:
            pieces = _chunk_text(s, max_u) or [s]
            out.append(pieces[0])
            j = i + 1
            for ov in pieces[1:]:
                if j < n and str(parts[j] or "") == ov: out.append(str(parts[j] or "")); j += 1
                else: out.append(ov)
            i = j
        else: out.append(s); i += 1
    return out if out else [""]

def _dedupe_overflow_tail(pieces: list[str], tail: list[str]) -> list[str]:
    if len(pieces) <= 1: return list(tail)
    rest = list(tail)
    for ov in pieces[1:]:
        if rest and str(rest[0] or "") == ov: rest.pop(0)
        else: break
    return rest

def _commit_enter_on_cell(kind: str, iso: str, entry_i: int, line_j: int, value: str) -> None:
    value = str(value or "")
    is_client = kind == "wl_ent_cl"
    max_u = _client_line_units() if is_client else _content_line_units()
    cur = _clients_from_widgets(iso, entry_i, keep_trailing_empty=True) if is_client else _lines_from_entry_widgets(iso, entry_i, keep_trailing_empty=True)
    while len(cur) <= line_j: cur.append("")
    if cur and cur[-1] == "": cur = cur[:-1]
    head, tail = cur[:line_j], cur[line_j + 1 :]
    pieces = _chunk_text(value, max_u) or [value] if _display_units(value) > max_u else [value]
    if len(pieces) == 1:
        new, focus = head + pieces + [""] + list(tail), line_j + 1
    else:
        new, focus = head + pieces + _dedupe_overflow_tail(pieces, list(tail)), line_j + 1
        if focus >= len(new): new.append("")
    if is_client: _apply_entry_clients(iso, entry_i, new, focus_j=min(focus, len(new) - 1))
    else: _apply_entry_lines(iso, entry_i, new, focus_j=min(focus, len(new) - 1), bump_gen=True)
    st.session_state.pop(f"wl_enter_done_{iso}", None)

def _mount_entry_client_editor(iso: str, entry_i: int, max_u: int) -> list[str]:
    ck, live_key, cs = _entry_clients_comp_key(iso, entry_i), _entry_clients_live_key(iso, entry_i), st.session_state.get(_entry_clients_comp_key(iso, entry_i))
    if isinstance(cs, dict) and isinstance(cs.get("lines"), list): lines, focus = [str(x or "") for x in cs.get("lines") or []], cs.get("focus", -1)
    elif isinstance(st.session_state.get(live_key), list): lines, focus = [str(x or "") for x in st.session_state.get(live_key) or []], -1
    elif int(st.session_state.get(_entry_client_count_key(iso, entry_i), 0) or 0) > 0: lines, focus = [str(st.session_state.get(_entry_client_key(iso, entry_i, j), "") or "") for j in range(int(st.session_state.get(_entry_client_count_key(iso, entry_i), 1) or 1))], -1
    else:
        raw = str(st.session_state.get(f"wl_ent_c_{iso}_{entry_i}", "") or "")
        if raw:
            _seed_entry_clients(iso, entry_i, raw)
            cs = st.session_state.get(ck)
            if isinstance(cs, dict) and isinstance(cs.get("lines"), list): lines, focus = [str(x or "") for x in cs.get("lines") or []], cs.get("focus", -1)
            else: lines, focus = [""], -1
        else: lines, focus = [""], -1
    if any(_display_units(p) > max_u for p in lines):
        fixed, focus = _split_overflow_parts(lines, max_u), 0
        for j, line in enumerate(fixed):
            if _display_units(line) >= max_u: focus = min(j + 1, len(fixed))
        if fixed != lines:
            _apply_entry_clients(iso, entry_i, fixed, focus_j=focus)
            cs = st.session_state.get(ck)
            if isinstance(cs, dict) and isinstance(cs.get("lines"), list): lines, focus = [str(x or "") for x in cs.get("lines") or []], cs.get("focus", focus)
            else: lines = fixed
    if not lines or lines[-1] != "": lines = list(lines) + [""]
    try: focus_n = int(focus)
    except (TypeError, ValueError): focus_n = -1

    def _on_clients_change() -> None:
        cur = st.session_state.get(ck)
        if isinstance(cur, dict) and isinstance(cur.get("lines"), list):
            synced = [str(x or "") for x in cur.get("lines") or []]
            st.session_state[f"wl_clients_user_edit_{iso}_{entry_i}"] = True
            st.session_state[live_key] = synced
            st.session_state[_entry_client_count_key(iso, entry_i)] = len(synced)
            for j, line in enumerate(synced): st.session_state[_entry_client_key(iso, entry_i, j)] = line
            filled = list(synced)
            while filled and filled[-1] == "": filled.pop()
            st.session_state[f"wl_ent_c_{iso}_{entry_i}"] = "\n".join(filled)
            lj = _last_used_line_index(synced)
            _remember_active_cell(iso, _entry_client_key(iso, entry_i, lj), len(str(synced[lj] or "")))

    def _on_clients_focus_change() -> None:
        _sync_editor_focus_from_comp(iso, entry_i, ck, _entry_client_key)

    result = _WL_LINES_EDITOR(
        key=ck,
        data={"lines": lines, "focus": focus_n, "max_u": int(max_u), "variant": "client", "rev": int(st.session_state.get(_entry_clients_rev_key(iso, entry_i), 0) or 0)},
        default={"lines": lines, "focus": focus_n}, 
        on_lines_change=_on_clients_change,
        on_focus_change=_on_clients_focus_change
    )
    
    forced = st.session_state.pop(f"wl_force_comp_clients_{iso}_{entry_i}", None)
    out = _pick_editor_out_lines(
        session_lines=lines,
        result=result,
        forced=forced,
        changed_key=f"wl_clients_user_edit_{iso}_{entry_i}",
    )
    st.session_state[live_key] = out
    old = int(st.session_state.get(_entry_client_count_key(iso, entry_i), 0) or 0)
    for j in range(max(old, len(out)) + 3): st.session_state.pop(_entry_client_key(iso, entry_i, j), None)
    st.session_state[_entry_client_count_key(iso, entry_i)] = len(out)
    for j, line in enumerate(out): st.session_state[_entry_client_key(iso, entry_i, j)] = line
    filled = list(out)
    while filled and filled[-1] == "": filled.pop()
    st.session_state[f"wl_ent_c_{iso}_{entry_i}"] = "\n".join(filled)
    return out

def _mount_entry_lines_editor(iso: str, entry_i: int, max_u: int) -> list[str]:
    ck, live_key, cs = _entry_lines_comp_key(iso, entry_i), _entry_lines_live_key(iso, entry_i), st.session_state.get(_entry_lines_comp_key(iso, entry_i))
    if isinstance(cs, dict) and isinstance(cs.get("lines"), list): lines, focus = [str(x or "") for x in cs.get("lines") or []], cs.get("focus", -1)
    elif isinstance(st.session_state.get(live_key), list): lines, focus = [str(x or "") for x in st.session_state.get(live_key) or []], -1
    else: lines, focus = [""], -1
    if any(_display_units(p) > max_u for p in lines):
        fixed, focus = _split_overflow_parts(lines, max_u), 0
        for j, line in enumerate(fixed):
            if _display_units(line) >= max_u: focus = min(j + 1, len(fixed))
        if fixed != lines:
            _apply_entry_lines(iso, entry_i, fixed, focus_j=focus)
            cs = st.session_state.get(ck)
            if isinstance(cs, dict) and isinstance(cs.get("lines"), list): lines, focus = [str(x or "") for x in cs.get("lines") or []], cs.get("focus", focus)
            else: lines = fixed
    if not lines or lines[-1] != "": lines = list(lines) + [""]
    try: focus_n = int(focus)
    except (TypeError, ValueError): focus_n = -1

    def _on_lines_change() -> None:
        cur = st.session_state.get(ck)
        if isinstance(cur, dict) and isinstance(cur.get("lines"), list):
            synced = [str(x or "") for x in cur.get("lines") or []]
            st.session_state[f"wl_lines_user_edit_{iso}_{entry_i}"] = True
            st.session_state[live_key] = synced
            st.session_state[_entry_line_count_key(iso, entry_i)] = len(synced)
            for j, line in enumerate(synced): st.session_state[_entry_line_key(iso, entry_i, j)] = line
            filled = list(synced)
            while filled and filled[-1] == "": filled.pop()
            st.session_state[f"wl_ent_t_{iso}_{entry_i}"] = "\n".join(filled)
            lj = _last_used_line_index(synced)
            _remember_active_cell(iso, _entry_line_key(iso, entry_i, lj), len(str(synced[lj] or "")))

    def _on_lines_focus_change() -> None:
        _sync_editor_focus_from_comp(iso, entry_i, ck, _entry_line_key)

    result = _WL_LINES_EDITOR(
        key=ck,
        data={"lines": lines, "focus": focus_n, "max_u": int(max_u), "variant": "content", "rev": int(st.session_state.get(_entry_lines_rev_key(iso, entry_i), 0) or 0)},
        default={"lines": lines, "focus": focus_n}, 
        on_lines_change=_on_lines_change,
        on_focus_change=_on_lines_focus_change
    )
    
    forced = st.session_state.pop(f"wl_force_comp_lines_{iso}_{entry_i}", None)
    out = _pick_editor_out_lines(
        session_lines=lines,
        result=result,
        forced=forced,
        changed_key=f"wl_lines_user_edit_{iso}_{entry_i}",
    )
    st.session_state[live_key] = out
    old = int(st.session_state.get(_entry_line_count_key(iso, entry_i), 0) or 0)
    for j in range(max(old, len(out)) + 3): st.session_state.pop(_entry_line_key(iso, entry_i, j), None)
    st.session_state[_entry_line_count_key(iso, entry_i)] = len(out)
    for j, line in enumerate(out): st.session_state[_entry_line_key(iso, entry_i, j)] = line
    filled = list(out)
    while filled and filled[-1] == "": filled.pop()
    st.session_state[f"wl_ent_t_{iso}_{entry_i}"] = "\n".join(filled)
    return out


def _remember_active_cell(iso: str, fk: str, pos: int) -> None:
    st.session_state["wl_active_cell_key"] = fk
    st.session_state[f"wl_focus_ln_{iso}"] = fk
    st.session_state["wl_active_cell_sel"] = (pos, pos)
    st.session_state[f"wl_focus_caret_{iso}"] = pos


def _read_cell_value(iso: str, fk: str) -> str:
    if fk.startswith("wl_next_area_") or fk.startswith("wl_notes_area_"):
        return str(st.session_state.get(fk, "") or "")
    m = re.match(r"^(wl_ent_ln|wl_ent_cl)_(\d{4}-\d{2}-\d{2})_(\d+)_(\d+)", fk)
    if not m or m.group(2) != iso:
        return ""
    kind, ei, lj = m.group(1), int(m.group(3)), int(m.group(4))
    if kind == "wl_ent_ln":
        lines = _lines_from_entry_widgets(iso, ei, keep_trailing_empty=True)
    else:
        lines = _clients_from_widgets(iso, ei, keep_trailing_empty=True)
    while len(lines) <= lj:
        lines.append("")
    return str(lines[lj] or "")


def _insert_special_char_at_active(iso: str, ch: str) -> bool:
    fk = str(st.session_state.get("wl_active_cell_key") or st.session_state.get(f"wl_focus_ln_{iso}") or "")
    if not (
        fk.startswith("wl_ent_ln_")
        or fk.startswith("wl_ent_cl_")
        or fk.startswith("wl_next_area_")
        or fk.startswith("wl_notes_area_")
    ):
        return False
    pos = int(st.session_state.get(f"wl_focus_caret_{iso}") or 0)
    sel = st.session_state.get("wl_active_cell_sel")
    if isinstance(sel, (list, tuple)) and len(sel) >= 1:
        try:
            pos = int(sel[0])
        except (TypeError, ValueError):
            pass
    val = _read_cell_value(iso, fk)
    pos = max(0, min(pos, len(val)))
    new_val = val[:pos] + ch + val[pos:]
    _apply_special_insert(iso, fk, new_val, pos + len(ch), ch)
    return True


def _sync_editor_focus_from_comp(iso: str, entry_i: int, comp_key: str, key_fn) -> None:
    cur = st.session_state.get(comp_key)
    if not isinstance(cur, dict):
        return
    try:
        fj = int(cur.get("focus", -1))
    except (TypeError, ValueError):
        fj = -1
    if fj < 0:
        return
    caret = cur.get("caret") if isinstance(cur.get("caret"), dict) else {}
    try:
        pos = int(caret.get("s", 0))
    except (TypeError, ValueError):
        pos = 0
    fk = key_fn(iso, entry_i, fj)
    _remember_active_cell(iso, fk, pos)
    # 같은 항목 안에서 줄(칸)이 바뀌면 왼쪽 요약 갱신 요청
    try:
        _request_left_preview_refresh(date.fromisoformat(iso), focus_sig=fk)
    except Exception:
        pass


def _last_used_line_index(lines: list[str]) -> int:
    for j in range(len(lines) - 1, -1, -1):
        if str(lines[j] or ""):
            return j
    return 0


def _resolve_special_entry_i(iso: str, entry_count: int) -> int:
    n = max(1, int(entry_count or 1))
    for i in range(n - 1, -1, -1):
        if st.session_state.get(f"wl_exp_{iso}_{i}"):
            return i
    return n - 1


def _insert_special_char_auto(iso: str, ch: str, entry_count: int) -> bool:
    """넣을 칸 선택 없이 삽입.
    1) 직전에 편집한 칸(active)이 있으면 그곳
    2) 없으면 펼쳐진(또는 마지막) 항목의 내용 칸 끝에 추가
    """
    ch = str(ch or "")
    if not ch:
        return False
    fk = str(st.session_state.get("wl_active_cell_key") or st.session_state.get(f"wl_focus_ln_{iso}") or "")
    if (
        fk.startswith(f"wl_ent_ln_{iso}_")
        or fk.startswith(f"wl_ent_cl_{iso}_")
        or fk == f"wl_next_area_{iso}"
        or fk == f"wl_notes_area_{iso}"
    ):
        if _insert_special_char_at_active(iso, ch):
            m = re.match(r"^wl_ent_(?:ln|cl)_\d{4}-\d{2}-\d{2}_(\d+)_", fk)
            if m:
                st.session_state[f"wl_force_expand_{iso}"] = int(m.group(1))
            return True
    ei = _resolve_special_entry_i(iso, entry_count)
    st.session_state[f"wl_force_expand_{iso}"] = ei
    lines = _lines_from_entry_widgets(iso, ei, keep_trailing_empty=True)
    if not lines:
        lines = [""]
    lj = _last_used_line_index(lines)
    while len(lines) <= lj:
        lines.append("")
    val = str(lines[lj] or "")
    new_val = val + ch
    _apply_special_insert(iso, f"wl_ent_ln_{iso}_{ei}_{lj}", new_val, len(new_val), ch)
    return True


def _process_pending_special_char(iso: str, entry_count: int) -> bool:
    """pending 특수기호 삽입. 성공하면 True(호출측에서 fragment rerun 권장)."""
    pending = st.session_state.pop(f"wl_pending_sp_{iso}", None)
    if not pending:
        return False
    ch = str(pending or "")
    if not ch:
        return False
    try:
        if _insert_special_char_auto(iso, ch, entry_count):
            st.session_state["wl_special_msg"] = f"「{ch}」삽입"
            # 바 리마운트(동일 기호 연속 클릭 인식)
            st.session_state[f"wl_sp_token_{iso}"] = str(int(time.time() * 1000) % 1_000_000_000)
            return True
        st.session_state["wl_special_msg"] = "특수기호를 넣지 못했습니다."
    except StreamlitAPIException:
        # 재시도 루프 방지 — pending을 되돌리지 않음
        st.session_state["wl_special_msg"] = "특수기호 적용에 실패했습니다. 칸을 선택한 뒤 다시 눌러 주세요."
    return False


def _queue_special_char(iso: str, ch: str) -> None:
    """특수기호 pending만 설정. 절대 여기서 rerun 하지 않음(재시도 오류 방지)."""
    ch = str(ch or "")
    if not ch:
        return
    st.session_state[f"wl_pending_sp_{iso}"] = ch


def _parse_special_pick_raw(raw: Any) -> tuple[str, str]:
    """컴포넌트 pick 트리거 → (문자, dedupe signature)."""
    if raw is None:
        return "", ""
    try:
        if isinstance(raw, str) and raw.startswith("{"):
            obj = json.loads(raw)
            ch = str(obj.get("ch") or "")
            sig = f"{ch}\0{obj.get('t')}"
            return ch, sig
        ch = str(raw)
        return ch, f"{ch}\0plain"
    except Exception:
        ch = str(raw or "")
        return ch, f"{ch}\0err"


def _consume_special_pick(iso: str, raw: Any) -> None:
    """pick 한 번만 pending에 넣고, 동일 시그니처 중복은 무시."""
    ch, sig = _parse_special_pick_raw(raw)
    if not ch or not sig:
        return
    done_k = f"wl_sp_pick_done_{iso}"
    if st.session_state.get(done_k) == sig:
        return
    st.session_state[done_k] = sig
    _queue_special_char(iso, ch)


def _render_worklog_special_chars(iso: str, entry_count: int = 1) -> bool:
    """날짜 아래: 접기 없이 한 줄 가로 스크롤. 삽입은 pending 경로.
    특수기호를 방금 넣었으면 True (호출측에서 fragment rerun).
    """
    st.markdown(
        """<style>
        div[class*="st-key-wl_sp_bar_"] {
          width: 100% !important;
          max-width: 100% !important;
          margin: 0.15rem 0 0.25rem !important;
          padding: 0 !important;
        }
        div[class*="st-key-wl_sp_bar_"] iframe {
          width: 100% !important;
          min-height: 2.35rem !important;
          height: 2.35rem !important;
          border: none !important;
          display: block !important;
        }
        </style>""",
        unsafe_allow_html=True,
    )

    def _on_pick() -> None:
        # 콜백 시점에는 session_state 뷰가 비어 있을 수 있어 반환값 경로를 우선한다.
        # 여기선 보조로 한 번 더 시도.
        stt = st.session_state.get(f"wl_sp_bar_{iso}")
        raw = None
        if isinstance(stt, dict):
            raw = stt.get("pick")
        else:
            raw = getattr(stt, "pick", None) if stt is not None else None
        _consume_special_pick(iso, raw)

    token = str(st.session_state.get(f"wl_sp_token_{iso}") or "2")
    result = _WL_SPECIAL_BAR(
        key=f"wl_sp_bar_{iso}",
        data={"iso": iso, "chars": list(_WL_SPECIAL_CHARS), "token": token},
        on_pick_change=_on_pick,
        width="stretch",
        height=40,
    )
    # 주 경로: 컴포넌트 반환값의 trigger(pick) — 콜백보다 안정적
    raw = None
    if result is not None:
        raw = getattr(result, "pick", None)
        if raw is None and isinstance(result, dict):
            raw = result.get("pick")
    _consume_special_pick(iso, raw)
    # 바로 삽입해 내용 칸에 반영 (이후 fragment rerun으로 UI 확정)
    return _process_pending_special_char(iso, entry_count)

def _apply_special_insert(iso: str, fk: str, val: str, pos: int, ch: str) -> None:
    val, pos = str(val or ""), max(0, min(int(pos or 0), len(str(val or ""))))
    if fk.startswith("wl_next_area_") or fk.startswith("wl_notes_area_"):
        if not fk.endswith(f"_{iso}"):
            return
        st.session_state[fk] = val
        st.session_state[f"wl_focus_ln_{iso}"] = fk
        st.session_state["wl_active_cell_key"] = fk
        st.session_state["wl_active_cell_sel"] = (pos, pos)
        st.session_state[f"wl_focus_caret_{iso}"] = pos
        st.session_state["wl_special_msg"] = f"「{ch}」삽입" if ch else "특수문자 삽입"
        return
    m = re.match(r"^(wl_ent_ln|wl_ent_cl)_(\d{4}-\d{2}-\d{2})_(\d+)_(\d+)", fk)
    if not m or m.group(2) != iso: return
    kind, ei, lj = m.group(1), int(m.group(3)), int(m.group(4))
    if kind == "wl_ent_ln":
        cur = _lines_from_entry_widgets(iso, ei, keep_trailing_empty=True)
        while len(cur) <= lj: cur.append("")
        cur[lj] = val
        _apply_entry_lines(iso, ei, cur, focus_j=lj, bump_gen=True)
    else:
        cur = _clients_from_widgets(iso, ei, keep_trailing_empty=True)
        while len(cur) <= lj: cur.append("")
        cur[lj] = val
        _apply_entry_clients(iso, ei, cur, focus_j=lj)
    st.session_state["wl_active_cell_key"] = st.session_state.get(f"wl_focus_ln_{iso}") or fk
    st.session_state["wl_active_cell_sel"] = (pos, pos)
    st.session_state[f"wl_focus_caret_{iso}"] = pos
    st.session_state["wl_special_msg"] = f"「{ch}」삽입" if ch else "특수문자 삽입"

def _seed_entry_lines(iso: str, entry_i: int, content: str, *, focus_j: int | None = None, focus_last: bool = False) -> None:
    max_u = _content_line_units()
    src = str(content or "").split('\n') if str(content or "") else [""]
    chunks = []
    for line in src:
        if not line.strip(): 
            chunks.append(line)
            continue
        chunks.extend(_chunk_text(line, max_u) or [line])
    if not chunks: chunks = [""]
    fj = focus_j
    if fj is None and focus_last:
        fj = max(0, len(chunks) - 1)
        for j, line in enumerate(chunks):
            if _display_units(line) >= max_u: fj = min(j + 1, len(chunks))
    _apply_entry_lines(iso, entry_i, chunks, focus_j=fj, remount_comp=True)

def _read_editor_entries(d: date) -> list[dict]:
    iso = d.isoformat()
    n = int(st.session_state.get(f"wl_entry_count_{iso}", 0) or 0)
    stored = st.session_state.get(_entries_key(d)) or [{"client": "", "content": "", "blank_after": 1}]
    if n <= 0: n = len(stored)
    out: list[dict] = []
    for i in range(n):
        ck, gk, lc = f"wl_ent_c_{iso}_{i}", f"wl_ent_gap_{iso}_{i}", int(st.session_state.get(_entry_line_count_key(iso, i), 0) or 0)
        if ck in st.session_state or lc > 0 or f"wl_ent_t_{iso}_{i}" in st.session_state or int(st.session_state.get(_entry_client_count_key(iso, i), 0) or 0) > 0:
            if int(st.session_state.get(_entry_client_count_key(iso, i), 0) or 0) > 0:
                client_lines = _clients_from_widgets(iso, i, keep_trailing_empty=False)
                client = "\n".join(client_lines)
            else:
                client = str(st.session_state.get(ck, "") or "")
                client_lines = _entry_client_lines({"client": client})
            lines = _lines_from_entry_widgets(iso, i, keep_trailing_empty=False)
            content = "\n".join(lines)
            blank_after = _entry_blank_after({"blank_after": st.session_state.get(gk)}, 1) if gk in st.session_state else (_entry_blank_after(stored[i], 1) if i < len(stored) else 1)
        elif i < len(stored):
            client = str(stored[i].get("client") or "")
            if not client and isinstance(stored[i].get("client_lines"), list): client = "\n".join(str(x or "") for x in stored[i].get("client_lines") or [])
            client_lines = _entry_client_lines(stored[i])
            lines = stored[i].get("lines")
            if not isinstance(lines, list): lines = _chunk_text(str(stored[i].get("content") or ""), _content_line_units()) or []
            content = ("\n".join(str(x or "") for x in lines) if lines else str(stored[i].get("content") or ""))
            blank_after = _entry_blank_after(stored[i], 1)
        else:
            client, content, blank_after, lines, client_lines = "", "", 1, [], []
        # 주의: 위젯이 비어 있다고 해서 stored(이전 저장값)로 되살리지 않음.
        # 사용자가 지운 뒤 저장하면 구값이 되살아나던 원인이었음.
        out.append({"client": client, "client_lines": client_lines, "content": content, "lines": lines, "blank_after": blank_after})
    return out or [{"client": "", "client_lines": [], "content": "", "lines": [], "blank_after": 1}]

def _cells_from_widgets(d: date) -> dict:
    entries = _read_editor_entries(d)
    nk, ok = f"wl_next_area_{d.isoformat()}", f"wl_notes_area_{d.isoformat()}"
    next_raw = str(st.session_state.get(nk) or st.session_state.get(_next_key(d), "") or "")
    notes_raw = str(st.session_state.get(ok) or st.session_state.get(_notes_key(d), "") or "")
    return _pack_entries_to_cells(d, entries, [x.strip() for x in next_raw.splitlines() if x.strip()], [x.strip() for x in notes_raw.splitlines() if x.strip()])

def _seed_day_entry_widgets(d: date, entries_list: list[dict], next_txt: str, notes_txt: str) -> None:
    """저장/추가/삭제 후 입력 위젯을 entries 기준으로 다시 심는다. CCv2 인스턴스 키도 갱신."""
    iso = d.isoformat()
    ek = _entries_key(d)
    old_n = int(st.session_state.get(f"wl_entry_count_{iso}", 0) or 0)
    for i in range(max(old_n, len(entries_list)) + 2):
        st.session_state.pop(f"wl_ent_c_{iso}_{i}", None)
        st.session_state.pop(f"wl_ent_t_{iso}_{i}", None)
        st.session_state.pop(f"wl_ent_gap_{iso}_{i}", None)
        st.session_state.pop(f"wl_exp_{iso}_{i}", None)
        st.session_state.pop(_entry_lines_comp_key(iso, i), None)
        st.session_state.pop(f"wl_lines_comp_{iso}_{i}", None)
        st.session_state.pop(_entry_lines_rev_key(iso, i), None)
        st.session_state.pop(_entry_lines_live_key(iso, i), None)
        st.session_state.pop(f"wl_force_comp_lines_{iso}_{i}", None)
        st.session_state.pop(f"wl_lines_user_edit_{iso}_{i}", None)
        st.session_state.pop(_entry_clients_comp_key(iso, i), None)
        st.session_state.pop(f"wl_clients_comp_{iso}_{i}", None)
        st.session_state.pop(_entry_clients_rev_key(iso, i), None)
        st.session_state.pop(_entry_clients_live_key(iso, i), None)
        st.session_state.pop(f"wl_force_comp_clients_{iso}_{i}", None)
        st.session_state.pop(f"wl_clients_user_edit_{iso}_{i}", None)
        old_lc = int(st.session_state.get(_entry_line_count_key(iso, i), 0) or 0)
        for j in range(old_lc + 3): st.session_state.pop(_entry_line_key(iso, i, j), None)
        st.session_state.pop(_entry_line_count_key(iso, i), None)
        old_cc = int(st.session_state.get(_entry_client_count_key(iso, i), 0) or 0)
        for j in range(old_cc + 3): st.session_state.pop(_entry_client_key(iso, i, j), None)
        st.session_state.pop(_entry_client_count_key(iso, i), None)
    st.session_state.pop(f"wl_next_area_{iso}", None)
    st.session_state.pop(f"wl_notes_area_{iso}", None)
    st.session_state[ek] = entries_list
    st.session_state[f"wl_entry_count_{iso}"] = len(entries_list)
    for i, ent in enumerate(entries_list):
        clines = ent.get("client_lines")
        if isinstance(clines, list) and clines: _seed_entry_clients(iso, i, clines)
        else: _seed_entry_clients(iso, i, ent.get("client") or "")
        st.session_state[f"wl_ent_gap_{iso}_{i}"] = _entry_blank_after(ent, 1)
        lines = ent.get("lines")
        if isinstance(lines, list): _apply_entry_lines(iso, i, [str(x or "") for x in lines], remount_comp=True)
        else: _seed_entry_lines(iso, i, ent.get("content") or "")
    st.session_state[f"wl_next_area_{iso}"] = next_txt
    st.session_state[f"wl_notes_area_{iso}"] = notes_txt
    st.session_state[_next_key(d)] = next_txt
    st.session_state[_notes_key(d)] = notes_txt


def _view_cells_key(d: date) -> str: return f"wl_view_cells_{d.isoformat()}"

def _publish_view_cells(d: date, cells: dict) -> None:
    iso = d.isoformat()
    cells = dict(cells or {})
    prev = st.session_state.get(_view_cells_key(d))
    if isinstance(prev, dict):
        try:
            if json.dumps(prev, ensure_ascii=False, sort_keys=True) == json.dumps(cells, ensure_ascii=False, sort_keys=True):
                return
        except Exception:
            pass
    st.session_state[_view_cells_key(d)] = cells
    for k in (f"wl_sum_sig_{iso}", f"wl_sum_html_{iso}", f"wl_left_excel_sig_v24_{iso}", f"wl_left_excel_html_v24_{iso}", f"wl_left_excel_h_v24_{iso}"): st.session_state.pop(k, None)

def _view_cells_for_preview(d: date) -> dict:
    key = _view_cells_key(d)
    if isinstance(st.session_state.get(key), dict) and st.session_state.get(key): return st.session_state.get(key)
    try:
        if os.path.exists(worklog_path(d)): cells = read_worklog_cells(d); st.session_state[key] = cells; return cells
    except Exception: pass
    cells = _empty_cells(d)
    st.session_state[key] = cells
    return cells

def _clear_date_widget_state(d: date) -> None:
    iso = d.isoformat()
    prefixes = (f"wl_ent_c_{iso}_", f"wl_ent_t_{iso}_", f"wl_ent_gap_{iso}_", f"wl_ent_ln_{iso}_", f"wl_ent_lc_{iso}_", f"wl_ent_gen_{iso}_", f"wl_ent_cl_{iso}_", f"wl_ent_clc_{iso}_", f"wl_ent_rev_{iso}_", f"wl_lines_comp_{iso}_", f"wl_lines_live_{iso}_", f"wl_lines_inst_{iso}_", f"wl_clients_comp_{iso}_", f"wl_clients_live_{iso}_", f"wl_clients_inst_{iso}_", f"wl_clients_rev_{iso}_", f"wl_force_comp_lines_{iso}_", f"wl_force_comp_clients_{iso}_", f"wl_lines_user_edit_{iso}_", f"wl_clients_user_edit_{iso}_", f"wl_exp_{iso}_", f"wl_entries_{iso}", f"wl_next_{iso}", f"wl_notes_{iso}", f"wl_next_area_{iso}", f"wl_notes_area_{iso}", f"wl_entry_count_{iso}", f"worklog_booted_{iso}", f"wl_save_btn_{iso}", f"wl_focus_ln_{iso}", f"wl_do_save_{iso}", f"wl_flash_save_{iso}", f"wl_view_cells_{iso}")
    for k in list(st.session_state.keys()):
        if not isinstance(k, str): continue
        if k in prefixes or any(k.startswith(p) for p in prefixes if p.endswith("_")): del st.session_state[k]
        elif k in {f"wl_entries_{iso}", f"wl_next_{iso}", f"wl_notes_{iso}", f"wl_entry_count_{iso}", f"worklog_booted_{iso}", f"wl_next_area_{iso}", f"wl_notes_area_{iso}", f"wl_pending_sync_{iso}", f"wl_do_add_{iso}", f"wl_do_del_{iso}", f"wl_focus_ln_{iso}", f"wl_do_save_{iso}", f"wl_flash_save_{iso}", f"wl_view_cells_{iso}", f"wl_open_ctx_{iso}", f"wl_saved_ok_{iso}"}: del st.session_state[k]

def _preview_path(d: date) -> str: return os.path.join(WORKLOG_DIR, f"_preview_{d.isoformat()}.xlsx")

def _build_preview_file(d: date, cells: dict) -> str:
    _ensure_dirs()
    if not os.path.exists(WORKLOG_TEMPLATE): raise FileNotFoundError("업무일지 템플릿이 없습니다.")
    dst = _preview_path(d)
    write_cells_to_path(dst, d, cells, force_template=True)
    return dst

def _excel_app_path() -> str | None:
    for p in ("/Applications/Microsoft Excel.app", os.path.expanduser("~/Applications/Microsoft Excel.app")):
        if os.path.isdir(p): return p
    return None

def _print_xlsx_path(d: date) -> str: return os.path.join(WORKLOG_DIR, f"일일업무일지_{d.isoformat()}_인쇄.xlsx")

def prepare_print_xlsx(d: date, cells: dict) -> str:
    _ensure_dirs()
    dst = _print_xlsx_path(d)
    write_cells_to_path(dst, d, cells, force_template=True)
    return os.path.abspath(dst)

def open_excel_print_preview(xlsx_path: str, *, prefer_print_dialog: bool = True) -> tuple[bool, str]:
    abs_path = os.path.abspath(xlsx_path)
    if not os.path.exists(abs_path): return False, "미리보기용 엑셀 파일이 없습니다."
    if platform.system() != "Darwin": return False, "Excel 인쇄 화면은 맥에서만 자동 연결됩니다."
    if not _excel_app_path():
        try: subprocess.Popen(["open", abs_path], start_new_session=True); return True, "파일을 열었습니다. Excel이 없다면 설치 후 다시 시도해 주세요."
        except Exception as e: return False, f"파일 열기 실패: {e}"
    ap = abs_path.replace("\\", "\\\\").replace('"', '\\"')
    if prefer_print_dialog:
        script = f'''set targetFile to POSIX file "{ap}"\ntell application "Microsoft Excel"\nactivate\nopen targetFile\ndelay 1.2\nend tell\ntell application "System Events"\nif exists process "Microsoft Excel" then\ntell process "Microsoft Excel"\nset frontmost to true\ndelay 0.4\nkeystroke "p" using {{command down}}\nend tell\nend if\nend tell\nreturn true'''
        ok_msg = "Excel에서 열어 인쇄(미리보기) 화면까지 연결했습니다."
    else:
        script = f'''set targetFile to POSIX file "{ap}"\nset previewDone to false\ntell application "Microsoft Excel"\nactivate\nopen targetFile\ndelay 1.3\ntry\nprint preview active sheet\nset previewDone to true\nend try\nif previewDone is false then\ntry\nprint preview\nset previewDone to true\nend try\nend if\nend tell\nif previewDone is false then\ntell application "System Events"\nif exists process "Microsoft Excel" then\ntell process "Microsoft Excel"\nset frontmost to true\ndelay 0.35\ntry\nclick menu item "인쇄 미리 보기" of menu "파일" of menu bar 1\nset previewDone to true\nend try\nif previewDone is false then\nkeystroke "p" using {{command down}}\nset previewDone to true\nend if\nend tell\nend if\nend tell\nend if\nreturn previewDone'''
        ok_msg = "Excel에서 열어 인쇄 미리보기까지 연결했습니다."
    try:
        r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=50)
        if r.returncode == 0: return True, ok_msg
        subprocess.Popen(["open", "-a", "Microsoft Excel", abs_path], start_new_session=True)
        err = (r.stderr or r.stdout or "").strip()
        hint = " (손쉬운 사용 권한을 허용하면 자동 연결됩니다)" if "Not authorized" in err or "assistive" in err.lower() or "1002" in err else ""
        return True, "Excel에서 파일을 열었습니다. ⌘P로 인쇄화면을 여세요." + hint
    except Exception as e:
        try: subprocess.Popen(["open", "-a", "Microsoft Excel", abs_path], start_new_session=True); return True, f"Excel에서 열었습니다. (자동화 실패: {e})"
        except Exception as e2: return False, f"실행 실패: {e2}"

def _launch_browser_print_dialog(xlsx_path: str) -> None:
    st.session_state["wl_print_panel"] = False
    abs_path = os.path.abspath(xlsx_path)
    if not os.path.exists(abs_path): st.error("인쇄용 파일이 없습니다."); return
    cache_k, meta_k = f"wl_print_html_cache_{abs_path}", f"wl_print_html_meta_{abs_path}"
    try: mtime = os.path.getmtime(abs_path)
    except OSError: mtime = 0.0
    cached, meta = st.session_state.get(cache_k), st.session_state.get(meta_k) or {}
    cache_ver = "v24"
    if isinstance(cached, str) and cached and meta.get("mtime") == mtime and meta.get("path") == abs_path and meta.get("ver") == cache_ver:
        stamped, preview_h = cached, int(meta.get("h") or 720)
    else:
        try:
            doc_html = render_worklog_view_html(abs_path, print_mode=True, auto_print=True, scale=1.0)
            _, raw_h = _worklog_sheet_pixel_size(abs_path)
            preview_h = min(920, max(420, int(raw_h * _a4_print_fit(*_worklog_sheet_pixel_size(abs_path), path=abs_path)) + 72))
        except Exception as e: st.error(f"인쇄 문서 준비 실패: {e}"); return
        nonce = int(st.session_state.get("wl_print_n", 0)) + 1
        st.session_state["wl_print_n"] = nonce
        stamped = doc_html.replace("<body>", f'<body data-wl-print="{nonce}">', 1)
        st.session_state[cache_k] = stamped
        st.session_state[meta_k] = {"mtime": mtime, "path": abs_path, "h": preview_h, "ver": cache_ver}
    components.html(stamped, height=preview_h, scrolling=True)
    st.caption("맞춤(줄여서)·축소·확대로 조절한 뒤 인쇄하세요.")

def _open_worklog_print_panel(xlsx_path: str, *, auto: bool = False) -> None:
    st.session_state["wl_print_panel"] = True
    st.session_state["wl_dialog_preview_path"] = os.path.abspath(xlsx_path)
    st.session_state["wl_print_auto_once"] = False

def _render_worklog_print_panel() -> bool:
    if not st.session_state.get("wl_print_panel"): return False
    path = st.session_state.get("wl_dialog_preview_path")
    st.markdown("""<style>.dashboard-filter-sticky, #dashboard-top-shield, #dashboard-sticky-spacer { display: none !important; } section.main .block-container { padding-top: 0.4rem !important; }</style>""", unsafe_allow_html=True)
    top1, top2 = st.columns([1.1, 3])
    with top1:
        if st.button("← 본화면으로", type="primary", width="stretch", key="wl_print_back_home"):
            st.session_state["wl_print_panel"] = False; st.session_state["wl_print_auto_once"] = False; _wl_rerun()
    with top2:
        st.markdown("##### 인쇄 미리보기")
        st.caption("자동 팝업 없음 · 「인쇄하기」만 누르면 인쇄 창이 열립니다.")
    if not path or not os.path.exists(str(path)):
        st.error("인쇄용 파일이 없습니다. 본화면으로 돌아가 다시 시도해 주세요.")
        return True
    path = str(path)
    try:
        print_html = render_worklog_view_html(path, print_mode=True, auto_print=False, scale=1.0)
        _, raw_h = _worklog_sheet_pixel_size(path)
        components.html(print_html, height=min(920, max(480, int(raw_h * _a4_print_fit(*_worklog_sheet_pixel_size(path), path=path)) + 72)), scrolling=True)
    except Exception as e: st.error(f"인쇄 미리보기 표시 실패: {e}")
    b1, b2 = st.columns(2)
    with b1:
        try:
            with open(path, "rb") as f: xbytes = f.read()
            st.download_button("엑셀 다운로드", data=xbytes, file_name=os.path.basename(path) or "일일업무일지.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch", key="wl_print_panel_dl")
        except Exception: pass
    with b2:
        if platform.system() == "Darwin":
            if st.button("Excel로 인쇄", width="stretch", key="wl_print_panel_excel"):
                ok, msg = open_excel_print_preview(path, prefer_print_dialog=True)
                (st.success if ok else st.warning)(msg)
    return True

@st.dialog("원본 엑셀 양식 미리보기", width="large")
def _worklog_form_preview_dialog() -> None:
    path = st.session_state.get("wl_dialog_preview_path")
    if not path or not os.path.exists(str(path)):
        st.error("미리보기 파일을 만들 수 없습니다. 템플릿·입력을 확인해 주세요.")
        return
    path = str(path)
    try:
        scale = _WL_PREVIEW_SCALE
        print_html = render_worklog_view_html(path, print_mode=False, auto_print=False, scale=scale)
        _, frame_h = _scaled_view_frame_size(path, scale)
        components.html(print_html, height=min(900, max(480, int(frame_h))), scrolling=True)
    except Exception as e:
        st.error(f"미리보기 표시 실패: {e}")
        return
    b1, b2, b3 = st.columns(3)
    with b1:
        try:
            with open(path, "rb") as f: xbytes = f.read()
            st.download_button("엑셀 다운로드", data=xbytes, file_name=os.path.basename(path) or "일일업무일지_미리보기.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", width="stretch", key="wl_dialog_dl_xlsx")
        except Exception: st.caption("엑셀 다운로드를 준비하지 못했습니다.")
    with b2:
        if st.button("인쇄창열기", width="stretch", key="wl_dialog_browser_print"): _launch_browser_print_dialog(path)
    with b3:
        if platform.system() == "Darwin":
            if st.button("Excel 인쇄 화면", width="stretch", key="wl_dialog_open_excel"):
                ok, msg = open_excel_print_preview(path, prefer_print_dialog=True)
                if ok: st.success(msg)
                else: st.warning(msg)

def _prepare_excel_preview(d: date, cells: dict) -> str:
    try: return prepare_print_xlsx(d, cells)
    except Exception: return _build_preview_file(d, cells)

def _render_month_calendar(selected: date, saved: set[str]) -> date | None:
    if "worklog_month" not in st.session_state: st.session_state["worklog_month"] = date(selected.year, selected.month, 1)
    month_anchor: date = st.session_state["worklog_month"]

    st.markdown("""<style>div[data-testid="stPopoverBody"] { max-width: 268px !important; width: 268px !important; padding: 0.35rem 0.45rem 0.5rem !important; } div[data-testid="stPopoverBody"] div[class*="st-key-wl_day_"] button, div[data-testid="stPopoverBody"] div[class*="st-key-wl_prev_month"] button, div[data-testid="stPopoverBody"] div[class*="st-key-wl_next_month"] button, div[data-testid="stPopoverBody"] div[class*="st-key-wl_today"] button { min-height: 1.55rem !important; height: 1.55rem !important; padding: 0 0.15rem !important; font-size: 0.7rem !important; line-height: 1.1 !important; border-radius: 5px !important; } div[data-testid="stPopoverBody"] [data-testid="stCaptionContainer"] { font-size: 0.65rem !important; margin-bottom: 0.15rem !important; } div[data-testid="stPopoverBody"] [data-testid="stHorizontalBlock"] { gap: 0.2rem !important; } div[data-testid="stPopoverBody"] [data-testid="column"] { padding: 0 !important; }</style>""", unsafe_allow_html=True)

    nav = st.columns([0.85, 0.85, 2.6, 1.2], gap="small")
    with nav[0]:
        if st.button("◀", key="wl_prev_month", width="stretch"):
            y, m = month_anchor.year, month_anchor.month - 1
            if m < 1: y, m = y - 1, 12
            st.session_state["worklog_month"] = date(y, m, 1)
            _wl_rerun()
    with nav[1]:
        if st.button("▶", key="wl_next_month", width="stretch"):
            y, m = month_anchor.year, month_anchor.month + 1
            if m > 12: y, m = y + 1, 1
            st.session_state["worklog_month"] = date(y, m, 1)
            _wl_rerun()
    with nav[2]:
        st.markdown(f"<div style='text-align:center;font-weight:700;font-size:12px;padding:2px 0;line-height:1.2;'>{month_anchor.year}년 {month_anchor.month}월</div>", unsafe_allow_html=True)
    with nav[3]:
        if st.button("오늘", key="wl_today", width="stretch"):
            today = date.today()
            st.session_state["worklog_selected"] = today
            st.session_state["worklog_month"] = date(today.year, today.month, 1)
            st.session_state["wl_date_sync"] = ""
            _wl_rerun()

    st.caption("• = 저장됨 · 날짜 탭 → 해당일")
    weeks = ["월", "화", "수", "목", "금", "토", "일"]
    head = st.columns(7, gap="small")
    for i, w in enumerate(weeks):
        color = "#DC2626" if i == 6 else ("#2563EB" if i == 5 else "#64748B")
        head[i].markdown(f"<div style='text-align:center;font-size:10px;font-weight:600;color:{color};line-height:1;padding:0 0 1px 0;'>{w}</div>", unsafe_allow_html=True)

    cal = calendar.Calendar(firstweekday=0)
    clicked = None
    for week in cal.monthdayscalendar(month_anchor.year, month_anchor.month):
        cols = st.columns(7, gap="small")
        for i, day in enumerate(week):
            with cols[i]:
                if day == 0:
                    st.markdown("<div style='height:1.55rem'></div>", unsafe_allow_html=True)
                    continue
                d = date(month_anchor.year, month_anchor.month, day)
                is_sel = d == selected
                has = d.isoformat() in saved
                label = f"{day}•" if has else f"{day}"
                if st.button(label, key=f"wl_day_{d.isoformat()}", width="stretch", type="primary" if is_sel else "secondary"): clicked = d
    return clicked



def _worklog_summary_html_body(full_html: str) -> str:
    if "<body>" in full_html:
        return full_html.split("<body>", 1)[1].rsplit("</body>", 1)[0].strip()
    return full_html


def _render_worklog_summary_block(selected: date, cells: dict) -> None:
    """요약 카드 — iframe(components.html) 대신 markdown으로 깜빡임 최소화."""
    draft_sig = json.dumps(cells, ensure_ascii=False, sort_keys=True)
    sum_sig_k = f"wl_sum_sig_{selected.isoformat()}"
    sum_html_k = f"wl_sum_html_{selected.isoformat()}"
    if st.session_state.get(sum_sig_k) != draft_sig:
        view_html = render_readable_preview_html(selected, cells)
        st.session_state[sum_sig_k] = draft_sig
        st.session_state[sum_html_k] = view_html
    else:
        view_html = st.session_state.get(sum_html_k)
        if not view_html:
            view_html = render_readable_preview_html(selected, cells)
            st.session_state[sum_html_k] = view_html
    body = _worklog_summary_html_body(view_html)
    st.markdown(
        f'<style>{_WL_SUMMARY_PREVIEW_CSS}</style>'
        f'<div class="wl-sum-preview" style="max-height:820px;overflow-y:auto;">{body}</div>',
        unsafe_allow_html=True,
    )


def _try_pull_remote_worklog_day(d: date) -> bool:
    """로컬에 없고 Gist에만 있을 때 해당 일자를 받아 달력·편집에 반영."""
    iso = d.isoformat()
    if os.path.isfile(worklog_path(d)):
        return False
    tried_k = f"wl_remote_pull_tried_{iso}"
    if st.session_state.get(tried_k):
        return False
    st.session_state[tried_k] = True
    try:
        from worklog_remote_sync import pull_worklog_day_from_remote

        if pull_worklog_day_from_remote(d, WORKLOG_DIR):
            _invalidate_saved_dates_cache()
            _invalidate_worklog_presence_cache(d)
            st.session_state.pop(_boot_key(d), None)
            st.session_state.pop(f"wl_open_ctx_{iso}", None)
            st.session_state.pop(f"wl_remote_pull_tried_{iso}", None)
            return True
    except Exception:
        pass
    return False


def _prepare_worklog_day_state(selected: date) -> None:
    """날짜별 위젯 초기화 + 저장 직후 pending 시드 (페이지 rerun 시 1회)."""
    iso = selected.isoformat()
    if not os.path.isfile(worklog_path(selected)):
        if _try_pull_remote_worklog_day(selected):
            iso = selected.isoformat()
    open_k = f"wl_open_ctx_{iso}"
    if open_k not in st.session_state:
        pres = detect_worklog_date_presence(selected, include_remote=False)
        st.session_state[open_k] = {"had_local": bool(pres.get("local")), "presence": pres}
    _init_widget_state(selected)
    pending = st.session_state.pop(f"wl_pending_sync_{iso}", None)
    if isinstance(pending, dict):
        _seed_day_entry_widgets(
            selected,
            pending.get("entries") or [{"client": "", "content": ""}],
            pending.get("next") or "",
            pending.get("notes") or "",
        )
        if pending.get("msg"):
            st.session_state[f"wl_flash_save_{iso}"] = {
                "msg": pending["msg"],
                "cloud_err": pending.get("cloud_err"),
            }


def _render_worklog_left_preview(selected: date) -> None:
    """왼쪽 요약/엑셀 — fragment 밖에서만 그림 (타이핑 시 재실행 안 됨)."""
    draft = _view_cells_for_preview(selected)
    st.markdown("##### 업무일지 보기")
    st.caption("입력은 바로 반영 · 왼쪽 요약은 칸 이동·저장·엑셀 미리보기 시 갱신")
    p1, p2 = st.columns(2)
    with p1:
        do_print = st.button(
            "엑셀 미리보기", width="stretch", key="wl_print_btn",
            help="현재 입력을 왼쪽에 원본 엑셀 양식으로 반영합니다.",
        )
    with p2:
        do_open_print = st.button(
            "인쇄창열기", width="stretch", key="wl_open_print_btn",
            help="브라우저 인쇄 창을 엽니다.", type="primary",
        )

    def _resolve_print_xlsx() -> str:
        cells_dl = _cells_from_widgets(selected)
        sig = json.dumps(cells_dl, ensure_ascii=False, sort_keys=True)
        sig_k = f"wl_print_cells_sig_{selected.isoformat()}"
        path_k = f"wl_print_cells_path_{selected.isoformat()}"
        out = os.path.abspath(_prepare_excel_preview(selected, cells_dl))
        prev_sig = st.session_state.get(sig_k)
        prev_path = st.session_state.get(path_k) or ""
        if prev_sig != sig or prev_path != out:
            st.session_state.pop(f"wl_print_html_cache_{out}", None)
            st.session_state.pop(f"wl_print_html_meta_{out}", None)
            if prev_path and prev_path != out:
                st.session_state.pop(f"wl_print_html_cache_{prev_path}", None)
                st.session_state.pop(f"wl_print_html_meta_{prev_path}", None)
        st.session_state[sig_k] = sig
        st.session_state[path_k] = out
        st.session_state[f"wl_left_excel_path_{selected.isoformat()}"] = out
        return out

    if do_open_print:
        try:
            xlsx_abs = _resolve_print_xlsx()
            _launch_browser_print_dialog(xlsx_abs)
        except Exception as e:
            st.error(f"인쇄 창을 열지 못했습니다: {e}")

    _left_excel_key = f"wl_left_excel_on_{selected.isoformat()}"
    _left_path_key = f"wl_left_excel_path_{selected.isoformat()}"
    if do_print:
        cells_now = _cells_from_widgets(selected)
        try:
            _publish_view_cells(selected, cells_now)
            xlsx_abs = _prepare_excel_preview(selected, cells_now)
            st.session_state[_left_excel_key] = True
            st.session_state[_left_path_key] = xlsx_abs
            st.session_state["wl_dialog_preview_path"] = xlsx_abs
            form_sig = json.dumps(cells_now, ensure_ascii=False)
            st.session_state[f"wl_form_sig_v14_{selected.isoformat()}"] = form_sig
            st.session_state["_wl_force_form_sig"] = form_sig
            st.success("✅ 엑셀 미리보기 화면이 갱신되었습니다.")
            draft = dict(cells_now)
        except Exception as e:
            st.error(f"미리보기 생성 중 오류가 발생했습니다: {e}")
            st.session_state[_left_excel_key] = False

    _show_excel_left = bool(st.session_state.get(_left_excel_key))
    if _show_excel_left:
        sw1, sw2 = st.columns([1, 1])
        with sw1:
            st.caption("원본 엑셀 양식 적용 중")
        with sw2:
            if st.button("요약 보기로", width="stretch", key=f"wl_left_to_summary_{selected.isoformat()}"):
                try:
                    _publish_view_cells(selected, _cells_from_widgets(selected))
                except Exception:
                    pass
                st.session_state[_left_excel_key] = False
        xlsx_left = st.session_state.get(_left_path_key) or ""
        if xlsx_left and os.path.exists(str(xlsx_left)):
            try:
                cells_view = _view_cells_for_preview(selected)
                live_sig = json.dumps(cells_view, ensure_ascii=False, sort_keys=True)
                sig_k = f"wl_left_excel_sig_v24_{selected.isoformat()}"
                html_k = f"wl_left_excel_html_v24_{selected.isoformat()}"
                h_k = f"wl_left_excel_h_v24_{selected.isoformat()}"
                scale_l = _WL_PREVIEW_SCALE
                if st.session_state.get(sig_k) != live_sig:
                    xlsx_left = _prepare_excel_preview(selected, cells_view)
                    st.session_state[_left_path_key] = xlsx_left
                    st.session_state[sig_k] = live_sig
                    excel_html = render_worklog_view_html(str(xlsx_left), print_mode=False, auto_print=False, scale=scale_l)
                    _, fh = _scaled_view_frame_size(str(xlsx_left), scale_l)
                    st.session_state[html_k] = excel_html
                    st.session_state[h_k] = fh
                else:
                    excel_html = st.session_state.get(html_k)
                    fh = st.session_state.get(h_k)
                    if not excel_html:
                        excel_html = render_worklog_view_html(str(xlsx_left), print_mode=False, auto_print=False, scale=scale_l)
                        _, fh = _scaled_view_frame_size(str(xlsx_left), scale_l)
                        st.session_state[html_k] = excel_html
                        st.session_state[h_k] = fh
                components.html(excel_html, height=min(1100, max(560, int(fh or 600))), scrolling=True)
                if st.button("크게 보기", width="stretch", key=f"wl_left_excel_big_{selected.isoformat()}"):
                    st.session_state["wl_dialog_preview_path"] = str(xlsx_left)
                    _worklog_form_preview_dialog()
            except Exception as e:
                st.warning(f"엑셀 양식 표시 실패: {e}")
                st.session_state[_left_excel_key] = False
        else:
            st.info("엑셀 미리보기 파일이 없습니다. 다시 「엑셀 미리보기」를 눌러 주세요.")
    else:
        try:
            _render_worklog_summary_block(selected, draft)
        except Exception as e:
            if _wl_quiet_ui():
                st.info("업무일지 요약을 표시하지 못했습니다. 입력 후 다시 확인해 주세요.")
            else:
                st.error(f"요약 보기 오류: {e}")


def _request_left_preview_refresh(d: date, *, focus_sig: str | None = None) -> None:
    """칸 이동·저장 등 — 왼쪽 요약 스냅샷 갱신 (fragment rerun만, 전체 앱 rerun 없음)."""
    iso = d.isoformat()
    if focus_sig is not None:
        prev = st.session_state.get(f"wl_left_focus_sig_{iso}")
        if prev == focus_sig:
            return
        st.session_state[f"wl_left_focus_sig_{iso}"] = focus_sig
    try:
        _publish_view_cells(d, _cells_from_widgets(d))
    except Exception:
        pass
    # 엑셀 미리보기 모드면 iframe 갱신 생략 (요약 보기로 돌아올 때 publish 반영)
    if st.session_state.get(f"wl_left_excel_on_{iso}"):
        return
    st.session_state["wl_need_left_refresh"] = True


def _wl_finish_edit_fragment() -> None:
    """편집 fragment 마무리 — 왼쪽 요약 갱신 시 fragment만 rerun (전체 앱·sync 생략)."""
    if st.session_state.pop("wl_need_left_refresh", None):
        _wl_rerun()



def _render_worklog_input_panel(selected: date) -> None:
    """오른쪽 게이지+입력. 칸 이동 시 published 스냅샷 갱신(동일 fragment rerun)."""
    saved = list_saved_worklog_dates()
    try:
        _gauge_usage = _content_row_usage(_read_editor_entries(selected))
    except Exception:
        _gauge_usage = _content_row_usage([])
    col_gauge, col_input = st.columns([0.14, 1], gap="small")
    with col_gauge:
            st.markdown("<div style='height:0.35rem'></div>", unsafe_allow_html=True)
            _render_row_remain_gauge(_gauge_usage, height_px=700)

    with col_input:
            st.markdown("##### 업무 입력")
            bar_date, bar_cal, bar_del = st.columns([2.4, 1.1, 0.7], gap="small")
            with bar_date:
                picked = st.date_input("업무일지 날짜", key="wl_date_pick", help="저장 후에도 날짜를 바꿀 수 있습니다.")
            with bar_cal:
                st.markdown("<div style='height:1.55rem'></div>", unsafe_allow_html=True)
                with st.popover("📅 달력", width="content"):
                    clicked = _render_month_calendar(selected, saved)
                    if clicked is not None and clicked != selected:
                        _clear_date_widget_state(selected)
                        st.session_state["worklog_selected"] = clicked
                        st.session_state["worklog_month"] = date(clicked.year, clicked.month, 1)
                        st.session_state["wl_date_sync"] = ""
                        st.rerun()
            with bar_del:
                st.markdown("<div style='height:1.55rem'></div>", unsafe_allow_html=True)
                with st.popover("삭제", width="content", key="wl_del_day_open"):
                    st.caption("이 날짜 일지 전체 삭제")
                    if st.button("확정", type="primary", width="content", key="wl_del_day_yes"):
                        delete_worklog_day(selected)
                        st.rerun()

            _iso_bar = selected.isoformat()
            _n_bar = int(st.session_state.get(f"wl_entry_count_{_iso_bar}", 1) or 1)
            if _render_worklog_special_chars(_iso_bar, _n_bar):
                _wl_rerun()

            if isinstance(picked, date) and picked != selected:
                if os.path.exists(worklog_path(picked)):
                    _clear_date_widget_state(selected)
                    st.session_state["worklog_selected"] = picked
                    st.session_state["worklog_month"] = date(picked.year, picked.month, 1)
                    st.session_state["wl_date_sync"] = ""
                    st.rerun()
                else:
                    try:
                        reassign_worklog_date(selected, picked)
                        st.session_state["wl_date_sync"] = ""
                        st.rerun()
                    except FileExistsError as e:
                        st.error(str(e))
                        st.session_state["wl_date_sync"] = ""
                        st.rerun()

            iso = selected.isoformat()
            ek = _entries_key(selected)
            if ek not in st.session_state or not st.session_state[ek]: st.session_state[ek] = [{"client": "", "content": ""}]

            def _seed_entry_widgets(entries_list: list[dict], next_txt: str, notes_txt: str) -> None:
                _seed_day_entry_widgets(selected, entries_list, next_txt, notes_txt)

            flash = st.session_state.pop(f"wl_flash_save_{iso}", None)
            if isinstance(flash, dict) and flash.get("msg"):
                if flash.get("cloud_err"):
                    st.warning(flash["msg"])
                else:
                    st.success(flash["msg"])

            add_clicked = st.session_state.pop(f"wl_do_add_{iso}", False)
            del_idx = st.session_state.pop(f"wl_do_del_{iso}", None)
            if add_clicked or isinstance(del_idx, int):
                entries = _read_editor_entries(selected)
                if isinstance(del_idx, int) and 0 <= del_idx < len(entries): entries.pop(del_idx)
                if add_clicked: entries.append({"client": "", "content": "", "lines": [""], "blank_after": 1})
                if not entries: entries = [{"client": "", "content": "", "blank_after": 1}]
                next_txt = str(st.session_state.get(f"wl_next_area_{iso}") or st.session_state.get(_next_key(selected), "") or "")
                notes_txt = str(st.session_state.get(f"wl_notes_area_{iso}") or st.session_state.get(_notes_key(selected), "") or "")
                _seed_entry_widgets(entries, next_txt, notes_txt)
                if add_clicked:
                    for _i in range(len(entries)): st.session_state[f"wl_exp_{iso}_{_i}"] = _i == len(entries) - 1
                    st.session_state[f"wl_force_expand_{iso}"] = len(entries) - 1
            else:
                entries = list(st.session_state[ek])
                for i, ent in enumerate(entries):
                    ck = f"wl_ent_c_{iso}_{i}"
                    gk = f"wl_ent_gap_{iso}_{i}"
                    if ck not in st.session_state: st.session_state[ck] = ent.get("client") or ""
                    if gk not in st.session_state: st.session_state[gk] = _entry_blank_after(ent, 1)
                    if int(st.session_state.get(_entry_client_count_key(iso, i), 0) or 0) <= 0:
                        cl0 = ent.get("client_lines")
                        if isinstance(cl0, list) and cl0: _seed_entry_clients(iso, i, cl0)
                        else: _seed_entry_clients(iso, i, ent.get("client") or "")
                    if int(st.session_state.get(_entry_line_count_key(iso, i), 0) or 0) <= 0:
                        lines0 = ent.get("lines")
                        if isinstance(lines0, list): _apply_entry_lines(iso, i, [str(x or "") for x in lines0], remount_comp=True)
                        else: _seed_entry_lines(iso, i, ent.get("content") or "")
                st.session_state[f"wl_entry_count_{iso}"] = len(entries)
                nk = f"wl_next_area_{iso}"
                ok = f"wl_notes_area_{iso}"
                if nk not in st.session_state: st.session_state[nk] = st.session_state.get(_next_key(selected), "")
                if ok not in st.session_state: st.session_state[ok] = st.session_state.get(_notes_key(selected), "")

            def _wl_entry_editor():
                d = st.session_state.get("worklog_selected") or selected
                iso2 = d.isoformat()
                n = int(st.session_state.get(f"wl_entry_count_{iso2}", 1) or 1)
                max_u = _content_line_units()
                # 저장은 2단계: 버튼 → flush 후 다음 런에서 실제 저장 (직전 입력 누락/구값 복원 방지)
                do_save = bool(st.session_state.pop(f"wl_do_save_{iso2}", None))

                _force_open = st.session_state.pop(f"wl_force_expand_{iso2}", None)
                if isinstance(_force_open, int): st.session_state[f"wl_exp_{iso2}_{_force_open}"] = True

                del_ln = st.session_state.pop(f"wl_do_del_ln_{iso2}", None)
                if isinstance(del_ln, (list, tuple)) and len(del_ln) == 2:
                    dei, dlj = int(del_ln[0]), int(del_ln[1])
                    cur = _lines_from_entry_widgets(iso2, dei, keep_trailing_empty=False)
                    if 0 <= dlj < len(cur): cur.pop(dlj)
                    if not cur: cur = [""]
                    _apply_entry_lines(iso2, dei, cur, focus_j=min(dlj, max(len(cur) - 1, 0)))

                ins_ln = st.session_state.pop(f"wl_do_insert_ln_{iso2}", None)
                if isinstance(ins_ln, (list, tuple)) and len(ins_ln) == 2:
                    _insert_line_after(iso2, int(ins_ln[0]), int(ins_ln[1]))

                del_cl = st.session_state.pop(f"wl_do_del_cl_{iso2}", None)
                if isinstance(del_cl, (list, tuple)) and len(del_cl) == 2:
                    dei, dlj = int(del_cl[0]), int(del_cl[1])
                    cur = _clients_from_widgets(iso2, dei, keep_trailing_empty=False)
                    if 0 <= dlj < len(cur): cur.pop(dlj)
                    if not cur: cur = [""]
                    _apply_entry_clients(iso2, dei, cur, focus_j=min(dlj, max(len(cur) - 1, 0)))

                ins_cl = st.session_state.pop(f"wl_do_insert_cl_{iso2}", None)
                if isinstance(ins_cl, (list, tuple)) and len(ins_cl) == 2:
                    _insert_client_after(iso2, int(ins_cl[0]), int(ins_cl[1]))

                ent_req = st.session_state.pop(f"wl_do_enter_cell_{iso2}", None)
                if isinstance(ent_req, dict):
                    _commit_enter_on_cell(str(ent_req.get("kind") or ""), iso2, int(ent_req.get("ei") or 0), int(ent_req.get("lj") or 0), str(ent_req.get("v") or ""))

                sp_req = st.session_state.pop(f"wl_do_special_{iso2}", None)
                if isinstance(sp_req, dict):
                    try: _sp_pos = int(sp_req.get("s") or 0)
                    except (TypeError, ValueError): _sp_pos = 0
                    _apply_special_insert(iso2, str(sp_req.get("key") or ""), str(sp_req.get("v") if sp_req.get("v") is not None else ""), _sp_pos, str(sp_req.get("ch") or ""))

                def _on_enter_trigger():
                    hook = st.session_state.get(f"wl_enter_hook_{iso2}") or {}
                    payload = str(hook.get("enter") or "") if isinstance(hook, dict) else ""
                    if not payload: return
                    done_k = f"wl_enter_done_{iso2}"
                    try:
                        obj = json.loads(payload)
                        key, val = str(obj.get("key") or ""), str(obj.get("v") or "")
                    except Exception:
                        key, val = payload.split(":", 1)[0], ""
                    sig = f"{key}\0{val}"
                    if st.session_state.get(done_k) == sig: return
                    st.session_state[done_k] = sig
                    m = re.match(r"^(wl_ent_ln|wl_ent_cl)_(\d{4}-\d{2}-\d{2})_(\d+)_(\d+)(?:_g\d+)?$", key)
                    if not m or m.group(2) != iso2: return
                    st.session_state[f"wl_do_enter_cell_{iso2}"] = {"kind": m.group(1), "ei": int(m.group(3)), "lj": int(m.group(4)), "v": val}

                def _on_focus_trigger():
                    hook = st.session_state.get(f"wl_enter_hook_{iso2}") or {}
                    if not isinstance(hook, dict):
                        return
                    fk = str(hook.get("focus") or "")
                    if fk.startswith("wl_next_area_") or fk.startswith("wl_notes_area_"):
                        st.session_state["wl_active_cell_key"] = fk
                        _request_left_preview_refresh(d, focus_sig=fk)
                        return
                    if fk.startswith("wl_ent_ln_") or fk.startswith("wl_ent_cl_"):
                        st.session_state["wl_active_cell_key"] = fk
                        st.session_state[f"wl_focus_ln_{iso2}"] = fk
                        _request_left_preview_refresh(d, focus_sig=fk)

                def _on_caret_trigger():
                    hook = st.session_state.get(f"wl_enter_hook_{iso2}") or {}
                    if not isinstance(hook, dict):
                        return
                    raw = hook.get("caret")
                    if not raw:
                        return
                    try:
                        obj = json.loads(str(raw))
                        fk, s, e = str(obj.get("key") or ""), int(obj.get("s") or 0), int(obj.get("e") or int(obj.get("s") or 0))
                    except Exception:
                        return
                    if fk.startswith("wl_next_area_") or fk.startswith("wl_notes_area_"):
                        st.session_state["wl_active_cell_key"] = fk
                        st.session_state["wl_active_cell_sel"] = (s, e)
                        return
                    if fk.startswith("wl_ent_ln_") or fk.startswith("wl_ent_cl_"):
                        st.session_state["wl_active_cell_key"] = fk
                        st.session_state[f"wl_focus_ln_{iso2}"] = fk
                        st.session_state["wl_active_cell_sel"] = (s, e)
                        st.session_state[f"wl_focus_caret_{iso2}"] = s

                st.caption("입력은 즉시 반영됩니다. 왼쪽 요약은 칸 이동·저장·엑셀 미리보기 시 갱신됩니다.")
                _live_entries = _read_editor_entries(d)
                _usage = _content_row_usage(_live_entries)
                _rem = _usage["remaining"]

                for i in range(n):
                    if int(st.session_state.get(_entry_client_count_key(iso2, i), 0) or 0) > 0:
                        _cl0 = _clients_from_widgets(iso2, i, keep_trailing_empty=False)
                        client_now = (_cl0[0] if _cl0 else "").strip()
                    else:
                        client_now = str(st.session_state.get(f"wl_ent_c_{iso2}_{i}", "") or "").strip()
                        if client_now: client_now = client_now.splitlines()[0].strip()
                    body_now = _content_from_entry_lines(iso2, i).strip().replace("\n", " ")
                    # 라벨은 현재 위젯/live만 사용 — stored 폴백은 지운 뒤에도 구제목 잔존/비어있음 오표시 원인
                    if len(body_now) > 24: body_now = body_now[:24] + "…"
                    label = f"항목 {i + 1}"
                    if client_now: label += f" · {client_now}"
                    elif body_now: label += f" · {body_now}"
                    else: label += " · (비어 있음)"

                    exp_key = f"wl_exp_{iso2}_{i}"
                    default_open = i == n - 1
                    if exp_key not in st.session_state: st.session_state[exp_key] = default_open

                    with st.expander(label, expanded=bool(st.session_state.get(exp_key)), key=exp_key):
                        if st.button("이 항목 삭제", key=f"wl_del_btn_{iso2}_{i}", use_container_width=True):
                            st.session_state[f"wl_do_del_{iso2}"] = i
                            _wl_rerun()

                        _cu = _client_line_units()
                        if int(st.session_state.get(_entry_client_count_key(iso2, i), 0) or 0) <= 0:
                            stored_e = st.session_state.get(_entries_key(d)) or []
                            if i < len(stored_e) and isinstance(stored_e[i].get("client_lines"), list):
                                _seed_entry_clients(iso2, i, stored_e[i].get("client_lines") or [""])
                            else:
                                _seed_entry_clients(iso2, i, str(st.session_state.get(f"wl_ent_c_{iso2}_{i}", "") or ""))
                        if int(st.session_state.get(_entry_line_count_key(iso2, i), 0) or 0) <= 0:
                            lines0 = None
                            stored_e = st.session_state.get(_entries_key(d)) or []
                            if i < len(stored_e) and isinstance(stored_e[i].get("lines"), list): lines0 = stored_e[i].get("lines")
                            if isinstance(lines0, list): _apply_entry_lines(iso2, i, [str(x or "") for x in lines0], remount_comp=True)
                            else: _seed_entry_lines(iso2, i, str(st.session_state.get(f"wl_ent_t_{iso2}_{i}", "") or ""))

                        if i == 0:
                            st.markdown(
                                """<style>div[class*="st-key-wl_clients_comp_"], div[class*="st-key-wl_lines_comp_"] { width: 100%; } div[class*="st-key-wl_clients_comp_"] .wl-lines, div[class*="st-key-wl_lines_comp_"] .wl-lines { margin: 0; } div[data-testid="stHorizontalBlock"]:has(div[class*="st-key-wl_clients_comp_"]) { align-items: flex-start !important; }</style>""",
                                unsafe_allow_html=True,
                            )

                        # 화면 비율 조절 (거래처 칸 넓힘)
                        col_client, col_content = st.columns([2.0, 6.0], gap="small")

                        with col_client: _mount_entry_client_editor(iso2, i, _cu)
                        with col_content: _mount_entry_lines_editor(iso2, i, max_u)

                        _filled = len(_lines_from_entry_widgets(iso2, i, keep_trailing_empty=False))
                        gap_key = f"wl_ent_gap_{iso2}_{i}"
                        if gap_key not in st.session_state: st.session_state[gap_key] = _entry_blank_after((_live_entries[i] if i < len(_live_entries) else None), 1)

                        st.markdown("<hr style='margin:16px 0 12px;border:none;border-top:1px dashed #E2E8F0;'>", unsafe_allow_html=True)
                        st.number_input("다음 항목 전 빈 칸 수", min_value=0, max_value=10, step=1, key=gap_key, help="이 항목 다음에 원본 엑셀에서 비워 둘 행 수.")
                        _ent_rows = (_usage["per_entry"][i] if i < len(_usage["per_entry"]) else max(_filled, 1))
                        st.caption(f"이 항목 약 {_ent_rows}행 사용 · 전체 남은 {_rem}행 (마지막 칸 G{_usage['last_row']})")

                if st.button("＋ 항목 추가", key=f"wl_add_btn_{iso2}", width="stretch"):
                    st.session_state[f"wl_do_add_{iso2}"] = True
                    _wl_rerun()

                st.markdown("<div style='font-size:12px;font-weight:700;color:#334155;margin:12px 0 4px;'>익일업무 <span style='font-weight:500;color:#94A3B8;'>(줄바꿈 = 항목 구분 · Enter=다음 줄)</span></div>", unsafe_allow_html=True)
                st.markdown(
                    f"""<style>
                    div[class*="st-key-wl_next_area_"] textarea,
                    div[class*="st-key-wl_notes_area_"] textarea {{
                      font-family: {_WL_FONT_STACK} !important;
                      font-size: 11pt !important;
                      line-height: 1.45 !important;
                    }}
                    </style>""",
                    unsafe_allow_html=True,
                )
                st.text_area("익일업무", key=f"wl_next_area_{iso2}", label_visibility="collapsed", height=110)
                st.markdown("<div style='font-size:12px;font-weight:700;color:#334155;margin:12px 0 4px;'>특 이 사 항 <span style='font-weight:500;color:#94A3B8;'>(줄바꿈 = 항목 구분 · Enter=다음 줄)</span></div>", unsafe_allow_html=True)
                st.text_area("특이사항", key=f"wl_notes_area_{iso2}", label_visibility="collapsed", height=100)

                if st.button("저장", type="primary", width="stretch", key=f"wl_save_btn_{iso2}"):
                    # 클릭 직후 한 번 더 그려 입력 컴포넌트 값을 확정한 뒤 저장
                    st.session_state[f"wl_do_save_{iso2}"] = True
                    _wl_rerun()
                elif do_save:
                    try:
                        entries_now = _read_editor_entries(d)
                        usage_now = _content_row_usage(entries_now)
                        if usage_now.get("overflow"):
                            st.error(f"내용칸 용량 초과: {usage_now['used']}/{usage_now['total']}행. 칸을 줄이거나 항목 사이 빈 칸 수를 낮춘 뒤 다시 저장하세요.")
                        else:
                            next_txt = str(st.session_state.get(f"wl_next_area_{iso2}", "") or "")
                            notes_txt = str(st.session_state.get(f"wl_notes_area_{iso2}", "") or "")
                            cells = _pack_entries_to_cells(
                                d, entries_now,
                                [x.strip() for x in next_txt.splitlines() if x.strip()],
                                [x.strip() for x in notes_txt.splitlines() if x.strip()],
                            )
                            open_ctx = st.session_state.get(f"wl_open_ctx_{iso2}") or {}
                            had_local = bool(open_ctx.get("had_local")) or bool(st.session_state.get(f"wl_saved_ok_{iso2}"))
                            # 저장 클릭 = 로컬·아카이브·Drive·Cloud 반영 (Cloud에만 있던 날도 맥에서 저장 가능)
                            path = save_worklog_cells(d, cells, force=True, allow_overwrite=had_local)
                            st.session_state[f"wl_saved_ok_{iso2}"] = True
                            ctx = dict(open_ctx)
                            ctx["had_local"] = True
                            st.session_state[f"wl_open_ctx_{iso2}"] = ctx
                            _publish_view_cells(d, cells)
                            # 시드는 방금 읽은 입력값 그대로 — cells 왕복으로 구형/변형 값이 되살아나지 않게
                            seed_entries = []
                            for ent in entries_now:
                                lines = ent.get("lines")
                                if not isinstance(lines, list):
                                    lines = _chunk_text(str(ent.get("content") or ""), _content_line_units()) or []
                                clines = ent.get("client_lines")
                                if not isinstance(clines, list):
                                    clines = _entry_client_lines(ent)
                                seed_entries.append(
                                    {
                                        "client": str(ent.get("client") or ""),
                                        "client_lines": list(clines),
                                        "content": str(ent.get("content") or ""),
                                        "lines": [str(x or "") for x in lines],
                                        "blank_after": _entry_blank_after(ent, 1),
                                    }
                                )
                            if not seed_entries:
                                seed_entries = [{"client": "", "content": "", "lines": [], "client_lines": [], "blank_after": 1}]
                            st.session_state[_entries_key(d)] = seed_entries
                            st.session_state[_next_key(d)] = next_txt
                            st.session_state[_notes_key(d)] = notes_txt
                            arch = (st.session_state.get("wl_last_archive_path") or "")
                            arch_target = st.session_state.get("wl_last_archive_target") or describe_worklog_archive_target(d)
                            drv = st.session_state.get("wl_last_drive_path") or ""
                            mdrv = st.session_state.get("wl_last_drive_month_path") or ""
                            gist = st.session_state.get("wl_last_cloud_gist") or ""
                            cerr = st.session_state.get("wl_last_cloud_err") or ""
                            drv_cf = st.session_state.get("wl_last_drive_conflict") or ""
                            msg = f"저장 완료: {os.path.basename(path)}"
                            if arch_target and not _wl_quiet_ui():
                                msg += f" · {arch_target}"
                            elif arch and not _wl_quiet_ui():
                                _sh = st.session_state.get("wl_last_archive_sheet") or worklog_archive_sheet_title(d)
                                msg += f" · 일지/{d.year}/{os.path.basename(arch)}#{_sh}"
                            if drv:
                                msg += " · Drive"
                            if mdrv and not _wl_quiet_ui():
                                msg += f" · Drive일지/{d.year}"
                            if gist:
                                msg += f" · Cloud OK (gist `{gist}`)"
                            elif cerr == "duplicate_date":
                                msg += " · Cloud: 선입력본 유지(덮어쓰기 안 함)"
                            elif cerr:
                                msg += f" · Cloud 실패: {cerr}"
                            elif not _wl_quiet_ui():
                                msg += " · Cloud미연동: secrets에 github_token"
                            if drv_cf and not drv:
                                msg += f" · Drive: 선입력본 유지({drv_cf})"
                            st.session_state[f"wl_pending_sync_{iso2}"] = {
                                "entries": seed_entries,
                                "next": next_txt,
                                "notes": notes_txt,
                                "msg": msg,
                                "cloud_err": cerr,
                            }
                            # 저장 직후 강제 pull은 구 원격본이 로컬을 덮을 수 있음 — push는 save_worklog_cells에서 이미 함
                            st.session_state["_wl_drive_sync_ts"] = time.time()
                            st.rerun()
                    except WorklogSaveBlockedError as e:
                        st.error(str(e))
                    except Exception as e:
                        if _wl_quiet_ui():
                            st.error("저장에 실패했습니다. 입력 내용을 확인한 뒤 다시 시도해 주세요.")
                        else:
                            st.error(f"저장 실패: {e}")

                focus_key = st.session_state.pop(f"wl_focus_ln_{iso2}", None)
                focus_caret = st.session_state.pop(f"wl_focus_caret_{iso2}", None)
                if isinstance(focus_key, str) and (focus_key.startswith("wl_ent_ln_") or focus_key.startswith("wl_ent_cl_")):
                    try:
                        _m = re.match(r"^wl_ent_(?:ln|cl)_\d{4}-\d{2}-\d{2}_(\d+)_", focus_key)
                        if _m: st.session_state[f"wl_exp_{iso2}_{int(_m.group(1))}"] = True
                    except Exception: pass
                else:
                    # 익일업무/특이사항 등은 포커스 강제 복원 금지
                    focus_key = None
                    focus_caret = None

                # 💡 [핵심] on_enter_change 이벤트 핸들러 형식 맞춤
                _WL_ENTER_HOOK(
                    key=f"wl_enter_hook_{iso2}",
                    data={"iso": iso2, "focus_key": focus_key if isinstance(focus_key, str) else "", "focus_caret": (int(focus_caret) if isinstance(focus_caret, (int, float)) else ""), "client_max_u": _client_line_units(), "content_max_u": _content_line_units()},
                    on_enter_change=_on_enter_trigger,
                    on_focus_change=_on_focus_trigger,
                    on_caret_change=_on_caret_trigger,
                )
                if st.session_state.get(f"wl_do_enter_cell_{iso2}"):
                    _request_left_preview_refresh(d)
                    _wl_finish_edit_fragment()
                    if not st.session_state.get("wl_need_left_refresh"):
                        _wl_rerun()
                ins_after = st.session_state.pop(f"wl_do_insert_ln_{iso2}", None)
                if isinstance(ins_after, (list, tuple)) and len(ins_after) == 2:
                    _insert_line_after(iso2, int(ins_after[0]), int(ins_after[1]))
                    _wl_rerun()
                ins_cl_after = st.session_state.pop(f"wl_do_insert_cl_{iso2}", None)
                if isinstance(ins_cl_after, (list, tuple)) and len(ins_cl_after) == 2:
                    _insert_client_after(iso2, int(ins_cl_after[0]), int(ins_cl_after[1]))
                    _wl_rerun()

            _n_now = int(st.session_state.get(f"wl_entry_count_{iso}", 1) or 1)
            # 날짜변경 등으로 위쪽 처리가 스킵된 pending 대비
            if _process_pending_special_char(iso, _n_now):
                _wl_rerun()
            _sp_msg = st.session_state.pop("wl_special_msg", None)
            if _sp_msg:
                st.caption(_sp_msg)
            _wl_entry_editor()


def _maybe_sync_worklog_remote() -> None:
    """페이지 전체 rerun 시에만 Drive/Gist 동기화 (fragment 입력 rerun 제외)."""
    if st.session_state.pop("wl_skip_sync_once", None):
        return
    try:
        from drive_autoload import sync_worklog_bidirectional
        from worklog_remote_sync import sync_worklog_remote

        _now = time.time()
        _prev = float(st.session_state.get("_wl_drive_sync_ts") or 0)
        _force = bool(st.session_state.pop("_wl_drive_sync_force", None))
        _on_cloud = _wl_is_streamlit_cloud()
        _sync_iv = 15 if _on_cloud else 120
        if not (_force or (_now - _prev >= _sync_iv)):
            return
        st.session_state["_wl_drive_sync_ts"] = _now
        _wl_sync: dict = {"ok": True, "skipped": True, "copied": [], "conflicts": []}
        _remote_sync: dict = {"ok": True, "skipped": True, "copied": [], "conflicts": []}
        if not _on_cloud:
            _wl_sync = sync_worklog_bidirectional(WORKLOG_DIR, force=_force)
        try:
            _remote_sync = sync_worklog_remote(WORKLOG_DIR, force=_force)
        except Exception as _re:
            _remote_sync = {
                "ok": False,
                "skipped": False,
                "copied": [],
                "conflicts": [],
                "error": str(_re),
            }
        st.session_state["_wl_last_wl_sync"] = _wl_sync
        st.session_state["_wl_last_remote_sync"] = _remote_sync
        _conflicts: list[str] = []
        for _src in (_wl_sync, _remote_sync):
            if isinstance(_src, dict):
                for _c in (_src.get("conflicts") or []):
                    if _c not in _conflicts:
                        _conflicts.append(_c)
        if _conflicts:
            st.session_state["_wl_sync_conflicts"] = _conflicts
        _copied_n = len((_wl_sync or {}).get("copied") or []) + len((_remote_sync or {}).get("copied") or [])
        if _copied_n:
            _invalidate_saved_dates_cache()
            try:
                from worklog_remote_sync import invalidate_gist_days_cache
                invalidate_gist_days_cache()
            except Exception:
                pass
    except Exception:
        pass


def _render_worklog_sync_ui() -> None:
    """동기화 결과·충돌 안내 (fragment 밖)."""
    try:
        from worklog_remote_sync import remote_sync_configured, resolve_gist_id

        _on_cloud = _wl_is_streamlit_cloud()
        _wl_sync = st.session_state.get("_wl_last_wl_sync") or {}
        _remote_sync = st.session_state.get("_wl_last_remote_sync") or {}
        _copied_n = len((_wl_sync or {}).get("copied") or []) + len((_remote_sync or {}).get("copied") or [])
        if _copied_n:
            st.caption(f"일지 동기화 · {_copied_n}개" + (" (Gist)" if _on_cloud else " (Drive/Cloud)"))
        elif _on_cloud and isinstance(_remote_sync, dict) and _remote_sync.get("error") and not _remote_sync.get("skipped"):
            st.warning(f"Gist 동기화 실패: {_remote_sync.get('error')}")
        if _on_cloud and remote_sync_configured():
            if st.button("↻ Gist에서 일지 가져오기", key="wl_gist_pull_btn", width="stretch"):
                st.session_state["_wl_drive_sync_force"] = True
                st.rerun()
        if st.session_state.get("_wl_sync_conflicts"):
            _cf = list(st.session_state.get("_wl_sync_conflicts") or [])
            st.warning(
                "로컬·클라우드 일지가 다릅니다(자동 덮어쓰기 안 함): "
                + ", ".join(_cf[:8])
                + ("…" if len(_cf) > 8 else "")
            )
            c1, c2, c3 = st.columns(3)
            with c1:
                if st.button("이 기기 → 클라우드", key="wl_cf_push_local", width="stretch", help="이 기기 내용으로 Drive·Cloud를 맞춥니다."):
                    try:
                        from drive_autoload import resolve_drive_conflict
                        from worklog_remote_sync import resolve_remote_conflict

                        for name in list(_cf):
                            try:
                                resolve_drive_conflict(name, WORKLOG_DIR, prefer="local")
                            except Exception:
                                pass
                            try:
                                resolve_remote_conflict(name, WORKLOG_DIR, prefer="local")
                            except Exception:
                                pass
                        st.session_state.pop("_wl_sync_conflicts", None)
                        st.session_state["_wl_drive_sync_force"] = True
                        _invalidate_saved_dates_cache()
                        st.rerun()
                    except Exception as e:
                        st.error(str(e) if not _wl_quiet_ui() else "동기화에 실패했습니다.")
            with c2:
                if st.button("클라우드 → 이 기기", key="wl_cf_pull_drive", width="stretch", help="Cloud·Drive 내용으로 이 기기를 맞춥니다."):
                    try:
                        from drive_autoload import resolve_drive_conflict
                        from worklog_remote_sync import resolve_remote_conflict

                        for name in list(_cf):
                            try:
                                resolve_remote_conflict(name, WORKLOG_DIR, prefer="cloud")
                            except Exception:
                                pass
                            try:
                                resolve_drive_conflict(name, WORKLOG_DIR, prefer="drive")
                            except Exception:
                                pass
                        st.session_state.pop("_wl_sync_conflicts", None)
                        st.session_state["_wl_drive_sync_force"] = True
                        _invalidate_saved_dates_cache()
                        st.rerun()
                    except Exception as e:
                        st.error(str(e) if not _wl_quiet_ui() else "동기화에 실패했습니다.")
            with c3:
                if st.button("나중에", key="wl_cf_dismiss", width="stretch"):
                    st.session_state.pop("_wl_sync_conflicts", None)
                    st.rerun()
        elif not remote_sync_configured():
            if not st.session_state.get("_wl_remote_setup_hint"):
                st.session_state["_wl_remote_setup_hint"] = True
                if _wl_quiet_ui():
                    st.caption("Cloud↔로컬 양방향: secrets에 github_token 을 넣으면 저장 시 서로 보입니다.")
                else:
                    st.info(
                        "로컬↔Cloud 양방향 연동: `.streamlit/secrets.toml` 에 "
                        "`github_token` (및 선택 `worklog_gist_id`) 을 넣으세요. "
                        "첫 저장 시 Gist가 만들어지고, 같은 값을 Cloud secrets에도 넣으면 "
                        "한쪽 저장이 다른쪽에 바로 보입니다."
                    )
        else:
            _gid = resolve_gist_id(WORKLOG_DIR)
            if _gid and not st.session_state.get("_wl_remote_ready_hint"):
                st.session_state["_wl_remote_ready_hint"] = True
                st.caption(f"로컬↔Cloud 양방향 연동 활성 · gist {_gid[:8]}…")
            elif not _gid and not st.session_state.get("_wl_remote_first_save_hint"):
                st.session_state["_wl_remote_first_save_hint"] = True
                st.caption("양방향 연동: 한 번 저장하면 Cloud Gist가 생성됩니다. 생성된 id를 Cloud secrets의 worklog_gist_id 에 넣으세요.")
    except Exception:
        pass


def render_worklog_tab(latest_update_str: str = "") -> None:
    if load_workbook is None:
        st.error("openpyxl 이 필요합니다. `pip install openpyxl` 후 다시 실행하세요.")
        return
    try:
        _ensure_dirs()
        if not os.path.exists(WORKLOG_TEMPLATE):
            if _wl_quiet_ui(): st.error("업무일지 템플릿이 없습니다. 사이드바에서 자료를 올린 뒤 `uploaded_cache/worklog/template.xlsx` 를 준비하세요.")
            else: st.error("업무일지 템플릿을 찾을 수 없습니다. `Desktop/업무일지.xlsx` 또는 `uploaded_cache/worklog/template.xlsx` 를 준비하세요.")
            return
    except Exception as e:
        if _wl_quiet_ui(): st.error("업무일지 템플릿을 준비하지 못했습니다. 파일을 다시 확인해 주세요.")
        else: st.error(f"템플릿 준비 실패: {e}")
        return

    # 반영 확인용 + 예전 미리보기 HTML 캐시 무효화(한 번)
    st.markdown(
        f"<link rel='stylesheet' href='https://fonts.googleapis.com/css2?family=Nanum+Myeongjo:wght@400;700&display=swap'>"
        f"<style>section.main {{ font-family:{_WL_FONT_STACK}; }}</style>",
        unsafe_allow_html=True,
    )
    st.caption(f"업무일지 빌드 {_WL_UI_BUILD}")
    _arch_root = resolve_worklog_archive_root()
    if _arch_root:
        st.caption(f"월별 저장 경로: `{_arch_root}/{{연도}}/{{N}}월.xlsx` (날짜=시트명)")
    else:
        st.caption("월별 저장 경로: `Desktop/업무/일지/{연도}/{N}월.xlsx` (Google Drive 동기화 시 「다른 컴퓨터/내 컴퓨터/Desktop/업무/일지」)")
    try:
        from worklog_remote_sync import cloud_sync_status

        _cs = cloud_sync_status(WORKLOG_DIR)
        if not _cs.get("token"):
            if _wl_is_streamlit_cloud():
                st.caption("☁ Gist **미연동** — Streamlit Cloud **Settings → Secrets** 에 `github_token`, `worklog_gist_id` 필요")
            else:
                st.caption("☁ Cloud: **미연동** — `.streamlit/secrets.toml` 에 `github_token` 넣고 Streamlit 재시작")
        elif _cs.get("gist_id"):
            st.caption(f"☁ Gist **연동됨** · `worklog_gist_id = \"{_cs['gist_id']}\"`")
        else:
            st.caption("☁ Gist: 토큰 OK · **저장**하면 gist id 생성")
    except Exception:
        pass
    if not st.session_state.get("_wl_cache_bust_v24"):
        st.session_state["_wl_cache_bust_v24"] = True
        for k in list(st.session_state.keys()):
            if not isinstance(k, str):
                continue
            if k.startswith("wl_left_excel_html_") or k.startswith("wl_left_excel_sig_") or k.startswith("wl_print_html_cache_"):
                st.session_state.pop(k, None)


    if "worklog_selected" not in st.session_state:
        st.session_state["worklog_selected"] = date.today()
    selected: date = st.session_state["worklog_selected"]

    _prepare_worklog_day_state(selected)
    _maybe_sync_worklog_remote()

    if st.session_state.get("wl_print_panel"):
        _render_worklog_print_panel()
        return

    if st.session_state.get("wl_date_sync") != selected.isoformat():
        st.session_state["wl_date_pick"] = selected
        st.session_state["wl_date_sync"] = selected.isoformat()

    st.markdown(
        """<style>div[class*="st-key-wl_del_day_open"] button, div[class*="st-key-wl_del_day_yes"] button { font-size: 0.72rem !important; padding: 0.12rem 0.4rem !important; min-height: 1.55rem !important; }</style>""",
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <style>
        div[data-testid="column"]:nth-of-type(2) {
            position: sticky;
            top: 4rem;
            align-self: flex-start;
            z-index: 99;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    _render_worklog_sync_ui()

    @st.fragment
    def _worklog_body() -> None:
        sel: date = st.session_state.get("worklog_selected") or selected
        col_preview, col_edit = st.columns([1, 1.14], gap="small")
        with col_preview:
            _render_worklog_left_preview(sel)
        with col_edit:
            _render_worklog_input_panel(sel)
        _wl_finish_edit_fragment()

    _worklog_body()
