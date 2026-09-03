"""왼쪽 업무일지 미리보기 — 입력 중 iframe 재부착 없이 칸만 패치."""
from __future__ import annotations

import os
import tempfile
import unittest
from datetime import date

from openpyxl import Workbook


class WorklogPreviewLiveTest(unittest.TestCase):
    def setUp(self):
        import worklog_tab as wt

        self.wt = wt

    def test_preview_cell_html_matches_excel_escape(self):
        self.assertEqual(self.wt._wl_preview_cell_html("G8", "상세하러간다"), "상세하러간다")
        self.assertEqual(self.wt._wl_preview_cell_html("G8", "A B"), "A&nbsp;B")
        self.assertEqual(self.wt._wl_preview_cell_html("G8", "A\nB"), "A<br>B")
        self.assertEqual(self.wt._wl_preview_cell_html("G8", "<x>"), "&lt;x&gt;")
        self.assertEqual(self.wt._wl_preview_cell_html("D40", ""), "&nbsp;")
        self.assertEqual(self.wt._wl_preview_cell_html("D40", self.wt._WL_SOFT_BLANK), "&nbsp;")
        self.assertEqual(self.wt._wl_preview_cell_html("G8", "   "), "")

    def test_preview_patches_cover_date_and_body_cells(self):
        d = date(2026, 9, 4)
        cells = self.wt._empty_cells(d)
        cells["G8"] = "상세하러간다"
        cells["C8"] = "거래처"
        cells["D40"] = "내일"
        patches = self.wt._wl_preview_patches(cells)
        self.assertEqual(patches["G8"], "상세하러간다")
        self.assertEqual(patches["C8"], "거래처")
        self.assertEqual(patches["D40"], "내일")
        self.assertIn("date", patches)
        self.assertIn("G39", patches)
        self.assertIn("C39", patches)
        self.assertEqual(patches["D41"], "&nbsp;")

    def test_workbook_html_marks_live_cells(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            wb = Workbook()
            ws = wb.active
            ws["C5"] = "2026-09-04 (금)"
            ws["C8"] = "거래처"
            ws["G8"] = "상세하러간다"
            ws["D40"] = "내일"
            wb.save(tmp.name)
            wb.close()
            html = self.wt.workbook_to_html(tmp.name, include_logo=False)
            self.assertIn('data-wl="date"', html)
            self.assertIn('data-wl="C8"', html)
            self.assertIn('data-wl="G8"', html)
            self.assertIn('data-wl="D40"', html)
            self.assertIn("상세하러간다", html)
        finally:
            os.unlink(tmp.name)

    def test_excel_host_html_is_fragment_not_document(self):
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        try:
            wb = Workbook()
            wb.save(tmp.name)
            wb.close()
            html = self.wt._excel_preview_host_html(tmp.name, scale=0.65)
            self.assertNotIn("<!DOCTYPE html>", html)
            self.assertIn("wl-sheet", html)
            self.assertIn("sheet-scale", html)
            self.assertIn("data-wl=", html)
        finally:
            os.unlink(tmp.name)


if __name__ == "__main__":
    unittest.main()
