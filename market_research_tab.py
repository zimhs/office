"""시장조사 탭 — Drive「Desktop/업무/시장조사」자료를 지역·산업단지·공급사로 정리해 조회."""
from __future__ import annotations

import io
import json
import os
import pickle
import re
import shutil
import unicodedata
import uuid
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    import market_research_cascade as mr_cascade
except Exception:  # pragma: no cover
    mr_cascade = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

MR_CACHE_DIR = os.path.join("uploaded_cache", "market_research")
MR_MANUAL_FILE = os.path.join(MR_CACHE_DIR, "manual_entries.json")
MR_UPLOAD_DIR = os.path.join(MR_CACHE_DIR, "uploads")
MR_UPLOAD_MANIFEST = os.path.join(MR_UPLOAD_DIR, "manifest.json")
MR_MANUAL_SOURCE = "직접입력"
MR_FRAME_PKL = os.path.join(MR_CACHE_DIR, "_merged_frame.pkl")
MR_FRAME_SIG = os.path.join(MR_CACHE_DIR, "_merged_frame.sig")
MR_BUNDLE_PKL = os.path.join(MR_CACHE_DIR, "_tab_bundle.pkl")
MR_BUNDLE_SIG = os.path.join(MR_CACHE_DIR, "_tab_bundle.sig")
_MR_DRIVE_SYNCED = False
MR_UPLOAD_KINDS = [
    ("자동", "파일명·양식으로 자동 감지"),
    ("지역시장조사", "시장조사(67) 스타일 — 상호/위치/현공급처"),
    ("방문조사", "김진혁·mail 스타일 — 업체명/공급처/사용가스"),
    ("LCO2경쟁사", "LCO2 경쟁사 취합본"),
    ("화성공장등록", "공장등록검색 (회사명·산업단지명)"),
    ("서진산업가스", "서진 거래처 시트"),
]
MR_DRIVE_CANDIDATES = [
    os.path.expanduser(
        "~/Library/CloudStorage/GoogleDrive-3023526@gmail.com/"
        "다른 컴퓨터/내 컴퓨터 (1)/Desktop/업무/시장조사"
    ),
    os.path.expanduser(
        "~/Library/CloudStorage/GoogleDrive-3023526@gmail.com/"
        "다른 컴퓨터/내 컴퓨터/업무/시장조사"
    ),
    os.path.join("Desktop", "업무", "시장조사"),
]

_SKIP_SHEET = re.compile(
    r"스케줄|원장|용기재고|재고현황|Sheet1\s*\(|^\d+$",
    re.I,
)

# 산업단지 분류 — 공장등록 DB 공식명 + 주소 키워드 + 읍면동/리 맵
# (업체마다 실시간 웹검색은 느리고 불안정 → 등록 DB·공개 단지 위치 규칙 사용)
_COMPLEX_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"동탄\s*도시\s*첨단|동탄도시첨단"), "화성동탄도시첨단산업단지"),
    (re.compile(r"동탄\s*일반|동탄산단|동탄\s*산업단지"), "화성동탄일반산업단지"),
    (re.compile(r"발안|팔탄.*발안|화성발안|발안리|구문천리|하길리"), "화성발안일반산업단지"),
    (re.compile(r"바이오\s*밸리|바이오밸리"), "화성바이오밸리일반산업단지"),
    (re.compile(r"전곡|전곡해양|전곡리|장외리"), "화성전곡해양일반산업단지"),
    (re.compile(r"마도\s*(일반|공단|산업)|화성마도|마도공단|마도안길|쌍송리|송정리|금당리|두곡리|석교리|해문리|백곡리"), "화성마도일반산업단지"),
    (re.compile(r"송산\s*테크노|테크노\s*파크|송산테크노|지화리|중송리"), "화성송산테크노파크일반산업단지"),
    (re.compile(r"송산\s*그린|그린\s*시티|송산그린|삼존리"), "송산그린시티 국가산업단지"),
    (re.compile(r"정남\s*(일반|산단|산업)|화성정남|음양리|귀래리|문학리|괘랑리|덕절리"), "화성정남일반산업단지"),
    (re.compile(r"향남\s*제약|제약\s*일반"), "화성향남제약일반산업단지"),
    (re.compile(r"향남\s*(지방|산단|산업|공단)|화성향남|동오리|백토리|화리현리|증거리|상신리|송곡리|길성리|상두리|수직리"), "화성향남지방산업단지"),
    (re.compile(r"주곡|화성주곡|주곡리"), "화성주곡일반산업단지"),
    (re.compile(r"화남|화성화남"), "화성화남일반산업단지"),
    (re.compile(r"장안\s*제?\s*2|장안2"), "장안제2첨단일반산업단지"),
    (re.compile(r"장안\s*제?\s*1|장안1|장안\s*첨단"), "장안제1첨단일반산업단지"),
    (re.compile(r"반월국가|반월\s*산단|반월동"), "반월국가산업단지"),
    (re.compile(r"시화|시화공단|시화MTV|정왕동"), "시화국가산업단지"),
    (re.compile(r"남동\s*(공단|산단|국가)|남동국가"), "남동국가산업단지"),
    (re.compile(r"포승|포승국가|평택.*포승"), "포승국가산업단지"),
    (re.compile(r"아산\s*국가|아산국가"), "아산국가산업단지"),
    (re.compile(r"평택\s*(산단|공단|일반)|평택공단"), "평택일반산업단지"),
    (re.compile(r"안성\s*(산단|공단|테크노)|안성공단"), "안성일반산업단지"),
    (re.compile(r"용인\s*(테크노|산단)|용인테크노"), "용인테크노밸리일반산업단지"),
    (re.compile(r"미음\s*산단|미음산단|과학산단"), "부산미음일반산업단지"),
    (re.compile(r"고덕\s*(산단|일반)|상몽산단|예산.*고덕"), "예산고덕일반산업단지"),
    (re.compile(r"증평\s*.*산단|증평2산단"), "증평일반산업단지"),
    (re.compile(r"메가폴리스|대소원"), "충주메가폴리스일반산업단지"),
    (re.compile(r"명지\s*(산단|국제)|녹산"), "부산명지녹산국가산업단지"),
    # 동 단위 (공장등록 고신뢰)
    (re.compile(r"영천동"), "화성동탄도시첨단산업단지"),
    (re.compile(r"방교동"), "화성동탄일반산업단지"),
]
# 주소에 「○○산단」만 있고 위 규칙에 안 걸릴 때 쓰는 보조 패턴
_COMPLEX_ADDR_TOKEN = re.compile(
    r"([가-힣A-Za-z0-9]{2,16})\s*산단(?:로|길|단지|공단)?"
)
_COMPLEX_AREA_ADDR = re.compile(
    r"(?:화성시|평택시|안산시|시흥시|안성시|용인시)?\s*"
    r"([가-힣]+(?:읍|면|동))"
    r"(?:\s*([가-힣]+리))?"
)

_REGION_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"화성|동탄|향남|팔탄|마도|봉담|우정|장안|비봉|양감|정남|서신"), "화성"),
    (re.compile(r"평택|송탄|팽성|청북|포승|안중"), "평택"),
    (re.compile(r"안성"), "안성"),
    (re.compile(r"오산"), "오산"),
    (re.compile(r"용인|처인|기흥|수지도"), "용인"),
    (re.compile(r"수원"), "수원"),
    (re.compile(r"안산|단원|상록"), "안산"),
    (re.compile(r"시흥|정왕|배곧"), "시흥"),
    (re.compile(r"광명"), "광명"),
    (re.compile(r"군포"), "군포"),
    (re.compile(r"의왕"), "의왕"),
    (re.compile(r"안양|동안|만안"), "안양"),
    (re.compile(r"과천"), "과천"),
    (re.compile(r"성남|분당|판교|수정구|중원구"), "성남"),
    (re.compile(r"광주(?:시|도)?(?!\s*광역시)|경기\s*광주"), "경기광주"),
    (re.compile(r"이천"), "이천"),
    (re.compile(r"여주"), "여주"),
    (re.compile(r"김포"), "김포"),
    (re.compile(r"부천"), "부천"),
    (re.compile(r"인천|남동|연수|부평|계양|서구\s*금곡|미추홀"), "인천"),
    (re.compile(r"서울|금천|가산|구로|영등포|강서|송파"), "서울"),
    (re.compile(r"당진"), "당진"),
    (re.compile(r"서산|태안|홍성|예산|아산|천안|공주|보령|논산|부여"), "충남"),
    (re.compile(r"음성|진천|증평|충주|청주|제천|옥천|영동|괴산"), "충북"),
    (re.compile(r"대전"), "대전"),
    (re.compile(r"세종"), "세종"),
    (re.compile(r"부산|강서구\s*미음|사하|사상"), "부산"),
    (re.compile(r"창원|진해|마산|김해|양산|함안|거제|통영|진주"), "경남"),
    (re.compile(r"울산"), "울산"),
    (re.compile(r"대구"), "대구"),
    (re.compile(r"광주광|광주시\s*서구|남구\s*서문대"), "광주"),
    (re.compile(r"여수|광양|순천|영암|무안|나주|목포"), "전남"),
    (re.compile(r"전주|군산|익산|완주"), "전북"),
    (re.compile(r"강원|원주|춘천|강릉"), "강원"),
    (re.compile(r"충남|충청남"), "충남"),
    (re.compile(r"충북|충청북"), "충북"),
    (re.compile(r"경기"), "경기기타"),
]


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    t = str(v).strip()
    if t.lower() in {"nan", "none", "null"}:
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", t))


def _company_key(name: str) -> str:
    """중복 병합용 업체 키 — ㈜/주식회사/공백·괄호 제거."""
    t = _s(name)
    t = re.sub(
        r"(주식회사|유한회사|유한책임회사|\(주\)|㈜|㈔|㈜)",
        "",
        t,
        flags=re.I,
    )
    t = re.sub(r"[\s\(\)（）\[\]【】·\.\-_/,，、xX×]+", "", t)
    return t.casefold()


def _merge_unique_text(values, *, sep: str = " · ", max_parts: int = 12) -> str:
    seen: set[str] = set()
    parts: list[str] = []
    for v in values:
        s = _s(v)
        if not s:
            continue
        for piece in re.split(r"\s*[|/·;；]\s*", s):
            piece = piece.strip()
            if not piece:
                continue
            key = piece.casefold()
            if key in seen:
                continue
            seen.add(key)
            parts.append(piece)
            if len(parts) >= max_parts:
                break
        if len(parts) >= max_parts:
            break
    return sep.join(parts)


def _best_text(values) -> str:
    """가장 길고 알찬 값 하나."""
    best = ""
    for v in values:
        s = _s(v)
        if len(s) > len(best):
            best = s
    return best


def _best_region(values) -> str:
    ranked = []
    for v in values:
        s = _s(v) or "미분류"
        ranked.append(s)
    for s in ranked:
        if s and s != "미분류":
            return s
    return ranked[0] if ranked else "미분류"


def _best_complex(values) -> str:
    """병합 시 산업단지 — 미분류가 아닌 값 중 가장 긴(공식명에 가까운) 것."""
    ranked = []
    for v in values:
        s = _s(v)
        if s and s != "미분류":
            ranked.append(s)
    if not ranked:
        return "미분류"
    return max(ranked, key=len)


def merge_duplicate_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """같은 업체키 행을 1건으로 병합. (병합 DF, 제거된 중복 건수).

    단일 행 그룹은 벡터로 통과시키고, 중복 키만 Python 병합한다.
    """
    if df.empty:
        return df, 0
    work = df.copy()
    if "산업단지" not in work.columns:
        work["산업단지"] = "미분류"
    work["업체키"] = work["업체명"].map(_company_key)
    work = work[work["업체키"].astype(str).str.len() >= 2]
    before = len(work)
    if before == 0:
        return work.reset_index(drop=True), 0
    vc = work["업체키"].value_counts(sort=False)
    dup_keys = set(vc[vc > 1].index)
    singles = work[~work["업체키"].isin(dup_keys)].copy()
    singles["병합건수"] = 1
    if not dup_keys:
        return singles.reset_index(drop=True), 0
    groups = []
    for key, g in work[work["업체키"].isin(dup_keys)].groupby("업체키", sort=False):
        groups.append(
            {
                "업체키": key,
                "업체명": _best_text(g["업체명"]),
                "지역": _best_region(g["지역"]),
                "산업단지": _best_complex(g["산업단지"]),
                "주소": _best_text(g["주소"]),
                "업종": _merge_unique_text(g["업종"], max_parts=6),
                "사용가스": _merge_unique_text(g["사용가스"], max_parts=8),
                "공급사": _merge_unique_text(g["공급사"], max_parts=8),
                "담당자": _merge_unique_text(g["담당자"], max_parts=6),
                "연락처": _merge_unique_text(g["연락처"], max_parts=6),
                "비고": _merge_unique_text(g["비고"], max_parts=10),
                "출처": _merge_unique_text(g["출처"], max_parts=8),
                "파일": _merge_unique_text(g["파일"], max_parts=6),
                "시트": _merge_unique_text(g["시트"], max_parts=8),
                "병합건수": int(len(g)),
            }
        )
    out = pd.concat([singles, pd.DataFrame(groups)], ignore_index=True, sort=False)
    removed = before - len(out)
    return out.reset_index(drop=True), removed


def _norm_header(v) -> str:
    t = _s(v)
    t = t.replace(" ", "").replace("\u3000", "")
    return t


def infer_region(*texts: str, sheet_hint: str = "") -> str:
    blob = " ".join([sheet_hint, *[t for t in texts if t]])
    if not blob.strip():
        return "미분류"
    for pat, name in _REGION_RULES:
        if pat.search(blob):
            return name
    return "미분류"


def infer_complex(*texts: str, explicit: str = "") -> str:
    """주소·비고·시트·공장등록 산업단지명으로 단지 추정."""
    ex = _s(explicit)
    if ex and ex not in {"-", "없음", "해당없음", "해당 없음"}:
        # 등록 DB 값이 이미 공식명이면 그대로
        return ex
    blob = " ".join(t for t in texts if t)
    if not blob.strip():
        return "미분류"
    for pat, name in _COMPLEX_RULES:
        if pat.search(blob):
            return name
    # 「○○산단로/길」 형태만 있으면 토큰으로 표기 (검색·필터용)
    m = _COMPLEX_ADDR_TOKEN.search(blob)
    if m:
        token = m.group(1).strip()
        if token and token not in {"일반", "지방", "국가", "도시", "첨단"}:
            return f"{token}산업단지"
    return "미분류"


def _complex_choices() -> list[str]:
    seen: list[str] = []
    for _, name in _COMPLEX_RULES:
        if name not in seen:
            seen.append(name)
    return seen


def _ms_default(prev, options: list[str]) -> list[str]:
    """Multiselect default는 options 부분집합이어야 Streamlit 오류가 안 남."""
    opts = set(options or [])
    return [x for x in (prev or []) if x in opts]


def _mr_is_cloud() -> bool:
    """Streamlit Community Cloud — Drive 로컬 동기화 생략."""
    try:
        env = (os.environ.get("STREAMLIT_RUNTIME_ENVIRONMENT") or "").strip().lower()
        if env == "cloud":
            return True
        return os.path.abspath(os.getcwd()).startswith("/mount/src")
    except Exception:
        return False


def _build_area_complex_lookup_from_records(fac: list[dict]) -> dict[str, str]:
    """공장등록 레코드 → 읍면동/리 산업단지 맵 (엑셀 재파싱 없음)."""
    from collections import Counter, defaultdict

    votes: dict[str, Counter] = defaultdict(Counter)
    for r in fac:
        cx = _s(r.get("산업단지"))
        if not cx or cx == "미분류":
            continue
        addr = _s(r.get("주소"))
        mm = _COMPLEX_AREA_ADDR.search(addr)
        if not mm:
            continue
        dong, ri = mm.group(1), mm.group(2) or ""
        if ri:
            votes[f"{dong}|{ri}"][cx] += 1
        votes[dong][cx] += 1
    out: dict[str, str] = {}
    for key, c in votes.items():
        top, n = c.most_common(1)[0]
        tot = sum(c.values())
        if "|" in key:
            ok = n >= 8 and n / tot >= 0.8
        else:
            ok = n >= 80 and n / tot >= 0.9
        if ok:
            out[key] = top
    return out


@st.cache_data(show_spinner=False, ttl=600)
def _area_complex_lookup(_cache_sig: str) -> dict[str, str]:
    """공장등록 주소(읍면동·리) → 산업단지 고신뢰 맵."""
    root = ensure_market_research_cache()
    fac: list[dict] = []
    for p in _list_xlsx(root):
        nfc = unicodedata.normalize("NFC", p.name)
        if "화성" in nfc and "공장" in nfc:
            fac = _parse_factory_registry(p)
            break
    return _build_area_complex_lookup_from_records(fac)


def _complex_from_area(addr: str, lookup: dict[str, str]) -> str:
    if not addr or not lookup:
        return ""
    mm = _COMPLEX_AREA_ADDR.search(addr)
    if not mm:
        return ""
    dong, ri = mm.group(1), mm.group(2) or ""
    if ri:
        hit = lookup.get(f"{dong}|{ri}")
        if hit:
            return hit
    return lookup.get(dong, "")


def enrich_unclassified_complexes(
    df: pd.DataFrame, lookup: dict[str, str]
) -> tuple[pd.DataFrame, int]:
    """미분류 행을 주소 읍면동/리 맵으로 재분류. (변경 DF, 재분류 건수)."""
    if df.empty or "산업단지" not in df.columns or not lookup:
        return df, 0
    out = df.copy()
    n = 0
    for i in out.index:
        if _s(out.at[i, "산업단지"]) not in {"", "미분류"}:
            continue
        cx = _complex_from_area(_s(out.at[i, "주소"]), lookup)
        if not cx:
            # 지역+주소 텍스트로 한 번 더
            cx2 = infer_complex(
                _s(out.at[i, "주소"]),
                _s(out.at[i, "지역"]),
                _s(out.at[i, "업체명"]),
                _s(out.at[i, "비고"]),
            )
            if cx2 != "미분류":
                cx = cx2
        if cx and cx != "미분류":
            out.at[i, "산업단지"] = cx
            n += 1
    return out, n


def _find_header_row(rows: list[tuple], keywords: list[str], scan: int = 12) -> int | None:
    keys = [k.replace(" ", "") for k in keywords]
    for i, row in enumerate(rows[:scan]):
        cells = [_norm_header(c) for c in row]
        hit = sum(1 for k in keys if any(k in c for c in cells if c))
        if hit >= max(2, min(3, len(keys) // 2 + 1)):
            return i
    return None


def _row_dict(header: list[str], row: tuple) -> dict[str, str]:
    out: dict[str, str] = {}
    for i, h in enumerate(header):
        if not h:
            continue
        val = _s(row[i]) if i < len(row) else ""
        if h in out and out[h] and val:
            out[h] = f"{out[h]} / {val}"
        elif h not in out or not out[h]:
            out[h] = val
    return out


def _pick(d: dict[str, str], *cands: str) -> str:
    for c in cands:
        c2 = c.replace(" ", "")
        for k, v in d.items():
            kn = k.replace(" ", "")
            if c2 == kn or c2 in kn or kn in c2:
                if v:
                    return v
    return ""


def ensure_market_research_cache(*, force: bool = False) -> str:
    """Drive 원본이 있으면 캐시로 동기화. 프로세스·세션당 1회만 시도."""
    global _MR_DRIVE_SYNCED
    os.makedirs(MR_CACHE_DIR, exist_ok=True)
    if not force:
        try:
            if _MR_DRIVE_SYNCED or st.session_state.get("_mr_cache_synced"):
                return MR_CACHE_DIR
        except Exception:
            if _MR_DRIVE_SYNCED:
                return MR_CACHE_DIR
    if _mr_is_cloud():
        try:
            st.session_state["_mr_cache_synced"] = True
        except Exception:
            pass
        _MR_DRIVE_SYNCED = True
        return MR_CACHE_DIR
    for cand in MR_DRIVE_CANDIDATES:
        if cand and os.path.isdir(cand):
            try:
                for root, _dirs, files in os.walk(cand):
                    rel = os.path.relpath(root, cand)
                    dest_root = (
                        MR_CACHE_DIR if rel == "." else os.path.join(MR_CACHE_DIR, rel)
                    )
                    os.makedirs(dest_root, exist_ok=True)
                    for name in files:
                        if name.startswith("~$") or name.startswith("."):
                            continue
                        if not name.lower().endswith((".xlsx", ".xls", ".csv")):
                            continue
                        src = os.path.join(root, name)
                        dst = os.path.join(dest_root, name)
                        try:
                            if (not os.path.exists(dst)) or (
                                os.path.getmtime(src) > os.path.getmtime(dst) + 1
                            ):
                                shutil.copy2(src, dst)
                        except Exception:
                            pass
                try:
                    st.session_state["_mr_cache_synced"] = True
                except Exception:
                    pass
                _MR_DRIVE_SYNCED = True
                return MR_CACHE_DIR
            except Exception:
                continue
    try:
        st.session_state["_mr_cache_synced"] = True
    except Exception:
        pass
    _MR_DRIVE_SYNCED = True
    return MR_CACHE_DIR


def _manual_entries_path() -> str:
    os.makedirs(MR_CACHE_DIR, exist_ok=True)
    return MR_MANUAL_FILE


def load_manual_entries() -> list[dict]:
    path = _manual_entries_path()
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [x for x in data if isinstance(x, dict)]
    except Exception:
        return []
    return []


def save_manual_entries(entries: list[dict]) -> None:
    path = _manual_entries_path()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)
    for cand in MR_DRIVE_CANDIDATES:
        if cand and os.path.isdir(cand):
            try:
                shutil.copy2(path, os.path.join(cand, "직접입력_시장조사.json"))
            except Exception:
                pass
            break


def _attach_complex(rec: dict, *, explicit: str = "") -> dict:
    """레코드에 산업단지 필드 보정."""
    ex = _s(explicit) or _s(rec.get("산업단지"))
    rec["산업단지"] = infer_complex(
        rec.get("주소", ""),
        rec.get("업체명", ""),
        rec.get("비고", ""),
        rec.get("시트", ""),
        explicit=ex,
    )
    return rec


def _manual_to_record(entry: dict) -> dict:
    name = _s(entry.get("업체명"))
    addr = _s(entry.get("주소"))
    region = _s(entry.get("지역")) or infer_region(addr, name)
    note = _s(entry.get("비고"))
    park = _s(entry.get("산업단지"))
    return _attach_complex(
        {
            "출처": MR_MANUAL_SOURCE,
            "파일": "manual_entries.json",
            "시트": "직접입력",
            "지역": region or "미분류",
            "업체명": name,
            "주소": addr,
            "업종": _s(entry.get("업종")),
            "사용가스": _s(entry.get("사용가스")),
            "공급사": _s(entry.get("공급사")),
            "담당자": _s(entry.get("담당자")),
            "연락처": _s(entry.get("연락처")),
            "비고": note,
            "산업단지": park,
        },
        explicit=park,
    )


def add_manual_entry(fields: dict) -> dict:
    name = _s(fields.get("업체명"))
    if len(name) < 2:
        raise ValueError("업체명을 2글자 이상 입력하세요.")
    entry = {
        "id": uuid.uuid4().hex[:12],
        "saved_at": datetime.now(timezone.utc)
        .astimezone()
        .strftime("%Y-%m-%d %H:%M:%S"),
        "업체명": name,
        "지역": _s(fields.get("지역")),
        "산업단지": _s(fields.get("산업단지")),
        "주소": _s(fields.get("주소")),
        "업종": _s(fields.get("업종")),
        "사용가스": _s(fields.get("사용가스")),
        "공급사": _s(fields.get("공급사")),
        "담당자": _s(fields.get("담당자")),
        "연락처": _s(fields.get("연락처")),
        "비고": _s(fields.get("비고")),
    }
    if not entry["지역"]:
        entry["지역"] = infer_region(entry["주소"], entry["업체명"])
    if not entry["산업단지"]:
        entry["산업단지"] = infer_complex(
            entry["주소"], entry["업체명"], entry["비고"]
        )
        if entry["산업단지"] == "미분류":
            entry["산업단지"] = ""
    entries = load_manual_entries()
    entries.insert(0, entry)
    save_manual_entries(entries)
    return entry


def delete_manual_entry(entry_id: str) -> bool:
    entries = load_manual_entries()
    n0 = len(entries)
    entries = [e for e in entries if str(e.get("id")) != str(entry_id)]
    if len(entries) == n0:
        return False
    save_manual_entries(entries)
    return True


def _invalidate_mr_loaded() -> None:
    for clear_fn in (
        load_market_research_frame,
        _cached_mr_cascade_index,
        _cached_mr_tab_bundle,
        _cached_cache_signature,
        _area_complex_lookup,
    ):
        try:
            clear_fn.clear()
        except Exception:
            pass
    for p in (MR_FRAME_PKL, MR_FRAME_SIG, MR_FRAME_PKL + ".tmp", MR_BUNDLE_PKL, MR_BUNDLE_SIG, MR_BUNDLE_PKL + ".tmp"):
        try:
            if os.path.exists(p):
                os.remove(p)
        except Exception:
            pass
    for k in (
        "_mr_data_warm",
        "_mr_data_sig",
        "_mr_df",
        "_mr_raw_n",
        "_mr_removed_n",
        "_mr_cascade",
        "_mr_ipad_bundle",
        "_mr_session_bundle",
    ):
        st.session_state.pop(k, None)


def _region_choices() -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for _pat, name in _REGION_RULES:
        if name not in seen and name != "경기기타":
            seen.add(name)
            names.append(name)
    names.append("미분류")
    return names


def _list_xlsx(root: str) -> list[Path]:
    out: list[Path] = []
    p = Path(root)
    if not p.is_dir():
        return out
    for f in sorted(p.rglob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        out.append(f)
    return out


def _load_upload_manifest() -> list[dict]:
    if not os.path.exists(MR_UPLOAD_MANIFEST):
        return []
    try:
        with open(MR_UPLOAD_MANIFEST, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [x for x in data if isinstance(x, dict)]
    except Exception:
        return []
    return []


def _save_upload_manifest(entries: list[dict]) -> None:
    os.makedirs(MR_UPLOAD_DIR, exist_ok=True)
    with open(MR_UPLOAD_MANIFEST, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)


def _guess_parse_kind(filename: str) -> str:
    nfc = unicodedata.normalize("NFC", filename or "")
    n = nfc.lower()
    if "lco2" in n or "경쟁사" in nfc:
        return "LCO2경쟁사"
    if "시장조사 (67)" in nfc or nfc.startswith("시장조사 (67)"):
        return "지역시장조사"
    if "화성" in nfc and "공장" in nfc:
        return "화성공장등록"
    if "김진혁" in nfc:
        return "방문조사"
    if "mail" in n or "시장조사ㅡ" in nfc or "시장조사-" in nfc:
        return "방문조사"
    if "서진" in nfc:
        return "서진산업가스"
    return "자동"


def _parse_by_kind(path: Path, kind: str) -> list[dict]:
    """업로드·기존 파일을 종류별로 파싱."""
    k = (kind or "자동").strip()
    if k == "자동":
        k = _guess_parse_kind(path.name)
        if k == "자동":
            # 양식 추정: 지역조사 → 방문조사 순
            rec = _parse_region_survey(path)
            if rec:
                return rec
            return _parse_visit_notes(path, "업로드조사")
    if k == "LCO2경쟁사":
        return _parse_lco2(path)
    if k == "지역시장조사":
        return _parse_region_survey(path)
    if k in {"방문조사", "방문조사(mail)", "방문조사(김진혁)"}:
        label = "업로드조사" if k == "방문조사" else k
        return _parse_visit_notes(path, label)
    if k == "화성공장등록":
        return _parse_factory_registry(path)
    if k == "서진산업가스":
        return _parse_seojin(path)
    return []


def _safe_upload_filename(original: str) -> str:
    base = unicodedata.normalize("NFC", Path(original or "upload.xlsx").name)
    base = re.sub(r"[^\w.\-가-힣()\[\] ]+", "_", base, flags=re.UNICODE)
    stem = Path(base).stem[:80] or "upload"
    suf = Path(base).suffix.lower()
    if suf not in {".xlsx", ".xls"}:
        suf = ".xlsx"
    stamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    return f"{stamp}_{stem}{suf}"


def save_uploaded_excel(file_obj, *, kind: str) -> dict:
    """탭에서 올린 엑셀을 uploads/에 저장하고 manifest에 기록."""
    os.makedirs(MR_UPLOAD_DIR, exist_ok=True)
    original = getattr(file_obj, "name", None) or "upload.xlsx"
    stored = _safe_upload_filename(original)
    dest = os.path.join(MR_UPLOAD_DIR, stored)
    data = file_obj.getbuffer() if hasattr(file_obj, "getbuffer") else file_obj.read()
    with open(dest, "wb") as f:
        f.write(data)
    # Drive에도 복사 (가능하면)
    for cand in MR_DRIVE_CANDIDATES:
        if cand and os.path.isdir(cand):
            try:
                drive_up = os.path.join(cand, "uploads")
                os.makedirs(drive_up, exist_ok=True)
                shutil.copy2(dest, os.path.join(drive_up, stored))
            except Exception:
                pass
            break
    entry = {
        "id": uuid.uuid4().hex[:12],
        "filename": stored,
        "original_name": unicodedata.normalize("NFC", Path(original).name),
        "kind": (kind or "자동").strip() or "자동",
        "uploaded_at": datetime.now(timezone.utc)
        .astimezone()
        .strftime("%Y-%m-%d %H:%M:%S"),
        "size": os.path.getsize(dest),
    }
    man = _load_upload_manifest()
    man.insert(0, entry)
    _save_upload_manifest(man)
    return entry


def delete_uploaded_excel(entry_id: str) -> bool:
    man = _load_upload_manifest()
    hit = None
    for e in man:
        if str(e.get("id")) == str(entry_id):
            hit = e
            break
    if not hit:
        return False
    path = os.path.join(MR_UPLOAD_DIR, hit.get("filename") or "")
    try:
        if os.path.isfile(path):
            os.remove(path)
    except OSError:
        pass
    man = [e for e in man if str(e.get("id")) != str(entry_id)]
    _save_upload_manifest(man)
    return True


def _read_sheet_rows(path: Path, sheet: str, max_rows: int = 20000) -> list[tuple]:
    if load_workbook is None:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        rows: list[tuple] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(tuple(row))
            if i >= max_rows:
                break
        return rows
    finally:
        wb.close()


def _parse_lco2(path: Path) -> list[dict]:
    if load_workbook is None or not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    try:
        for sn in wb.sheetnames:
            rows = []
            ws = wb[sn]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(tuple(row))
                if i >= 5000:
                    break
            hi = _find_header_row(
                rows, ["상호명", "소재지", "공급사", "위치", "제품명"]
            )
            if hi is None:
                continue
            header = [_norm_header(c) for c in rows[hi]]
            last_region = sn.strip() or "미분류"
            for row in rows[hi + 1 :]:
                d = _row_dict(header, row)
                name = _pick(d, "상호명", "상호", "업체명")
                if not name or len(name) < 2:
                    continue
                loc = _pick(d, "위치", "지역")
                addr = _pick(d, "소재지", "주소")
                if _pick(d, "지역") and len(_pick(d, "지역")) <= 6:
                    last_region = _pick(d, "지역")
                region = infer_region(loc, addr, name, sheet_hint=last_region)
                note = _pick(d, "비고", "대납업체", "월사용량")
                records.append(
                    _attach_complex(
                        {
                            "출처": "LCO2경쟁사",
                            "파일": path.name,
                            "시트": sn.strip(),
                            "지역": region,
                            "업체명": name,
                            "주소": addr or loc,
                            "업종": _pick(d, "고압가스종류", "업종"),
                            "사용가스": _pick(d, "제품명", "사용가스"),
                            "공급사": _pick(d, "공급사", "현공급처"),
                            "담당자": _pick(d, "담당", "담당자"),
                            "연락처": _pick(d, "연락처", "전화"),
                            "비고": note,
                        }
                    )
                )
    finally:
        wb.close()
    return records


def _parse_region_survey(path: Path) -> list[dict]:
    """시장조사 (67).xlsx — 지역 시트별 표준 양식."""
    if load_workbook is None or not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    try:
        for sn in wb.sheetnames:
            rows = []
            ws = wb[sn]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(tuple(row))
                if i >= 5000:
                    break
            hi = _find_header_row(
                rows, ["상호", "위치", "현공급처", "용기보유", "업종"]
            )
            if hi is None:
                # 헤더 없이 데이터만 있는 음성 시트 등
                hi = _find_header_row(rows, ["상호", "위치", "연락처"])
            header = (
                [_norm_header(c) for c in rows[hi]]
                if hi is not None
                else ["NO", "상호", "위치", "연락처", "담당", "업종", "용기보유현황", "현공급처", "비고"]
            )
            start = (hi + 1) if hi is not None else 4
            # 헤더가 약하면 고정 매핑
            if hi is None or not any("상호" in h for h in header):
                for row in rows[start:]:
                    vals = [_s(c) for c in row[:9]]
                    if len(vals) < 3 or not vals[1]:
                        continue
                    if vals[1] in {"상호", "상 호", "NO"}:
                        continue
                    addr = vals[2]
                    region = infer_region(addr, sn, sheet_hint=sn)
                    note = vals[8] if len(vals) > 8 else ""
                    records.append(
                        _attach_complex(
                            {
                                "출처": "지역시장조사",
                                "파일": path.name,
                                "시트": sn,
                                "지역": region,
                                "업체명": vals[1],
                                "주소": addr,
                                "업종": vals[5] if len(vals) > 5 else "",
                                "사용가스": vals[6] if len(vals) > 6 else "",
                                "공급사": vals[7] if len(vals) > 7 else "",
                                "담당자": vals[4] if len(vals) > 4 else "",
                                "연락처": vals[3] if len(vals) > 3 else "",
                                "비고": note,
                            }
                        )
                    )
                continue
            for row in rows[start:]:
                d = _row_dict(header, row)
                name = _pick(d, "상호", "상호명", "업체명")
                if not name or name in {"상호", "상 호"}:
                    continue
                addr = _pick(d, "위치", "주소", "소재지", "지역")
                region = infer_region(addr, sn, sheet_hint=sn)
                note = _pick(d, "비고", "비 고")
                records.append(
                    _attach_complex(
                        {
                            "출처": "지역시장조사",
                            "파일": path.name,
                            "시트": sn,
                            "지역": region,
                            "업체명": name,
                            "주소": addr,
                            "업종": _pick(d, "업종", "생산품목"),
                            "사용가스": _pick(d, "용기보유현황", "사용가스", "용기"),
                            "공급사": _pick(d, "현공급처", "공급처", "공급사"),
                            "담당자": _pick(d, "담당", "담당자"),
                            "연락처": _pick(d, "연락처", "전화", "전화번호"),
                            "비고": note,
                        }
                    )
                )
    finally:
        wb.close()
    return records


def _parse_visit_notes(path: Path, source_label: str) -> list[dict]:
    """김진혁/mail 등 — 시트명≈지역, 업체명·공급처·사용가스."""
    if load_workbook is None or not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    try:
        for sn in wb.sheetnames:
            if _SKIP_SHEET.search(sn):
                continue
            if "통합" in sn and "모든내용" in sn:
                pass  # keep
            rows = []
            ws = wb[sn]
            # 넓은 시트(화성)는 앞 열만
            for i, row in enumerate(ws.iter_rows(values_only=True, max_col=14)):
                rows.append(tuple(row))
                if i >= 8000:
                    break
            hi = _find_header_row(
                rows, ["업체명", "지역", "공급처", "사용가스", "담당자"]
            )
            if hi is None:
                continue
            header = [_norm_header(c) for c in rows[hi]]
            for row in rows[hi + 1 :]:
                d = _row_dict(header, row)
                name = _pick(d, "업체명", "상호", "입체명")
                if not name or len(name) < 2:
                    continue
                if name.endswith("x") and len(name) <= 3:
                    continue
                addr = _pick(d, "지역", "주소", "위치", "회사위치")
                region = infer_region(addr, sn, sheet_hint=sn)
                note = _pick(d, "비고", "특이사항", "세부사항")
                records.append(
                    _attach_complex(
                        {
                            "출처": source_label,
                            "파일": path.name,
                            "시트": sn,
                            "지역": region,
                            "업체명": name.rstrip("x").strip() or name,
                            "주소": addr,
                            "업종": _pick(d, "생산품목", "업종", "종목"),
                            "사용가스": _pick(d, "사용가스", "가스"),
                            "공급사": _pick(d, "공급처", "현공급처", "공급사"),
                            "담당자": _pick(d, "담당자", "담당"),
                            "연락처": "",
                            "비고": note,
                        }
                    )
                )
    finally:
        wb.close()
    return records


def _parse_factory_registry(path: Path) -> list[dict]:
    """화성공장등록검색 — 대규모 공장 DB (샘플/필터용)."""
    if not path.exists():
        return []
    try:
        df = pd.read_excel(path, sheet_name=0, header=0, dtype=str)
    except Exception:
        return []
    cols = {str(c).strip(): c for c in df.columns}
    def col(*names):
        for n in names:
            for k, c in cols.items():
                if n in k.replace(" ", ""):
                    return c
        return None

    c_name = col("회사명")
    c_addr = col("공장주소", "주소")
    c_prod = col("생산품")
    c_ind = col("업종명", "대표업종")
    c_tel = col("전화번호")
    c_park = col("산업단지")
    if c_name is None:
        return []
    records: list[dict] = []
    for _, r in df.iterrows():
        name = _s(r.get(c_name))
        if not name:
            continue
        addr = _s(r.get(c_addr)) if c_addr is not None else ""
        region = infer_region(addr, sheet_hint="화성")
        park = _s(r.get(c_park)) if c_park is not None else ""
        prod = _s(r.get(c_prod)) if c_prod is not None else ""
        records.append(
            _attach_complex(
                {
                    "출처": "화성공장등록",
                    "파일": path.name,
                    "시트": "등록공장",
                    "지역": region if region != "미분류" else "화성",
                    "업체명": name,
                    "주소": addr,
                    "업종": _s(r.get(c_ind)) if c_ind is not None else "",
                    "사용가스": "",
                    "공급사": "",
                    "담당자": "",
                    "연락처": _s(r.get(c_tel)) if c_tel is not None else "",
                    "비고": prod,
                    "산업단지": park,
                },
                explicit=park,
            )
        )
    return records


def _parse_seojin(path: Path) -> list[dict]:
    if load_workbook is None or not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    try:
        for sn in wb.sheetnames:
            if _SKIP_SHEET.search(sn) and "화성" not in sn:
                continue
            rows = []
            ws = wb[sn]
            for i, row in enumerate(ws.iter_rows(values_only=True, max_col=12)):
                rows.append(tuple(row))
                if i >= 3000:
                    break
            hi = _find_header_row(rows, ["입체명", "주소", "담당자", "업체명"])
            if hi is None:
                hi = _find_header_row(rows, ["업체명", "지역", "공급처"])
            if hi is None:
                continue
            header = [_norm_header(c) for c in rows[hi]]
            for row in rows[hi + 1 :]:
                d = _row_dict(header, row)
                name = _pick(d, "입체명", "업체명", "상호")
                if not name:
                    continue
                addr = _pick(d, "주소", "지역", "위치")
                note = _pick(d, "특이사항", "비고")
                records.append(
                    _attach_complex(
                        {
                            "출처": "서진산업가스",
                            "파일": path.name,
                            "시트": sn,
                            "지역": infer_region(addr, sn, sheet_hint=sn),
                            "업체명": name,
                            "주소": addr,
                            "업종": _pick(d, "종목", "업종", "생산품목"),
                            "사용가스": "",
                            "공급사": "서진산업가스",
                            "담당자": _pick(d, "담당자"),
                            "연락처": _pick(d, "전화번호", "연락처"),
                            "비고": note,
                        }
                    )
                )
    finally:
        wb.close()
    return records


def _bundle_disk_load(sig: str):
    if not sig or not os.path.isfile(MR_BUNDLE_PKL) or not os.path.isfile(MR_BUNDLE_SIG):
        return None
    try:
        with open(MR_BUNDLE_SIG, "r", encoding="utf-8") as f:
            if f.read().strip() != sig:
                return None
        with open(MR_BUNDLE_PKL, "rb") as f:
            payload = pickle.load(f)
        if not isinstance(payload, (tuple, list)) or len(payload) != 4:
            return None
        slim, raw_n, removed, cascade = payload
        if not isinstance(slim, pd.DataFrame) or not isinstance(cascade, dict):
            return None
        return slim, int(raw_n), int(removed), cascade
    except Exception:
        return None


def _bundle_disk_save(
    sig: str, slim: pd.DataFrame, raw_n: int, removed: int, cascade: dict
) -> None:
    if not sig:
        return
    try:
        os.makedirs(MR_CACHE_DIR, exist_ok=True)
        tmp = MR_BUNDLE_PKL + ".tmp"
        with open(tmp, "wb") as f:
            pickle.dump((slim, int(raw_n), int(removed), cascade), f, protocol=4)
        os.replace(tmp, MR_BUNDLE_PKL)
        with open(MR_BUNDLE_SIG, "w", encoding="utf-8") as f:
            f.write(sig)
    except Exception:
        try:
            if os.path.exists(MR_BUNDLE_PKL + ".tmp"):
                os.remove(MR_BUNDLE_PKL + ".tmp")
        except Exception:
            pass


def _frame_disk_load(sig: str):
    if not sig or not os.path.isfile(MR_FRAME_PKL) or not os.path.isfile(MR_FRAME_SIG):
        return None
    try:
        with open(MR_FRAME_SIG, "r", encoding="utf-8") as f:
            if f.read().strip() != sig:
                return None
        with open(MR_FRAME_PKL, "rb") as f:
            payload = pickle.load(f)
        if not isinstance(payload, (tuple, list)) or len(payload) != 3:
            return None
        df, raw_n, removed = payload
        if not isinstance(df, pd.DataFrame):
            return None
        return df, int(raw_n), int(removed)
    except Exception:
        return None


def _frame_disk_save(sig: str, df: pd.DataFrame, raw_n: int, removed: int) -> None:
    if not sig:
        return
    try:
        os.makedirs(MR_CACHE_DIR, exist_ok=True)
        tmp = MR_FRAME_PKL + ".tmp"
        with open(tmp, "wb") as f:
            pickle.dump((df, int(raw_n), int(removed)), f, protocol=4)
        os.replace(tmp, MR_FRAME_PKL)
        with open(MR_FRAME_SIG, "w", encoding="utf-8") as f:
            f.write(sig)
    except Exception:
        try:
            if os.path.exists(MR_FRAME_PKL + ".tmp"):
                os.remove(MR_FRAME_PKL + ".tmp")
        except Exception:
            pass


@st.cache_data(show_spinner=False, ttl=600)
def load_market_research_frame(_cache_sig: str) -> tuple[pd.DataFrame, int, int]:
    """모든 시장조사 엑셀을 합치고 중복 업체를 병합.

    Returns
    -------
    (merged_df, raw_count, removed_dup_count)
    """
    disk = _frame_disk_load(_cache_sig)
    if disk is not None:
        return disk
    root = ensure_market_research_cache()
    files = list(_list_xlsx(root))
    kind_map = {
        str(e.get("filename")): str(e.get("kind") or "자동")
        for e in _load_upload_manifest()
        if e.get("filename")
    }
    records: list[dict] = []
    fac_for_lookup: list[dict] = []
    seen_paths: set[str] = set()

    for p in files:
        try:
            key = str(p.resolve())
        except Exception:
            key = str(p)
        if key in seen_paths:
            continue
        seen_paths.add(key)
        name = p.name
        nfc = unicodedata.normalize("NFC", name)
        # uploads/ 는 manifest 종류 우선
        if "uploads" in {x.casefold() for x in p.parts} and name in kind_map:
            records.extend(_parse_by_kind(p, kind_map[name]))
            continue
        n = nfc.lower()
        if "스페셜" in nfc or "스폐셜" in nfc:
            continue
        if "lco2" in n or "경쟁사" in nfc:
            records.extend(_parse_lco2(p))
        elif "시장조사 (67)" in nfc or nfc.startswith("시장조사 (67)"):
            records.extend(_parse_region_survey(p))
        elif "mail" in n or "시장조사ㅡ" in nfc or "시장조사-" in nfc:
            records.extend(_parse_visit_notes(p, "방문조사(mail)"))
        elif "화성" in nfc and "공장" in nfc:
            fac_recs = _parse_factory_registry(p)
            fac_for_lookup.extend(fac_recs)
            records.extend(fac_recs)
        elif "김진혁" in nfc:
            records.extend(_parse_visit_notes(p, "방문조사(김진혁)"))
        elif "서진" in nfc:
            records.extend(_parse_seojin(p))
        elif name in kind_map:
            records.extend(_parse_by_kind(p, kind_map[name]))
        # 그 외 이름 미매칭 파일은 무시 (업로드는 manifest로만)

    for ent in load_manual_entries():
        rec = _manual_to_record(ent)
        if len(rec["업체명"]) >= 2:
            records.append(rec)

    empty_cols = [
        "출처",
        "파일",
        "시트",
        "지역",
        "산업단지",
        "업체명",
        "주소",
        "업종",
        "사용가스",
        "공급사",
        "담당자",
        "연락처",
        "비고",
        "업체키",
        "병합건수",
    ]
    if not records:
        return pd.DataFrame(columns=empty_cols), 0, 0

    df = pd.DataFrame(records)
    df = df[df["업체명"].map(lambda x: len(_s(x)) >= 2)].copy()
    raw_n = len(df)
    merged, removed = merge_duplicate_rows(df)
    # 검색·필터 가속용 사전계산 컬럼
    for c in (
        "업체명",
        "주소",
        "업종",
        "사용가스",
        "공급사",
        "담당자",
        "연락처",
        "비고",
        "출처",
        "시트",
        "지역",
        "산업단지",
    ):
        if c not in merged.columns:
            merged[c] = ""
        else:
            merged[c] = merged[c].fillna("").astype(str)
    merged.loc[merged["산업단지"].isin(["", "nan"]), "산업단지"] = "미분류"
    # 미분류 → 공장등록 기반 읍면동/리 맵 + 주소 키워드로 재분류 (엑셀 재파싱 없음)
    lookup = _build_area_complex_lookup_from_records(fac_for_lookup)
    merged, _ = enrich_unclassified_complexes(merged, lookup)
    merged["_search"] = (
        merged["업체명"]
        + " "
        + merged["주소"]
        + " "
        + merged["업종"]
        + " "
        + merged["사용가스"]
        + " "
        + merged["공급사"]
        + " "
        + merged["담당자"]
        + " "
        + merged["비고"]
        + " "
        + merged["출처"]
        + " "
        + merged["산업단지"]
    ).str.casefold()
    merged["_factory_only"] = merged["출처"].eq("화성공장등록")
    merged["_has_factory"] = merged["출처"].str.contains(
        "화성공장등록", regex=False, na=False
    )
    _frame_disk_save(_cache_sig, merged, raw_n, removed)
    return merged, raw_n, removed


@st.cache_data(show_spinner=False, ttl=600)
def _cached_cache_signature() -> str:
    """xlsx·manifest 시그니처.

    name+size(+소파일 content hash)만 사용 — mtime 제외.
    Cloud git checkout 시 mtime이 바뀌어도 커밋된 `_tab_bundle.pkl`이
    그대로 hit 되게 한다 (콜드 파싱 ~8초 방지).
    """
    import hashlib

    root = MR_CACHE_DIR
    if not os.path.isdir(root):
        return "empty"
    parts = []
    for p in _list_xlsx(root):
        try:
            parts.append(f"{p.name}:{os.path.getsize(p)}")
        except OSError:
            parts.append(p.name)
    for extra in (_manual_entries_path(), MR_UPLOAD_MANIFEST):
        try:
            if not os.path.exists(extra):
                continue
            raw = Path(extra).read_bytes()
            digest = hashlib.md5(raw).hexdigest()[:12]
            parts.append(f"{Path(extra).name}:{len(raw)}:{digest}")
        except OSError:
            parts.append(f"{Path(extra).name}:0")
    return "|".join(parts) or "empty"


def _cache_signature() -> str:
    return _cached_cache_signature()


_MR_SHOW_COLS = [
    "지역",
    "산업단지",
    "업체명",
    "주소",
    "업종",
    "사용가스",
    "공급사",
    "담당자",
    "연락처",
    "비고",
    "출처",
    "시트",
    "병합건수",
]
_MR_DISPLAY_LIMIT = 100
_MR_DISPLAY_LIMIT_MAX = 400
_MR_EXCEL_LIMIT = 10000


@st.cache_data(show_spinner=False)
def _mr_filter_excel_bytes(df: pd.DataFrame) -> bytes:
    """검색 결과 DataFrame → .xlsx 바이트."""
    buf = io.BytesIO()
    out = df.copy() if df is not None else pd.DataFrame()
    if out.empty:
        out = pd.DataFrame({"알림": ["검색 결과가 없습니다."]})
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="검색결과", index=False)
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def _filter_frame(
    df: pd.DataFrame,
    *,
    regions: tuple,
    complexes: tuple,
    suppliers: tuple,
    query: str,
    include_factory: bool,
    hide_unclassified: bool = False,
    _logic_ver: int = 11,
) -> pd.DataFrame:
    """가벼운 필터 — Pandas str.contains 대신 초고속 List Comprehension 적용"""
    view = df
    if not include_factory:
        view = view[~view["_factory_only"]]
    if hide_unclassified and "산업단지" in view.columns:
        view = view[view["산업단지"] != "미분류"]
    if regions:
        reg_set = {str(r).strip() for r in regions}
        view = view[view["지역"].astype(str).str.strip().isin(reg_set)]
    if complexes:
        cx_set = {str(c).strip() for c in complexes}
        view = view[view["산업단지"].astype(str).str.strip().isin(cx_set)]
        
    # 1. 공급사 검색 속도 극대화 (List Comprehension 적용)
    if suppliers:
        sup_lowers = [s.casefold() for s in suppliers]
        sup_list = view["공급사"].str.casefold().tolist()
        mask = [any(sl in v for sl in sup_lowers) for v in sup_list]
        view = view[mask]

    # 2. 띄어쓰기 기반 스마트 다중 검색 및 속도 극대화
    q = (query or "").strip()
    if q:
        # 띄어쓰기 기준으로 키워드 분리 (예: "동탄 산소" -> "동탄" AND "산소")
        keywords = [k.casefold() for k in q.split()]
        search_list = view["_search"].tolist()
        
        # 정규식(Regex) 엔진을 거치지 않는 Python 순수 in 연산자 사용 (압도적 속도)
        mask = [all(k in s for k in keywords) for s in search_list]
        view = view[mask]
        
    return view


def _metric_box(label: str, value: str) -> str:
    return (
        f"<div class='metric-box'><div class='metric-label'>{label}</div>"
        f"<div class='metric-value'>{value}</div></div>"
    )


def _mr_extract_suppliers(series, *, limit: int = 300) -> list[str]:
    """공급사 셀(·|/ 구분)에서 옵션 목록 추출."""
    seen: set[str] = set()
    out: list[str] = []
    for v in series.astype(str):
        for p in re.split(r"\s*[·|/]\s*", v):
            p = p.strip()
            if not p or p in seen:
                continue
            seen.add(p)
            out.append(p)
            if len(out) >= limit:
                return sorted(out)
    return sorted(out)


def _mr_cascade_base(
    df: pd.DataFrame,
    *,
    regions: list[str] | None = None,
    complexes: list[str] | None = None,
    include_factory: bool = False,
    hide_unclassified: bool = False,
) -> pd.DataFrame:
    """종속 옵션 계산용 범위 (검색어·공급사 필터는 제외)."""
    view = df
    if not include_factory and "_factory_only" in view.columns:
        view = view[~view["_factory_only"]]
    if hide_unclassified and "산업단지" in view.columns:
        view = view[view["산업단지"] != "미분류"]
    if regions:
        reg_set = {str(r).strip() for r in regions if r}
        view = view[view["지역"].astype(str).str.strip().isin(reg_set)]
    if complexes:
        cx_set = {str(c).strip() for c in complexes if c}
        view = view[view["산업단지"].astype(str).str.strip().isin(cx_set)]
    return view


def _mr_complex_options(base: pd.DataFrame) -> list[str]:
    """검색용 산업단지 — '미분류'는 옵션에서 제외 (없으면 빈 목록)."""
    if base is None or base.empty or "산업단지" not in base.columns:
        return []
    vals = base["산업단지"].dropna().astype(str).unique().tolist()
    return sorted([c for c in vals if c and c != "미분류"])


def _mr_widget_key(prefix: str, parts: list[str]) -> str:
    """상위 선택이 바뀌면 위젯을 새로 만들어 옵션이 확실히 갱신되게 함."""
    tail = "|".join(parts) if parts else "ALL"
    tail = re.sub(r"[^\w가-힣|.\-]+", "_", tail)[:80]
    return f"{prefix}__{tail}"


@st.cache_data(show_spinner=False, ttl=600)
def _cached_mr_cascade_index(_cache_sig: str) -> dict:
    if mr_cascade is None:
        return {"survey": {"cxr": {}, "cxa": [], "sr": {}, "src": {}, "sa": []}, "all": {"cxr": {}, "cxa": [], "sr": {}, "src": {}, "sa": []}}
    frame, _, _ = load_market_research_frame(_cache_sig)
    return mr_cascade.build_cascade_index(frame)


_MR_SLIM_COLS = list(dict.fromkeys(
    list(_MR_SHOW_COLS) + ["_factory_only", "_has_factory", "_search"]
))


def _mr_slim_frame(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in _MR_SLIM_COLS if c in df.columns]
    return df.loc[:, cols]


@st.cache_data(show_spinner=False, ttl=600)
def _cached_mr_tab_bundle(_cache_sig: str) -> tuple[pd.DataFrame, int, int, dict]:
    """병합 DF(slim) + cascade — disk·cache 1회, cascade 중복 빌드 없음."""
    disk = _bundle_disk_load(_cache_sig)
    if disk is not None:
        return disk
    df, raw_n, removed_n = load_market_research_frame(_cache_sig)
    slim = _mr_slim_frame(df)
    if mr_cascade is None:
        cascade: dict = {
            "survey": {"cxr": {}, "cxa": [], "sr": {}, "src": {}, "sa": []},
            "all": {"cxr": {}, "cxa": [], "sr": {}, "src": {}, "sa": []},
        }
    else:
        cascade = mr_cascade.build_cascade_index(df)
    _bundle_disk_save(_cache_sig, slim, raw_n, removed_n, cascade)
    return slim, raw_n, removed_n, cascade


def _mr_load_tab_bundle() -> tuple[pd.DataFrame, int, int, dict]:
    """세션 warm + st.cache_data — 적용 rerun 시 목록 재로딩 최소화."""
    ensure_market_research_cache()
    sig = _cache_signature()
    warm = st.session_state.get("_mr_session_bundle")
    if isinstance(warm, dict) and warm.get("sig") == sig and warm.get("df") is not None:
        return (
            warm["df"],
            int(warm.get("raw_n") or 0),
            int(warm.get("removed_n") or 0),
            warm.get("cascade") or {},
        )
    df, raw_n, removed_n, cascade = _cached_mr_tab_bundle(sig)
    st.session_state["_mr_session_bundle"] = {
        "sig": sig,
        "df": df,
        "raw_n": raw_n,
        "removed_n": removed_n,
        "cascade": cascade,
    }
    return df, raw_n, removed_n, cascade


def _mr_market_tab_keep_visible() -> None:
    """상단 필터 full rerun 시 시장조사 탭이 defer stub으로 바뀌지 않게."""
    mounted = st.session_state.setdefault("_dash_heavy_mounted", {})
    for idx in (9, 10, 11):
        mounted[idx] = True
        st.session_state[f"_dash_force_tab_{idx}"] = True
    st.session_state["_dash_filter_changed_flag"] = False


def _mr_ensure_draft_widgets() -> None:
    """적용 전 초안(mr_w_*) — 최초 1회 applied(mr_v6_*)에서 복사."""
    if "mr_w_region" not in st.session_state:
        st.session_state["mr_w_region"] = list(st.session_state.get("mr_v6_region") or [])
    if "mr_w_complex" not in st.session_state:
        st.session_state["mr_w_complex"] = list(st.session_state.get("mr_v6_complex") or [])
    if "mr_w_sup" not in st.session_state:
        st.session_state["mr_w_sup"] = list(st.session_state.get("mr_v6_sup") or [])
    if "mr_w_q" not in st.session_state:
        st.session_state["mr_w_q"] = str(st.session_state.get("mr_v6_q") or "")
    if "mr_w_fac" not in st.session_state:
        st.session_state["mr_w_fac"] = bool(st.session_state.get("mr_v6_fac", False))
    if "mr_w_hide" not in st.session_state:
        st.session_state["mr_w_hide"] = bool(st.session_state.get("mr_v6_hide", False))


def _mr_apply_draft_filters(cascade: dict) -> dict:
    """초안 → 적용 상태. 유효하지 않은 단지·공급사는 자동 정리."""
    r_in = list(st.session_state.get("mr_w_region") or [])
    c_in = list(st.session_state.get("mr_w_complex") or [])
    s_in = list(st.session_state.get("mr_w_sup") or [])
    fac_in = bool(st.session_state.get("mr_w_fac", False))
    hide_in = bool(st.session_state.get("mr_w_hide", False))
    q_in = str(st.session_state.get("mr_w_q") or "").strip()

    st.session_state["mr_v6_region"] = r_in
    st.session_state["mr_v6_q"] = q_in
    st.session_state["mr_v6_fac"] = fac_in
    st.session_state["mr_v6_hide"] = hide_in
    new_cx = mr_cascade.complex_opts(cascade, r_in, include_factory=fac_in)
    kept_c = [x for x in c_in if x in new_cx]
    st.session_state["mr_v6_complex"] = kept_c
    new_sup = mr_cascade.supplier_opts(cascade, r_in, kept_c, include_factory=fac_in)
    st.session_state["mr_v6_sup"] = [x for x in s_in if x in new_sup]
    return {"regions": r_in, "complexes": len(kept_c), "query": q_in}


@st.fragment
def _mr_filter_and_results(
    df: pd.DataFrame,
    regions: list[str],
    cascade: dict,
    latest_update_str: str,
) -> None:
    """필터+결과 전체 fragment — 적용·보기 전환 시 Tab1~8 재실행 없음."""
    if mr_cascade is None:
        st.error("market_research_cascade 모듈이 없습니다.")
        return

    _mr_ensure_draft_widgets()

    st.caption(
        "필터 **v11** · 지역→단지→공급사 종속 · "
        "**적용** 클릭 시 이 영역만 갱신 (전체 대시보드 rerun 없음)."
    )

    draft_r = list(st.session_state.get("mr_w_region") or [])
    draft_c = list(st.session_state.get("mr_w_complex") or [])
    draft_fac = bool(st.session_state.get("mr_w_fac", False))

    form_cx_opts = mr_cascade.complex_opts(cascade, draft_r, include_factory=draft_fac)
    form_sup_opts = mr_cascade.supplier_opts(
        cascade, draft_r, draft_c, include_factory=draft_fac
    )

    reg_set = set(regions)
    cx_set = set(form_cx_opts)
    sup_set = set(form_sup_opts)
    st.session_state["mr_w_region"] = [
        x for x in st.session_state.get("mr_w_region", []) if x in reg_set
    ]
    st.session_state["mr_w_complex"] = [
        x for x in st.session_state.get("mr_w_complex", []) if x in cx_set
    ]
    st.session_state["mr_w_sup"] = [
        x for x in st.session_state.get("mr_w_sup", []) if x in sup_set
    ]

    with st.container(border=True):
        c_fac, c_hide = st.columns(2)
        with c_fac:
            st.checkbox("화성공장 DB(단독) 포함", key="mr_w_fac")
        with c_hide:
            st.checkbox("산업단지 미분류 제외", key="mr_w_hide")

        f1, f2, f3, f4 = st.columns([1.2, 1.4, 1.2, 1.6])
        with f1:
            st.multiselect(
                "지역",
                options=regions,
                key="mr_w_region",
                placeholder="전체 지역",
            )
        with f2:
            st.multiselect(
                "산업단지" + (f" ({len(form_cx_opts)}개)" if draft_r else ""),
                options=form_cx_opts,
                key="mr_w_complex",
                placeholder=(
                    "데이터 없음"
                    if (draft_r and not form_cx_opts)
                    else "전체 산업단지"
                ),
            )
            if draft_r and not form_cx_opts:
                st.caption("선택 지역에 산업단지 데이터 없음")
        with f3:
            st.multiselect(
                "공급사",
                options=form_sup_opts,
                key="mr_w_sup",
                placeholder=(
                    "데이터 없음"
                    if (draft_c and not form_sup_opts)
                    else "전체 공급사"
                ),
            )
        with f4:
            st.text_input("검색 (업체·주소·단지·가스·비고)", key="mr_w_q")

        applied = st.button(
            "🔍 적용",
            type="primary",
            use_container_width=True,
            key="mr_apply_btn",
        )

    if applied:
        flash = _mr_apply_draft_filters(cascade)
        _rg = flash.get("regions") or []
        st.success(
            "필터 적용됨 · "
            + (f"지역 {', '.join(_rg)}" if _rg else "전체")
            + (f" · 검색「{flash.get('query')}」" if flash.get("query") else "")
        )

    app_r = list(st.session_state.get("mr_v6_region") or [])
    app_c = list(st.session_state.get("mr_v6_complex") or [])
    app_s = list(st.session_state.get("mr_v6_sup") or [])
    app_q = str(st.session_state.get("mr_v6_q") or "")
    app_fac = bool(st.session_state.get("mr_v6_fac", False))
    app_hide = bool(st.session_state.get("mr_v6_hide", False))

    _applied_parts = []
    if app_r:
        _applied_parts.append("지역=" + "·".join(app_r))
    if app_c:
        _applied_parts.append("단지=" + "·".join(app_c[:3]) + ("…" if len(app_c) > 3 else ""))
    if app_s:
        _applied_parts.append("공급사=" + "·".join(app_s[:2]) + ("…" if len(app_s) > 2 else ""))
    if app_q:
        _applied_parts.append(f"검색={app_q}")
    if app_fac:
        _applied_parts.append("화성공장DB포함")
    if app_hide:
        _applied_parts.append("미분류제외")
    if _applied_parts:
        st.caption("✅ 적용됨 · " + " · ".join(_applied_parts))
    else:
        st.caption("ℹ️ 적용 조건 없음 — 전체 목록 표시")

    view = _filter_frame(
        df,
        regions=tuple(app_r),
        complexes=tuple(app_c),
        suppliers=tuple(app_s),
        query=app_q,
        include_factory=app_fac,
        hide_unclassified=app_hide,
    )

    show_cols = [c for c in _MR_SHOW_COLS if c in view.columns]
    limit = int(st.session_state.get("mr_display_limit") or _MR_DISPLAY_LIMIT)
    limit = min(max(limit, _MR_DISPLAY_LIMIT), _MR_DISPLAY_LIMIT_MAX)
    n_view = len(view)
    cap_c, dl_c = st.columns([2.4, 1])
    with cap_c:
        st.caption(f"검색 결과 **{n_view:,}**건 · 화면에 {min(n_view, limit):,}건 표시")
        if n_view > limit and limit < _MR_DISPLAY_LIMIT_MAX:
            if st.button(
                f"더 보기 (최대 {_MR_DISPLAY_LIMIT_MAX:,}건)",
                key="mr_show_more_rows",
            ):
                st.session_state["mr_display_limit"] = _MR_DISPLAY_LIMIT_MAX
                st.rerun(scope="fragment")
    with dl_c:
        if n_view > 0:
            xl_n = min(n_view, _MR_EXCEL_LIMIT)
            xl_df = view[show_cols].head(xl_n) if show_cols else view.head(xl_n)
            st.session_state.pop("mr_dl_xlsx_search", None)
            st.download_button(
                f"엑셀 다운로드 ({xl_n:,}건)",
                data=_mr_filter_excel_bytes(xl_df),
                file_name=f"시장조사_검색결과_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
    mode = st.radio(
        "보기",
        options=["통합 목록", "지역별", "공급사별", "화성공장 DB", "산업단지별"],
        horizontal=True,
        key="mr_view_v6",
    )

    if mode == "통합 목록":
        st.dataframe(
            view[show_cols].head(limit),
            width="stretch",
            hide_index=True,
            height=480,
        )
    elif mode == "지역별":
        region_counts = (
            view.groupby("지역", dropna=False)
            .size()
            .rename("건수")
            .sort_values(ascending=False)
            .reset_index()
        )
        left, right = st.columns([1, 1.4])
        with left:
            st.dataframe(region_counts, width="stretch", hide_index=True, height=400)
        with right:
            opts = region_counts["지역"].tolist() or ["미분류"]
            pick = st.selectbox("지역 상세", options=opts, key="mr_region_pick_v6")
            sub = view[view["지역"] == pick][show_cols]
            st.caption(f"{pick} · {len(sub):,}건")
            st.dataframe(sub.head(limit), width="stretch", hide_index=True, height=400)
    elif mode == "공급사별":
        cnt: Counter[str] = Counter()
        for s in view["공급사"].astype(str):
            for p in re.split(r"\s*[·|/]\s*", s):
                p = p.strip()
                if p:
                    cnt[p] += 1
        if not cnt:
            st.info("공급사 정보가 있는 행이 없습니다.")
        else:
            sc = (
                pd.DataFrame({"공급사": list(cnt.keys()), "건수": list(cnt.values())})
                .sort_values("건수", ascending=False)
                .head(50)
                .reset_index(drop=True)
            )
            a, b = st.columns([1, 1.4])
            with a:
                st.dataframe(sc, width="stretch", hide_index=True, height=400)
            with b:
                pick_s = st.selectbox(
                    "공급사 상세", options=sc["공급사"].tolist(), key="mr_sup_pick_v6"
                )
                sub = view[
                    view["공급사"].str.contains(
                        re.escape(pick_s), regex=True, na=False
                    )
                ][show_cols]
                st.caption(f"{pick_s} · {len(sub):,}건")
                st.dataframe(
                    sub.head(limit), width="stretch", hide_index=True, height=400
                )
    elif mode == "화성공장 DB":
        fac = (
            df[df["_has_factory"]]
            if "_has_factory" in df.columns
            else df.iloc[0:0]
        )
        st.caption(f"화성공장 관련 **{len(fac):,}**건")
        q2 = st.text_input("공장 DB 검색", key="mr_fac_q_v6")
        fac2 = fac
        qq = (q2 or "").strip()
        if qq and "_search" in fac2.columns:
            fac2 = fac[
                fac["_search"].str.contains(qq.casefold(), regex=False, na=False)
            ]
        st.dataframe(
            fac2[show_cols].head(limit),
            width="stretch",
            hide_index=True,
            height=440,
        )
    else:
        cx_counts = (
            view.groupby("산업단지", dropna=False)
            .size()
            .rename("건수")
            .sort_values(ascending=False)
            .reset_index()
        )
        left, right = st.columns([1, 1.4])
        with left:
            st.dataframe(cx_counts, width="stretch", hide_index=True, height=400)
        with right:
            opts_cx = cx_counts["산업단지"].tolist() or ["미분류"]
            pick_cx = st.selectbox(
                "산업단지 상세", options=opts_cx, key="mr_complex_pick_v6"
            )
            sub = view[view["산업단지"] == pick_cx][show_cols]
            st.caption(f"{pick_cx} · {len(sub):,}건")
            st.dataframe(sub.head(limit), width="stretch", hide_index=True, height=400)
    if latest_update_str:
        st.caption(f"대시보드 기준 시각: {latest_update_str}")


def _mr_filter_results(
    df: pd.DataFrame,
    regions: list[str],
    cascade: dict,
    latest_update_str: str,
) -> None:
    try:
        _mr_filter_and_results(df, regions, cascade, latest_update_str)
    except Exception as e:
        st.error(f"시장조사 필터 오류: {e}")
        st.exception(e)



@st.fragment
def _mr_loaded_panel(latest_update_str: str = "") -> None:
    """데이터 로드·지표·필터 — fragment로 셸(헤더·업로드) 먼저 그린 뒤 비동기 채움."""
    with st.spinner("시장조사 데이터 불러오는 중…"):
        df, raw_n, removed_n, cascade = _mr_load_tab_bundle()

    if df.empty:
        st.info(
            "아직 목록이 비어 있습니다. 위에서 **새 시장조사 입력**으로 첫 건을 넣거나, "
            "Drive「업무/시장조사」엑셀을 동기화하세요."
        )
        return

    n_all = len(df)
    n_survey = int((~df["_factory_only"]).sum())
    n_merged_rows = int((df["병합건수"] > 1).sum()) if "병합건수" in df.columns else 0
    n_complex = int((df["산업단지"] != "미분류").sum()) if "산업단지" in df.columns else 0
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(_metric_box("병합 후 업체", f"{n_all:,}"), unsafe_allow_html=True)
    c2.markdown(_metric_box("조사·경쟁사", f"{n_survey:,}"), unsafe_allow_html=True)
    c3.markdown(_metric_box("산업단지 분류", f"{n_complex:,}"), unsafe_allow_html=True)
    c4.markdown(
        _metric_box("원본→병합", f"{raw_n:,}→{n_all:,}"),
        unsafe_allow_html=True,
    )
    if removed_n:
        st.caption(
            f"중복 {removed_n:,}건 병합 · 병합 업체 {n_merged_rows:,}곳. "
            "단지명은 공장등록 DB + 주소 키워드로 붙입니다 (웹검색 없음)."
        )

    regions = sorted(
        [r for r in df["지역"].dropna().unique().tolist() if r],
        key=lambda x: (x == "미분류", x),
    )
    try:
        _mr_filter_results(df, regions, cascade, latest_update_str)
    except Exception as e:
        st.error(f"시장조사 필터 표시 실패: {e}")
        st.exception(e)


def render_market_research_tab(latest_update_str: str = "") -> None:
    """시장조사 탭 UI."""
    st.markdown(
        "<div class='sub-header dashboard-tab-panel-head'>🔎 시장조사</div>",
        unsafe_allow_html=True,
    )
    st.caption(
        "경로: Google Drive › Desktop › 업무 › 시장조사  ·  "
        "엑셀 업로드·직접입력 가능 · 검색은 「적용」"
    )

    with st.expander("📁 엑셀 업로드", expanded=False):
        st.caption(
            "파일을 `uploaded_cache/market_research/uploads/`에 저장한 뒤 목록에 합칩니다. "
            "양식이 다르면 아래 **파싱 형식**을 지정하세요."
        )
        with st.form("mr_upload_form", clear_on_submit=True):
            up_files = st.file_uploader(
                "엑셀 파일 (.xlsx)",
                type=["xlsx", "xls"],
                accept_multiple_files=True,
                key="mr_upload_files",
                help="여러 개 선택 가능. Cloud에서도 업로드한 파일은 앱 캐시에 남습니다.",
            )
            kind_labels = [f"{k} — {d}" for k, d in MR_UPLOAD_KINDS]
            kind_pick = st.selectbox(
                "파싱 형식",
                options=list(range(len(MR_UPLOAD_KINDS))),
                format_func=lambda i: kind_labels[i],
                key="mr_upload_kind_form",
            )
            up_go = st.form_submit_button(
                "저장하고 목록에 반영", type="primary", width="stretch"
            )
        if up_go:
            if not up_files:
                st.warning("파일을 선택한 뒤 다시 눌러 주세요.")
            else:
                kind = MR_UPLOAD_KINDS[int(kind_pick)][0]
                saved_n = 0
                errors: list[str] = []
                for uf in up_files:
                    try:
                        ent = save_uploaded_excel(uf, kind=kind)
                        # 바로 파싱 스모크 (실패해도 파일은 저장됨)
                        path = Path(MR_UPLOAD_DIR) / ent["filename"]
                        n_rec = len(_parse_by_kind(path, ent["kind"]))
                        saved_n += 1
                        if n_rec == 0:
                            errors.append(
                                f"{ent['original_name']}: 저장됨 · 파싱 0건 "
                                f"(형식을 바꿔 다시 올려 보세요)"
                            )
                        else:
                            st.success(
                                f"업로드: **{ent['original_name']}** → "
                                f"{ent['kind']} · {n_rec:,}건"
                            )
                    except Exception as e:
                        errors.append(f"{getattr(uf, 'name', '?')}: {e}")
                if saved_n:
                    _invalidate_mr_loaded()
                    if errors:
                        for msg in errors:
                            st.warning(msg)
                    st.rerun()
                elif errors:
                    for msg in errors:
                        st.error(msg)

        uploads = _load_upload_manifest()
        if uploads:
            st.markdown(f"**업로드된 파일** ({len(uploads)}건)")
            for ent in uploads[:12]:
                c_a, c_b = st.columns([5, 1])
                with c_a:
                    sz = ent.get("size") or 0
                    st.caption(
                        f"`{ent.get('uploaded_at', '')}` · "
                        f"**{ent.get('original_name', ent.get('filename'))}** · "
                        f"{ent.get('kind', '자동')} · "
                        f"{round(sz / 1024, 1)} KB"
                    )
                with c_b:
                    if st.button(
                        "삭제",
                        key=f"mr_up_del_{ent.get('id')}",
                        width="stretch",
                    ):
                        delete_uploaded_excel(str(ent.get("id")))
                        _invalidate_mr_loaded()
                        st.rerun()
            st.caption(f"저장 폴더: `{MR_UPLOAD_DIR}`")

    with st.expander("✍️ 새 시장조사 입력", expanded=False):
        with st.form("mr_new_entry_form", clear_on_submit=True):
            r1c1, r1c2, r1c3 = st.columns([1.3, 0.9, 1.2])
            with r1c1:
                name_in = st.text_input("업체명 *", placeholder="예: ○○엔지니어링")
            with r1c2:
                region_in = st.selectbox(
                    "지역",
                    options=[""] + _region_choices(),
                    format_func=lambda x: "(주소로 자동)" if x == "" else x,
                )
            with r1c3:
                complex_in = st.selectbox(
                    "산업단지",
                    options=[""] + _complex_choices(),
                    format_func=lambda x: "(주소로 자동)" if x == "" else x,
                )
            addr_in = st.text_input(
                "주소 / 위치",
                placeholder="예: 경기도 화성시 팔탄면 …",
            )
            r2c1, r2c2, r2c3 = st.columns(3)
            with r2c1:
                industry_in = st.text_input("업종 / 생산품목")
            with r2c2:
                gas_in = st.text_input("사용가스", placeholder="예: LCO2, LN2")
            with r2c3:
                supplier_in = st.text_input("현 공급사")
            r3c1, r3c2 = st.columns(2)
            with r3c1:
                person_in = st.text_input("담당자")
            with r3c2:
                phone_in = st.text_input("연락처")
            note_in = st.text_area("비고", height=70)
            saved = st.form_submit_button("💾 저장하고 목록에 반영", type="primary")
        if saved:
            try:
                ent = add_manual_entry(
                    {
                        "업체명": name_in,
                        "지역": region_in,
                        "산업단지": complex_in,
                        "주소": addr_in,
                        "업종": industry_in,
                        "사용가스": gas_in,
                        "공급사": supplier_in,
                        "담당자": person_in,
                        "연락처": phone_in,
                        "비고": note_in,
                    }
                )
                _invalidate_mr_loaded()
                st.success(
                    f"저장됨: **{ent['업체명']}** ({ent.get('지역') or '미분류'})  "
                    f"· {ent.get('산업단지') or '단지미분류'} · 출처「{MR_MANUAL_SOURCE}」"
                )
                st.rerun()
            except ValueError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"저장 실패: {e}")

        manual = load_manual_entries()
        if manual:
            st.markdown(f"**최근 직접입력** ({len(manual)}건)")
            show_n = min(8, len(manual))
            for ent in manual[:show_n]:
                c_a, c_b = st.columns([5, 1])
                with c_a:
                    st.caption(
                        f"`{ent.get('saved_at', '')}` · "
                        f"**{ent.get('업체명', '')}** · "
                        f"{ent.get('지역', '')} · "
                        f"{ent.get('공급사', '') or '-'}"
                    )
                with c_b:
                    if st.button(
                        "삭제",
                        key=f"mr_del_{ent.get('id')}",
                        width="stretch",
                    ):
                        delete_manual_entry(str(ent.get("id")))
                        _invalidate_mr_loaded()
                        st.rerun()
            st.caption(
                f"저장 위치: `{MR_MANUAL_FILE}` "
                "(맥이면 Drive「시장조사/직접입력_시장조사.json」에도 복사)"
            )

    _mr_loaded_panel(latest_update_str)