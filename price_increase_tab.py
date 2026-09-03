"""단가인상 공문 탭 — 거래처·메일·품목단가·공문작성·발송.

다른 탭과 공유 상태를 바꾸지 않음. sales_df / address는 인자로만 받음.
"""
from __future__ import annotations

import io
import json
import os
import re
import smtplib
import ssl
from copy import copy
from datetime import date, datetime
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from typing import Any, Optional
from urllib.parse import quote
import sys

import pandas as pd
import streamlit as st

from dev_mode import dev_caption

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore
    PatternFill = None  # type: ignore

PI_DIR = os.path.join("uploaded_cache", "price_increase")
PI_MAIL_CSV = os.path.join(PI_DIR, "mail_contacts.csv")
PI_SMTP_LOCAL = os.path.join(PI_DIR, "smtp_local.toml")
PI_TEMPLATE = os.path.join(PI_DIR, "공문양식.xlsx")
PI_DRAFTS = os.path.join(PI_DIR, "drafts")
PI_SENT_LOG = os.path.join(PI_DRAFTS, "sent_log.jsonl")
PI_UI_BUILD = "2026-09-03e · 로컬SMTP설정"
PI_FONTS_DIR = os.path.join(PI_DIR, "fonts")
_KR_FONT_CANDIDATES = (
    os.path.join(PI_FONTS_DIR, "NotoSansKR-Regular.ttf"),
    os.path.join(PI_FONTS_DIR, "NanumGothic.ttf"),
    os.path.join(PI_FONTS_DIR, "wqy-microhei.ttf"),
    # macOS
    "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
    "/Library/Fonts/AppleGothic.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/System/Library/Fonts/AppleSDGothicNeo.ttc",
    "/Library/Fonts/NanumGothic.ttf",
    "/Library/Fonts/NanumGothicBold.ttf",
    os.path.expanduser("~/Library/Fonts/NanumGothic.ttf"),
    os.path.expanduser("~/Library/Fonts/NotoSansKR-Regular.ttf"),
    os.path.expanduser("~/Library/Fonts/AppleGothic.ttf"),
    # Linux
    "/usr/share/fonts/truetype/noto/NotoSansKR-Regular.ttf",
    "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
    "/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
)

# 기본 공문 후보 (캐시 원본 우선)
_TEMPLATE_CANDIDATES = (
    os.path.join(PI_DIR, "탄산단가인상공문.xlsx"),
    os.path.join("uploaded_cache", "price_increase", "탄산단가인상공문.xlsx"),
    "탄산단가인상공문.xlsx",
    os.path.expanduser("~/Desktop/탄산단가인상공문.xlsx"),
    os.path.expanduser("~/Desktop/업무/탄산단가인상공문.xlsx"),
    os.path.expanduser("~/Desktop/dashboard/탄산단가인상공문.xlsx"),
)

_NOISE_ITEM = re.compile(
    r"입금|이월|단가차액|잔액정리|기화기|공사|작업비|임대|운반비|회수|보증|수수료|운임",
    re.I,
)
_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")


def _ensure_dirs() -> None:
    os.makedirs(PI_DIR, exist_ok=True)
    os.makedirs(PI_DRAFTS, exist_ok=True)


def _pi_is_streamlit_cloud() -> bool:
    env = (os.environ.get("STREAMLIT_RUNTIME_ENVIRONMENT") or "").strip().lower()
    if env in ("cloud", "streamlit_cloud"):
        return True
    try:
        if os.path.isfile("/home/appuser/.streamlit/secrets.toml"):
            # Streamlit Cloud 이미지에서 흔한 경로 (로컬 Mac은 아님)
            if not sys.platform.startswith("darwin"):
                return "STREAMLIT_SHARING" in os.environ or "STREAMLIT_CLOUD" in os.environ
    except Exception:
        pass
    return bool(os.environ.get("STREAMLIT_CLOUD") or os.environ.get("STREAMLIT_SHARING"))


def _parse_simple_toml_flat(text: str) -> dict[str, str]:
    """평면 key = \"value\" 만 파싱 (smtp_local.toml용)."""
    out: dict[str, str] = {}
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or line.startswith("["):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        key = k.strip()
        val = v.strip()
        if len(val) >= 2 and val[0] == val[-1] and val[0] in ("'", '"'):
            val = val[1:-1]
        out[key] = val
    return out


def load_local_smtp() -> dict[str, str]:
    """로컬 Mac 전용 SMTP 저장본 (secrets.toml 없이도 연동)."""
    path = PI_SMTP_LOCAL
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return _parse_simple_toml_flat(f.read())
    except Exception:
        return {}


def save_local_smtp(
    *,
    user: str,
    password: str,
    provider: str = "daum",
    from_name: str = "신일가스",
    ssl_verify: str = "0",
) -> str:
    """smtp_local.toml 저장. 반환: 저장 경로."""
    _ensure_dirs()
    user = str(user or "").replace("\xa0", "").replace("\u200b", "").strip()
    password = (
        str(password or "")
        .replace("\xa0", "")
        .replace("\u200b", "")
        .replace(" ", "")
        .strip()
    )
    provider = (provider or "daum").strip().lower() or "daum"
    from_name = str(from_name or "신일가스").strip() or "신일가스"
    ssl_verify = "0" if str(ssl_verify).strip() in ("0", "false", "no", "off") else "1"
    body = (
        "# 로컬 공문 SMTP (git 무시 · 비밀번호 포함 — 공유 금지)\n"
        f'smtp_provider = "{provider}"\n'
        f'smtp_user = "{user}"\n'
        f'smtp_password = "{password}"\n'
        f'smtp_from_name = "{from_name}"\n'
        f'smtp_ssl_verify = "{ssl_verify}"\n'
    )
    with open(PI_SMTP_LOCAL, "w", encoding="utf-8") as f:
        f.write(body)
    # Streamlit secrets에도 맞춰 두면 재시작 후 st.secrets 경로도 동작
    try:
        _upsert_streamlit_secrets_smtp(
            user=user,
            password=password,
            provider=provider,
            from_name=from_name,
            ssl_verify=ssl_verify,
        )
    except Exception:
        pass
    return PI_SMTP_LOCAL


def _upsert_streamlit_secrets_smtp(
    *,
    user: str,
    password: str,
    provider: str,
    from_name: str,
    ssl_verify: str,
) -> None:
    secrets_path = os.path.join(".streamlit", "secrets.toml")
    os.makedirs(".streamlit", exist_ok=True)
    existing = ""
    if os.path.isfile(secrets_path):
        with open(secrets_path, "r", encoding="utf-8") as f:
            existing = f.read()
    flat = _parse_simple_toml_flat(existing)
    # 기존 비SMTP 키 유지: 원본에서 SMTP 관련 줄만 교체/추가
    smtp_keys = {
        "smtp_provider",
        "smtp_host",
        "smtp_port",
        "smtp_user",
        "smtp_password",
        "smtp_from_name",
        "smtp_ssl_verify",
        "smtp_from",
        "smtp_ssl",
    }
    lines_out: list[str] = []
    for raw in existing.splitlines():
        s = raw.strip()
        if s and not s.startswith("#") and "=" in s and not s.startswith("["):
            k = s.split("=", 1)[0].strip()
            if k in smtp_keys:
                continue
        lines_out.append(raw)
    # 끝에 평면 SMTP 블록 추가
    while lines_out and lines_out[-1].strip() == "":
        lines_out.pop()
    if lines_out and lines_out[-1].strip():
        lines_out.append("")
    host = {
        "daum": "smtp.daum.net",
        "kakao": "smtp.daum.net",
        "naver": "smtp.naver.com",
        "gmail": "smtp.gmail.com",
    }.get(provider, "smtp.daum.net")
    lines_out.extend(
        [
            "# --- 공문 SMTP (로컬 저장) ---",
            f'smtp_provider = "{provider}"',
            f'smtp_host = "{host}"',
            "smtp_port = 465",
            f'smtp_user = "{user}"',
            f'smtp_password = "{password}"',
            f'smtp_from_name = "{from_name}"',
            f'smtp_ssl_verify = "{ssl_verify}"',
            "",
        ]
    )
    with open(secrets_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines_out))
    _ = flat  # keep parse for future merge needs


def _local_smtp_get(*keys: str) -> str:
    data = load_local_smtp()
    if not data:
        return ""
    for k in keys:
        if not k:
            continue
        v = str(data.get(k) or "").strip()
        if v:
            return v
        low = str(k).strip().lower()
        # SMTP_USER → smtp_user
        if low.startswith("smtp_") or low in (
            "mail_user",
            "mail_password",
            "mail_app_password",
            "daum_mail_id",
            "daum_mail_password",
        ):
            aliases = {
                "smtp_user": ("smtp_user", "mail_user", "daum_mail_id"),
                "SMTP_USER": ("smtp_user",),
                "mail_user": ("smtp_user", "mail_user"),
                "daum_mail_id": ("smtp_user", "daum_mail_id"),
                "smtp_password": ("smtp_password", "mail_password", "mail_app_password", "daum_mail_password"),
                "SMTP_PASSWORD": ("smtp_password",),
                "mail_password": ("smtp_password", "mail_password"),
                "mail_app_password": ("smtp_password", "mail_app_password"),
                "daum_mail_password": ("smtp_password", "daum_mail_password"),
                "smtp_provider": ("smtp_provider",),
                "SMTP_PROVIDER": ("smtp_provider",),
                "smtp_host": ("smtp_host",),
                "SMTP_HOST": ("smtp_host",),
                "smtp_port": ("smtp_port",),
                "SMTP_PORT": ("smtp_port",),
                "smtp_from": ("smtp_from",),
                "SMTP_FROM": ("smtp_from",),
                "smtp_from_name": ("smtp_from_name",),
                "SMTP_FROM_NAME": ("smtp_from_name",),
                "smtp_ssl": ("smtp_ssl",),
                "SMTP_SSL": ("smtp_ssl",),
                "smtp_ssl_verify": ("smtp_ssl_verify",),
                "SMTP_SSL_VERIFY": ("smtp_ssl_verify",),
            }
            for alt in aliases.get(k, ()):
                vv = str(data.get(alt) or "").strip()
                if vv:
                    return vv
    return ""


def _pi_rerun(*, full: bool = False) -> None:
    """버튼·저장 후 갱신. 기본은 공문 fragment만(전체 앱 로딩 생략).

    dialog 닫기 등 앱 전역 상태가 필요할 때만 full=True.
    """
    if full:
        st.rerun()
    try:
        st.rerun(scope="fragment")
    except TypeError:
        st.rerun()


def _secret_get(*keys: str) -> str:
    """secrets.toml / Cloud secrets / env / 로컬 smtp_local.toml 조회.

    평면 키(`smtp_user`)와 중첩(`[smtp].user`, `[mail].password`) 모두 지원.
    """
    nested_sections = ("smtp", "mail", "email", "daum", "gmail")

    def _from_mapping(obj: Any, key: str) -> str:
        if obj is None:
            return ""
        try:
            if hasattr(obj, "get"):
                v = obj.get(key)
            else:
                v = obj[key]
        except Exception:
            v = None
        if v is None:
            return ""
        # nested section object — not a leaf value
        if hasattr(v, "get") and not isinstance(v, (str, bytes, int, float, bool)):
            return ""
        s = str(v).strip()
        return s

    try:
        secrets = getattr(st, "secrets", None)
    except Exception:
        secrets = None
    if secrets is not None:
        for k in keys:
            got = _from_mapping(secrets, k)
            if got:
                return got
            # smtp_user → section smtp, key user
            low = str(k or "").strip().lower()
            for sec in nested_sections:
                prefix = f"{sec}_"
                if low.startswith(prefix) and len(low) > len(prefix):
                    leaf = low[len(prefix) :]
                    try:
                        section = secrets.get(sec) if hasattr(secrets, "get") else secrets[sec]
                    except Exception:
                        section = None
                    got = _from_mapping(section, leaf)
                    if got:
                        return got
                    # aliases inside section
                    for alt in (leaf, "user", "username", "id", "password", "pass", "app_password"):
                        if alt == leaf:
                            continue
                        if leaf in ("user", "username", "id") and alt in ("user", "username", "id"):
                            got = _from_mapping(section, alt)
                            if got:
                                return got
                        if leaf in ("password", "pass", "app_password") and alt in (
                            "password",
                            "pass",
                            "app_password",
                        ):
                            got = _from_mapping(section, alt)
                            if got:
                                return got
            # bare leaf inside [smtp]: user / password
            if low in ("user", "username", "password", "host", "port", "from", "from_name", "provider"):
                for sec in nested_sections:
                    try:
                        section = secrets.get(sec) if hasattr(secrets, "get") else secrets[sec]
                    except Exception:
                        section = None
                    got = _from_mapping(section, low)
                    if got:
                        return got
    for k in keys:
        v = (os.environ.get(k) or "").strip()
        if v:
            return v
    # 로컬 저장본 (Mac · secrets 없을 때)
    got = _local_smtp_get(*keys)
    if got:
        return got
    return ""


def _norm_name(s: Any) -> str:
    t = str(s or "").strip()
    t = re.sub(r"\s+", "", t)
    t = t.rstrip(".")
    return t.lower()


def _core_name(s: Any) -> str:
    """매칭용 핵심명: (주)/주식회사/공백 제거, 흔한 접미 완화."""
    t = _norm_name(s)
    t = re.sub(r"^\(주\)|^㈜|^주식회사", "", t)
    t = re.sub(r"주식회사$", "", t)
    # 끝이 '상사'면 '상'까지도 허용 비교용으로 원본 core 유지
    return t


@st.cache_data(show_spinner=False)
def _read_contacts_csv_bytes(data: bytes) -> pd.DataFrame:
    """다음/아웃룩/엑셀 주소록 CSV — 인코딩·구분자 자동 시도."""
    if not data:
        return pd.DataFrame()
    # UTF-16 BOM ( occasional Outlook exports )
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        try:
            return pd.read_csv(io.BytesIO(data), encoding="utf-16")
        except Exception:
            pass
    encodings = ("utf-8-sig", "utf-8", "cp949", "euc-kr", "latin1")
    seps = (",", ";", "\t")
    last_err: Exception | None = None
    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(io.BytesIO(data), encoding=enc, sep=sep, engine="python")
            except Exception as e:
                last_err = e
                continue
            if df is None or df.empty:
                continue
            # 열이 1개뿐이면 구분자 오판 가능성 → 다른 sep 계속
            if len(df.columns) == 1 and sep != "\t":
                only = str(df.columns[0])
                if "," in only or ";" in only or "\t" in only:
                    continue
            return df
    if last_err:
        raise last_err
    return pd.DataFrame()


def _read_contacts_csv_path(path: str) -> pd.DataFrame:
    with open(path, "rb") as f:
        return _read_contacts_csv_bytes(f.read())


def _load_mail_contacts_cached(path: str, mtime: float) -> pd.DataFrame:
    """mtime을 키에 넣어 파일 변경 시에만 다시 읽음."""
    try:
        df = _read_contacts_csv_path(path)
    except Exception:
        return pd.DataFrame(columns=["거래처", "이메일", "비고"])
    return _normalize_mail_df(df)


def load_mail_contacts(path: str = PI_MAIL_CSV) -> pd.DataFrame:
    if not os.path.isfile(path):
        return pd.DataFrame(columns=["거래처", "이메일", "비고"])
    try:
        mtime = float(os.path.getmtime(path))
    except OSError:
        mtime = 0.0
    return _load_mail_contacts_cached(path, mtime)


def _mail_contact_candidate_paths() -> list[str]:
    """로컬/Cloud에서 자동 적재할 주소록 CSV 후보."""
    home = os.path.expanduser("~")
    desk = os.path.join(home, "Desktop")
    names = (
        "mail_contacts.csv",
        "daum-addrbook.csv",
        "daum_addrbook.csv",
        "주소록.csv",
        "다음주소록.csv",
        "카카오주소록.csv",
        "contacts.csv",
        "email_contacts.csv",
    )
    dirs = (
        PI_DIR,
        os.path.join("uploaded_cache", "price_increase"),
        desk,
        os.path.join(desk, "dashboard"),
        os.path.join(desk, "업무"),
        os.path.join(home, "Downloads"),
        os.path.join(home, "Desktop", "dashboard", "uploaded_cache", "price_increase"),
    )
    out: list[str] = []
    seen: set[str] = set()
    for d in dirs:
        for n in names:
            p = os.path.join(d, n)
            ap = os.path.abspath(p)
            if ap in seen:
                continue
            seen.add(ap)
            out.append(p)
    return out


def ensure_mail_contacts_autoload(path: str = PI_MAIL_CSV) -> tuple[pd.DataFrame, str]:
    """저장된 연락처를 읽고, 비어 있으면 Desktop/캐시 CSV를 자동 적재.

    한 번 업로드·저장되면 다음부터는 저장본을 자동 사용(재업로드 불필요).
    Returns: (mail_df, note) — note는 UI용 짧은 안내(없으면 '').
    """
    _ensure_dirs()
    cur = load_mail_contacts(path)
    if not cur.empty:
        if not st.session_state.get("_pi_mail_saved_noted"):
            st.session_state["_pi_mail_saved_noted"] = True
            return cur, f"연락처 저장본 자동사용 · {len(cur)}건 (재업로드 불필요)"
        return cur, ""
    # 이미 세션에서 자동적재 시도했으면 반복 스킵(업로더는 계속 가능)
    if st.session_state.get("_pi_mail_autoload_done"):
        return cur, ""
    st.session_state["_pi_mail_autoload_done"] = True
    best_df = pd.DataFrame(columns=["거래처", "이메일", "비고"])
    best_src = ""
    for cand in _mail_contact_candidate_paths():
        if not os.path.isfile(cand):
            continue
        if os.path.abspath(cand) == os.path.abspath(path):
            continue
        try:
            raw = _read_contacts_csv_path(cand)
            merged = _normalize_mail_df(raw)
        except Exception:
            continue
        if len(merged) > len(best_df):
            best_df = merged
            best_src = cand
    if best_df.empty:
        return cur, ""
    try:
        save_mail_contacts(best_df, path)
    except Exception:
        return best_df, f"자동 적재(메모리): `{os.path.basename(best_src)}` {len(best_df)}건"
    st.session_state["_pi_mail_saved_noted"] = True
    return load_mail_contacts(path), f"연락처 자동 적재: `{os.path.basename(best_src)}` → {len(best_df)}건"


def _col_key(label: Any) -> str:
    return str(label or "").strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def _pick_mail_columns(df: pd.DataFrame) -> tuple[Any, Any, Any, Any]:
    """(회사열, 이름열, 메일열, 비고열) — 다음/아웃룩 CSV 대응."""
    company_c = None
    name_c = None
    mail_c = None
    note_c = None
    # 점수제로 우선순위 높은 열 선택 (부분일치 오탐 완화)
    company_score = -1
    name_score = -1
    mail_score = -1
    for c in df.columns:
        low = _col_key(c)
        if not low or low.startswith("unnamed"):
            continue
        # 회사/상호 (매출 거래처 매칭용) — 회사전화·회사주소 제외
        if any(x in low for x in ("회사전화", "회사주소", "회사fax", "회사팩스", "businessphone", "companymain")):
            pass
        else:
            cs = -1
            if low in ("회사", "회사명", "상호", "업체", "거래처", "company", "companyname", "organization", "org"):
                cs = 100
            elif any(k in low for k in ("회사명", "companyname", "거래처", "상호")):
                cs = 80
            elif low == "회사" or low.endswith("회사") or low.startswith("회사"):
                if "전화" not in low and "주소" not in low and "팩스" not in low:
                    cs = 70
            elif "company" in low and "phone" not in low and "fax" not in low and "street" not in low:
                cs = 60
            if cs > company_score:
                company_score = cs
                company_c = c
        # 이름/표시명
        ns = -1
        if low in ("이름", "성명", "표시이름", "displayname", "fullname", "fullname", "연락처이름"):
            ns = 100
        elif low in ("firstname", "first"):
            ns = 50
        elif low in ("lastname", "last"):
            ns = 40
        elif any(k in low for k in ("표시이름", "displayname", "fullname", "성명", "이름")):
            if "파일" not in low and "username" not in low:
                ns = 70
        elif "name" in low and "username" not in low and "filename" not in low:
            # First Name / Last Name / Full Name
            if low in ("fullname", "fullname") or "full" in low or "display" in low:
                ns = 90
            elif "first" in low:
                ns = 55
            elif "last" in low:
                ns = 45
            else:
                ns = 30
        if ns > name_score:
            name_score = ns
            name_c = c
        # 이메일 (E-mail Address, 전자 메일 등)
        ms = -1
        if low in ("이메일", "메일", "email", "e-mail", "mail", "전자우편", "전자메일", "전자메일주소"):
            ms = 100
        elif any(k in low for k in ("emailaddress", "e-mailaddress", "전자메일", "이메일")):
            ms = 90
        elif low.startswith("email") or low.startswith("e-mail") or "메일주소" in low:
            ms = 80
        elif "email" in low or "e-mail" in low or (low.endswith("메일") and "스팸" not in low):
            ms = 60
        elif low == "mail" or low.endswith("mail"):
            ms = 50
        if ms > mail_score:
            mail_score = ms
            mail_c = c
        if note_c is None and any(k in low for k in ("비고", "메모", "소속", "그룹", "note", "부서", "department")):
            note_c = c
    if mail_c is None:
        for c in df.columns:
            sample = df[c].astype(str).head(50).str.cat(sep=" ")
            if _EMAIL_RE.search(sample):
                mail_c = c
                break
    return company_c, name_c, mail_c, note_c


def _extract_email_cell(val: Any) -> str:
    s = str(val or "").strip()
    if not s or s.lower() in ("nan", "none", "null"):
        return ""
    m = _EMAIL_RE.search(s)
    return m.group(0) if m else ""


def _normalize_mail_df(df: pd.DataFrame) -> pd.DataFrame:
    """다음·아웃룩 주소록 → 거래처/이메일/비고.

    매출 거래처는 보통 회사명이므로 회사열이 있으면 회사명을 우선 키로 쓰고,
    이름(담당자)도 별도 행으로 넣어 둘 다 매칭 가능하게 한다.
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=["거래처", "이메일", "비고"])
    # 헤더가 없는 경우 대비: 첫 행이 이메일처럼 보이면 그대로 진행
    company_c, name_c, mail_c, note_c = _pick_mail_columns(df)
    if mail_c is None and name_c is None and company_c is None:
        return pd.DataFrame(columns=["거래처", "이메일", "비고"])

    # First+Last Name 결합 (아웃룩)
    first_c = last_c = None
    for c in df.columns:
        low = _col_key(c)
        if low in ("firstname", "first", "이름(이름)", "이름"):
            if first_c is None and ("last" not in low):
                # '이름'은 전체 이름일 수도 있음 — First Name만 first_c
                if low in ("firstname", "first"):
                    first_c = c
        if low in ("lastname", "last", "성"):
            last_c = c

    rows: list[dict[str, str]] = []
    for _, r in df.iterrows():
        em = _extract_email_cell(r.get(mail_c)) if mail_c is not None else ""
        if not em:
            # 행 전체에서 이메일 스캔
            for c in df.columns:
                em = _extract_email_cell(r.get(c))
                if em:
                    break
        if not em:
            continue
        note = ""
        if note_c is not None:
            note = str(r.get(note_c) or "").strip()
            if note.lower() == "nan":
                note = ""
        company = ""
        if company_c is not None:
            company = str(r.get(company_c) or "").strip()
            if company.lower() == "nan":
                company = ""
        person = ""
        # 아웃룩 First+Last 가 있으면 결합명을 우선
        if first_c is not None:
            fn = str(r.get(first_c) or "").strip()
            ln = str(r.get(last_c) or "").strip() if last_c is not None else ""
            if fn.lower() == "nan":
                fn = ""
            if ln.lower() == "nan":
                ln = ""
            if fn or ln:
                # 한글 성은 앞, 영문은 First Last
                if fn and ln and re.search(r"[가-힣]", fn + ln):
                    person = f"{ln}{fn}".strip()
                else:
                    person = f"{fn} {ln}".strip()
        if not person and name_c is not None:
            person = str(r.get(name_c) or "").strip()
            if person.lower() == "nan":
                person = ""
        # 회사 우선 + 이름 보조 (둘 다 있으면 두 키로 저장)
        keys: list[str] = []
        if company:
            keys.append(company)
        if person and _norm_name(person) not in {_norm_name(k) for k in keys}:
            keys.append(person)
        if not keys:
            continue
        for i, key in enumerate(keys):
            n = note
            if i == 0 and company and person and _norm_name(company) != _norm_name(person):
                n = (f"{person} · {note}".strip(" ·") if note else person)
            elif i > 0 and company:
                n = (f"{company} · {note}".strip(" ·") if note else company)
            rows.append({"거래처": key, "이메일": em, "비고": n})

    out = pd.DataFrame(rows, columns=["거래처", "이메일", "비고"])
    if out.empty:
        return out
    out = out[(out["거래처"] != "") & (out["이메일"] != "")]
    out = out.drop_duplicates(subset=["거래처", "이메일"], keep="last")
    return out.reset_index(drop=True)


def save_mail_contacts(df: pd.DataFrame, path: str = PI_MAIL_CSV) -> None:
    _ensure_dirs()
    df.to_csv(path, index=False, encoding="utf-8-sig")
    try:
        _load_mail_contacts_cached.clear()
    except Exception:
        pass


def lookup_email(client: str, mail_df: pd.DataFrame) -> str:
    """연락처에서 이메일 찾기. 정확·포함·핵심명(상사/(주) 완화) 순."""
    hit, _matched_as = lookup_email_with_meta(client, mail_df)
    return hit


def lookup_email_with_meta(client: str, mail_df: pd.DataFrame) -> tuple[str, str]:
    """(이메일, 매칭된연락처명). 없으면 ('', '')."""
    if not client or mail_df is None or mail_df.empty:
        return "", ""
    key = _norm_name(client)
    core = _core_name(client)
    # 1) 정확 일치
    for _, row in mail_df.iterrows():
        n = _norm_name(row.get("거래처"))
        if n == key:
            return str(row.get("이메일") or "").strip(), str(row.get("거래처") or "")
    # 2) 한쪽이 다른 쪽을 포함 (대영가스상 ⊂ 대영가스상사)
    best = ("", "", 0)  # email, name, score
    for _, row in mail_df.iterrows():
        raw = str(row.get("거래처") or "")
        n = _norm_name(raw)
        c = _core_name(raw)
        em = str(row.get("이메일") or "").strip()
        if not n or not em:
            continue
        score = 0
        if key in n or n in key:
            score = max(len(n), len(key))
        elif core and c and (core in c or c in core):
            score = max(len(c), len(core)) - 1
        elif core and c and (core.startswith(c) or c.startswith(core)) and min(len(c), len(core)) >= 4:
            score = min(len(c), len(core))
        if score > best[2]:
            best = (em, raw, score)
    if best[2] > 0:
        return best[0], best[1]
    return "", ""


def suggest_mail_matches(client: str, mail_df: pd.DataFrame, limit: int = 30) -> list[dict]:
    """비슷한 연락처 후보 (사용자가 목록에서 선택)."""
    if not client or mail_df is None or mail_df.empty:
        return []
    core = _core_name(client)
    prefix2 = core[:2] if len(core) >= 2 else core
    prefix3 = core[:3] if len(core) >= 3 else core
    rows: list[dict] = []
    seen: set[str] = set()
    for _, row in mail_df.iterrows():
        raw = str(row.get("거래처") or "").strip()
        c = _core_name(raw)
        em = str(row.get("이메일") or "").strip()
        if not em or not raw:
            continue
        uid = f"{raw}|{em}"
        if uid in seen:
            continue
        keep = False
        if prefix3 and prefix3 in c:
            keep = True
        elif prefix2 and c.startswith(prefix2):
            keep = True
        elif core and c and (core in c or c in core):
            keep = True
        if keep:
            seen.add(uid)
            rows.append({"거래처": raw, "이메일": em})

    def _pref(r: dict) -> tuple[int, int]:
        a, b = core, _core_name(r["거래처"])
        i = 0
        while i < min(len(a), len(b)) and a[i] == b[i]:
            i += 1
        # 포함이면 가산
        contain = 1 if (a in b or b in a) else 0
        return (contain, i)

    rows.sort(key=_pref, reverse=True)
    return rows[:limit]


def exact_mail_match(client: str, mail_df: pd.DataFrame) -> tuple[str, str]:
    """이름 정규화 후 완전 일치만."""
    if not client or mail_df is None or mail_df.empty:
        return "", ""
    key = _norm_name(client)
    for _, row in mail_df.iterrows():
        if _norm_name(row.get("거래처")) == key:
            return str(row.get("이메일") or "").strip(), str(row.get("거래처") or "")
    return "", ""


def list_staff_options(sales_df: pd.DataFrame) -> list[str]:
    if sales_df is None or sales_df.empty or "담당자" not in sales_df.columns:
        return []
    vals = sorted(
        {
            str(x).strip()
            for x in sales_df["담당자"].dropna().unique()
            if str(x).strip() and str(x).strip() not in ("nan", "거래종료")
        }
    )
    return vals


def list_clients_for_staff(sales_df: pd.DataFrame, staff: str) -> list[str]:
    if sales_df is None or sales_df.empty:
        return []
    df = sales_df
    if staff and staff != "전체" and "담당자" in df.columns:
        df = df[df["담당자"].astype(str).str.strip() == staff]
    names: set[str] = set()
    if "거래처" in df.columns:
        names |= set(df["거래처"].dropna().astype(str).str.strip().unique())
    if "거래처_원본" in df.columns:
        names |= set(df["거래처_원본"].dropna().astype(str).str.strip().unique())
    return sorted(n for n in names if n and n != "nan")


def classify_client_kind(name: str, sales_df: pd.DataFrame) -> str:
    """종속 | 단독(부모) | 단독."""
    if not name:
        return "단독"
    if "(" in name and name.endswith(")"):
        return "종속"
    if sales_df is not None and not sales_df.empty and "거래처_원본" in sales_df.columns:
        base = name.rstrip(".")
        kids = sales_df[
            sales_df["거래처"].astype(str).str.contains(re.escape(base) + r"\(", regex=True, na=False)
        ]
        if len(kids) > 0:
            return "단독(부모)"
    return "단독"


def latest_unit_prices(sales_df: pd.DataFrame, client: str) -> pd.DataFrame:
    """거래처 선택 시 최근 적용 단가 (품목별). session 캐시로 반복 계산·일괄 발송 부하 완화."""
    empty = pd.DataFrame(columns=["품목명", "기존단가", "최근매출일", "출고량합"])
    if sales_df is None or sales_df.empty or not client:
        return empty
    token = str(st.session_state.get("pi_sales_cache_token") or "")
    bucket = st.session_state.setdefault("_pi_unit_price_cache", {})
    if bucket.get("_token") != token:
        bucket.clear()
        bucket["_token"] = token
    cached = bucket.get(client)
    if isinstance(cached, pd.DataFrame):
        return cached.copy()
    out = _compute_latest_unit_prices(sales_df, client)
    bucket[client] = out
    return out.copy()


def _compute_latest_unit_prices(sales_df: pd.DataFrame, client: str) -> pd.DataFrame:
    """거래처별 최근 단가 실계산 (캐시 미스 시)."""
    empty = pd.DataFrame(columns=["품목명", "기존단가", "최근매출일", "출고량합"])
    df = sales_df
    # 부모 선택 시 종속 포함
    if "거래처_원본" in df.columns:
        m = (df["거래처"].astype(str).str.strip() == client) | (
            df["거래처_원본"].astype(str).str.strip() == client
        )
    else:
        m = df["거래처"].astype(str).str.strip() == client
    sub = df.loc[m]
    if sub.empty:
        return empty
    if "품목명" not in sub.columns or "단가" not in sub.columns:
        return empty
    sub = sub.copy()
    sub["품목명"] = sub["품목명"].astype(str).str.strip()
    sub = sub[~sub["품목명"].str.contains(_NOISE_ITEM, na=False)]
    sub["단가"] = pd.to_numeric(sub["단가"], errors="coerce").fillna(0)
    if "출고량" in sub.columns:
        sub["출고량"] = pd.to_numeric(sub["출고량"], errors="coerce").fillna(0)
    else:
        sub["출고량"] = 0.0
    date_col = "매출일_dt" if "매출일_dt" in sub.columns else None
    rows = []
    for item, g in sub.groupby("품목명", dropna=True):
        g2 = g[g["단가"] > 0]
        if g2.empty:
            continue
        if date_col and g2[date_col].notna().any():
            g2 = g2.sort_values(date_col)
            last = g2.iloc[-1]
            price = float(last["단가"])
            last_d = last[date_col]
            last_s = (
                last_d.strftime("%Y-%m-%d")
                if hasattr(last_d, "strftime")
                else str(last_d)[:10]
            )
        else:
            price = float(g2["단가"].mode().iloc[0]) if not g2["단가"].mode().empty else float(g2["단가"].iloc[-1])
            last_s = ""
        rows.append(
            {
                "품목명": str(item),
                "기존단가": round(price, 2),
                "최근매출일": last_s,
                "출고량합": float(g["출고량"].sum()),
            }
        )
    if not rows:
        return empty
    return pd.DataFrame(rows).sort_values("출고량합", ascending=False).reset_index(drop=True)


def default_increase_price(old: float, pct: float) -> float:
    try:
        return round(float(old) * (1.0 + float(pct) / 100.0), 1)
    except Exception:
        return float(old or 0)


def apply_increase_by_amount(old: float, amount: float) -> float:
    try:
        return round(float(old) + float(amount), 1)
    except Exception:
        return float(old or 0)


# 탄산단가인상공문 기본 본문 (양식 셀 유지, 내용만 편집)
_DEFAULT_LETTER_PARAS: dict[str, str] = {
    "C16": "1.귀사의 무궁한 발전을 기원하며, 평소 당사에 보내주시는 신뢰와 협력에 깊은 감사를 드립니다.",
    "C18": "2.당사는 귀사와의 지속적인 파트너십 유지를 최우선 가치로 삼아, 대내외적인 원가 상승 압박 속에서도 경영 ",
    "C19": "효율화를 통해 단가 인상을 최대한 억제해 왔습니다.",
    "C21": "3.그러나 최근 지정학적 리스크 심화와 글로벌 에너지 공급망의 불안정성으로 인해 당사가 감내할 수 있는 . ",
    "C22": "임계치를 상회하는 제조 원가 상승이 발생하였습니다.",
    "C23": "이에 안정적인 품질 유지와 지속 가능한 공급 체계 확보를 위해 부득이하게 아래와 같이 단가 인상을 요청드리오니 널리 양해하여 ",
    "C24": "양해하여 주시기 바랍니다. ",
    "C26": "4. 주요 인상 요인 분석",
    "C27": "• 중동 분쟁 장기화에 따른 에너지 비용 급등: 중동 지역의 지정학적 불안정 지속으로 국제 유가 및 천연가스 가격의 변동성이 확대되었으며, 탄산 원료 가스(Raw Gas) 수급 비용 및 생산 설비 가동 에너지 비용이 급격히 상승함.",
    "C29": "• 원료 가스 확보 단가 및 제조 원가 상승: 석유화학 플랜트 가동률 변화와 원료 공급원의 제한적 수급 상황이 맞물려 원료 가스 매입 단가가 폭등하였으며, 정제 및 액화 과정의 부자재 가격 상승이 직접적인 원가 부담으로 작용함.",
    "C31": "• 물류망 불안정에 따른 운반비 가중: 유가 상승 및 요소수 등 차량 유지비 증가로 인해 내륙 운송 단가가 상향 평준화되었으며, 특히 특수 고압 탱크로리 운영비용 상승이 전체 공급가에 심대한 영향을 미침.",
    "C33": "5. 당사는 이번 조정을 바탕으로 더욱 철저한 품질 관리와 원활한 공급 체계를 구축하여 귀사의 기대에 보답할 것을 약속드립니다.",
    "C35": "                                      단가 조정 내용",
}

# 한 칸 공문입력 → 엑셀 본문 셀 매핑 순서 (C35 제목 제외)
_BODY_CELL_ORDER = [
    "C16",
    "C18",
    "C19",
    "C21",
    "C22",
    "C23",
    "C24",
    "C26",
    "C27",
    "C29",
    "C31",
    "C33",
]


def _default_letter_body_text() -> str:
    return "\n".join(_DEFAULT_LETTER_PARAS[c] for c in _BODY_CELL_ORDER if c in _DEFAULT_LETTER_PARAS)


def _body_text_to_paras(text: str) -> dict[str, str]:
    """한 칸 본문 → 양식 셀 dict. 기본 양식 제목(C35)은 유지."""
    paras = dict(_DEFAULT_LETTER_PARAS)
    for c in _BODY_CELL_ORDER:
        paras[c] = ""
    lines = [ln.rstrip() for ln in str(text or "").splitlines()]
    if not any(ln.strip() for ln in lines):
        return dict(_DEFAULT_LETTER_PARAS)
    for i, line in enumerate(lines):
        if i < len(_BODY_CELL_ORDER):
            paras[_BODY_CELL_ORDER[i]] = line
        else:
            paras[_BODY_CELL_ORDER[-1]] = (paras[_BODY_CELL_ORDER[-1]] + "\n" + line).strip()
    paras["C35"] = _DEFAULT_LETTER_PARAS.get("C35", "                                      단가 조정 내용")
    return paras


def ensure_default_template(path: str = PI_TEMPLATE) -> str:
    """원본 공문 양식이 없으면 기본 양식 생성."""
    _ensure_dirs()
    if os.path.isfile(path) and load_workbook is not None:
        return path
    if Workbook is None:
        return path
    wb = Workbook()
    ws = wb.active
    ws.title = "단가인상공문"
    thin = Border(
        left=Side(style="thin", color="94A3B8"),
        right=Side(style="thin", color="94A3B8"),
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="thin", color="94A3B8"),
    )
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 28

    ws.merge_cells("B2:F2")
    ws["B2"] = "신일가스(주)"
    ws["B2"].font = Font(name="맑은 고딕", size=16, bold=True, color="0F766E")
    ws["B2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:F3")
    ws["B3"] = "단 가 인 상 안 내 공 문"
    ws["B3"].font = Font(name="맑은 고딕", size=18, bold=True)
    ws["B3"].alignment = Alignment(horizontal="center")

    ws["B5"] = "문서번호"
    ws["C5"] = "{{DOC_NO}}"
    ws["B6"] = "작성일"
    ws["C6"] = "{{DATE}}"
    ws["B7"] = "수신"
    ws["C7"] = "{{CLIENT}} 귀중"
    ws["B8"] = "참조"
    ws["C8"] = "{{EMAIL}}"
    ws["B9"] = "제목"
    ws.merge_cells("C9:F9")
    ws["C9"] = "{{TITLE}}"

    ws.merge_cells("B11:F11")
    ws["B11"] = (
        "항상 저희 신일가스를 이용해 주셔서 감사드립니다. "
        "원자재·물류비 상승에 따라 아래와 같이 공급 단가 조정을 안내드립니다."
    )
    ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[11].height = 48

    headers = ["순번", "품목명", "기존단가(원)", "인상적용단가(원)", "비고"]
    for i, h in enumerate(headers, start=2):
        cell = ws.cell(14, i, h)
        cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
        cell.fill = __import__("openpyxl.styles", fromlist=["PatternFill"]).PatternFill(
            "solid", fgColor="0F766E"
        )
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    # 플레이스홀더 행 (최대 20)
    for r in range(15, 35):
        for c in range(2, 7):
            cell = ws.cell(r, c, "")
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 2, f"{{{{ITEM_{r-14}_NO}}}}")
        ws.cell(r, 3, f"{{{{ITEM_{r-14}_NAME}}}}")
        ws.cell(r, 4, f"{{{{ITEM_{r-14}_OLD}}}}")
        ws.cell(r, 5, f"{{{{ITEM_{r-14}_NEW}}}}")
        ws.cell(r, 6, f"{{{{ITEM_{r-14}_NOTE}}}}")

    ws.merge_cells("B36:F36")
    ws["B36"] = "{{BODY}}"
    ws["B36"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[36].height = 80

    ws.merge_cells("B38:F38")
    ws["B38"] = "적용예정일: {{EFFECTIVE}}"
    ws.merge_cells("B40:F40")
    ws["B40"] = "신일가스(주) 화성공장"
    ws["B40"].alignment = Alignment(horizontal="right")
    ws.merge_cells("B41:F41")
    ws["B41"] = "문의: {{CONTACT}}"
    ws["B41"].alignment = Alignment(horizontal="right")

    wb.save(path)
    wb.close()
    return path


def _is_real_letter_template(ws) -> bool:
    v9 = str(ws["C9"].value or "")
    v10 = str(ws["C10"].value or "")
    return "문서번호" in v9 or "발송일자" in v10


def _format_doc_no(client: str, doc_no: str = "") -> str:
    if doc_no:
        return doc_no
    return f"신일(1)-{date.today().strftime('%y%m')}-{abs(hash(client)) % 10000:04d}"


def _format_korean_effective(effective: str) -> str:
    t = str(effective or "").strip()
    if not t:
        return f"{date.today().strftime('%Y년 %m월 %d일')} 출고분부터"
    m = re.match(r"^(\d{4})[-./](\d{1,2})[-./](\d{1,2})", t)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y}년 {mo:02d}월 {d:02d}일 출고분부터"
    if "년" in t and "월" in t:
        return t if "출고" in t else f"{t} 출고분부터"
    return t


def _format_c37_target_items(items: list[dict]) -> str:
    names = [str(it.get("품목명") or "").strip() for it in items if str(it.get("품목명") or "").strip()]
    joined = ", ".join(names) if names else "-"
    return f"                           • 대상 품목 : {joined} "


def _format_c38_price_lines(items: list[dict]) -> str:
    parts: list[str] = []
    for it in items:
        name = str(it.get("품목명") or "").strip()
        if not name:
            continue
        old = float(it.get("기존단가") or 0)
        new = float(it.get("인상적용단가") or 0)
        diff = new - old
        if len(items) == 1:
            parts.append(f"기존단가 {name} {old:,.0f}원 ➠ {new:,.0f}원 (+{diff:,.0f}원)")
        else:
            parts.append(f"{name} {old:,.0f}원 ➠ {new:,.0f}원 (+{diff:,.0f}원)")
    body = " / ".join(parts) if parts else "별도 협의"
    return f"                           • 단가 인상 금액 : {body} "


def _unmerge_row(ws, row: int) -> None:
    from openpyxl.utils import range_boundaries

    merged = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row <= row <= mr.max_row]
    for mr in merged:
        rng = str(mr)
        try:
            min_c, min_r, max_c, max_r = range_boundaries(rng)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if (r, c) not in ws._cells:
                        ws.cell(row=r, column=c, value=None)
            ws.unmerge_cells(rng)
        except Exception:
            try:
                ws.merged_cells.ranges.remove(mr)
            except Exception:
                pass


def _shrink_letter_body_for_table(ws, n_items: int) -> None:
    """표 칸이 모자라면 공문 본문 문단 행 높이를 줄임."""
    if n_items <= 2:
        return
    # 본문·인상요인 구간 압축
    for r in range(16, 34):
        h = ws.row_dimensions[r].height
        if h is None:
            continue
        if n_items >= 5:
            ws.row_dimensions[r].height = max(8, float(h) * 0.55)
        elif n_items >= 3:
            ws.row_dimensions[r].height = max(10, float(h) * 0.75)
    if n_items >= 5:
        for coord in ("C27", "C29", "C31"):
            v = ws[coord].value
            if isinstance(v, str) and len(v) > 50:
                ws[coord].value = v[:50].rstrip() + "…"


def _write_price_adjust_table(ws, items: list[dict], effective: str) -> None:
    """단가 조정 내용: 제품명 | 기존단가 | 인상단가 | 비고 표."""
    clean = [it for it in items if str(it.get("품목명") or "").strip()]
    n = len(clean)
    _shrink_letter_body_for_table(ws, max(n, 1))

    need_rows = 1 + max(n, 1) + 1  # header + data + 시행일
    avail = 5  # rows 36..40
    if need_rows > avail:
        try:
            ws.insert_rows(41, need_rows - avail)
        except Exception:
            # 삽입 실패 시 품목 수를 가용 칸에 맞춤
            clean = clean[: max(1, avail - 2)]
            n = len(clean)
            need_rows = 1 + max(n, 1) + 1

    table_end = 35 + need_rows  # inclusive-ish
    for r in range(36, table_end + 1):
        _unmerge_row(ws, r)
        for c in range(3, 28):
            try:
                ws.cell(r, c).value = None
            except Exception:
                pass

    thin = Border(
        left=Side(style="thin", color="64748B"),
        right=Side(style="thin", color="64748B"),
        top=Side(style="thin", color="64748B"),
        bottom=Side(style="thin", color="64748B"),
    )
    header_fill = PatternFill("solid", fgColor="0F766E") if PatternFill else None
    spans = {"제품명": (3, 7), "기존단가": (8, 12), "인상단가": (13, 17), "비고": (18, 27)}

    header_row = 36
    for title, (c1, c2) in spans.items():
        try:
            if c2 > c1:
                ws.merge_cells(start_row=header_row, start_column=c1, end_row=header_row, end_column=c2)
        except Exception:
            pass
        cell = ws.cell(header_row, c1, title)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if header_fill is not None:
            cell.fill = header_fill
        for c in range(c1, c2 + 1):
            try:
                ws.cell(header_row, c).border = thin
            except Exception:
                pass
    ws.row_dimensions[header_row].height = 18

    data_start = 37
    rows_src = clean if clean else [{"품목명": "-", "기존단가": 0, "인상적용단가": 0, "비고": ""}]
    for i, it in enumerate(rows_src):
        r = data_start + i
        vals = {
            "제품명": str(it.get("품목명") or ""),
            "기존단가": f"{float(it.get('기존단가') or 0):,.0f}",
            "인상단가": f"{float(it.get('인상적용단가') or 0):,.0f}",
            "비고": str(it.get("비고") or ""),
        }
        for title, (c1, c2) in spans.items():
            try:
                if c2 > c1:
                    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            except Exception:
                pass
            cell = ws.cell(r, c1, vals[title])
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(
                horizontal="center" if title != "제품명" else "left",
                vertical="center",
            )
            for c in range(c1, c2 + 1):
                try:
                    ws.cell(r, c).border = thin
                except Exception:
                    pass
        ws.row_dimensions[r].height = 18

    eff_row = data_start + len(rows_src)
    _unmerge_row(ws, eff_row)
    try:
        ws.merge_cells(start_row=eff_row, start_column=3, end_row=eff_row, end_column=27)
    except Exception:
        pass
    cell = ws.cell(
        eff_row,
        3,
        f"                           • 시행 일자 : {_format_korean_effective(effective)}",
    )
    cell.font = Font(name="맑은 고딕", size=10)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[eff_row].height = 18


def _fill_real_template_cells(
    ws,
    *,
    client: str,
    title: str,
    effective: str,
    items: list[dict],
    doc_no: str,
    send_date: Optional[date] = None,
    letter_paras: Optional[dict[str, str]] = None,
    c37_override: str = "",
    c38_override: str = "",
) -> None:
    """양식(레이아웃·로고)은 유지하고 내용 셀만 채움. 단가조정은 표로 기록."""
    del c37_override, c38_override  # 문구 대신 표 사용
    send_date = send_date or date.today()
    ws["C9"].value = f"문서번호 : {doc_no}"
    ws["C10"].value = f"발송일자 : {send_date.strftime('%Y.%m.%d')}"
    ws["C11"].value = f"수    신 : {client}"
    ws["C12"].value = f"제    목 : {title}"
    paras = letter_paras if letter_paras is not None else _DEFAULT_LETTER_PARAS
    for coord, text in paras.items():
        if coord in ws:
            ws[coord].value = text
    # C35 이하 단가표: 품목이 있을 때만 (공문 종류에 따라 선택)
    clean_items = [it for it in items if str(it.get("품목명") or "").strip()]
    if clean_items:
        if not str(ws["C35"].value or "").strip():
            ws["C35"].value = "                                      단가 조정 내용"
        _write_price_adjust_table(ws, clean_items, effective)
    else:
        # 단가표 미사용 — 양식 기본 단가 문구/표 자리를 비움
        for coord in ("C35", "C36", "C37", "C38", "C39", "C40"):
            try:
                if coord in ws:
                    ws[coord].value = ""
            except Exception:
                pass
        for r in range(35, 41):
            _unmerge_row(ws, r)
            for c in range(3, 28):
                try:
                    ws.cell(r, c).value = None
                except Exception:
                    pass


def fill_letter_workbook(
    template_path: str,
    *,
    client: str,
    email: str,
    title: str,
    body: str,
    effective: str,
    contact: str,
    items: list[dict],
    doc_no: str = "",
    send_date: Optional[date] = None,
    letter_paras: Optional[dict[str, str]] = None,
    c37_override: str = "",
    c38_override: str = "",
) -> bytes:
    if load_workbook is None:
        raise RuntimeError("openpyxl 필요")
    ensure_default_template(template_path)
    wb = load_workbook(template_path)
    ws = wb.active
    today = send_date or date.today()
    today_s = today.strftime("%Y-%m-%d") if hasattr(today, "strftime") else str(today)[:10]
    doc_no = _format_doc_no(client, doc_no)

    if _is_real_letter_template(ws):
        _fill_real_template_cells(
            ws,
            client=client,
            title=title,
            effective=effective,
            items=items,
            doc_no=doc_no,
            send_date=today if isinstance(today, date) else date.today(),
            letter_paras=letter_paras,
            c37_override=c37_override,
            c38_override=c38_override,
        )
    else:
        mapping = {
            "{{DOC_NO}}": doc_no,
            "{{DATE}}": today_s,
            "{{CLIENT}}": client,
            "{{EMAIL}}": email or "-",
            "{{TITLE}}": title,
            "{{BODY}}": body,
            "{{EFFECTIVE}}": effective,
            "{{CONTACT}}": contact,
        }
        for i in range(1, 21):
            if i <= len(items):
                it = items[i - 1]
                mapping[f"{{{{ITEM_{i}_NO}}}}"] = str(i)
                mapping[f"{{{{ITEM_{i}_NAME}}}}"] = str(it.get("품목명") or "")
                mapping[f"{{{{ITEM_{i}_OLD}}}}"] = f"{float(it.get('기존단가') or 0):,.1f}"
                mapping[f"{{{{ITEM_{i}_NEW}}}}"] = f"{float(it.get('인상적용단가') or 0):,.1f}"
                mapping[f"{{{{ITEM_{i}_NOTE}}}}"] = str(it.get("비고") or "")
            else:
                mapping[f"{{{{ITEM_{i}_NO}}}}"] = ""
                mapping[f"{{{{ITEM_{i}_NAME}}}}"] = ""
                mapping[f"{{{{ITEM_{i}_OLD}}}}"] = ""
                mapping[f"{{{{ITEM_{i}_NEW}}}}"] = ""
                mapping[f"{{{{ITEM_{i}_NOTE}}}}"] = ""

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    for k, rep in mapping.items():
                        if k in v:
                            v = v.replace(k, rep)
                    cell.value = v

        for r in range(15, 35):
            if not ws.cell(r, 3).value:
                for c in range(2, 7):
                    ws.cell(r, c).value = None

    bio = io.BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue()


def resolve_letter_template() -> str:
    """탄산단가인상공문.xlsx 우선, 없으면 기본 생성 양식."""
    try:
        custom = st.session_state.get("pi_template_path")
    except Exception:
        custom = None
    if custom and os.path.isfile(custom):
        return custom
    for p in _TEMPLATE_CANDIDATES:
        if p and os.path.isfile(p):
            return p
    return ensure_default_template(PI_TEMPLATE)


def smtp_settings() -> dict:
    """다음메일(Daum) SMTP 기본. secrets로 덮어씀."""
    provider = (_secret_get("smtp_provider", "SMTP_PROVIDER") or "daum").strip().lower()
    defaults = {
        "daum": {"host": "smtp.daum.net", "port": 465, "ssl": True},
        "kakao": {"host": "smtp.daum.net", "port": 465, "ssl": True},
        "naver": {"host": "smtp.naver.com", "port": 465, "ssl": True},
        "gmail": {"host": "smtp.gmail.com", "port": 465, "ssl": True},
    }
    base = dict(defaults.get(provider, defaults["daum"]))
    host = _secret_get("smtp_host", "SMTP_HOST") or base["host"]
    port_s = _secret_get("smtp_port", "SMTP_PORT")
    port = int(port_s) if port_s else int(base["port"])
    ssl_s = (_secret_get("smtp_ssl", "SMTP_SSL") or ("1" if base["ssl"] else "0")).strip().lower()
    use_ssl = ssl_s in ("1", "true", "yes", "ssl")
    user = _secret_get("smtp_user", "SMTP_USER", "mail_user", "daum_mail_id")
    password = _secret_get(
        "smtp_password",
        "SMTP_PASSWORD",
        "mail_password",
        "mail_app_password",
        "daum_mail_password",
    )
    # 복사 붙여넣기 NBSP 등 제거 (Gmail 앱 비밀번호에서 자주 발생)
    user = str(user or "").replace("\xa0", "").replace("\u200b", "").strip()
    password = str(password or "").replace("\xa0", "").replace("\u200b", "").replace(" ", "").strip()
    from_addr = _secret_get("smtp_from", "SMTP_FROM") or user
    from_name = _secret_get("smtp_from_name", "SMTP_FROM_NAME") or "신일가스"
    from_addr = str(from_addr or "").replace("\xa0", "").strip()
    from_name = str(from_name or "신일가스").replace("\xa0", " ").strip()
    return {
        "provider": provider,
        "host": host,
        "port": port,
        "ssl": use_ssl,
        "user": user,
        "password": password,
        "from_addr": from_addr,
        "from_name": from_name,
        "ready": bool(user and password),
    }


def smtp_status_label(cfg: Optional[dict] = None) -> str:
    cfg = cfg or smtp_settings()
    if cfg.get("ready"):
        src = "로컬저장" if os.path.isfile(PI_SMTP_LOCAL) else "secrets"
        return f"연동됨 · {cfg.get('user')} → {cfg.get('host')}:{cfg.get('port')} ({src})"
    has_user = bool(cfg.get("user"))
    has_pw = bool(cfg.get("password"))
    if not has_user and not has_pw:
        return "미연동 — 아래「SMTP 계정 설정」에서 다음메일 저장"
    if not has_user:
        return "미연동 — 메일 아이디 없음"
    return "미연동 — 메일 비밀번호 없음"


def _smtp_ssl_contexts() -> list[ssl.SSLContext]:
    """기본 검증 → (실패 시) 검증 완화. Mac/프록시 self-signed 체인 대응."""
    contexts: list[ssl.SSLContext] = []
    try:
        contexts.append(ssl.create_default_context())
    except Exception:
        pass
    # secrets로 강제 완화
    relax = (_secret_get("smtp_ssl_verify", "SMTP_SSL_VERIFY") or "1").strip().lower()
    if relax in ("0", "false", "no", "off"):
        return [ssl._create_unverified_context()]
    unverified = ssl._create_unverified_context()
    if not contexts:
        contexts.append(unverified)
    else:
        contexts.append(unverified)
    return contexts


def _smtp_run(
    cfg: dict,
    *,
    timeout: int = 45,
    send_fn: Optional[Any] = None,
) -> tuple[bool, str]:
    """로그인(+선택 발송). CERTIFICATE_VERIFY_FAILED 시 검증 완화로 1회 재시도."""
    host, port = cfg["host"], int(cfg["port"])
    user, password = cfg["user"], cfg["password"]
    use_ssl = bool(cfg.get("ssl") or port == 465)
    last_err: Optional[BaseException] = None
    contexts = _smtp_ssl_contexts()
    for i, context in enumerate(contexts):
        try:
            if use_ssl:
                with smtplib.SMTP_SSL(host, port, context=context, timeout=timeout) as server:
                    server.login(user, password)
                    if send_fn is not None:
                        send_fn(server)
            else:
                with smtplib.SMTP(host, port, timeout=timeout) as server:
                    server.ehlo()
                    server.starttls(context=context)
                    server.ehlo()
                    server.login(user, password)
                    if send_fn is not None:
                        send_fn(server)
            note = ""
            if i > 0:
                note = " (SSL 인증서 검증 완화로 연결)"
            return True, note
        except smtplib.SMTPAuthenticationError:
            raise
        except Exception as e:
            last_err = e
            msg = str(e).lower()
            # 인증서 문제만 다음 컨텍스트로 재시도
            if "certificate" in msg or "cert verify" in msg or "ssl:" in msg:
                continue
            break
    assert last_err is not None
    raise last_err


def test_smtp_connection(cfg: Optional[dict] = None) -> tuple[bool, str]:
    cfg = cfg or smtp_settings()
    if not cfg.get("ready"):
        return False, "메일 아이디/비밀번호가 없습니다. 위「SMTP 계정 설정」에서 저장하세요."
    try:
        ok, note = _smtp_run(cfg, timeout=20, send_fn=None)
        host, port = cfg["host"], int(cfg["port"])
        return True, f"로그인 성공 ({host}:{port}){note}"
    except smtplib.SMTPAuthenticationError:
        return (
            False,
            "로그인 실패 — 다음메일: 메일설정→POP3/IMAP 사용 ON, 아이디는 전체메일(예: id@daum.net)",
        )
    except Exception as e:
        return False, str(e)


def _mail_clean_text(value: Any, *, keep_newlines: bool = False) -> str:
    """복사/붙여넣기 NBSP(\\xa0) 등 비가시 문자 제거 — ascii encode 오류 방지."""
    text = str(value or "")
    # non-breaking space 및 zero-width 문자
    for ch in ("\xa0", "\u202f", "\u2007", "\u200b", "\u200c", "\u200d", "\ufeff"):
        text = text.replace(ch, " " if ch in ("\xa0", "\u202f", "\u2007") else "")
    if keep_newlines:
        # 본문: 줄바꿈 유지, 양끝만 정리
        return text.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"[ \t]+", " ", text).strip()


def send_mail_smtp(
    *,
    to_addr: str,
    subject: str,
    body: str,
    attachment_bytes: Optional[bytes] = None,
    attachment_name: str = "단가인상공문.pdf",
    cc: str = "",
) -> tuple[bool, str]:
    cfg = smtp_settings()
    if not cfg.get("ready"):
        return (
            False,
            "메일 미연동: `.streamlit/secrets.toml` 또는 Cloud Secrets에 "
            "smtp_user / smtp_password 를 넣으세요. (Gmail: smtp.gmail.com:465)",
        )
    if not to_addr:
        return False, "수신 메일 없음"

    # secrets·본문에 섞인 NBSP 정리 (앱 비밀번호 복사 시 자주 발생)
    from_name = _mail_clean_text(cfg.get("from_name") or "신일가스")
    from_addr = _mail_clean_text(cfg.get("from_addr") or cfg.get("user") or "")
    user = _mail_clean_text(cfg.get("user") or "")
    password = _mail_clean_text(cfg.get("password") or "")
    cfg = dict(cfg)
    cfg["user"] = user
    cfg["password"] = password
    cfg["from_addr"] = from_addr
    cfg["from_name"] = from_name

    subject = _mail_clean_text(subject)
    body = _mail_clean_text(body, keep_newlines=True)
    to_addr = _mail_clean_text(to_addr)
    cc = _mail_clean_text(cc)
    attachment_name = _mail_clean_text(attachment_name) or "letter.pdf"

    recipients = [a for a in re.split(r"[;,]", to_addr) if a]
    cc_list = [a for a in re.split(r"[;,]", cc or "") if a]

    msg = MIMEMultipart()
    # 한글 From/Subject는 Header로 UTF-8 인코딩 (ascii codec 오류 방지)
    msg["From"] = formataddr((str(Header(from_name, "utf-8")), from_addr))
    msg["To"] = ", ".join(recipients)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg["Subject"] = str(Header(subject, "utf-8"))
    msg.attach(MIMEText(body, "plain", "utf-8"))
    if attachment_bytes:
        safe_name = attachment_name or "letter.pdf"
        subtype = "pdf" if str(safe_name).lower().endswith(".pdf") else "octet-stream"
        part = MIMEApplication(attachment_bytes, _subtype=subtype)
        # RFC2231 파일명 (한글 첨부명)
        part.add_header(
            "Content-Disposition",
            "attachment",
            filename=("utf-8", "", safe_name),
        )
        msg.attach(part)

    all_rcpt = recipients + cc_list

    def _do_send(server: smtplib.SMTP) -> None:
        # as_string()의 기본 ascii 인코딩을 피하기 위해 bytes로 전송
        server.sendmail(from_addr, all_rcpt, msg.as_bytes())

    try:
        ok, note = _smtp_run(cfg, timeout=45, send_fn=_do_send)
        return True, f"발송 완료 → {', '.join(all_rcpt)}{note}"
    except smtplib.SMTPAuthenticationError:
        return (
            False,
            "SMTP 인증 실패 — Gmail은 앱 비밀번호(16자리)를 쓰세요. "
            "secrets의 smtp_password를 다시 붙여넣고(공백/특수공백 주의), "
            "https://myaccount.google.com/apppasswords",
        )
    except Exception as e:
        err = str(e)
        if "CERTIFICATE_VERIFY_FAILED" in err or "certificate verify failed" in err.lower():
            return (
                False,
                "발송 실패: SSL 인증서 검증 오류입니다. "
                "Mac은 Python 인증서 설치(Install Certificates) 후 재시도하거나, "
                "secrets에 smtp_ssl_verify = \"0\" 을 넣고 다시 보내 보세요. "
                f"({err})",
            )
        if "ascii" in err.lower() and "encode" in err.lower():
            return (
                False,
                "발송 실패: 문자 인코딩 오류입니다. "
                "메일 제목/본문 또는 secrets(비밀번호)에 특수 공백이 있는지 확인하고 "
                "앱 비밀번호를 다시 입력해 주세요. "
                f"({err})",
            )
        return False, f"발송 실패: {e}"


def append_sent_log(
    *,
    client: str,
    email: str,
    subject: str,
    ok: bool,
    mode: str = "single",
    staff: str = "",
    items: int = 0,
    msg: str = "",
) -> None:
    _ensure_dirs()
    row = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "date": date.today().isoformat(),
        "client": client,
        "email": email,
        "subject": subject,
        "ok": bool(ok),
        "mode": mode,
        "staff": staff,
        "items": int(items or 0),
        "msg": msg or "",
    }
    try:
        with open(PI_SENT_LOG, "a", encoding="utf-8") as f:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    except OSError:
        pass
    try:
        _load_sent_log_cached.clear()
    except Exception:
        pass


@st.cache_data(show_spinner=False)
def _load_sent_log_cached(path: str, mtime: float) -> pd.DataFrame:
    cols = ["ts", "date", "client", "email", "subject", "ok", "mode", "staff", "items", "msg"]
    rows = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except Exception:
                    continue
    except OSError:
        return pd.DataFrame(columns=cols)
    if not rows:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(rows)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    if "date" not in df.columns or df["date"].isna().all():
        df["date"] = df["ts"].astype(str).str[:10]
    if "subject" not in df.columns:
        df["subject"] = ""
    return df[cols]


def load_sent_log() -> pd.DataFrame:
    cols = ["ts", "date", "client", "email", "subject", "ok", "mode", "staff", "items", "msg"]
    if not os.path.isfile(PI_SENT_LOG):
        return pd.DataFrame(columns=cols)
    try:
        mtime = float(os.path.getmtime(PI_SENT_LOG))
    except OSError:
        mtime = 0.0
    return _load_sent_log_cached(PI_SENT_LOG, mtime)


def last_sent_for_client(client: str, log_df: Optional[pd.DataFrame] = None) -> dict:
    """성공 발송만, 가장 최근 1건."""
    if not client:
        return {}
    df = log_df if log_df is not None else load_sent_log()
    if df is None or df.empty:
        return {}
    sub = df[df["client"].astype(str).str.strip() == str(client).strip()]
    if "ok" in sub.columns:
        sub = sub[sub["ok"].astype(str).isin(["True", "true", "1"]) | (sub["ok"] == True)]  # noqa: E712
    if sub.empty:
        return {}
    sub = sub.sort_values("ts")
    last = sub.iloc[-1].to_dict()
    return {
        "date": str(last.get("date") or str(last.get("ts") or "")[:10]),
        "subject": str(last.get("subject") or ""),
        "email": str(last.get("email") or ""),
        "ts": str(last.get("ts") or ""),
    }


def sent_summary_by_client(log_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    df = log_df if log_df is not None else load_sent_log()
    if df is None or df.empty:
        return pd.DataFrame(columns=["거래처", "최근발송일", "최근제목", "발송횟수"])
    ok = df[df["ok"].astype(str).isin(["True", "true", "1"]) | (df["ok"] == True)].copy()  # noqa: E712
    if ok.empty:
        return pd.DataFrame(columns=["거래처", "최근발송일", "최근제목", "발송횟수"])
    ok = ok.sort_values("ts")
    rows = []
    for client, g in ok.groupby(ok["client"].astype(str).str.strip()):
        last = g.iloc[-1]
        rows.append(
            {
                "거래처": client,
                "최근발송일": str(last.get("date") or str(last.get("ts") or "")[:10]),
                "최근제목": str(last.get("subject") or ""),
                "발송횟수": int(len(g)),
            }
        )
    return pd.DataFrame(rows).sort_values("최근발송일", ascending=False).reset_index(drop=True)


def _items_key(client: str) -> str:
    return f"pi_items_{_norm_name(client) or 'none'}"


def _init_items_from_prices(client: str, price_df: pd.DataFrame, pct: float = 0.0) -> list[dict]:
    """거래처 선택 시: 거래 품목명 + 마지막 거래 단가. 인상단가는 일단 동일(계산 UI에서 적용)."""
    del pct  # 공문·초기표에 %를 자동 반영하지 않음
    items = []
    for _, row in price_df.iterrows():
        old = float(row.get("기존단가") or 0)
        items.append(
            {
                "선택": True,
                "품목명": str(row.get("품목명") or ""),
                "기존단가": old,
                "인상적용단가": old,
                "비고": "",
                "최근매출일": str(row.get("최근매출일") or ""),
            }
        )
    return items


def _normalize_items_selection(items: list[dict]) -> list[dict]:
    """기존 session 품목에 선택 플래그가 없으면 기본 True."""
    out: list[dict] = []
    for it in items or []:
        row = dict(it)
        row["선택"] = bool(row.get("선택", True))
        out.append(row)
    return out


def _selected_items(items: list[dict]) -> list[dict]:
    """공문·요율 적용 대상: 선택 체크된 품목만."""
    return [it for it in (items or []) if bool(it.get("선택", True))]


def _items_df_for_editor(items: list[dict]) -> pd.DataFrame:
    """편집용 표: 선택 · 제품명 · 기존단가 · 인상단가 · 비고."""
    rows = []
    for it in _normalize_items_selection(items):
        old = float(it.get("기존단가") or 0)
        new = float(it.get("인상적용단가") or 0)
        rows.append(
            {
                "선택": bool(it.get("선택", True)),
                "제품명": str(it.get("품목명") or ""),
                "기존단가": old,
                "인상단가": new,
                "비고": str(it.get("비고") or ""),
            }
        )
    if not rows:
        return pd.DataFrame(columns=["선택", "제품명", "기존단가", "인상단가", "비고"])
    return pd.DataFrame(rows)


def _items_from_editor_df(df: pd.DataFrame) -> list[dict]:
    if df is None or df.empty:
        return []
    out: list[dict] = []
    for _, row in df.iterrows():
        name = ""
        for k in ("제품명", "기존거래제품명", "품목명"):
            if k in row.index and str(row.get(k) or "").strip():
                name = str(row.get(k) or "").strip()
                break
        if not name:
            continue
        if row.get("삭제") is True:
            continue
        if "기존단가" in row.index:
            old = row.get("기존단가")
        else:
            old = row.get("단가")
        new = row.get("인상단가") if "인상단가" in row.index else row.get("인상적용단가")
        selected = True
        if "선택" in row.index:
            selected = bool(row.get("선택"))
        out.append(
            {
                "선택": selected,
                "품목명": name,
                "기존단가": float(old or 0),
                "인상적용단가": float(new or 0),
                "비고": str(row.get("비고") or ""),
                "최근매출일": str(row.get("최근매출일") or ""),
            }
        )
    return out


def _editor_widget_key(editor_key: str) -> str:
    ver = int(st.session_state.get(f"{editor_key}_ver", 0))
    return f"{editor_key}_editor_v{ver}"


def _bump_editor_widget(editor_key: str) -> None:
    """data_editor 위젯 상태를 강제로 갱신 (일괄적용·삭제·재불러오기)."""
    vk = f"{editor_key}_ver"
    st.session_state[vk] = int(st.session_state.get(vk, 0)) + 1
    # 이전 위젯 키 잔여 상태 제거
    prev = int(st.session_state[vk]) - 1
    st.session_state.pop(f"{editor_key}_editor_v{prev}", None)
    st.session_state.pop(f"{editor_key}_editor", None)


def _apply_pct_to_items(items: list[dict], pct: float, *, only_selected: bool = True) -> list[dict]:
    out = []
    for it in _normalize_items_selection(items):
        row = dict(it)
        if only_selected and not bool(row.get("선택", True)):
            out.append(row)
            continue
        old = float(row.get("기존단가") or 0)
        row["인상적용단가"] = default_increase_price(old, pct)
        out.append(row)
    return out


def _apply_amount_to_items(items: list[dict], amount: float, *, only_selected: bool = True) -> list[dict]:
    out = []
    for it in _normalize_items_selection(items):
        row = dict(it)
        if only_selected and not bool(row.get("선택", True)):
            out.append(row)
            continue
        old = float(row.get("기존단가") or 0)
        row["인상적용단가"] = apply_increase_by_amount(old, amount)
        out.append(row)
    return out


def _default_letter_title() -> str:
    return "액체탄산 공급단가 인상 협조 요청의 건"


def _default_doc_no(client: str = "") -> str:
    return _format_doc_no(client, "")


def _default_mail_body(client: str, effective: str, items: list[dict]) -> str:
    lines = [
        f"{client} 귀중",
        "",
        "항상 저희 신일가스를 이용해 주셔서 감사드립니다.",
        "첨부와 같이 공급 단가 조정을 안내드립니다.",
        "",
        f"시행일: {_format_korean_effective(effective)}",
        "",
    ]
    for it in items:
        name = str(it.get("품목명") or "")
        old = float(it.get("기존단가") or 0)
        new = float(it.get("인상적용단가") or 0)
        lines.append(f"· {name}: {old:,.0f}원 ➠ {new:,.0f}원")
    lines.extend(["", "신일가스(주) 화성공장", "문의: 031-366-0799"])
    return "\n".join(lines)


def _build_letter_bytes(
    *,
    client: str,
    email: str,
    title: str,
    body: str,
    effective: str,
    contact: str,
    items: list[dict],
    doc_no: str = "",
    send_date: Optional[date] = None,
    letter_paras: Optional[dict[str, str]] = None,
    c37_override: str = "",
    c38_override: str = "",
) -> bytes:
    tpl = resolve_letter_template()
    return fill_letter_workbook(
        tpl,
        client=client,
        email=email,
        title=title,
        body=body,
        effective=effective,
        contact=contact,
        items=items,
        doc_no=doc_no,
        send_date=send_date,
        letter_paras=letter_paras,
        c37_override=c37_override,
        c38_override=c38_override,
    )


def _ensure_kr_font_ttf() -> str:
    """한글 TTF 경로. TTC면 첫 폰트를 캐시 폴더에 풀어 씀."""
    _ensure_dirs()
    os.makedirs(PI_FONTS_DIR, exist_ok=True)
    for path in _KR_FONT_CANDIDATES:
        if not path or not os.path.isfile(path):
            continue
        low = path.lower()
        if low.endswith(".ttf") or low.endswith(".otf"):
            return path
        if low.endswith(".ttc"):
            out = os.path.join(PI_FONTS_DIR, "kr-from-ttc.ttf")
            if os.path.isfile(out) and os.path.getsize(out) > 1000:
                return out
            try:
                from fontTools.ttLib import TTCollection

                ttc = TTCollection(path)
                ttc.fonts[0].save(out)
                if os.path.isfile(out) and os.path.getsize(out) > 1000:
                    return out
            except Exception:
                continue
    # 시스템/캐시 없을 때 NanumGothic 자동 다운로드
    dl = os.path.join(PI_FONTS_DIR, "NanumGothic.ttf")
    if os.path.isfile(dl) and os.path.getsize(dl) > 100_000:
        return dl
    urls = (
        "https://cdn.jsdelivr.net/gh/google/fonts@main/ofl/nanumgothic/NanumGothic-Regular.ttf",
        "https://cdn.jsdelivr.net/npm/@fontsource/nanum-gothic@5.0.0/files/nanum-gothic-korean-400-normal.woff",
        "https://github.com/google/fonts/raw/main/ofl/nanumgothic/NanumGothic-Regular.ttf",
    )
    try:
        import urllib.request

        for url in urls:
            # woff는 fpdf2 미지원 — ttf/otf만
            if url.lower().endswith(".woff") or url.lower().endswith(".woff2"):
                continue
            tmp = dl + ".part"
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=60) as resp:
                    data = resp.read()
                if not data or len(data) < 100_000:
                    continue
                with open(tmp, "wb") as f:
                    f.write(data)
                os.replace(tmp, dl)
                if os.path.isfile(dl) and os.path.getsize(dl) > 100_000:
                    return dl
            except Exception:
                try:
                    if os.path.isfile(tmp):
                        os.remove(tmp)
                except Exception:
                    pass
                continue
    except Exception:
        pass
    raise RuntimeError(
        "한글 PDF 폰트를 찾을 수 없습니다. "
        f"`{PI_FONTS_DIR}`에 NotoSansKR-Regular.ttf 또는 NanumGothic.ttf 를 두세요. "
        "또는 `python3 -m pip install -r requirements.txt` 후 앱을 재시작하세요."
    )


def _extract_template_images(template_path: str) -> dict[str, str]:
    """양식 xlsx 이미지 → PNG 캐시. header / footer / stamp."""
    out: dict[str, str] = {}
    os.makedirs(PI_FONTS_DIR, exist_ok=True)
    cached = {
        "header": os.path.join(PI_FONTS_DIR, "logo_header.png"),
        "footer": os.path.join(PI_FONTS_DIR, "logo_footer.png"),
        "stamp": os.path.join(PI_FONTS_DIR, "stamp.png"),
    }
    if all(os.path.isfile(p) for p in cached.values()):
        return cached
    if load_workbook is None or not os.path.isfile(template_path):
        return {k: v for k, v in cached.items() if os.path.isfile(v)}
    try:
        from PIL import Image as PILImage

        wb = load_workbook(template_path)
        ws = wb.active
        blobs: list[bytes] = []
        for img in getattr(ws, "_images", []) or []:
            try:
                blobs.append(img._data())
            except Exception:
                continue
        wb.close()
        # 크기 기준으로 분류: 가로 긴 로고 2개 + 정사각 도장
        decoded: list[tuple[int, int, bytes, str]] = []
        for raw in blobs:
            ext = "png"
            if raw[:6] in (b"GIF87a", b"GIF89a"):
                ext = "gif"
            elif raw[:2] == b"\xff\xd8":
                ext = "jpg"
            elif raw[:8] == b"\x89PNG\r\n\x1a\n":
                ext = "png"
            else:
                continue
            tmp = os.path.join(PI_FONTS_DIR, f"_tmp_img.{ext}")
            with open(tmp, "wb") as f:
                f.write(raw)
            im = PILImage.open(tmp).convert("RGBA")
            decoded.append((im.width, im.height, raw, ext))
            im.close()
            try:
                os.remove(tmp)
            except OSError:
                pass
        logos = sorted(
            [d for d in decoded if d[0] >= d[1] * 1.5],
            key=lambda x: -(x[0] * x[1]),
        )
        stamps = sorted(
            [d for d in decoded if d[0] < d[1] * 1.5],
            key=lambda x: -(x[0] * x[1]),
        )
        mapping = []
        if logos:
            mapping.append(("header", logos[0]))
        if len(logos) > 1:
            mapping.append(("footer", logos[1]))
        elif logos:
            mapping.append(("footer", logos[0]))
        if stamps:
            mapping.append(("stamp", stamps[0]))
        for key, (w, h, raw, ext) in mapping:
            path = cached[key]
            if ext == "png":
                with open(path, "wb") as f:
                    f.write(raw)
            else:
                tmp = os.path.join(PI_FONTS_DIR, f"_conv.{ext}")
                with open(tmp, "wb") as f:
                    f.write(raw)
                PILImage.open(tmp).convert("RGBA").save(path)
                try:
                    os.remove(tmp)
                except OSError:
                    pass
            out[key] = path
            del w, h
    except Exception:
        pass
    for k, v in cached.items():
        if k not in out and os.path.isfile(v):
            out[k] = v
    return out


def _pdf_wrap_lines(pdf, text: str, max_w: float) -> list[str]:
    """폭에 맞춰 줄바꿈 (공백·한글 혼합)."""
    t = str(text or "").replace("\t", " ")
    if not t.strip():
        return [""]
    out: list[str] = []
    for para in t.split("\n"):
        if not para:
            out.append("")
            continue
        line = ""
        for ch in para:
            trial = line + ch
            if pdf.get_string_width(trial) <= max_w:
                line = trial
            else:
                if line:
                    out.append(line)
                line = ch
        out.append(line)
    return out if out else [""]


def build_letter_pdf(
    *,
    client: str,
    title: str,
    effective: str,
    items: list[dict],
    doc_no: str = "",
    send_date: Optional[date] = None,
    letter_paras: Optional[dict[str, str]] = None,
    contact: str = "031-366-0799",
    email: str = "",
    body: str = "",
    **_extra: Any,
) -> bytes:
    """업체 전송용 공문 PDF (A4). 인상율/% 문구는 넣지 않음.

    단가표가 1장에 안 들어가면 본문에 「후면첨부」를 적고,
    2장에 동일 양식(헤더·하단 로고/대표자/도장/회사명)으로 단가표를 표시한다.
    """
    del email, body  # 메일본문·주소는 PDF 헤더에 불필요
    try:
        from fpdf import FPDF
    except Exception as e:  # pragma: no cover
        raise RuntimeError("fpdf2 패키지가 필요합니다. pip install fpdf2") from e

    font_path = _ensure_kr_font_ttf()
    tpl = resolve_letter_template()
    imgs = _extract_template_images(tpl)
    send_date = send_date or date.today()
    doc_no = _format_doc_no(client, doc_no)
    paras = letter_paras if letter_paras is not None else dict(_DEFAULT_LETTER_PARAS)
    send_s = send_date.strftime("%Y.%m.%d") if hasattr(send_date, "strftime") else str(send_date)
    clean = [it for it in items if str(it.get("품목명") or "").strip()]

    class LetterPDF(FPDF):
        def footer(self) -> None:  # noqa: N802
            pass

    pdf = LetterPDF(orientation="P", unit="mm", format="A4")
    # 페이지 넘김은 직접 제어 (푸터·양식 유지)
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    pdf.add_font("kr", "", font_path)
    bold_path = font_path
    for cand in (
        font_path.replace("Regular", "Bold"),
        os.path.join(PI_FONTS_DIR, "NotoSansKR-Bold.ttf"),
        os.path.join(PI_FONTS_DIR, "NanumGothicBold.ttf"),
    ):
        if cand != font_path and os.path.isfile(cand):
            bold_path = cand
            break
    try:
        pdf.add_font("kr", "B", bold_path)
    except Exception:
        pdf.add_font("kr", "B", font_path)

    left = 18.0
    usable = 210.0 - left - 18.0
    footer_top = 248.0  # 하단 로고·대표자 시작 Y (양면 공통)
    content_max = footer_top - 6.0

    def _rule_bar(y0: float, *, height: float = 1.6, gray: int = 200) -> float:
        pdf.set_fill_color(gray, gray, gray)
        pdf.rect(left, y0, usable, height, style="F")
        return y0 + height

    def _draw_letterhead(*, include_meta: bool = True) -> float:
        """상단 양식(로고·주소·문서정보). 1·2장 공통."""
        y = 8.0
        header = imgs.get("header")
        if header and os.path.isfile(header):
            try:
                logo_w = 72.0
                pdf.image(header, x=left + (usable - logo_w) / 2.0, y=y, w=logo_w)
                y += 26.0
            except Exception:
                y += 6.0
        else:
            y += 6.0

        y = _rule_bar(y, height=1.4, gray=205)
        pdf.set_xy(left, y + 1.2)
        pdf.set_font("kr", "", 8)
        pdf.set_text_color(30, 30, 30)
        addr = (
            "우 18524 경기도 화성시 팔탄면 서해로 1327-17"
            f"    /    전화 {contact or '031-366-0799'}    /    FAX 031-366-5633"
        )
        pdf.multi_cell(usable, 4.0, addr, align="C")
        y = pdf.get_y() + 1.2
        y = _rule_bar(y, height=1.4, gray=205)
        y += 3.5

        if include_meta:
            pdf.set_xy(left, y)
            pdf.set_font("kr", "", 12)
            for line in (
                f"문서번호 : {doc_no}",
                f"발송일자 : {send_s}",
                f"수    신 : {client}",
                f"제    목 : {title}",
            ):
                pdf.set_x(left)
                pdf.cell(usable, 7.0, line, new_x="LMARGIN", new_y="NEXT")
            y = pdf.get_y() + 2.0
            y = _rule_bar(y, height=2.0, gray=190)
            y += 4.0
        pdf.set_y(y)
        return y

    def _draw_page_footer() -> None:
        """하단 큰 로고·대표자+도장·회사명 — 매 페이지."""
        footer_y = footer_top
        footer_logo = imgs.get("footer") or imgs.get("header")
        stamp = imgs.get("stamp")
        if footer_logo and os.path.isfile(footer_logo):
            try:
                pdf.image(footer_logo, x=left, y=footer_y, w=78.0)
            except Exception:
                pass

        rep = "대표자 : 유 봉 래"
        pdf.set_font("kr", "B", 12)
        pdf.set_text_color(30, 30, 30)
        rep_w = pdf.get_string_width(rep)
        rep_x = left + usable - max(rep_w + 4, 42)
        rep_y = footer_y + 8.0
        pdf.set_xy(rep_x, rep_y)
        pdf.cell(rep_w + 2, 7, rep, align="L")
        if stamp and os.path.isfile(stamp):
            try:
                stamp_x = rep_x + max(rep_w - 18, 22)
                pdf.image(stamp, x=stamp_x, y=rep_y - 6.0, w=22.0)
            except Exception:
                pass

        line_y = footer_y + 26.0
        pdf.set_draw_color(180, 180, 180)
        pdf.set_line_width(0.45)
        pdf.line(left, line_y, left + usable, line_y)
        pdf.set_xy(left, line_y + 3.0)
        pdf.set_font("kr", "B", 12)
        pdf.cell(usable, 7, "(주) 신 일 가 스", align="C")

    def _draw_price_table(start_y: Optional[float] = None) -> float:
        """단가 조정 표 + 시행일. 반환: 끝난 Y."""
        if start_y is not None:
            pdf.set_y(start_y)
        col_w = [usable * 0.34, usable * 0.20, usable * 0.20, usable * 0.26]
        headers = ["제품명", "기존단가", "인상단가", "비고"]
        row_h = 7.0

        pdf.set_font("kr", "B", 11)
        pdf.set_text_color(15, 23, 42)
        pdf.set_x(left)
        pdf.cell(usable, 7, "단가 조정 내용", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1.5)

        pdf.set_font("kr", "B", 9)
        pdf.set_fill_color(15, 118, 110)
        pdf.set_text_color(255, 255, 255)
        pdf.set_x(left)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], row_h, h, border=1, fill=True, align="C")
        pdf.ln(row_h)
        pdf.set_text_color(15, 23, 42)
        pdf.set_font("kr", "", 9)

        for it in clean:
            # 2장에서도 푸터를 침범하지 않도록 페이지 넘김(양식 유지)
            if pdf.get_y() + row_h > content_max:
                _draw_page_footer()
                pdf.add_page()
                _draw_letterhead(include_meta=True)
                pdf.set_font("kr", "B", 11)
                pdf.set_x(left)
                pdf.cell(usable, 7, "단가 조정 내용 (계속)", align="C", new_x="LMARGIN", new_y="NEXT")
                pdf.ln(1.0)
                pdf.set_font("kr", "B", 9)
                pdf.set_fill_color(15, 118, 110)
                pdf.set_text_color(255, 255, 255)
                pdf.set_x(left)
                for i, h in enumerate(headers):
                    pdf.cell(col_w[i], row_h, h, border=1, fill=True, align="C")
                pdf.ln(row_h)
                pdf.set_text_color(15, 23, 42)
                pdf.set_font("kr", "", 9)

            vals = [
                str(it.get("품목명") or ""),
                f"{float(it.get('기존단가') or 0):,.0f}",
                f"{float(it.get('인상적용단가') or 0):,.0f}",
                str(it.get("비고") or ""),
            ]
            aligns = ["L", "C", "C", "L"]
            pdf.set_x(left)
            for i, v in enumerate(vals):
                while v and pdf.get_string_width(v) > col_w[i] - 2:
                    v = v[:-1]
                pdf.cell(col_w[i], row_h, v, border=1, align=aligns[i])
            pdf.ln(row_h)

        pdf.ln(2)
        pdf.set_font("kr", "", 10)
        pdf.set_x(left)
        pdf.cell(
            usable,
            6,
            f"                           • 시행 일자 : {_format_korean_effective(effective)}",
            new_x="LMARGIN",
            new_y="NEXT",
        )
        return float(pdf.get_y())

    def _price_block_height() -> float:
        """단가표 블록 예상 높이(mm)."""
        n = max(len(clean), 1)
        return 7.0 + 1.5 + 7.0 + 7.0 * n + 2.0 + 6.0 + 2.0

    # ── 1장: 헤더 + 본문 ──
    _draw_letterhead(include_meta=True)
    pdf.set_font("kr", "", 10)
    pdf.set_text_color(15, 23, 42)
    body_order = list(_BODY_CELL_ORDER)
    for coord in body_order:
        text = str(paras.get(coord) or "")
        if not text.strip():
            pdf.set_y(pdf.get_y() + 1.5)
            continue
        lines = _pdf_wrap_lines(pdf, text, usable)
        for ln in lines:
            if pdf.get_y() + 5.0 > content_max:
                break
            pdf.set_x(left)
            pdf.cell(usable, 5.0, ln, new_x="LMARGIN", new_y="NEXT")
        if coord in ("C16", "C19", "C24", "C26", "C27", "C29", "C31"):
            pdf.ln(1.2)

    attach_to_back = False
    if clean:
        pdf.ln(3)
        need = _price_block_height()
        if pdf.get_y() + need > content_max:
            attach_to_back = True
            pdf.set_font("kr", "B", 12)
            pdf.set_x(left)
            pdf.cell(usable, 8, "※ 단가 조정 내용 : 후면첨부", align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
            pdf.set_font("kr", "", 10)
            pdf.set_x(left)
            pdf.cell(
                usable,
                6,
                f"                           • 시행 일자 : {_format_korean_effective(effective)}",
                new_x="LMARGIN",
                new_y="NEXT",
            )
        else:
            _draw_price_table()

    _draw_page_footer()

    # ── 2장(+): 동일 양식 + 단가표 ──
    if attach_to_back and clean:
        pdf.add_page()
        _draw_letterhead(include_meta=True)
        pdf.set_font("kr", "", 10)
        pdf.set_x(left)
        pdf.cell(usable, 6, "【후면첨부】 단가 조정 내용", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        _draw_price_table()
        _draw_page_footer()

    bio = io.BytesIO()
    pdf.output(bio)
    return bio.getvalue()


def _build_letter_pdf_bytes(**kwargs: Any) -> bytes:
    """session kwargs → 업체 전송용 PDF."""
    return build_letter_pdf(**kwargs)


def _show_pdf_preview(pdf_bytes: bytes, *, height: int = 640, key: str = "pi_pdf_view") -> None:
    """st.pdf 우선, 실패 시 HTML 양식 + 다운로드로 폴백."""
    if not pdf_bytes:
        st.warning("PDF가 비어 있습니다.")
        return
    try:
        st.pdf(pdf_bytes, height=height, key=key)
        return
    except Exception as e:
        st.caption(f"PDF 뷰어 표시 실패 → 아래 양식·다운로드로 확인 ({e})")


@st.dialog("단가인상공문 미리보기 (업체 전송 양식)", width="large")
def _pi_letter_preview_dialog() -> None:
    """업무일지처럼 큰 팝업으로 업체 전송용 PDF를 보여 줌."""
    pdf_bytes = st.session_state.get("pi_pdf_bytes")
    pdf_name = st.session_state.get("pi_pdf_name") or "단가인상공문.pdf"
    xlsx = st.session_state.get("pi_dl_bytes")
    xlsx_name = st.session_state.get("pi_dl_name") or "단가인상공문.xlsx"
    meta = st.session_state.get("pi_preview_meta") or {}

    pdf_error = str(st.session_state.get("pi_pdf_error") or "")
    st.caption("메일 첨부와 동일한 PDF입니다. 인상율·인상금액은 포함되지 않습니다.")
    if pdf_bytes:
        _show_pdf_preview(pdf_bytes, height=720, key="pi_dialog_pdf")
    else:
        if pdf_error:
            st.error(pdf_error)
        else:
            st.error("PDF를 만들 수 없습니다. 품목·공문 입력을 확인하세요.")

    # PDF 뷰어가 막히는 환경 대비 — 동일 내용 HTML 양식
    if meta:
        with st.expander("화면 양식(보조)", expanded=not bool(pdf_bytes)):
            _render_letter_form_preview(
                client=str(meta.get("client") or ""),
                title=str(meta.get("title") or ""),
                doc_no=str(meta.get("doc_no") or ""),
                send_date=meta.get("send_date") or date.today(),
                effective_s=str(meta.get("effective") or ""),
                letter_body=str(meta.get("letter_body") or ""),
                items=list(meta.get("items") or []),
            )

    d1, d2 = st.columns(2)
    with d1:
        if pdf_bytes:
            st.download_button(
                "📥 공문 PDF 다운로드",
                data=pdf_bytes,
                file_name=pdf_name,
                mime="application/pdf",
                use_container_width=True,
                key="pi_dialog_dl_pdf",
            )
    with d2:
        if xlsx:
            st.download_button(
                "📥 엑셀(참고) 다운로드",
                data=xlsx,
                file_name=xlsx_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="pi_dialog_dl_xlsx",
            )


def _prepare_letter_preview(
    *,
    letter_kwargs: dict,
    letter_body: str,
    pdf_name: str,
    xlsx_name: str,
) -> None:
    """PDF·엑셀 생성 후 인라인 미리보기용 session 상태만 채움 (팝업 없음)."""
    st.session_state.pop("pi_pdf_error", None)
    try:
        pdf_bytes = _build_letter_pdf_bytes(**letter_kwargs)
        st.session_state["pi_pdf_bytes"] = pdf_bytes
        st.session_state["pi_pdf_name"] = pdf_name
    except Exception as e:
        st.session_state.pop("pi_pdf_bytes", None)
        st.session_state["pi_pdf_error"] = (
            "PDF 생성 라이브러리(fpdf2)가 현재 서버에 없어 PDF를 만들지 못했습니다. "
            "requirements 반영 후 재배포하거나, 로컬에서 `pip install -r requirements.txt`를 실행하세요. "
            f"(원인: {e})"
        )
    try:
        xlsx = _build_letter_bytes(**letter_kwargs)
        st.session_state["pi_dl_bytes"] = xlsx
        st.session_state["pi_dl_name"] = xlsx_name
    except Exception:
        st.session_state.pop("pi_dl_bytes", None)
    st.session_state["pi_show_dl"] = True
    st.session_state["pi_left_mode"] = "pdf"
    st.session_state["pi_preview_meta"] = {
        "client": letter_kwargs.get("client") or "",
        "title": letter_kwargs.get("title") or "",
        "doc_no": letter_kwargs.get("doc_no") or "",
        "send_date": letter_kwargs.get("send_date") or date.today(),
        "effective": letter_kwargs.get("effective") or "",
        "letter_body": letter_body,
        "items": list(letter_kwargs.get("items") or []),
    }
    st.session_state.pop("pi_preview_force", None)


def _open_letter_preview_dialog(
    *,
    letter_kwargs: dict,
    letter_body: str,
    pdf_name: str,
    xlsx_name: str,
) -> None:
    """PDF·엑셀 생성 후 dialog 팝업 오픈 (크게 보기용)."""
    _prepare_letter_preview(
        letter_kwargs=letter_kwargs,
        letter_body=letter_body,
        pdf_name=pdf_name,
        xlsx_name=xlsx_name,
    )
    _pi_letter_preview_dialog()


@st.dialog("메일 작성 · 최종 발송", width="large")
def _pi_mail_compose_dialog() -> None:
    """메일본문 확인·수정 후 최종 발송. 바로 보내지 않음."""
    draft = st.session_state.get("pi_mail_draft") or {}
    to_addr = str(draft.get("to") or "")
    pdf_name = str(draft.get("pdf_name") or "공문.pdf")
    pdf_bytes = st.session_state.get("pi_pdf_bytes")
    client = str(draft.get("client") or "")
    staff = str(draft.get("staff") or "")
    items_n = int(draft.get("items") or 0)

    st.caption("바로 발송되지 않습니다. 본문을 확인·수정한 뒤 「최종 발송」을 누르세요.")
    st.text_input("수신", value=to_addr, disabled=True, key="pi_mail_compose_to_view")
    if "pi_mail_compose_subject" not in st.session_state:
        st.session_state["pi_mail_compose_subject"] = str(draft.get("subject") or "")
    if "pi_mail_compose_body" not in st.session_state:
        st.session_state["pi_mail_compose_body"] = str(draft.get("body") or "")
    st.text_input("메일 제목", key="pi_mail_compose_subject")
    st.text_area("메일 본문", height=280, key="pi_mail_compose_body")
    st.caption(f"첨부 PDF: `{pdf_name}`" + (" · 준비됨" if pdf_bytes else " · 없음(발송 불가)"))

    c1, c2 = st.columns(2)
    with c1:
        if st.button("닫기", use_container_width=True, key="pi_mail_compose_close"):
            st.session_state.pop("pi_mail_draft", None)
            _pi_rerun(full=True)
    with c2:
        do_final = st.button(
            "최종 발송",
            type="primary",
            use_container_width=True,
            key="pi_mail_compose_send",
            disabled=not (to_addr and pdf_bytes),
        )
    if do_final:
        subject = str(st.session_state.get("pi_mail_compose_subject") or "").strip()
        body = str(st.session_state.get("pi_mail_compose_body") or "")
        if not subject:
            st.error("메일 제목을 입력하세요.")
            return
        try:
            ok, msg = send_mail_smtp(
                to_addr=to_addr,
                subject=subject,
                body=body,
                attachment_bytes=pdf_bytes,
                attachment_name=pdf_name,
            )
            append_sent_log(
                client=client,
                email=to_addr,
                subject=subject,
                ok=ok,
                mode="single",
                staff=staff,
                items=items_n,
                msg=msg,
            )
            if ok:
                st.session_state.pop("pi_mail_draft", None)
                st.success(msg)
            else:
                st.error(msg)
        except Exception as e:
            em = str(e)
            if "fpdf2" in em.lower():
                st.error(
                    "발송 실패: PDF 생성 라이브러리(fpdf2)가 없습니다. "
                    "`python3 -m pip install -r requirements.txt` 후 재시도하세요."
                )
            else:
                st.error(f"발송 실패: {e}")


def _open_mail_compose_dialog(
    *,
    client: str,
    email: str,
    title: str,
    body: str,
    pdf_name: str,
    letter_kwargs: dict,
    staff: str = "",
    items_n: int = 0,
) -> None:
    """PDF 준비 후 메일 작성 팝업 오픈 (즉시 발송 안 함)."""
    try:
        pdf_bytes = _build_letter_pdf_bytes(**letter_kwargs)
        st.session_state["pi_pdf_bytes"] = pdf_bytes
        st.session_state["pi_pdf_name"] = pdf_name
        st.session_state.pop("pi_pdf_error", None)
    except Exception as e:
        st.session_state.pop("pi_pdf_bytes", None)
        st.session_state["pi_pdf_error"] = str(e)
        st.error(f"첨부 PDF 생성 실패: {e}")
        return
    st.session_state["pi_mail_draft"] = {
        "to": email,
        "subject": title,
        "body": body,
        "pdf_name": pdf_name,
        "client": client,
        "staff": staff,
        "items": items_n,
    }
    # 위젯 초기값 (키로 제어)
    st.session_state["pi_mail_compose_subject"] = title
    st.session_state["pi_mail_compose_body"] = body
    _pi_mail_compose_dialog()


def _render_items_table(
    *,
    editor_key: str,
    items: list[dict],
    pct_default: float,
) -> list[dict]:
    st.markdown("##### 단가 조정 내용")
    st.caption(
        "「선택」체크한 제품만 요율 적용·공문 단가표에 들어갑니다. "
        "거래처 선택 시 품목·최종단가 자동 반영. %/인상금액은 계산용(공문 미표시)."
    )

    mode = st.radio(
        "인상 적용 방식",
        ["퍼센테이지(%)", "인상금액(원)"],
        horizontal=True,
        key=f"{editor_key}_mode",
        help="선택한(체크) 제품에만 적용해 인상단가를 계산합니다. 공문에는 %/금액이 들어가지 않습니다.",
    )
    c_val, c_apply, c_reload = st.columns([1.4, 1.1, 1.2])
    with c_val:
        if mode.startswith("퍼센트"):
            apply_val = st.number_input(
                "적용 퍼센테이지(%)",
                min_value=0.0,
                max_value=100.0,
                value=float(pct_default),
                step=0.5,
                key=f"{editor_key}_pct",
            )
        else:
            apply_val = st.number_input(
                "인상금액(원)",
                min_value=0.0,
                max_value=1_000_000.0,
                value=float(st.session_state.get(f"{editor_key}_amt", 0.0)),
                step=100.0,
                key=f"{editor_key}_amt",
            )
    with c_apply:
        st.write("")
        if st.button("인상단가에 적용", key=f"{editor_key}_apply", use_container_width=True, type="primary"):
            base = _normalize_items_selection(st.session_state.get(editor_key, items))
            if mode.startswith("퍼센트"):
                st.session_state[editor_key] = _apply_pct_to_items(base, apply_val, only_selected=True)
            else:
                st.session_state[editor_key] = _apply_amount_to_items(base, apply_val, only_selected=True)
            _bump_editor_widget(editor_key)
            _pi_rerun()
    with c_reload:
        st.write("")
        if st.button("매출 최종단가 다시 불러오기", key=f"{editor_key}_reload", use_container_width=True):
            st.session_state.pop(editor_key, None)
            _bump_editor_widget(editor_key)
            _pi_rerun()

    if editor_key not in st.session_state:
        st.session_state[editor_key] = _normalize_items_selection(items)
    else:
        st.session_state[editor_key] = _normalize_items_selection(st.session_state[editor_key])

    cur_items = st.session_state[editor_key]
    edit_df = _items_df_for_editor(cur_items)

    edited = st.data_editor(
        edit_df,
        column_config={
            "선택": st.column_config.CheckboxColumn(
                "선택",
                help="체크한 제품만 요율 적용 · 공문 단가표에 포함",
                default=True,
                width="small",
            ),
            "제품명": st.column_config.TextColumn("제품명", width="medium"),
            "기존단가": st.column_config.NumberColumn(
                "기존단가", min_value=0.0, format="%.1f", width="small"
            ),
            "인상단가": st.column_config.NumberColumn(
                "인상단가", min_value=0.0, format="%.1f", width="small"
            ),
            "비고": st.column_config.TextColumn("비고", width="small"),
        },
        column_order=["선택", "제품명", "기존단가", "인상단가", "비고"],
        num_rows="dynamic",
        hide_index=True,
        use_container_width=True,
        key=_editor_widget_key(editor_key),
    )

    b1, b2, b3 = st.columns(3)
    with b1:
        if st.button("빈 행 모두 제거", key=f"{editor_key}_drop_empty", use_container_width=True):
            cleaned = _items_from_editor_df(edited)
            st.session_state[editor_key] = cleaned
            _bump_editor_widget(editor_key)
            _pi_rerun()
    with b2:
        if st.button("표 초기화(전체 삭제)", key=f"{editor_key}_clear", use_container_width=True):
            st.session_state[editor_key] = []
            _bump_editor_widget(editor_key)
            _pi_rerun()
    with b3:
        if st.button("표 내용 저장", key=f"{editor_key}_save", type="primary", use_container_width=True):
            st.session_state[editor_key] = _items_from_editor_df(edited)
            st.success("품목표 저장됨 (거래처별 유지)")
            # session_state에 이미 반영됨 — 전체/fragment rerun 없이 확인 메시지만 표시

    result = _normalize_items_selection(_items_from_editor_df(edited))
    st.session_state[editor_key] = result
    chosen = _selected_items(result)
    if result:
        preview = " · ".join(
            f"{it['품목명']} {float(it['기존단가']):,.0f}➠{float(it['인상적용단가']):,.0f}"
            for it in chosen[:5]
        )
        if len(chosen) > 5:
            preview += f" 외 {len(chosen) - 5}건"
        st.caption(
            f"전체 {len(result)}개 · 공문 반영 {len(chosen)}개"
            + (f" — {preview}" if chosen else " — (선택 없음)")
        )
    return result


def _render_letter_content_editor(client: str, items: list[dict], effective_s: str) -> dict:
    """호환용 — 새 레이아웃에서는 사용하지 않음."""
    del client, items, effective_s
    return {
        "doc_no": "",
        "send_date": date.today(),
        "letter_paras": dict(_DEFAULT_LETTER_PARAS),
        "c37_override": "",
        "c38_override": "",
    }


def _render_pi_left_summary(
    *,
    client: str,
    email: str,
    title: str,
    effective_s: str,
    items: list[dict],
    last: Optional[dict],
    letter_body: str,
) -> None:
    """왼쪽: 요약 보기 (업무일지 요약 칸에 해당)."""
    st.markdown(
        f"""
<div style="border:1px solid #E2E8F0;border-radius:12px;padding:14px 16px;background:linear-gradient(180deg,#F8FAFC,#fff);">
  <div style="font-size:13px;color:#64748B;margin-bottom:6px;">요약</div>
  <div style="font-size:16px;font-weight:700;color:#0F172A;">{client}</div>
  <div style="margin-top:8px;font-size:13px;color:#334155;line-height:1.55;">
    <div>수신 메일: <b>{email or "(미입력)"}</b></div>
    <div>제목: {title}</div>
    <div>시행일: {_format_korean_effective(effective_s)}</div>
    <div>품목: {len(items)}건</div>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )
    if last:
        st.caption(f"최근 발송: {last.get('date')} · {last.get('subject') or '-'}")
    else:
        st.caption("최근 발송 이력 없음")
    if items:
        st.markdown("###### 단가 조정 표")
        st.dataframe(_items_df_for_editor(items), use_container_width=True, hide_index=True)
    else:
        st.caption("단가표 없음 · 공문 본문만 발송")
    with st.expander("공문 본문 미리보기", expanded=False):
        st.text(letter_body or "(본문 없음)")


def _render_letter_form_preview(
    *,
    client: str,
    title: str,
    doc_no: str,
    send_date: date,
    effective_s: str,
    letter_body: str,
    items: list[dict],
) -> None:
    """엑셀 양식 적용 결과를 왼쪽 칸에 바로 보여 줌."""
    rows_html = ""
    for it in items:
        name = str(it.get("품목명") or "")
        old = float(it.get("기존단가") or 0)
        new = float(it.get("인상적용단가") or 0)
        note = str(it.get("비고") or "")
        rows_html += (
            f"<tr><td>{name}</td><td style='text-align:right'>{old:,.0f}</td>"
            f"<td style='text-align:right'>{new:,.0f}</td><td>{note}</td></tr>"
        )
    if not rows_html:
        rows_html = "<tr><td colspan='4'>(품목 없음)</td></tr>"
    body_html = "<br/>".join(
        line.replace(" ", "&nbsp;") if line.strip() == "" else line
        for line in str(letter_body or "").splitlines()
    )
    send_s = send_date.strftime("%Y.%m.%d") if hasattr(send_date, "strftime") else str(send_date)
    price_block = ""
    if any(str(it.get("품목명") or "").strip() for it in items):
        price_block = f"""
  <div style="margin-top:14px;font-weight:700;text-align:center;">단가 조정 내용</div>
  <table style="width:100%;border-collapse:collapse;margin-top:8px;font-size:12px;">
    <thead>
      <tr style="background:#0F766E;color:#fff;">
        <th style="padding:6px;border:1px solid #0F766E;">제품명</th>
        <th style="padding:6px;border:1px solid #0F766E;">기존단가</th>
        <th style="padding:6px;border:1px solid #0F766E;">인상단가</th>
        <th style="padding:6px;border:1px solid #0F766E;">비고</th>
      </tr>
    </thead>
    <tbody>{rows_html}</tbody>
  </table>
  <div style="margin-top:10px;font-size:13px;">• 시행 일자 : {_format_korean_effective(effective_s)}</div>
"""
    st.markdown(
        f"""
<div style="border:1px solid #CBD5E1;border-radius:10px;padding:14px 16px;background:#fff;max-height:640px;overflow:auto;">
  <div style="text-align:center;font-weight:800;font-size:16px;color:#0F766E;margin-bottom:8px;">공문 미리보기</div>
  <div style="height:2px;background:#D1D5DB;margin:6px 0;"></div>
  <div style="text-align:center;font-size:11px;color:#475569;padding:4px 0;">
    우 18524 경기도 화성시 팔탄면 서해로 1327-17 &nbsp;/&nbsp; 전화 031-366-0799 &nbsp;/&nbsp; FAX 031-366-5633
  </div>
  <div style="height:2px;background:#D1D5DB;margin:6px 0 10px;"></div>
  <div style="font-size:14px;line-height:1.7;color:#0F172A;">
    <div>문서번호 : {doc_no}</div>
    <div>발송일자 : {send_s}</div>
    <div>수&nbsp;&nbsp;&nbsp;&nbsp;신 : {client}</div>
    <div>제&nbsp;&nbsp;&nbsp;&nbsp;목 : {title}</div>
  </div>
  <div style="height:2px;background:#C4C4C4;margin:12px 0;"></div>
  <div style="font-size:13px;line-height:1.55;color:#1E293B;white-space:normal;">{body_html}</div>
  {price_block}
</div>
""",
        unsafe_allow_html=True,
    )


def _collect_letter_kwargs(
    *,
    client: str,
    email: str,
    body_key: str,
    items_key: str,
) -> tuple[dict, str, list[dict], str]:
    """오른쪽 입력(session_state) 기준으로 공문 생성 인자 수집."""
    title = str(st.session_state.get("pi_single_title") or _default_letter_title())
    effective = st.session_state.get("pi_single_eff") or date.today().replace(day=1)
    effective_s = (
        effective.strftime("%Y-%m-%d") if hasattr(effective, "strftime") else str(effective)
    )
    contact = str(st.session_state.get("pi_single_contact") or "031-366-0799")
    doc_no = str(st.session_state.get("pi_letter_doc_no") or _default_doc_no(client))
    send_date = st.session_state.get("pi_letter_send_date") or date.today()
    if not isinstance(send_date, date):
        send_date = date.today()
    letter_body = str(st.session_state.get(body_key) or _default_letter_body_text())
    items_all = _normalize_items_selection(list(st.session_state.get(items_key) or []))
    items = _selected_items(items_all)
    letter_paras = _body_text_to_paras(letter_body)
    mail_body = _default_mail_body(client, effective_s, items)
    kwargs = dict(
        client=client,
        email=email,
        title=title,
        body=mail_body,
        effective=effective_s,
        contact=contact,
        items=items,
        doc_no=doc_no,
        send_date=send_date,
        letter_paras=letter_paras,
    )
    return kwargs, letter_body, items, effective_s


def _pi_on_client_change() -> None:
    """거래처 select 변경 시 수신메일 세션값 즉시 갱신."""
    client = str(st.session_state.get("pi_single_client") or "").strip()
    st.session_state["pi_email_client"] = client
    mail_df = st.session_state.get("_pi_mail_df_cache")
    hit = ""
    if client and isinstance(mail_df, pd.DataFrame) and not mail_df.empty:
        hit, matched = lookup_email_with_meta(client, mail_df)
        st.session_state["pi_email_matched_as"] = matched
    else:
        st.session_state["pi_email_matched_as"] = ""
    st.session_state["pi_single_email"] = hit


def _render_email_row(client: str, mail_df: pd.DataFrame) -> str:
    """거래처 선택 시 연락처에서 메일 자동반영(유사명 포함)."""
    st.session_state["_pi_mail_df_cache"] = mail_df
    auto_email, matched_as = lookup_email_with_meta(client, mail_df)
    try:
        mail_mtime = float(os.path.getmtime(PI_MAIL_CSV)) if os.path.isfile(PI_MAIL_CSV) else 0.0
    except OSError:
        mail_mtime = 0.0
    prev_client = st.session_state.get("pi_email_client")
    prev_mtime = st.session_state.get("pi_email_mail_mtime")
    cur_email = str(st.session_state.get("pi_single_email") or "").strip()
    client_changed = prev_client != client
    contacts_changed = prev_mtime != mail_mtime
    should_fill = False
    if client_changed:
        should_fill = True
    elif contacts_changed:
        should_fill = True
    elif auto_email and not cur_email:
        should_fill = True
    if should_fill:
        st.session_state["pi_email_client"] = client
        st.session_state["pi_email_mail_mtime"] = mail_mtime
        # 위젯 키 갱신 전 기존 값 제거 → Streamlit이 이전 빈칸을 붙잡는 문제 방지
        st.session_state.pop("pi_single_email", None)
        st.session_state["pi_single_email"] = auto_email
        st.session_state["pi_email_matched_as"] = matched_as
    email = st.text_input("수신 이메일", key="pi_single_email", placeholder="name@example.com")
    if auto_email:
        if matched_as and _norm_name(matched_as) != _norm_name(client):
            st.caption(f"연락처 자동반영: `{auto_email}` ← {matched_as}")
        else:
            st.caption(f"연락처 자동반영: `{auto_email}`")
    elif mail_df is None or mail_df.empty:
        st.warning(
            "연락처 CSV가 없어 자동반영할 수 없습니다. "
            "위 **📇 메일 연락처 관리**에서 CSV를 한 번 업로드하거나, "
            "아래에 이메일을 입력 후 「연락처에 저장」하세요. (다음부터 자동반영)"
        )
    else:
        st.caption("이름 불일치 시 「연락처에서 메일 고르기」에서 선택 · 또는 아래 저장")
    need_pick = (not bool(auto_email)) and mail_df is not None and not mail_df.empty
    with st.expander("연락처에서 메일 고르기", expanded=need_pick):
        cands = suggest_mail_matches(client, mail_df, limit=40)
        if cands:
            pick_labels = ["— 유사 연락처 선택 —"] + [f"{r['거래처']}  ·  {r['이메일']}" for r in cands]
            pick_map = {f"{r['거래처']}  ·  {r['이메일']}": str(r["이메일"]) for r in cands}
            chosen = st.selectbox("유사 연락처", pick_labels, key=f"pi_mail_pick_{_norm_name(client)}")
            if chosen in pick_map and st.session_state.get("pi_single_email") != pick_map[chosen]:
                st.session_state["pi_single_email"] = pick_map[chosen]
                _pi_rerun()
        all_labels = ["— 전체 연락처 선택 —"]
        all_map: dict[str, str] = {}
        if mail_df is not None and not mail_df.empty:
            for _, row in mail_df.sort_values("거래처").iterrows():
                nm = str(row.get("거래처") or "").strip()
                em = str(row.get("이메일") or "").strip()
                if nm and em:
                    lab = f"{nm}  ·  {em}"
                    all_labels.append(lab)
                    all_map[lab] = em
        if len(all_labels) > 1:
            chosen_all = st.selectbox("전체 연락처", all_labels, key=f"pi_mail_all_{_norm_name(client)}")
            if chosen_all in all_map and st.session_state.get("pi_single_email") != all_map[chosen_all]:
                st.session_state["pi_single_email"] = all_map[chosen_all]
                _pi_rerun()
        quick = st.text_input("직접 입력 후 이 거래처에 저장", key="pi_quick_email")
        if st.button("연락처에 저장", key="pi_quick_save"):
            q = str(quick or "").strip()
            if not q:
                q = str(st.session_state.get("pi_single_email") or "").strip()
            if q and "@" in q:
                add = pd.DataFrame([{"거래처": client, "이메일": q, "비고": ""}])
                out = pd.concat([mail_df, add], ignore_index=True) if mail_df is not None and not mail_df.empty else add
                out = out.drop_duplicates(subset=["거래처"], keep="last")
                save_mail_contacts(out)
                st.session_state["_pi_mail_df_cache"] = out
                st.session_state["pi_single_email"] = q
                st.session_state["pi_email_mail_mtime"] = (
                    float(os.path.getmtime(PI_MAIL_CSV)) if os.path.isfile(PI_MAIL_CSV) else 0.0
                )
                st.session_state.pop("_pi_mail_autoload_done", None)
                st.session_state["_pi_mail_saved_noted"] = False
                st.success(f"{client} → {q} 저장됨 (다음부터 자동반영)")
                _pi_rerun()
            else:
                st.error("올바른 이메일을 입력하세요.")
    return str(st.session_state.get("pi_single_email") or email or "")


def _render_mail_settings_expander(mail_df: pd.DataFrame) -> pd.DataFrame:
    with st.expander("📇 메일 연락처 관리", expanded=mail_df.empty):
        st.caption(
            f"CSV: `{PI_MAIL_CSV}` · 거래처명과 이메일을 등록해야 "
            "**수신 이메일이 자동 반영**됩니다. "
            "(저장본 자동 사용 · Desktop/다음 주소록 CSV도 자동 적재 · 수동 업로드 가능)"
        )
        up = st.file_uploader("연락처 CSV 업로드", type=["csv"], key="pi_mail_upload")
        if up is not None:
            raw_bytes = up.getvalue()
            file_id = f"{up.name}:{len(raw_bytes)}:{hash(raw_bytes)}"
            already = st.session_state.get("_pi_mail_upload_id") == file_id
            if not already:
                try:
                    raw = _read_contacts_csv_bytes(raw_bytes)
                    merged = _normalize_mail_df(raw)
                except Exception as e:
                    st.error(f"CSV 읽기 실패: {e}")
                    raw = pd.DataFrame()
                    merged = pd.DataFrame(columns=["거래처", "이메일", "비고"])
                if not merged.empty:
                    save_mail_contacts(merged)
                    st.session_state["_pi_mail_upload_id"] = file_id
                    st.session_state.pop("_pi_mail_autoload_done", None)
                    st.session_state["_pi_mail_saved_noted"] = False
                    st.session_state["pi_email_mail_mtime"] = (
                        float(os.path.getmtime(PI_MAIL_CSV)) if os.path.isfile(PI_MAIL_CSV) else 0.0
                    )
                    st.success(
                        f"`{up.name}` → {len(merged)}건 저장 "
                        "(회사명·이름 모두 매칭키로 등록)"
                    )
                    mail_df = load_mail_contacts()
                    # 현재 선택 거래처 메일 즉시 재반영
                    cur_client = str(st.session_state.get("pi_single_client") or "").strip()
                    if cur_client:
                        hit, matched = lookup_email_with_meta(cur_client, mail_df)
                        st.session_state["pi_email_client"] = cur_client
                        st.session_state.pop("pi_single_email", None)
                        st.session_state["pi_single_email"] = hit
                        st.session_state["pi_email_matched_as"] = matched
                else:
                    cols = list(raw.columns) if raw is not None and not raw.empty else []
                    st.error(
                        "이메일·이름/회사 열을 찾지 못했습니다. "
                        "다음 주소록은 「아웃룩 주소록(.csv)」으로 내보내기 하세요."
                    )
                    if cols:
                        st.caption("감지된 열: " + ", ".join(str(c) for c in cols[:20]))
            elif not mail_df.empty:
                st.caption(f"업로드 반영됨 · 등록 {len(mail_df)}건")
        if not mail_df.empty:
            st.dataframe(mail_df.head(200), use_container_width=True, hide_index=True)
            st.caption(f"등록 {len(mail_df)}건 · 거래처 선택 시 회사명·유사명 자동 매칭")
        else:
            st.info("연락처 CSV가 없습니다. 업로드하거나 직접 입력하세요.")
            with st.form("pi_mail_manual"):
                n = st.text_input("거래처")
                e = st.text_input("이메일")
                if st.form_submit_button("추가"):
                    if n and e:
                        add = pd.DataFrame([{"거래처": n, "이메일": e, "비고": ""}])
                        out = pd.concat([mail_df, add], ignore_index=True)
                        save_mail_contacts(out)
                        st.session_state.pop("_pi_mail_autoload_done", None)
                        _pi_rerun()
    return mail_df


def _pi_clear_widget_session_keys() -> None:
    """defer 탭 복원 시 button 등 충돌 키만 제거.

    주의: pi_mail_upload / pi_single_* / 연락처·본문 키는 지우면
    업로드·자동완성·선택값이 매 렌더마다 사라진다.
    """
    for k in list(st.session_state.keys()):
        if not isinstance(k, str):
            continue
        if k.startswith(
            (
                "pi_dialog_",
                "pi_mail_compose_",
                "pi_smtp_",
                "pi_pdf_preview_",
                "pi_pdf_help",
                "pi_body_reset",
                "pi_left_preview",
                "pi_left_send",
                "pi_left_summary",
                "pi_left_big",
            )
        ):
            st.session_state.pop(k, None)


def _render_smtp_bar() -> dict:
    cfg = smtp_settings()
    c1, c2, c3 = st.columns([2.5, 1.2, 1.2])
    with c1:
        st.caption(f"SMTP · {smtp_status_label(cfg)}")
    with c2:
        st.session_state.pop("pi_smtp_test", None)
        if st.button("SMTP 연결 테스트", use_container_width=True):
            ok, msg = test_smtp_connection(cfg)
            (st.success if ok else st.error)(msg)
    with c3:
        tpl = resolve_letter_template()
        st.caption(f"양식: `{os.path.basename(tpl)}`")

    # 로컬: secrets 없어도 화면에서 다음메일 계정 저장
    show_setup = (not cfg.get("ready")) or (not _pi_is_streamlit_cloud())
    if show_setup:
        with st.expander(
            "🔐 SMTP 계정 설정 (다음메일)",
            expanded=not bool(cfg.get("ready")),
        ):
            if _pi_is_streamlit_cloud():
                st.caption(
                    "Cloud는 Streamlit Secrets에 `smtp_user` / `smtp_password` 를 넣으세요."
                )
            else:
                st.caption(
                    "다음메일 → 설정 → **POP3/IMAP 사용 ON** 후, "
                    "아이디(전체메일)와 비밀번호를 저장하세요. "
                    f"저장 위치: `{PI_SMTP_LOCAL}` (이 Mac에만 보관)"
                )
            local = load_local_smtp()
            u0 = str(cfg.get("user") or local.get("smtp_user") or "")
            p0 = str(local.get("smtp_password") or "")
            with st.form("pi_local_smtp_form"):
                provider = st.selectbox(
                    "메일",
                    ["daum", "gmail", "naver"],
                    index=["daum", "gmail", "naver"].index(
                        str(local.get("smtp_provider") or cfg.get("provider") or "daum")
                    )
                    if str(local.get("smtp_provider") or cfg.get("provider") or "daum")
                    in ("daum", "gmail", "naver")
                    else 0,
                )
                user = st.text_input(
                    "메일 아이디",
                    value=u0,
                    placeholder="아이디@daum.net",
                )
                password = st.text_input(
                    "메일 비밀번호",
                    value=p0,
                    type="password",
                    placeholder="다음메일 비밀번호",
                )
                from_name = st.text_input(
                    "보내는 이름",
                    value=str(local.get("smtp_from_name") or cfg.get("from_name") or "신일가스"),
                )
                relax_ssl = st.checkbox(
                    "Mac SSL 오류 시 검증 완화 (CERTIFICATE_VERIFY_FAILED)",
                    value=str(local.get("smtp_ssl_verify") or "0") in ("0", "false", "off"),
                )
                saved = st.form_submit_button("저장 · 연동", use_container_width=True)
            if saved:
                if not user or not password:
                    st.error("아이디와 비밀번호를 모두 입력하세요.")
                else:
                    path = save_local_smtp(
                        user=user,
                        password=password,
                        provider=str(provider),
                        from_name=from_name,
                        ssl_verify="0" if relax_ssl else "1",
                    )
                    st.success(f"저장됨 → `{path}` · 연결 테스트로 확인하세요.")
                    _pi_rerun()
            if cfg.get("ready"):
                st.caption("이미 연동됨. 비밀번호를 바꾸려면 다시 저장하세요.")
    return smtp_settings()


@st.fragment
def render_price_increase_tab(sales_df: pd.DataFrame, latest_update_str: str = "") -> None:
    """공문 탭 — 업무일지형 좌우 레이아웃 · 개별·일괄·이력.

    fragment: 저장·버튼 클릭 시 공문 탭만 갱신(다른 탭·사이드바 전체 로딩 생략).
    """
    _ensure_dirs()
    _pi_clear_widget_session_keys()
    # 매출 스냅샷 토큰 — 바뀌면 단가 캐시 무효화
    _n = 0 if sales_df is None else int(len(sales_df))
    st.session_state["pi_sales_cache_token"] = f"{latest_update_str}|{_n}"

    st.markdown(
        "<div class='sub-header dashboard-tab-panel-head'>📨 공문</div>",
        unsafe_allow_html=True,
    )
    cap = f"빌드 {PI_UI_BUILD}"
    if latest_update_str:
        cap += f" · 매출 {latest_update_str}"
    dev_caption(cap)

    mail_df, autoload_note = ensure_mail_contacts_autoload()
    st.session_state["_pi_mail_df_cache"] = mail_df
    smtp_cfg = _render_smtp_bar()
    if autoload_note:
        st.caption(autoload_note)
    mail_df = _render_mail_settings_expander(mail_df)
    st.session_state["_pi_mail_df_cache"] = mail_df

    tab_single, tab_bulk, tab_hist = st.tabs(["개별 발송", "일괄 발송", "발송 이력"])

    # ── 개별 발송 (업무일지형: 왼쪽 요약·미리보기·메일 / 오른쪽 공문·단가) ──
    with tab_single:
        staff_opts = ["전체"] + list_staff_options(sales_df)
        r1c1, r1c2, r1c3 = st.columns([1.2, 2, 1.2])
        with r1c1:
            staff = st.selectbox("담당자", staff_opts, key="pi_single_staff")
        clients = list_clients_for_staff(sales_df, staff)
        with r1c2:
            client = st.selectbox(
                "거래처",
                clients or [""],
                key="pi_single_client",
                on_change=_pi_on_client_change,
            )
        with r1c3:
            kind = classify_client_kind(client, sales_df) if client else ""
            st.caption(f"유형: **{kind}**" if kind else "")

        if not client:
            st.info("담당자·거래처를 선택하세요.")
        else:
            last = last_sent_for_client(client)
            email = _render_email_row(client, mail_df)

            price_df = latest_unit_prices(sales_df, client)
            pct_key = "pi_global_pct"
            if pct_key not in st.session_state:
                st.session_state[pct_key] = 5.0
            items_key = _items_key(client)
            if items_key not in st.session_state:
                st.session_state[items_key] = _init_items_from_prices(
                    client, price_df, st.session_state[pct_key]
                )
            st.session_state["pi_last_client"] = client

            # 공문 본문 키 (거래처별)
            body_key = f"pi_letter_body_{_norm_name(client)}"
            if body_key not in st.session_state:
                st.session_state[body_key] = _default_letter_body_text()
            if st.session_state.get("pi_body_client") != client:
                st.session_state["pi_body_client"] = client
                if body_key not in st.session_state:
                    st.session_state[body_key] = _default_letter_body_text()

            col_left, col_right = st.columns([1, 1], gap="medium")

            # 오른쪽 입력을 먼저 실행해 session_state를 갱신한 뒤 왼쪽 미리보기에 반영
            # (화면 배치는 columns 생성 순서대로 왼쪽|오른쪽 유지)
            with col_right:
                st.markdown("##### 공문 입력")
                m1, m2 = st.columns(2)
                with m1:
                    st.text_input("문서번호", value=_default_doc_no(client), key="pi_letter_doc_no")
                with m2:
                    st.date_input("발송일자", value=date.today(), key="pi_letter_send_date")
                m3, m4, m5 = st.columns([2, 1, 1])
                with m3:
                    st.text_input("제목", value=_default_letter_title(), key="pi_single_title")
                with m4:
                    st.date_input("시행일", value=date.today().replace(day=1), key="pi_single_eff")
                with m5:
                    st.text_input("문의", value="031-366-0799", key="pi_single_contact")

                st.markdown("**1. 공문내용**")
                st.caption("업무일지 입력처럼 한 칸에 작성 · 기본은 단가인상공문 양식 · 다른 공문도 여기서 수정 가능")
                bbar1, bbar2 = st.columns([1, 1])
                with bbar1:
                    if st.button("기본공문양식 적용", key="pi_body_reset", use_container_width=True):
                        st.session_state[body_key] = _default_letter_body_text()
                        _pi_rerun()
                with bbar2:
                    st.caption("미리보기·메일은 PDF(업체 전송 양식)")
                st.text_area(
                    "공문 본문",
                    height=300,
                    key=body_key,
                    label_visibility="collapsed",
                )

                include_price = st.checkbox(
                    "2. 단가적용 포함 (선택)",
                    value=bool(st.session_state.get("pi_include_price", True)),
                    key="pi_include_price",
                    help="체크 해제 시 단가표 없이 1. 공문내용만으로 미리보기·발송합니다.",
                )
                if include_price:
                    st.caption("단가표가 공문에 들어갑니다. 필요 없으면 위 체크를 해제하세요.")
                    items = _render_items_table(
                        editor_key=items_key,
                        items=st.session_state[items_key],
                        pct_default=st.session_state[pct_key],
                    )
                    st.session_state[pct_key] = float(
                        st.session_state.get(f"{items_key}_pct", st.session_state[pct_key])
                    )
                    if not items:
                        st.info("품목이 비어 있으면 단가표 없이 본문만 발송됩니다.")
                else:
                    items = list(st.session_state.get(items_key) or [])
                    st.caption("단가적용 생략 · 1. 공문내용만 사용합니다. (표 데이터는 유지됨)")

            with col_left:
                st.markdown("##### 미리보기")
                p1, p2, p3 = st.columns([1, 1, 1])
                with p1:
                    do_preview = st.button(
                        "엑셀 미리보기",
                        use_container_width=True,
                        key="pi_left_preview",
                        help="업체 전송용 공문 PDF를 이 화면에 표시합니다. 메일 첨부와 동일합니다.",
                    )
                with p2:
                    do_send = st.button(
                        "메일보내기",
                        type="primary",
                        use_container_width=True,
                        key="pi_left_send",
                        help="메일본문 작성 팝업을 연 뒤, 최종 발송합니다.",
                    )
                with p3:
                    if st.button("요약 보기", use_container_width=True, key="pi_left_summary_btn"):
                        st.session_state["pi_left_mode"] = "summary"
                        _pi_rerun()

                letter_kwargs, letter_body, items_now, effective_s = _collect_letter_kwargs(
                    client=client,
                    email=email,
                    body_key=body_key,
                    items_key=items_key,
                )
                # 선택 해제 시 단가표를 공문/메일에 넣지 않음
                if not st.session_state.get("pi_include_price", True):
                    letter_kwargs = dict(letter_kwargs)
                    letter_kwargs["items"] = []
                    items_now = []
                title = str(letter_kwargs.get("title") or "")
                day_tag = date.today().strftime("%Y%m%d")
                pdf_name = f"공문_{client}_{day_tag}.pdf"
                xlsx_name = f"공문_{client}_{day_tag}.xlsx"
                mode = st.session_state.get("pi_left_mode") or "summary"

                if do_preview:
                    _prepare_letter_preview(
                        letter_kwargs=letter_kwargs,
                        letter_body=letter_body,
                        pdf_name=pdf_name,
                        xlsx_name=xlsx_name,
                    )
                    mode = "pdf"
                    if not st.session_state.get("pi_pdf_bytes") and st.session_state.get("pi_pdf_error"):
                        st.warning("PDF 생성에 실패했습니다. 아래 안내를 확인하세요.")

                if mode in ("pdf", "excel"):
                    pdf_bytes = st.session_state.get("pi_pdf_bytes")
                    xlsx = st.session_state.get("pi_dl_bytes")
                    if not pdf_bytes:
                        try:
                            pdf_bytes = _build_letter_pdf_bytes(**letter_kwargs)
                            st.session_state["pi_pdf_bytes"] = pdf_bytes
                            st.session_state["pi_pdf_name"] = pdf_name
                            st.session_state.pop("pi_pdf_error", None)
                        except Exception as e:
                            st.session_state["pi_pdf_error"] = (
                                "PDF 생성 라이브러리(fpdf2)가 현재 서버에 없어 PDF를 만들지 못했습니다. "
                                "requirements 반영 후 재배포하거나, 로컬에서 `pip install -r requirements.txt`를 실행하세요. "
                                f"(원인: {e})"
                            )
                            pdf_bytes = None
                    if not xlsx:
                        try:
                            xlsx = _build_letter_bytes(**letter_kwargs)
                            st.session_state["pi_dl_bytes"] = xlsx
                            st.session_state["pi_dl_name"] = xlsx_name
                        except Exception:
                            xlsx = st.session_state.get("pi_dl_bytes")
                    if items_now:
                        st.caption("업체 전송 양식(PDF) · 단가표 포함")
                    else:
                        st.caption("업체 전송 양식(PDF) · 본문만(단가표 없음)")
                    # 미리보기 안 버튼: 크게 보기 / PDF / 엑셀
                    b_big, b_dl1, b_dl2 = st.columns([1, 1, 1])
                    with b_big:
                        if pdf_bytes:
                            if st.button("크게 보기", use_container_width=True, key="pi_left_big"):
                                try:
                                    _open_letter_preview_dialog(
                                        letter_kwargs=letter_kwargs,
                                        letter_body=letter_body,
                                        pdf_name=pdf_name,
                                        xlsx_name=xlsx_name,
                                    )
                                except Exception as e:
                                    st.error(f"크게 보기 실패: {e}")
                        elif st.session_state.get("pi_pdf_error"):
                            if st.button("설치 가이드 보기", use_container_width=True, key="pi_pdf_help"):
                                st.info("로컬: python3 -m pip install -r requirements.txt / Cloud: 재배포")
                    with b_dl1:
                        if pdf_bytes:
                            st.download_button(
                                "📥 PDF",
                                data=pdf_bytes,
                                file_name=pdf_name,
                                mime="application/pdf",
                                key="pi_single_dl_pdf",
                                use_container_width=True,
                            )
                    with b_dl2:
                        if xlsx:
                            st.download_button(
                                "📥 엑셀",
                                data=xlsx,
                                file_name=xlsx_name,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="pi_single_dl_xlsx",
                                use_container_width=True,
                            )
                    if pdf_bytes:
                        _show_pdf_preview(pdf_bytes, height=520, key="pi_pdf_preview_left")
                    elif st.session_state.get("pi_pdf_error"):
                        st.error(str(st.session_state.get("pi_pdf_error")))
                        if st.button("보조 양식 크게 보기", use_container_width=True, key="pi_left_big_fallback"):
                            _pi_letter_preview_dialog()
                else:
                    _render_pi_left_summary(
                        client=client,
                        email=email,
                        title=title,
                        effective_s=effective_s,
                        items=items_now,
                        last=last,
                        letter_body=letter_body,
                    )
                    if st.session_state.get("pi_pdf_bytes"):
                        st.download_button(
                            "📥 최근 생성 PDF 다운로드",
                            data=st.session_state["pi_pdf_bytes"],
                            file_name=st.session_state.get("pi_pdf_name") or pdf_name,
                            mime="application/pdf",
                            key="pi_single_dl_pdf_keep",
                            use_container_width=True,
                        )

                if do_send:
                    if not email:
                        st.error("수신 이메일을 입력하세요.")
                    elif not smtp_cfg.get("ready"):
                        st.error("SMTP secrets 미설정")
                    else:
                        mail_body = _default_mail_body(client, effective_s, items_now)
                        _open_mail_compose_dialog(
                            client=client,
                            email=email,
                            title=title,
                            body=mail_body,
                            pdf_name=pdf_name,
                            letter_kwargs=letter_kwargs,
                            staff=staff if staff != "전체" else "",
                            items_n=len(items_now),
                        )
    # ── 일괄 발송 (담당자 단위 · 거래처는 한 곳씩 개별 공문) ──
    with tab_bulk:
        st.caption(
            "담당자를 먼저 고른 뒤, 해당 담당자 거래처를 **기본 전체 선택**합니다. "
            "제외할 곳만 체크 해제하세요. 발송은 **한 거래처씩** 개별 공문으로 진행됩니다. "
            "(회사 전체 일괄 발송은 없습니다.)"
        )
        staff_list = list_staff_options(sales_df)
        if not staff_list:
            st.warning("담당자 목록이 없습니다. 매출 데이터를 확인하세요.")
            b_staff = ""
        else:
            b_staff = st.selectbox(
                "담당자 (필수)",
                staff_list,
                key="pi_bulk_staff",
                help="일괄은 담당자 단위입니다. 전체 거래처 일괄은 지원하지 않습니다.",
            )
        bulk_clients = list_clients_for_staff(sales_df, b_staff) if b_staff else []

        mode_bulk = st.radio(
            "공통 인상 적용 (표 계산용 · 공문에 %/금액 미표시)",
            ["퍼센테이지(%)", "인상금액(원)"],
            horizontal=True,
            key="pi_bulk_mode",
        )
        if mode_bulk.startswith("퍼센트"):
            pct_bulk = st.number_input(
                "공통 적용 퍼센테이지(%)",
                min_value=0.0,
                max_value=100.0,
                value=float(st.session_state.get("pi_global_pct", 5.0)),
                step=0.5,
                key="pi_bulk_pct",
            )
            amt_bulk = 0.0
            st.session_state["pi_global_pct"] = pct_bulk
        else:
            amt_bulk = st.number_input(
                "공통 인상금액(원)",
                min_value=0.0,
                max_value=1_000_000.0,
                value=0.0,
                step=100.0,
                key="pi_bulk_amt",
            )
            pct_bulk = 0.0

        title_bulk = st.text_input("공통 제목", value=_default_letter_title(), key="pi_bulk_title")
        eff_bulk = st.date_input("공통 시행일", value=date.today().replace(day=1), key="pi_bulk_eff")
        eff_bulk_s = eff_bulk.strftime("%Y-%m-%d") if hasattr(eff_bulk, "strftime") else str(eff_bulk)

        # 담당자 바뀌면 선택 위젯 리셋 → 다시 전체 선택
        prev_staff = st.session_state.get("pi_bulk_staff_prev")
        if prev_staff != b_staff:
            st.session_state["pi_bulk_staff_prev"] = b_staff
            st.session_state.pop("pi_bulk_select", None)
            ver = int(st.session_state.get("pi_bulk_sel_ver", 0)) + 1
            st.session_state["pi_bulk_sel_ver"] = ver

        rows = []
        for cl in bulk_clients:
            em = lookup_email(cl, mail_df)
            last = last_sent_for_client(cl)
            rows.append(
                {
                    "발송": True,  # 담당자 내 거래처 전체 선택 기본
                    "거래처": cl,
                    "이메일": em or "(미등록)",
                    "유형": classify_client_kind(cl, sales_df),
                    "최근발송일": last.get("date", "") if last else "",
                    "최근제목": last.get("subject", "") if last else "",
                }
            )
        bulk_df = pd.DataFrame(rows)
        if not b_staff:
            st.info("담당자를 선택하세요.")
        elif bulk_df.empty:
            st.info("해당 담당자 거래처가 없습니다.")
        else:
            sel1, sel2, sel3 = st.columns(3)
            with sel1:
                if st.button("담당자 거래처 전체 선택", key="pi_bulk_all_on", use_container_width=True):
                    st.session_state["pi_bulk_force_all"] = True
                    st.session_state["pi_bulk_force_none"] = False
                    st.session_state["pi_bulk_sel_ver"] = int(st.session_state.get("pi_bulk_sel_ver", 0)) + 1
                    st.session_state.pop("pi_bulk_select", None)
                    _pi_rerun()
            with sel2:
                if st.button("전체 제외", key="pi_bulk_all_off", use_container_width=True):
                    st.session_state["pi_bulk_force_all"] = False
                    st.session_state["pi_bulk_force_none"] = True
                    st.session_state["pi_bulk_sel_ver"] = int(st.session_state.get("pi_bulk_sel_ver", 0)) + 1
                    st.session_state.pop("pi_bulk_select", None)
                    _pi_rerun()
            with sel3:
                st.caption("체크 해제로 개별 제외")

            if st.session_state.get("pi_bulk_force_none"):
                bulk_df["발송"] = False
            elif st.session_state.get("pi_bulk_force_all", True):
                bulk_df["발송"] = True
            # force 플래그는 한 번 반영 후 유지 위해 데이터에만 적용; 위젯 키는 ver로 갱신

            edited_bulk = st.data_editor(
                bulk_df,
                column_config={
                    "발송": st.column_config.CheckboxColumn(
                        "발송",
                        help="기본 전체 선택. 제외할 거래처만 체크 해제",
                        default=True,
                    ),
                    "거래처": st.column_config.TextColumn("거래처", disabled=True),
                    "이메일": st.column_config.TextColumn("이메일", disabled=True),
                    "유형": st.column_config.TextColumn("유형", disabled=True),
                    "최근발송일": st.column_config.TextColumn("최근발송일", disabled=True),
                    "최근제목": st.column_config.TextColumn("최근제목", disabled=True),
                },
                hide_index=True,
                use_container_width=True,
                key=f"pi_bulk_select_v{int(st.session_state.get('pi_bulk_sel_ver', 0))}",
            )
            # 발송 체크된 곳 — 이메일은 실제 발송 시 검증 (한 거래처씩)
            checked = edited_bulk[edited_bulk["발송"] == True].copy()  # noqa: E712
            has_mail = checked["이메일"].astype(str).str.contains("@", na=False)
            targets = checked[has_mail]
            no_mail = checked[~has_mail]
            st.caption(
                f"담당자 **{b_staff}** · 선택 {len(checked)}/{len(bulk_df)}곳 "
                f"· 발송가능(메일) **{len(targets)}**곳 · 메일미등록 {len(no_mail)}곳"
            )
            if not no_mail.empty:
                st.warning(
                    "메일 미등록 거래처는 건너뜁니다: "
                    + ", ".join(no_mail["거래처"].astype(str).head(8).tolist())
                    + (" …" if len(no_mail) > 8 else "")
                )

            confirm = st.checkbox(
                f"선택 거래처를 **한 곳씩** 개별 공문으로 발송합니다 ({len(targets)}곳, 되돌릴 수 없음)",
                key="pi_bulk_confirm",
            )
            if st.button(
                "📧 담당자 일괄 발송 (거래처별 개별)",
                type="primary",
                key="pi_bulk_send",
                disabled=not confirm or targets.empty,
            ):
                if not smtp_cfg.get("ready"):
                    st.error("SMTP secrets 미설정")
                else:
                    prog = st.progress(0.0)
                    status = st.empty()
                    results: list[str] = []
                    n = len(targets)
                    for i, (_, row) in enumerate(targets.iterrows()):
                        cl = str(row["거래처"])
                        em = str(row["이메일"]).strip()
                        status.info(f"({i + 1}/{n}) **{cl}** 개별 공문 발송 중…")
                        pdf = latest_unit_prices(sales_df, cl)
                        items_b = _init_items_from_prices(cl, pdf, 0.0)
                        if mode_bulk.startswith("퍼센트"):
                            items_b = _apply_pct_to_items(items_b, pct_bulk)
                        else:
                            items_b = _apply_amount_to_items(items_b, amt_bulk)
                        body_b = _default_mail_body(cl, eff_bulk_s, items_b)
                        dl = f"단가인상_{cl}_{date.today().strftime('%Y%m%d')}.pdf"
                        pdf_b = _build_letter_pdf_bytes(
                            client=cl,
                            email=em,
                            title=title_bulk,
                            body=body_b,
                            effective=eff_bulk_s,
                            contact="031-366-0799",
                            items=items_b,
                            letter_paras=dict(_DEFAULT_LETTER_PARAS),
                        )
                        ok, msg = send_mail_smtp(
                            to_addr=em,
                            subject=title_bulk,
                            body=body_b,
                            attachment_bytes=pdf_b,
                            attachment_name=dl,
                        )
                        append_sent_log(
                            client=cl,
                            email=em,
                            subject=title_bulk,
                            ok=ok,
                            mode="bulk",
                            staff=b_staff,
                            items=len(items_b),
                            msg=msg,
                        )
                        results.append(f"{'✅' if ok else '❌'} {cl}: {msg}")
                        prog.progress((i + 1) / max(n, 1))
                    status.empty()
                    for line in results:
                        st.write(line)
                    st.success(f"담당자 [{b_staff}] 거래처별 개별 발송 완료 ({n}건)")

    # ── 발송 이력 ──
    with tab_hist:
        log_df = load_sent_log()
        st.caption("거래처별 최근 발송일·제목 및 전체 로그")
        summary = sent_summary_by_client(log_df)
        if summary.empty:
            st.info("발송 이력이 없습니다.")
        else:
            st.markdown("##### 거래처별 최근 발송")
            st.dataframe(summary, use_container_width=True, hide_index=True)

        if not log_df.empty:
            show = log_df.copy()
            show["ok"] = show["ok"].map(lambda x: "✅" if str(x) in ("True", "true", "1") or x is True else "❌")
            show = show.rename(
                columns={
                    "date": "발송일",
                    "client": "거래처",
                    "email": "이메일",
                    "subject": "제목",
                    "mode": "모드",
                    "staff": "담당자",
                    "items": "품목수",
                    "msg": "결과",
                    "ts": "시각",
                }
            )
            st.markdown("##### 전체 이력")
            st.dataframe(
                show[["발송일", "거래처", "제목", "이메일", "ok", "모드", "담당자", "품목수", "결과", "시각"]],
                use_container_width=True,
                hide_index=True,
            )

